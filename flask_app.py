from datetime import datetime, timedelta, date, time, timezone
from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_security import Security, SQLAlchemyUserDatastore, UserMixin, RoleMixin, roles_required, current_user, login_required, logout_user
from flask import jsonify
from flask_login import login_user, login_required, current_user


from pytz import timezone as pytz_timezone
from flask_babel import Babel
from flask_mail import Message
import schedule
import threading
from flask_wtf import FlaskForm
from wtforms import StringField, DateField, IntegerField, SubmitField
from wtforms.validators import DataRequired
from flask_mail import Mail

from wtforms import StringField, PasswordField
from flask_security import login_user
from flask_security.utils import verify_password
from wtforms import SelectField, TextAreaField, FileField
from flask_wtf.file import FileField, FileAllowed
from wtforms.validators import InputRequired
import os

from wtforms.validators import DataRequired, EqualTo, Length
from wtforms.validators import DataRequired, Email
from itsdangerous import URLSafeTimedSerializer as Serializer
from itsdangerous import  BadSignature
from werkzeug.utils import secure_filename
from werkzeug.utils import secure_filename
from io import BytesIO
import base64
from base64 import b64encode, b64decode
from flask_wtf.file import FileField, FileAllowed
from flask import send_file
import io
from wtforms import FloatField, StringField, SubmitField
from sqlalchemy import extract
from flask_migrate import Migrate
import json
import pytz
from sqlalchemy.sql import text
from flask import make_response, request
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from sqlalchemy.types import Enum
from wtforms import BooleanField, FloatField, StringField, SubmitField, TimeField
from flask_wtf import FlaskForm
from wtforms.validators import Optional
import random
import string
from flask import request, jsonify
from werkzeug.security import generate_password_hash
from flask import Response
from flask import render_template, jsonify
from flask_cors import CORS
from sqlalchemy import DateTime
# Resto de tu código


# Configurar la zona horaria para España
timezone_spain = pytz_timezone('Europe/Madrid')

# Obtener la hora actual directamente en España
hora_espana = datetime.now(timezone_spain)

# Formatear la hora en el estilo español
hora_espana_str = hora_espana.strftime("%d-%m-%y %H:%M:%S")

print("Hora actual en España:", hora_espana_str)
app = Flask(__name__)
CORS(app, origins=["https://demo2.gestionaempresa.es"], supports_credentials=True)
url_app="https://proyectohombre.gestionaempresa.es/"
babel = Babel(app)

basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, 'vacation2.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{db_path}"

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'secret'
app.config['SECURITY_PASSWORD_SALT'] = '12345'
app.config['MAIL_SERVER'] = 'smtp.strato.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'notificaciones@gestionaempresa.es'
app.config['MAIL_PASSWORD'] = 'NotiF24@gES%'
app.config['MAIL_DEFAULT_SENDER'] = 'notificaciones@gestionaempresa.es'
UPLOAD_FOLDER = os.path.join('static', 'uploads')
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER2 = os.path.join(BASE_DIR, 'static')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)  # Asegura que la carpeta exista
app.config['WTF_CSRF_ENABLED'] = False



mail = Mail(app)

db = SQLAlchemy(app)
migrate = Migrate(app, db)


non_workdays = [
    date(datetime.now().year, 1, 1),   # Año Nuevo
    date(datetime.now().year, 12, 25),  # Navidad
    date(datetime.now().year, 1, 6),  # Reyes
    date(datetime.now().year, 3, 28),  # Marzo
    date(datetime.now().year, 3, 29),  # Marzo

    date(datetime.now().year, 5, 1),  # Mayo
    date(datetime.now().year, 5, 30),  # Mayo
    date(datetime.now().year, 5, 31),  # Mayo
    date(datetime.now().year, 8, 15),  # Agosto
    date(datetime.now().year, 10, 12),  # Hispanidad
    date(datetime.now().year, 11, 1),  # todos los santos
    date(datetime.now().year, 12, 6),#Diciembre
    date(datetime.now().year, 9, 12),#local
    date(datetime.now().year, 9, 13), # local
    # Agrega más días festivos según sea necesario
]

non_workdays_2026 = [
    date(2026, 1, 1),   # Año Nuevo

    date(2026, 1, 6),  # Reyes

    date(2026, 4, 2),
    date(2026, 4, 3),
    date(2026, 4, 6),
    date(2026, 5, 1),
    date(2026, 6, 4),
    date(2026, 8, 15),
    date(2026, 9, 8),#guada
    date(2026, 9, 18),#guada
    date(2026, 10, 12),
    date(2026, 11, 2),
    date(2026, 12, 8),
    #date(2026, 12, 24),
    #date(2026, 12, 31),
    date(2026, 12, 25),



]

#weekly_non_workdays = [
    #5,  # Viernes
    #6,  # Sábado
    # Agrega más días según sea necesario
#]


# Define User and Role models
roles_users = db.Table('roles_users',
                       db.Column('user_id', db.Integer(), db.ForeignKey('user.id')),
                       db.Column('role_id', db.Integer(), db.ForeignKey('role.id')))





class Role(db.Model, RoleMixin):
    __tablename__ = 'role'
    id = db.Column(db.Integer(), primary_key=True)
    name = db.Column(db.String(80), unique=True)
    description = db.Column(db.String(255))

def verify_reset_token(token):
    # Configura el serializador con la misma clave secreta que se usó para generar el token
    s = Serializer(app.config['SECRET_KEY'])

    try:
        # Deserializa el token y obtiene el ID de usuario
        user_id = s.loads(token, max_age=3600)  # El token es válido durante 1 hora (3600 segundos)

        # Recupera el usuario desde la base de datos usando el ID
        user = User.query.get(user_id)
        return user
    except BadSignature:
        # El token es inválido o ha caducado
        return None

def generate_reset_token(user):
    s = Serializer(app.config['SECRET_KEY'])
    user_id_str = str(user.id)  # Convertir user.id a cadena (string)
    return s.dumps({'user_id': user_id_str})

class Absence(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    absence_type = db.Column(db.String(50), nullable=False)
    description = db.Column(db.Text, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # Clave foránea a la tabla User
    start_time = db.Column(db.Time, nullable=False)
    type = db.Column(db.String(20), nullable=True)
    duration_hours = db.Column(db.String(8), nullable=True)
    file_name = db.Column(db.String(255))  # Nombre del archivo
    file_path = db.Column(db.String(255))  # Ruta del archivo
    mime_type = db.Column(db.String(50))   # Tipo MIME

from wtforms import FloatField
from wtforms.validators import DataRequired, NumberRange

from wtforms import Form, BooleanField, TimeField, SelectField, SubmitField, FloatField
from wtforms.validators import Optional, NumberRange, ValidationError

class ConfigureUserForm(FlaskForm):
    # Campos de días restantes
    dias_restantes_este = FloatField('Días restantes de este año (2024):', validators=[Optional(), NumberRange(min=0)])
    dias_restantes_pasado = FloatField('Días restantes del año pasado (2023):', validators=[NumberRange(min=0)])

    # Campos de horarios para cada día de la semana
    trabaja_lunes = BooleanField('Trabaja lunes?')
    entrada_lunes = TimeField('Entrada lunes:', format='%H:%M', validators=[Optional()])
    salida_comida_lunes = TimeField('Salida comida lunes:', format='%H:%M', validators=[Optional()])
    entrada_comida_lunes = TimeField('Entrada comida lunes:', format='%H:%M', validators=[Optional()])
    salida_lunes = TimeField('Salida lunes:', format='%H:%M', validators=[Optional()])

    trabaja_martes = BooleanField('Trabaja martes?')
    entrada_martes = TimeField('Entrada martes:', format='%H:%M', validators=[Optional()])
    salida_comida_martes = TimeField('Salida comida martes:', format='%H:%M', validators=[Optional()])
    entrada_comida_martes = TimeField('Entrada comida martes:', format='%H:%M', validators=[Optional()])
    salida_martes = TimeField('Salida martes:', format='%H:%M', validators=[Optional()])

    trabaja_miercoles = BooleanField('Trabaja miércoles?')
    entrada_miercoles = TimeField('Entrada miércoles:', format='%H:%M', validators=[Optional()])
    salida_comida_miercoles = TimeField('Salida comida miércoles:', format='%H:%M', validators=[Optional()])
    entrada_comida_miercoles = TimeField('Entrada comida miércoles:', format='%H:%M', validators=[Optional()])
    salida_miercoles = TimeField('Salida miércoles:', format='%H:%M', validators=[Optional()])

    trabaja_jueves = BooleanField('Trabaja jueves?')
    entrada_jueves = TimeField('Entrada jueves:', format='%H:%M', validators=[Optional()])
    salida_comida_jueves = TimeField('Salida comida jueves:', format='%H:%M', validators=[Optional()])
    entrada_comida_jueves = TimeField('Entrada comida jueves:', format='%H:%M', validators=[Optional()])
    salida_jueves = TimeField('Salida jueves:', format='%H:%M', validators=[Optional()])

    trabaja_viernes = BooleanField('Trabaja viernes?')
    entrada_viernes = TimeField('Entrada viernes:', format='%H:%M', validators=[Optional()])
    salida_comida_viernes = TimeField('Salida comida viernes:', format='%H:%M', validators=[Optional()])
    entrada_comida_viernes = TimeField('Entrada comida viernes:', format='%H:%M', validators=[Optional()])
    salida_viernes = TimeField('Salida viernes:', format='%H:%M', validators=[Optional()])

    trabaja_sabado = BooleanField('Trabaja sábado?')
    entrada_sabado = TimeField('Entrada sábado:', format='%H:%M', validators=[Optional()])
    salida_comida_sabado = TimeField('Salida comida sábado:', format='%H:%M', validators=[Optional()])
    entrada_comida_sabado = TimeField('Entrada comida sábado:', format='%H:%M', validators=[Optional()])
    salida_sabado = TimeField('Salida sábado:', format='%H:%M', validators=[Optional()])

    trabaja_domingo = BooleanField('Trabaja domingo?')
    entrada_domingo = TimeField('Entrada domingo:', format='%H:%M', validators=[Optional()])
    salida_comida_domingo = TimeField('Salida comida domingo:', format='%H:%M', validators=[Optional()])
    entrada_comida_domingo = TimeField('Entrada comida domingo:', format='%H:%M', validators=[Optional()])
    salida_domingo = TimeField('Salida domingo:', format='%H:%M', validators=[Optional()])


    # Botón de submit
    submit = SubmitField('Save Configuration')

    def validate(self):
        """
        Validación personalizada que verifica si se han asignado días de vacaciones
        y si es así, no valida los campos de tiempo.
        """
        if self.dias_restantes_este.data > 0 or self.dias_restantes_pasado.data > 0:
            # Si hay días de vacaciones asignados, no validamos los campos de tiempo
            return True

        # Si no hay días de vacaciones asignados, realizamos la validación normal
        return super().validate()




from sqlalchemy import PickleType

from sqlalchemy.types import Enum

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(255), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    active = db.Column(db.Boolean(), default=True)

    roles = db.relationship('Role', secondary=roles_users, backref=db.backref('users', lazy='dynamic'))
    role = db.Column(db.String(20))  # 'admin', 'user', etc.

    # Datos personales
    nombre = db.Column(db.String(50), nullable=True)
    primer_apellido = db.Column(db.String(50), nullable=True)
    DNI = db.Column(db.String(50), nullable=True)
    Domicilio = db.Column(db.String(100), nullable=True)
    Centro_trabajo = db.Column(db.String(50), nullable=True)
    Departamento = db.Column(db.String(50), nullable=True)
    Provincia = db.Column(db.String(50), nullable=True)
    Localidad = db.Column(db.String(50), nullable=True)
    CP = db.Column(db.Integer, nullable=True)
    Telefono = db.Column(db.Integer, nullable=True)
    Fecha_nacimiento = db.Column(db.Date, nullable=True)

    # Vacaciones
    dias_restantes_este = db.Column(db.Float, default=0.0)
    dias_restantes_pasado = db.Column(db.Integer, default=0)
    start_date = db.Column(db.Date)

    # Estado de la cuenta
    approved = db.Column(db.Boolean, default=False)

    # Seguridad y acceso
    intentos_fallidos = db.Column(db.Integer, default=0)           # Número de intentos fallidos de login
    bloqueos_temporales = db.Column(db.Integer, default=0)         # Ciclos de bloqueo temporal
    bloqueado = db.Column(db.Boolean, default=False)               # Si la cuenta está bloqueada (temporal o permanente)
    bloqueado_hasta = db.Column(db.DateTime, nullable=True)        # Fecha y hora del desbloqueo automático si es temporal

    # Autenticación en dos pasos (2FA)
    two_factor_enabled = db.Column(db.Boolean, default=False)
    two_factor_secret = db.Column(db.String(32), nullable=True)

    # Métodos auxiliares

    def desbloquear(self):
        """Desbloquea al usuario completamente (temporal o permanente)."""
        self.bloqueado = 0
        self.bloqueado_hasta = None
        self.intentos_fallidos = 0
        self.bloqueos_temporales = 0

    def esta_bloqueado_temporalmente(self):
        """Devuelve True si el usuario está bloqueado temporalmente y el tiempo no ha expirado."""
        return self.bloqueado and self.bloqueado_hasta and datetime.now(timezone.utc) < self.bloqueado_hasta

    def esta_bloqueado_permanentemente(self):
        """Devuelve True si el usuario está bloqueado sin fecha de desbloqueo (bloqueo manual/definitivo)."""
        return self.bloqueado and self.bloqueado_hasta is None

    # Horarios para cada día de la semana
    trabaja_lunes = db.Column(db.Boolean, default=False)
    entrada_lunes = db.Column(db.Time, nullable=True)
    salida_comida_lunes = db.Column(db.Time, nullable=True)
    entrada_comida_lunes = db.Column(db.Time, nullable=True)
    salida_lunes = db.Column(db.Time, nullable=True)

    trabaja_martes = db.Column(db.Boolean, default=False)
    entrada_martes = db.Column(db.Time, nullable=True)
    salida_comida_martes = db.Column(db.Time, nullable=True)
    entrada_comida_martes = db.Column(db.Time, nullable=True)
    salida_martes = db.Column(db.Time, nullable=True)

    trabaja_miercoles = db.Column(db.Boolean, default=False)
    entrada_miercoles = db.Column(db.Time, nullable=True)
    salida_comida_miercoles = db.Column(db.Time, nullable=True)
    entrada_comida_miercoles = db.Column(db.Time, nullable=True)
    salida_miercoles = db.Column(db.Time, nullable=True)

    trabaja_jueves = db.Column(db.Boolean, default=False)
    entrada_jueves = db.Column(db.Time, nullable=True)
    salida_comida_jueves = db.Column(db.Time, nullable=True)
    entrada_comida_jueves = db.Column(db.Time, nullable=True)
    salida_jueves = db.Column(db.Time, nullable=True)

    trabaja_viernes = db.Column(db.Boolean, default=False)
    entrada_viernes = db.Column(db.Time, nullable=True)
    salida_comida_viernes = db.Column(db.Time, nullable=True)
    entrada_comida_viernes = db.Column(db.Time, nullable=True)
    salida_viernes = db.Column(db.Time, nullable=True)

    trabaja_sabado = db.Column(db.Boolean, default=False)
    entrada_sabado = db.Column(db.Time, nullable=True)
    salida_comida_sabado = db.Column(db.Time, nullable=True)
    entrada_comida_sabado = db.Column(db.Time, nullable=True)
    salida_sabado = db.Column(db.Time, nullable=True)

    trabaja_domingo = db.Column(db.Boolean, default=False)
    entrada_domingo = db.Column(db.Time, nullable=True)
    salida_comida_domingo = db.Column(db.Time, nullable=True)
    entrada_comida_domingo = db.Column(db.Time, nullable=True)
    salida_domingo = db.Column(db.Time, nullable=True)

    # Horarios de verano
    trabaja_verano_lunes = db.Column(db.Boolean, default=False)
    entrada_verano_lunes = db.Column(db.Time, nullable=True)
    salida_comida_verano_lunes = db.Column(db.Time, nullable=True)
    entrada_comida_verano_lunes = db.Column(db.Time, nullable=True)
    salida_verano_lunes = db.Column(db.Time, nullable=True)

    trabaja_verano_martes = db.Column(db.Boolean, default=False)
    entrada_verano_martes = db.Column(db.Time, nullable=True)
    salida_comida_verano_martes = db.Column(db.Time, nullable=True)
    entrada_comida_verano_martes = db.Column(db.Time, nullable=True)
    salida_verano_martes = db.Column(db.Time, nullable=True)

    trabaja_verano_miercoles = db.Column(db.Boolean, default=False)
    entrada_verano_miercoles = db.Column(db.Time, nullable=True)
    salida_comida_verano_miercoles = db.Column(db.Time, nullable=True)
    entrada_comida_verano_miercoles = db.Column(db.Time, nullable=True)
    salida_verano_miercoles = db.Column(db.Time, nullable=True)

    trabaja_verano_jueves = db.Column(db.Boolean, default=False)
    entrada_verano_jueves = db.Column(db.Time, nullable=True)
    salida_comida_verano_jueves = db.Column(db.Time, nullable=True)
    entrada_comida_verano_jueves = db.Column(db.Time, nullable=True)
    salida_verano_jueves = db.Column(db.Time, nullable=True)

    trabaja_verano_viernes = db.Column(db.Boolean, default=False)
    entrada_verano_viernes = db.Column(db.Time, nullable=True)
    salida_comida_verano_viernes = db.Column(db.Time, nullable=True)
    entrada_comida_verano_viernes = db.Column(db.Time, nullable=True)
    salida_verano_viernes = db.Column(db.Time, nullable=True)

    trabaja_verano_sabado = db.Column(db.Boolean, default=False)
    entrada_verano_sabado = db.Column(db.Time, nullable=True)
    salida_comida_verano_sabado = db.Column(db.Time, nullable=True)
    entrada_comida_verano_sabado = db.Column(db.Time, nullable=True)
    salida_verano_sabado = db.Column(db.Time, nullable=True)

    trabaja_verano_domingo = db.Column(db.Boolean, default=False)
    entrada_verano_domingo = db.Column(db.Time, nullable=True)
    salida_comida_verano_domingo = db.Column(db.Time, nullable=True)
    entrada_comida_verano_domingo = db.Column(db.Time, nullable=True)
    salida_verano_domingo = db.Column(db.Time, nullable=True)

    # Indicar si es invierno o verano
    es_invierno = db.Column(db.Boolean, default=True)

    tipo_jornada = db.Column(db.Enum('verano', 'invierno', name='tipo_jornada'), nullable=True)

    nominas = db.relationship('Nomina', backref='user', lazy=True)
    contratos = db.relationship('Contrato', backref='user', lazy=True)
    certificados = db.relationship('Certificado', backref='user', lazy=True)
    absences = db.relationship('Absence', backref='user', lazy=True)  # Relación con la tabla Absence

    two_factor_secret = db.Column(db.String(32), nullable=True)
    two_factor_enabled = db.Column(db.Boolean, default=False)
    two_factor_verified = db.Column(db.Boolean, default=False)
    horarios = db.relationship('RegistroHorario', backref='user', lazy=True)  # Relación con la tabla RegistroHorario
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=True)
    empresa = db.relationship('Empresa', backref=db.backref('usuarios', lazy=True))

    vacaciones = db.Column(db.Boolean, default=False)
    registros_horarios = db.Column(db.Boolean, default=False)
    ausencias = db.Column(db.Boolean, default=False)
    qr = db.relationship('QR', backref='usuario', uselist=False)
    bloquear_por_ip = db.Column(db.Boolean, default=False)
    bajas = db.relationship('Baja', back_populates='usuario', cascade='all, delete-orphan')

    @property
    def baja_activa(self):
        return any(b.fecha_fin is None for b in self.bajas)


class Baja(db.Model):
    __tablename__ = 'baja'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    fecha_inicio = db.Column(db.Date, nullable=False)
    fecha_fin = db.Column(db.Date, nullable=True)


    # Relación con User usando backref
    usuario = db.relationship('User', back_populates='bajas')


class QR(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    imagen = db.Column(db.String(255), nullable=False)  # Ruta de la imagen del QR
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    centro_trabajo_id = db.Column(db.Integer, db.ForeignKey('centro_trabajo.id'), nullable=False)  # Relación con el centro de trabajo
    centro_trabajo = db.relationship('CentroTrabajo', backref='qrs')  # Relación con el modelo CentroTrabajo



class RegistroHorario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    entrada = db.Column(db.Time, nullable=True)
    salida = db.Column(db.Time, nullable=True)
    entrada_comida = db.Column(db.Time, nullable=True)  # Nuevo campo para entrada comida
    salida_comida = db.Column(db.Time, nullable=True)  # Nuevo campo para salida comida
    validada = db.Column(db.Boolean, default=False)

class ExcepcionSalidaTardia(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    registro_id = db.Column(db.Integer, db.ForeignKey('registro.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    hora_salida_real = db.Column(db.Time, nullable=False)  # Hora que el usuario registró
    comentarios = db.Column(db.Text, nullable=True)
    aceptada = db.Column(db.Boolean, default=None)  # None = pendiente, True = aceptada, False = rechazada

    registro = db.relationship('Registro', backref='excepcion_salida_tardia')


class SolicitudCambioHorario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    registro_id = db.Column(db.Integer, db.ForeignKey('registro.id'), nullable=False)
    registro = db.relationship('Registro', backref='solicitudes')
    tipo = db.Column(db.String(10))  # 'entrada' o 'salida'
    nueva_fecha_hora = db.Column(db.DateTime, nullable=False)
    motivo = db.Column(db.Text, nullable=False)
    estado = db.Column(db.String(10), default='pendiente')  # 'pendiente', 'aprobada', 'rechazada'
    token = db.Column(db.String(100), unique=True, nullable=False)
    fecha_solicitud = db.Column(db.DateTime, default=datetime.utcnow)


from wtforms.validators import Optional

class UserProfileForm(FlaskForm):
    nombre = StringField('Nombre*')
    primer_apellido = StringField('Apellidos*')
    segundo_apellido = StringField('Segundo Apellido', validators=[Optional()])
    email = StringField('Correo*')
    DNI = StringField('DNI', validators=[Optional()])
    Domicilio = StringField('Domicilio', validators=[Optional()])

    # Crea los campos Centro_trabajo y Departamento como SelectField sin opciones iniciales
    Centro_trabajo = SelectField('Centro de Trabajo', choices=[], coerce=int)
    Departamento = SelectField('Departamento', choices=[], coerce=int)

    Provincia = StringField('Provincia', validators=[Optional()])
    Localidad = StringField('Localidad', validators=[Optional()])
    CP = IntegerField('Código Postal', validators=[Optional()])
    Telefono = IntegerField('Nº Teléfono', validators=[Optional()])
    Fecha_nacimiento = DateField('Fecha de nacimiento', validators=[Optional()])

    submit = SubmitField('Guardar Cambios')

class UserProfileForm2(FlaskForm):
    nombre = StringField('Nombre*')
    primer_apellido = StringField('Apellidos*')
    segundo_apellido = StringField('Segundo Apellido', validators=[Optional()])
    email = StringField('Correo*')
    DNI = StringField('DNI', validators=[Optional()])
    Domicilio = StringField('Domicilio', validators=[Optional()])

    # Crea los campos Centro_trabajo y Departamento como SelectField sin opciones iniciales
    Centro_trabajo = SelectField('Centro de Trabajo', choices=[], coerce=int)
    Departamento = SelectField('Departamento', choices=[], coerce=int)

    Provincia = StringField('Provincia', validators=[Optional()])
    Localidad = StringField('Localidad', validators=[Optional()])
    CP = IntegerField('Código Postal', validators=[Optional()])
    Telefono = IntegerField('Nº Teléfono', validators=[Optional()])
    Fecha_nacimiento = DateField('Fecha de nacimiento', validators=[Optional()])
    empresa_id = SelectField('Empresa', coerce=int)
    submit = SubmitField('Guardar Cambios')

# Clase Nomina
class Nomina(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    ruta = db.Column(db.String(255))

# Clase Contrato
class Contrato(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    archivo = db.Column(db.String(255), nullable=False)

    # Clase Certificado
class Certificado(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    archivo = db.Column(db.LargeBinary, nullable=False)


class BankProofForm(FlaskForm):
    bank_proof = FileField('Comprobante de Cuenta Bancaria', validators=[DataRequired()])
    submit = SubmitField('Adjuntar Comprobante')


# Flask-Security configuration
user_datastore = SQLAlchemyUserDatastore(db, User, Role)
security = Security(app, user_datastore)

# Vacation Request model
class VacationRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='PENDIENTE')
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    is_accepted = db.Column(db.Boolean, default=False)
    rejection_reason = db.Column(db.String(255))
    cancel_reason = db.Column(db.String(255))
    user_email = db.Column(db.String(255), nullable=False)
    user_name = db.Column(db.String(255), nullable=False)
    user_apellido = db.Column(db.String(255), nullable=False)
    dias_vacas=db.Column(db.Integer)
    dias_utilizados_pasado = db.Column(db.Float, default=0.0)
    dias_utilizados_este = db.Column(db.Float, default=0.0)
    dia_solicitud = db.Column(db.Date, nullable=False)
    dia_decision = db.Column(db.Date)
    centro = db.Column(db.String(50), nullable=True)


class VacationHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20))
    is_accepted = db.Column(db.Boolean, default=False)
    rejection_reason = db.Column(db.String(255))
    cancel_reason = db.Column(db.String(255))
    dia_solicitud = db.Column(db.Date, nullable=False)
    dia_decision = db.Column(db.Date)

class CompanyConfigurationForm(FlaskForm):
    rest_day_name = StringField('Rest Day Name', validators=[DataRequired()])
    holiday_name = StringField('Holiday Name', validators=[DataRequired()])
    rest_day_submit = SubmitField('Add Rest Day')
    holiday_submit = SubmitField('Add Holiday')

class MyForm(FlaskForm):
    start_date = DateField('Desde:', validators=[DataRequired()], format='%Y-%m-%d')
    end_date = DateField('Hasta:', validators=[DataRequired()], format='%Y-%m-%d')
    submit = SubmitField('SOLICITAR')

class ReportAbsenceForm(FlaskForm):
    start_date = DateField('DESDE:', format='%d/%m/%Y')
    end_date = DateField('HASTA:', format='%d/%m/%Y')

    absence_type = SelectField('TIPO:', choices=[('option1', 'Option 1'), ('option2', 'Option 2')])
    description = TextAreaField('DESCRIPCIÓN:')
    type = SelectField('TIPO:', choices=[('Completa', 'Completa'), ('Parcial', 'Parcial')])
    duration_hours = StringField('Duración:', render_kw={'readonly': True})
    file = FileField('Archivo Adjunto:', validators=[FileAllowed(['pdf', 'doc', 'docx', 'txt', 'jpg', 'jpeg', 'png'])])
    submit = SubmitField('INFORMAR AUSENCIA')


class ChangePasswordForm(FlaskForm):
    current_password = PasswordField('Contraseña actual', validators=[DataRequired()])
    new_password = PasswordField('Nueva contraseña', validators=[
        DataRequired(),
        Length(min=6, message='La contraseña debe tener al menos 6 caracteres')
    ])
    confirm_password = PasswordField('Confirmar nueva contraseña', validators=[
        DataRequired(),
        EqualTo('new_password', message='Las contraseñas deben coincidir')
    ])
    submit = SubmitField('Cambiar Contraseña')

class ForgotPasswordForm(FlaskForm):
    email = StringField('Correo Electrónico', validators=[DataRequired(), Email()])
    submit = SubmitField('Recuperar Contraseña')

class ResetPasswordForm(FlaskForm):
    email = StringField('Correo Electrónico', validators=[DataRequired(), Email()])
    new_password = PasswordField('Nueva Contraseña', validators=[DataRequired()])
    confirm_password = PasswordField('Confirmar Contraseña', validators=[DataRequired(), EqualTo('new_password')])
    submit = SubmitField('Cambiar Contraseña')

class Registro(db.Model):
    __tablename__ = 'registro'

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    entrada = db.Column(db.DateTime, nullable=False)
    salida = db.Column(db.DateTime)
    ip_address_entrada = db.Column(db.String(45), nullable=False)
    ip_address_salida = db.Column(db.String(45))
    latitud = db.Column(db.Float)
    longitud = db.Column(db.Float)
    latitud_salida = db.Column(db.Float)
    longitud_salida = db.Column(db.Float)
    observaciones = db.Column(db.Text)
    aprobado_modificacion = db.Column(db.Boolean, default=False)  # ✅ Nuevo campo

    modificado = db.Column(db.String(50), nullable=True)  # 'administrador', 'usuario', o 'null'
    fecha_modificacion = db.Column(db.DateTime, nullable=True)
    campo_modificado = db.Column(db.String(255), nullable=True)  # Nombre del campo modificado
    valor_anterior = db.Column(db.Text, nullable=True)  # Valor previo antes de modificar
    centro_trabajo_id = db.Column(db.Integer, db.ForeignKey('centro_trabajo.id'))
    centro_trabajo = db.relationship('CentroTrabajo', backref='registros')


    user = db.relationship('User', backref='registros')
    # No es necesario definir el backref aquí, ya se definió en Pausa

    @property
    def duracion(self):
        if self.salida:
            return self.salida - self.entrada
        else:
            return None

    @property
    def duracion_formateada(self):
        duracion = self.duracion
        if duracion:
            horas, segundos = divmod(duracion.seconds, 3600)
            minutos, segundos = divmod(segundos, 60)
            return '{:02}:{:02}:{:02}'.format(horas, minutos, segundos)
        else:
            return 'N/A'

    @property
    def duracion_total_pausas(self):
        total = timedelta()
        for pausa in self.pausas_registro:  # Usar el nuevo nombre del backref aquí
            if pausa.duracion:
                total += pausa.duracion
        return total

    @property
    def duracion_neta(self):
        if self.duracion:
            return self.duracion - self.duracion_total_pausas
        else:
            return None

    @property
    def duracion_neta_formateada(self):
        duracion = self.duracion_neta
        if duracion:
            horas, segundos = divmod(duracion.seconds, 3600)
            minutos, segundos = divmod(segundos, 60)
            return '{:02}:{:02}:{:02}'.format(horas, minutos, segundos)
        else:
            return 'N/A'

    def __repr__(self):
        return f"Registro(user_id={self.user_id}, entrada={self.entrada}, salida={self.salida})"






class Pausa(db.Model):
    __tablename__ = 'pausa'

    id = db.Column(db.Integer, primary_key=True)
    registro_id = db.Column(db.Integer, db.ForeignKey('registro.id'), nullable=False)
    inicio = db.Column(db.DateTime, nullable=False)
    fin = db.Column(db.DateTime)
    tipo = db.Column(db.String(100), nullable=False)

    registro = db.relationship('Registro', backref=db.backref('pausas_registro', lazy=True))  # Renombrar el backref aquí

    @property
    def duracion(self):
        if self.fin:
            return self.fin - self.inicio
        else:
            return None

    @property
    def duracion_minutos(self):
        if self.fin:
            delta = self.fin - self.inicio
            return int(delta.total_seconds() // 60)  # minutos enteros
        else:
            return None

    @property
    def duracion_formateada(self):
        duracion = self.duracion
        if duracion:
            horas, segundos = divmod(duracion.seconds, 3600)
            minutos, segundos = divmod(segundos, 60)
            return '{:02}:{:02}:{:02}'.format(horas, minutos, segundos)
        else:
            return 'Duración no disponible'

    def __repr__(self):
        return f"Pausa(registro_id={self.registro_id}, inicio={self.inicio}, fin={self.fin})"






class Notificacion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    concepto = db.Column(db.String(255), nullable=False)
    fecha_hora = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    leida = db.Column(db.Boolean, default=False)  # Campo para marcar si la notificación ha sido leída

    usuario = db.relationship('User', backref=db.backref('notificaciones', lazy=True))



class Mensaje(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    remitente_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    destinatario_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    contenido = db.Column(db.Text, nullable=False)
    fecha_envio = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    leido = db.Column(db.Integer, default=0)  # Cuenta de mensajes no leídos
    remitente = db.relationship('User', foreign_keys=[remitente_id], backref='mensajes_enviados')
    destinatario = db.relationship('User', foreign_keys=[destinatario_id], backref='mensajes_recibidos')




class Empresa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), nullable=True)
    CIF = db.Column(db.String(50), nullable=True)
    Domicilio = db.Column(db.String(100), nullable=True)
    Provincia = db.Column(db.String(50), nullable=True)
    Localidad = db.Column(db.String(50), nullable=True)
    CP = db.Column(db.String(10), nullable=True)
    Telefono = db.Column(db.String(20), nullable=True)
    email = db.Column(db.String(255), unique=True, nullable=True)
    horas_totales_laborables = db.Column(db.Integer, nullable=True)
    dias_vacaciones = db.Column(db.Integer, nullable=True)
    ver_vacaciones = db.Column(db.Boolean, default=True)  # True para mostrar, False para ocultar
    ver_pausas = db.Column(db.Boolean, default=True)

    # Horarios de verano
    trabaja_verano_lunes = db.Column(db.Boolean, default=False)
    entrada_verano_lunes = db.Column(db.Time, nullable=True)
    salida_comida_verano_lunes = db.Column(db.Time, nullable=True)
    entrada_comida_verano_lunes = db.Column(db.Time, nullable=True)
    salida_verano_lunes = db.Column(db.Time, nullable=True)
    # ... Repetir para martes, miércoles, jueves, viernes, sábado y domingo

    trabaja_verano_martes = db.Column(db.Boolean, default=False)
    entrada_verano_martes = db.Column(db.Time, nullable=True)
    salida_comida_verano_martes = db.Column(db.Time, nullable=True)
    entrada_comida_verano_martes = db.Column(db.Time, nullable=True)
    salida_verano_martes = db.Column(db.Time, nullable=True)

    trabaja_verano_miercoles = db.Column(db.Boolean, default=False)
    entrada_verano_miercoles = db.Column(db.Time, nullable=True)
    salida_comida_verano_miercoles = db.Column(db.Time, nullable=True)
    entrada_comida_verano_miercoles = db.Column(db.Time, nullable=True)
    salida_verano_miercoles = db.Column(db.Time, nullable=True)

    trabaja_verano_jueves = db.Column(db.Boolean, default=False)
    entrada_verano_jueves = db.Column(db.Time, nullable=True)
    salida_comida_verano_jueves = db.Column(db.Time, nullable=True)
    entrada_comida_verano_jueves = db.Column(db.Time, nullable=True)
    salida_verano_jueves = db.Column(db.Time, nullable=True)

    trabaja_verano_viernes = db.Column(db.Boolean, default=False)
    entrada_verano_viernes = db.Column(db.Time, nullable=True)
    salida_comida_verano_viernes = db.Column(db.Time, nullable=True)
    entrada_comida_verano_viernes = db.Column(db.Time, nullable=True)
    salida_verano_viernes = db.Column(db.Time, nullable=True)

    trabaja_verano_sabado = db.Column(db.Boolean, default=False)
    entrada_verano_sabado = db.Column(db.Time, nullable=True)
    salida_comida_verano_sabado = db.Column(db.Time, nullable=True)
    entrada_comida_verano_sabado = db.Column(db.Time, nullable=True)
    salida_verano_sabado = db.Column(db.Time, nullable=True)

    trabaja_verano_domingo = db.Column(db.Boolean, default=False)
    entrada_verano_domingo = db.Column(db.Time, nullable=True)
    salida_comida_verano_domingo = db.Column(db.Time, nullable=True)
    entrada_comida_verano_domingo = db.Column(db.Time, nullable=True)
    salida_verano_domingo = db.Column(db.Time, nullable=True)

    # Horarios de invierno
    trabaja_invierno_lunes = db.Column(db.Boolean, default=False)
    entrada_invierno_lunes = db.Column(db.Time, nullable=True)
    salida_comida_invierno_lunes = db.Column(db.Time, nullable=True)
    entrada_comida_invierno_lunes = db.Column(db.Time, nullable=True)
    salida_invierno_lunes = db.Column(db.Time, nullable=True)
    # ... Repetir para martes, miércoles, jueves, viernes, sábado y domingo

    trabaja_invierno_martes = db.Column(db.Boolean, default=False)
    entrada_invierno_martes = db.Column(db.Time, nullable=True)
    salida_comida_invierno_martes = db.Column(db.Time, nullable=True)
    entrada_comida_invierno_martes = db.Column(db.Time, nullable=True)
    salida_invierno_martes = db.Column(db.Time, nullable=True)

    trabaja_invierno_miercoles = db.Column(db.Boolean, default=False)
    entrada_invierno_miercoles = db.Column(db.Time, nullable=True)
    salida_comida_invierno_miercoles = db.Column(db.Time, nullable=True)
    entrada_comida_invierno_miercoles = db.Column(db.Time, nullable=True)
    salida_invierno_miercoles = db.Column(db.Time, nullable=True)

    trabaja_invierno_jueves = db.Column(db.Boolean, default=False)
    entrada_invierno_jueves = db.Column(db.Time, nullable=True)
    salida_comida_invierno_jueves = db.Column(db.Time, nullable=True)
    entrada_comida_invierno_jueves = db.Column(db.Time, nullable=True)
    salida_invierno_jueves = db.Column(db.Time, nullable=True)

    trabaja_invierno_viernes = db.Column(db.Boolean, default=False)
    entrada_invierno_viernes = db.Column(db.Time, nullable=True)
    salida_comida_invierno_viernes = db.Column(db.Time, nullable=True)
    entrada_comida_invierno_viernes = db.Column(db.Time, nullable=True)
    salida_invierno_viernes = db.Column(db.Time, nullable=True)

    trabaja_invierno_sabado = db.Column(db.Boolean, default=False)
    entrada_invierno_sabado = db.Column(db.Time, nullable=True)
    salida_comida_invierno_sabado = db.Column(db.Time, nullable=True)
    entrada_comida_invierno_sabado = db.Column(db.Time, nullable=True)
    salida_invierno_sabado = db.Column(db.Time, nullable=True)

    trabaja_invierno_domingo = db.Column(db.Boolean, default=False)
    entrada_invierno_domingo = db.Column(db.Time, nullable=True)
    salida_comida_invierno_domingo = db.Column(db.Time, nullable=True)
    entrada_comida_invierno_domingo = db.Column(db.Time, nullable=True)
    salida_invierno_domingo = db.Column(db.Time, nullable=True)

    # Fechas de inicio y fin del horario de verano
    fecha_inicio_verano = db.Column(db.Date, nullable=True)
    fecha_fin_verano = db.Column(db.Date, nullable=True)

    tipo_jornada = db.Column(db.Enum('intensiva', 'partida', name='tipo_jornada'), nullable=False)

    centros_trabajo = db.relationship('CentroTrabajo', back_populates='empresa', cascade='all, delete-orphan')
    departamentos = db.relationship('Departamento', back_populates='empresa', cascade='all, delete-orphan')
    tiempo_cortesia = db.Column(db.Integer, nullable=True, default=15)


    logo_favicon = db.Column(db.String(255), nullable=True)
    logo_login = db.Column(db.String(255), nullable=True)
    logo_menu = db.Column(db.String(255), nullable=True)

    ip_empresa = db.Column(db.String(45), nullable=True)

    def __repr__(self):
        return f'<Empresa {self.nombre}>'

class CentroTrabajo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    empresa = db.relationship('Empresa', back_populates='centros_trabajo')

    def __repr__(self):
        return f'<CentroTrabajo {self.nombre}>'

class Departamento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False, unique=True)
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)
    empresa = db.relationship('Empresa', back_populates='departamentos')

    def __repr__(self):
        return f'<Departamento {self.nombre}>'


class Festivo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    dia = db.Column(db.Integer, nullable=False)
    mes = db.Column(db.Integer, nullable=False)
    año = db.Column(db.Integer, nullable=True)   # ← AÑADIDO
    empresa_id = db.Column(db.Integer, db.ForeignKey('empresa.id'), nullable=False)

    empresa = db.relationship('Empresa', backref='festivos')

    def __repr__(self):
        return f'<Festivo {self.dia}/{self.mes} - Empresa {self.empresa_id}>'



from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, IntegerField, TimeField, SubmitField, FormField, FieldList
from wtforms.validators import DataRequired, Length, Optional

class CentroTrabajoForm(FlaskForm):
    nombre = StringField('Nombre del Centro', validators=[DataRequired(), Length(max=50)])

class DepartamentoForm(FlaskForm):
    nombre = StringField('Nombre del Departamento', validators=[DataRequired(), Length(max=50)])

from flask_wtf import FlaskForm
from wtforms import StringField, SelectField, IntegerField, TimeField, DateField, FieldList, FormField, SubmitField
from wtforms.validators import DataRequired, Length, Optional
from wtforms import StringField
from wtforms.validators import Optional, IPAddress
class EmpresaForm(FlaskForm):
    nombre = StringField('Nombre', validators=[Length(max=100)])
    CIF = StringField('CIF', validators=[Length(max=20)])
    Domicilio = StringField('Domicilio', validators=[Length(max=150)])
    Provincia = SelectField('Provincia', choices=[
        ('', 'Selecciona una provincia'),
        ('Álava', 'Álava'), ('Albacete', 'Albacete'), ('Alicante', 'Alicante'),
        ('Almería', 'Almería'), ('Asturias', 'Asturias'), ('Ávila', 'Ávila'),
        ('Badajoz', 'Badajoz'), ('Barcelona', 'Barcelona'), ('Burgos', 'Burgos'),
        ('Cáceres', 'Cáceres'), ('Cádiz', 'Cádiz'), ('Cantabria', 'Cantabria'),
        ('Castellón', 'Castellón'), ('Ceuta', 'Ceuta'), ('Ciudad Real', 'Ciudad Real'),
        ('Córdoba', 'Córdoba'), ('La Coruña', 'La Coruña'), ('Cuenca', 'Cuenca'),
        ('Gerona', 'Gerona'), ('Granada', 'Granada'), ('Guadalajara', 'Guadalajara'),
        ('Guipúzcoa', 'Guipúzcoa'), ('Huelva', 'Huelva'), ('Huesca', 'Huesca'),
        ('Islas Baleares', 'Islas Baleares'), ('Jaén', 'Jaén'), ('La Rioja', 'La Rioja'),
        ('Las Palmas', 'Las Palmas'), ('León', 'León'), ('Lérida', 'Lérida'),
        ('Madrid', 'Madrid'), ('Málaga', 'Málaga'), ('Murcia', 'Murcia'),
        ('Navarra', 'Navarra'), ('Ourense', 'Ourense'), ('Palencia', 'Palencia'),
        ('Pontevedra', 'Pontevedra'), ('Salamanca', 'Salamanca'), ('Segovia', 'Segovia'),
        ('Sevilla', 'Sevilla'), ('Soria', 'Soria'), ('Tarragona', 'Tarragona'),
        ('Teruel', 'Teruel'), ('Toledo', 'Toledo'), ('Valencia', 'Valencia'),
        ('Valladolid', 'Valladolid'), ('Vizcaya', 'Vizcaya'), ('Zamora', 'Zamora'),
        ('Zaragoza', 'Zaragoza')
    ], validators=[DataRequired()])
    Localidad = StringField('Localidad')
    CP = IntegerField('Código Postal')
    Telefono = IntegerField('Teléfono')
    email = StringField('Correo Electrónico')
    centro_trabajo = FieldList(FormField(CentroTrabajoForm), min_entries=0, max_entries=10)
    departamento = FieldList(FormField(DepartamentoForm), min_entries=0, max_entries=10)

    tipo_jornada = SelectField('Tipo de Jornada', choices=[('intensiva', 'Intensiva'), ('partida', 'Partida')], validators=[Optional()])
    horas_totales_laborables = IntegerField('Horas Totales Laborables', validators=[Optional()])

    # Fechas de horario de verano
    fecha_inicio_verano = DateField('Inicio Horario de Verano', format='%Y-%m-%d', validators=[Optional()])
    fecha_fin_verano = DateField('Fin Horario de Verano', format='%Y-%m-%d', validators=[Optional()])

     # Horario de verano
    trabaja_verano_lunes = BooleanField('Trabaja Lunes', validators=[Optional()])
    entrada_verano_lunes = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_verano_lunes = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_verano_lunes = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_verano_lunes = TimeField('Salida Lunes', format='%H:%M', validators=[Optional()])

    trabaja_verano_martes = BooleanField('Trabaja Martes', validators=[Optional()])
    entrada_verano_martes = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_verano_martes = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_verano_martes = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_verano_martes = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_verano_miercoles = BooleanField('Trabaja Miércoles', validators=[Optional()])
    entrada_verano_miercoles = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_verano_miercoles = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_verano_miercoles = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_verano_miercoles = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_verano_jueves = BooleanField('Trabaja Jueves', validators=[Optional()])
    entrada_verano_jueves = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_verano_jueves = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_verano_jueves = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_verano_jueves = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_verano_viernes = BooleanField('Trabaja Viernes', validators=[Optional()])
    entrada_verano_viernes = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_verano_viernes = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_verano_viernes = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_verano_viernes = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_verano_sabado = BooleanField('Trabaja Sábado', validators=[Optional()])
    entrada_verano_sabado = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_verano_sabado = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_verano_sabado = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_verano_sabado = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_verano_domingo = BooleanField('Trabaja Domingo', validators=[Optional()])
    entrada_verano_domingo = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_verano_domingo = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_verano_domingo = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_verano_domingo = TimeField('Salida', format='%H:%M', validators=[Optional()])

    # Horarios de invierno
    trabaja_invierno_lunes = BooleanField('Trabaja Lunes', validators=[Optional()])
    entrada_invierno_lunes = TimeField('Entrada', validators=[Optional()])
    salida_comida_invierno_lunes = TimeField('Salida Comida', validators=[Optional()])
    entrada_comida_invierno_lunes = TimeField('Entrada Comida', validators=[Optional()])
    salida_invierno_lunes = TimeField('Salida', validators=[Optional()])

    trabaja_invierno_martes = BooleanField('Trabaja Martes', validators=[Optional()])
    entrada_invierno_martes = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_invierno_martes = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_invierno_martes = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_invierno_martes = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_invierno_miercoles = BooleanField('Trabaja Miércoles', validators=[Optional()])
    entrada_invierno_miercoles = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_invierno_miercoles = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_invierno_miercoles = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_invierno_miercoles = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_invierno_jueves = BooleanField('Trabaja Jueves', validators=[Optional()])
    entrada_invierno_jueves = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_invierno_jueves = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_invierno_jueves = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_invierno_jueves = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_invierno_viernes = BooleanField('Trabaja Viernes', validators=[Optional()])
    entrada_invierno_viernes = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_invierno_viernes = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_invierno_viernes = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_invierno_viernes = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_invierno_sabado = BooleanField('Trabaja Sábado', validators=[Optional()])
    entrada_invierno_sabado = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_invierno_sabado = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_invierno_sabado = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_invierno_sabado = TimeField('Salida', format='%H:%M', validators=[Optional()])

    trabaja_invierno_domingo = BooleanField('Trabaja Domingo', validators=[Optional()])
    entrada_invierno_domingo = TimeField('Entrada', format='%H:%M', validators=[Optional()])
    salida_comida_invierno_domingo = TimeField('Salida Comida', format='%H:%M', validators=[Optional()])
    entrada_comida_invierno_domingo = TimeField('Entrada Comida', format='%H:%M', validators=[Optional()])
    salida_invierno_domingo = TimeField('Salida', format='%H:%M', validators=[Optional()])

    ip_empresa = StringField('Dirección IP', validators=[Optional(), IPAddress(ipv4=True, ipv6=True, message="IP no válida")])

    submit = SubmitField('Guardar Cambios')


class AdminAssignVacationForm(FlaskForm):
    user_id = SelectField("Empleado", coerce=int, validators=[DataRequired()])
    start_date = DateField("Fecha inicio", format="%Y-%m-%d", validators=[DataRequired()])
    end_date = DateField("Fecha fin", format="%Y-%m-%d", validators=[DataRequired()])
    submit = SubmitField("Asignar vacaciones")


contador_global = 0
@app.route('/incrementar_contador')
def incrementar_contador():
    global contador_global
    contador_global += 1
    return f'Contador incrementado: {contador_global}'
def reset_contador():
    global contador_global
    contador_global== 0
    return f'Contador Reiniciado: {contador_global}'
def obtener_contador_global():
    global contador_global
    return contador_global


from functools import wraps

from functools import wraps
from flask import flash, redirect, url_for
from flask_login import current_user

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        print("→ Entrando en admin_required")
        print("→ current_user:", current_user)

        if not current_user.is_authenticated:
            print("⚠️ Usuario no autenticado")
            flash("Debes iniciar sesión para acceder a esta página.", "warning")
            return redirect(url_for("login"))

        try:
            roles = current_user.roles
            print("→ current_user.roles:", roles)
        except Exception as e:
            print("⚠️ Error accediendo a current_user.roles:", e)
            flash("Error interno al verificar permisos.", "danger")
            return redirect(url_for("login"))

        if not any(role.name in ["admin", "gestor", "encargado"] for role in roles):
            print("⛔ Acceso denegado: el usuario no tiene un rol autorizado.")
            flash("Acceso denegado. Se requiere rol de administrador, gestor o encargado.", "danger")
            return redirect(url_for("login"))

        print("✅ Acceso concedido.")
        return f(*args, **kwargs)
    return decorated_function


@app.route('/approve_user/<int:user_id>', methods=['GET', 'POST'])
def approve_user(user_id):
    # Obtén el usuario de la base de datos
    user = User.query.get_or_404(user_id)

    # Asegúrate de que solo los administradores pueden aprobar usuarios
    # Puedes implementar tu propia lógica de autenticación aquí

    # Cambia el estado del usuario a aprobado
    user.approved = True

    # Guarda los cambios en la base de datos
    db.session.commit()
    send_email_to_user(user.email, 'Usuario validado', f'Su usuario ha sido validado por el responsable')

    # Redirige a la página de administración de usuarios
    return redirect(url_for('admin_users'))



from werkzeug.security import check_password_hash
#función login nueva con control de intentos bloqueo temporal y bloqueo permanente
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        otp_code = request.form.get('otp_code', None)

        print(f"Email recibido: {email}")
        # ⚠️ No imprimas contraseñas en producción

        user = User.query.filter_by(email=email).first()

        if user:
            print(f"Usuario encontrado: {user.email}")

            timezone_spain = pytz_timezone("Europe/Madrid")
            now = datetime.now().astimezone(timezone_spain)

            # 🔒 Bloqueo permanente
            if user.bloqueado and not user.bloqueado_hasta:
                print("Cuenta bloqueada permanentemente")
                return jsonify({
                    "status": "error",
                    "message": "Tu cuenta ha sido bloqueada permanentemente. Contacta al administrador."
                })

            # 🔒 Bloqueo temporal
            if user.bloqueado and user.bloqueado_hasta and now < user.bloqueado_hasta:
                tiempo_restante = int((user.bloqueado_hasta - now).total_seconds() // 60) + 1
                print(f"Cuenta bloqueada temporalmente hasta {user.bloqueado_hasta}")
                return jsonify({
                    "status": "error",
                    "message": f"Tu cuenta está bloqueada. Intenta de nuevo en {tiempo_restante} minutos."
                })

            # 🔓 Desbloqueo automático si el tiempo ya pasó
            if user.bloqueado and user.bloqueado_hasta and now >= user.bloqueado_hasta:
                print("Desbloqueo automático activado")
                user.bloqueado = False
                user.bloqueado_hasta = None
                user.intentos_fallidos = 0
                db.session.commit()

            # ✅ Verificar contraseña
            if check_password_hash(user.password, password):
                print("Contraseña correcta")

                if user.approved:
                    print("Cuenta aprobada")

                    # 🔐 2FA
                    if user.two_factor_enabled:
                        print("2FA habilitado")
                        if not otp_code:
                            return jsonify({
                                "status": "2fa_required",
                                "message": "El código 2FA es necesario para completar el inicio de sesión."
                            })

                        if not pyotp.TOTP(user.two_factor_secret).verify(otp_code):
                            print("Código 2FA incorrecto")
                            return jsonify({
                                "status": "error",
                                "message": "Código 2FA incorrecto. Por favor, inténtalo de nuevo."
                            })

                    # 🔄 Resetear estado de bloqueo
                    user.intentos_fallidos = 0
                    user.bloqueado_hasta = None
                    user.bloqueado = False
                    user.bloqueos_temporales = 0
                    db.session.commit()

                    login_user(user)
                    log_login_event(email)

                    # 🎯 Redirección por rol
                    roles = [role.name for role in user.roles]
                    if any(r in roles for r in ['admin', 'gestor', 'encargado']):
                        return jsonify({"status": "success", "redirect_url": url_for('admin_page', _external=True)})
                    elif 'user' in roles:
                        return jsonify({"status": "success", "redirect_url": url_for('user_page', _external=True)})
                    elif 'inspector' in roles:
                        return jsonify({"status": "success", "redirect_url": url_for('obtener_registros', _external=True)})

                else:
                    print("Cuenta no aprobada")
                    return jsonify({
                        "status": "error",
                        "message": "Tu cuenta aún no ha sido aprobada. Espera la validación de un administrador."
                    })

            else:
                print("Contraseña incorrecta")
                user.intentos_fallidos += 1

                if user.intentos_fallidos >= 3:
                    user.intentos_fallidos = 0
                    user.bloqueos_temporales += 1

                    if user.bloqueos_temporales >= 3:
                        # Bloqueo permanente
                        user.bloqueado = True
                        user.bloqueado_hasta = None
                        user.approved = 0
                        print("Cuenta bloqueada permanentemente")
                        db.session.commit()
                        return jsonify({
                            "status": "error",
                            "message": "Has excedido el número máximo de intentos. Tu cuenta ha sido bloqueada permanentemente."
                        })
                    else:
                        # Bloqueo temporal

                        user.bloqueado_hasta = now + timedelta(minutes=5)
                        print(f"Cuenta bloqueada temporalmente hasta {user.bloqueado_hasta}")
                        db.session.commit()
                        return jsonify({
                            "status": "error",
                            "message": "Has excedido el número de intentos. Tu cuenta ha sido bloqueada por 5 minutos."
                        })

                db.session.commit()
                return jsonify({"status": "error", "message": "Correo electrónico o contraseña incorrectos"})

        else:
            print(f"No se encontró el usuario con email: {email}")
            return jsonify({
                "status": "error",
                "message": "Correo electrónico o contraseña incorrectos"
            })

    return render_template('login.html')


# Funcion para crear log de inicios de sesion
def log_login_event(email, success=True):
    # Ruta del archivo log
    log_dir = os.path.join('/home/adaptasystem/gestiona_empresa/matriz/log')
    log_file = os.path.join(log_dir, 'log.txt')

    # Asegúrate de que el directorio log exista
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

        # Escribir en el archivo log.txt
    with open(log_file, 'a') as f:
        if success:
            f.write(f"Usuario: {email} - Inicio de sesión exitoso - Fecha y Hora: {hora_espana}\n")
        else:
            f.write(f"Usuario: {email} - Error de autenticación - Fecha y Hora: {hora_espana}\n")

# Función para crear log de cierres de sesión
def log_logout_event(email):
    # Ruta del archivo log
    log_dir = os.path.join('/home/adaptasystem/gestiona_empresa/matriz/log')
    log_file = os.path.join(log_dir, 'log.txt')


    # Asegúrate de que el directorio log exista
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    # Escribir en el archivo log.txt
    with open(log_file, 'a') as f:
        f.write(f"Usuario: {email} - Cierre de sesión - Fecha y Hora: {hora_espana}\n")



# Función para enviar el correo con archivo adjunto
def send_email_to_me(email, subject, body, attachment=None):
    msg = Message(subject, recipients=[email], body=body)

    # Adjunta el archivo si está presente
    if attachment:
        msg.attach(attachment.filename, "image/*", attachment.read())

    # Envía el correo
    mail.send(msg)

# Ruta para procesar el formulario y enviar el correo
@app.route('/report_problem', methods=['GET', 'POST'])
def report_problem():
    current_user_email = current_user.email
    if request.method == 'POST':
        # Obtén los datos del formulario
        issue_title = request.form['issue_title']
        description = request.form['description']

        # Verifica si se cargó un archivo
        attachment = None
        if 'image' in request.files:
            image = request.files['image']
            if image.filename != '':
                # Adjunta el archivo directamente al correo
                send_email_to_me('incidencias@gestionaempresa.es', issue_title, description, image)
                return jsonify({"status": "success", "message": "Reporte enviado exitosamente."})

    # Si la petición no es POST, simplemente renderiza el formulario
    return render_template('report_problem.html',current_user_email=current_user_email, active_menu='report_problem')


@app.route('/view_user_details/<int:user_id>')
def view_user_details(user_id):
    user = User.query.get_or_404(user_id)  # Debes tener una función para obtener los detalles del usuario
    fecha_nacimiento_formateada = user.Fecha_nacimiento.strftime('%d/%m/%Y')
    return render_template('user_details.html', user=user, fecha_nacimiento_formateada = fecha_nacimiento_formateada )

@app.route('/user_details/<int:user_id>')
def user_details(user_id):
    user = User.query.get_or_404(user_id)
    user_data = {
        "email": user.email,
        "nombre": user.nombre,
        "primer_apellido": user.primer_apellido,
        "segundo_apellido": user.segundo_apellido,
        "DNI": user.DNI,
        "Domicilio": user.Domicilio,
        "Provincia": user.Provincia,
        "Localidad": user.Localidad,
        "CP": user.CP,
        "Telefono": user.Telefono,
        "Fecha_nacimiento": user.Fecha_nacimiento,
        "Centro_trabajo": user.Centro_trabajo
    }
    return jsonify(user_data)


# Función para guardar el archivo cargado en una ubicación temporal
def save_uploaded_file(file):
    # Ajusta la ubicación y el nombre del archivo según tu necesidad
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(file_path)
    return file_path

def obtener_nombre_usuario_por_id(user_id):
    # Aquí iría tu código para consultar la base de datos y obtener el nombre del usuario por su ID
    # Por ejemplo, si estás utilizando SQLAlchemy podrías hacer algo así:

    usuario = User.query.filter_by(id=user_id).first()
    if usuario:
        return f"{usuario.nombre} {usuario.primer_apellido}"
    else:
        return "Usuario no encontrado"



@app.route('/user')
def user_page():
    print("[DEBUG] Iniciando función user_page")

    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    print(f"[DEBUG] Notificaciones: {len(notificaciones)}")

    today = datetime.today().date()
    print(f"[DEBUG] Fecha actual: {today}")

    # Obtener la hora actual
    zona_es = pytz_timezone('Europe/Madrid')
    now = datetime.now(zona_es).time()
    print(f"[DEBUG] Hora actual: {now}")

    # Convertir a datetime y sumar una hora
    now_datetime = datetime.combine(datetime.today(), now)  # Combinar la fecha de hoy con la hora actual
    new_time = now_datetime + timedelta(hours=1)  # Sumar 1 hora

    # Obtener solo la hora sin la fecha
    new_time_only = new_time.time()
    print(f"[DEBUG] Hora después de sumar 1 hora: {new_time_only}")



    current_user_email = current_user.email if current_user.is_authenticated else None
    print(f"[DEBUG] Correo del usuario actual: {current_user_email}")

    empresa = Empresa.query.first()
    print(f"[DEBUG] Empresa: {empresa.nombre if empresa else 'No definida'}")

    user = current_user
    useri = current_user.id
    users = User.query.all()
    print(f"[DEBUG] Total de usuarios: {len(users)}")

    # Filtrar solicitudes pendientes y aceptadas
    # Filtrar solicitudes solo de la misma empresa
    pending_requests_query = VacationRequest.query.join(User).filter(
        VacationRequest.status == 'PENDIENTE',
        User.empresa_id == current_user.empresa_id
    )

    accepted_requests_query = VacationRequest.query.join(User).filter(
        VacationRequest.is_accepted == True,
        User.empresa_id == current_user.empresa_id
    )


    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        print("[DEBUG] Filtro para 'CEEI Guadalajara'")
        pending_requests_query = pending_requests_query.join(User).filter(User.Centro_trabajo == 'CEEI Guadalajara')
        accepted_requests_query = accepted_requests_query.join(User).filter(User.Centro_trabajo == 'CEEI Guadalajara')
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        print("[DEBUG] Filtro para 'CEOE Formación'")
        pending_requests_query = pending_requests_query.join(User).filter(User.Centro_trabajo == 'CEOE Formación')
        accepted_requests_query = accepted_requests_query.join(User).filter(User.Centro_trabajo == 'CEOE Formación')

    pending_requests = pending_requests_query.all()
    accepted_requests = accepted_requests_query.all()
    print(f"[DEBUG] Solicitudes pendientes: {len(pending_requests)}")
    print(f"[DEBUG] Solicitudes aceptadas: {len(accepted_requests)}")

    # Eventos de vacaciones para FullCalendar
    events = []
    for request in accepted_requests:
        if empresa.ver_vacaciones or request.user_id == current_user.id:
            start_date = request.start_date.strftime('%Y-%m-%d')
            end_date = (request.end_date + timedelta(days=1)).strftime('%Y-%m-%d')
            events.append({
                'start': start_date,
                'end': end_date,
                'rendering': 'background',
                'color': '#6173D6',
                'title': f"{request.user_name} {request.user_apellido}",
                'user_id': request.user_id
            })
    print(f"[DEBUG] Eventos generados para el calendario: {len(events)}")

    # Días festivos
    festivos = Festivo.query.all()
    current_year = today.year
    non_workdays = [f"{festivo.año}-{festivo.mes:02d}-{festivo.dia:02d}" for festivo in festivos]
    print(f"[DEBUG] Días festivos: {non_workdays}")

    # Estado del botón de entrada/salida
    es_entrada = True
    if current_user.is_authenticated:
        registro_entrada = Registro.query.filter_by(user_id=current_user.id, salida=None).first()
        es_entrada = registro_entrada is None
    print(f"[DEBUG] Estado de entrada: {'Entrada' if es_entrada else 'Salida'}")

    # Estadísticas de ausencias
    absences_count_mensual = Absence.query.filter(
        Absence.end_date.between(today.replace(day=1), (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1))
    ).count()
    absences_count_anual = Absence.query.filter(
        Absence.end_date.between(today.replace(month=1, day=1), today.replace(month=12, day=31))
    ).count()
    print(f"[DEBUG] Ausencias mensuales: {absences_count_mensual}")
    print(f"[DEBUG] Ausencias anuales: {absences_count_anual}")

    # Vacaciones de hoy
    current_day_vacations = VacationRequest.query.filter_by(is_accepted=True).filter(
        extract('year', VacationRequest.start_date) == today.year,
        extract('month', VacationRequest.start_date) == today.month,
        extract('day', VacationRequest.start_date) == today.day
    ).all()
    current_day_vacation_names = [f"{request.user_name} {request.user_apellido}" for request in current_day_vacations]
    print(f"[DEBUG] Vacaciones del día de hoy: {current_day_vacation_names}")

    # Verificar pausa activa
    registro = Registro.query.filter_by(user_id=current_user.id, salida=None).first()
    pausa_activa = None
    if registro:
        pausa_activa = Pausa.query.filter_by(registro_id=registro.id, fin=None).first()
    print(f"[DEBUG] Pausa activa: {pausa_activa}")

    # Última entrada
    hora_entrada = None
    if registro:
        hora_entrada = registro.entrada.strftime('%Y-%m-%d %H:%M:%S')
    print(f"[DEBUG] Hora de entrada: {hora_entrada}")

    # Determinar si es verano
    if empresa:

        es_verano = empresa.fecha_inicio_verano <= today <= empresa.fecha_fin_verano if empresa.fecha_inicio_verano else False
    else:
        es_verano= False
    print(f"[DEBUG] Es verano: {'Sí' if es_verano else 'No'}")

    # Día de la semana en español
    dias_semana = {
        "monday": "lunes", "tuesday": "martes", "wednesday": "miercoles",
        "thursday": "jueves", "friday": "viernes", "saturday": "sabado", "sunday": "domingo", "lunes": "lunes", "martes": "martes", "miércoles": "miercoles", "miercoles": "miercoles", "jueves": "jueves", "viernes": "viernes", "sabado": "sabado", "sábado": "sabado", "domingo": "domingo"
    }

    # Imprimir el valor de today.strftime("%A").lower() para ver qué devuelve
    dia_str = today.strftime("%A").lower()
    print(f"[DEBUG] Día en inglés: {dia_str}")

    # Obtener el día de la semana en español
    dia_semana = dias_semana.get(dia_str, "dia_desconocido")
    print(f"[DEBUG] Día de la semana: {dia_semana}")
    # Claves dinámicas para entrada/salida según temporada
    entrada_key = f"entrada_verano_{dia_semana}" if es_verano else f"entrada_{dia_semana}"
    salida_key = f"salida_verano_{dia_semana}" if es_verano else f"salida_{dia_semana}"
    print("Salida key: ", salida_key)
    salida_comida_key = f"salida_comida_verano_{dia_semana}" if es_verano else f"salida_comida_{dia_semana}"
    print("Salida_comida key: ", salida_comida_key)

    # Obtener excepciones
    excepcion = RegistroHorario.query.filter_by(user_id=user.id, fecha=today).first()
    print(f"[DEBUG] Excepción encontrada: {'Sí' if excepcion else 'No'}")

    # Determinar la hora de entrada
    hora_entrar = excepcion.entrada.strftime("%H:%M:%S") if excepcion and excepcion.entrada else (
        getattr(user, entrada_key, None).strftime("%H:%M:%S") if getattr(user, entrada_key, None) else None
    )
    print(f"[DEBUG] Hora de entrar (con excepción si aplica): {hora_entrar}")

        # Determinar la hora de salida
    hora_salida_final = getattr(user, salida_key, None)
    hora_salida_comida = getattr(user, salida_comida_key, None)
    hora_salida=None
    hora_salida_str=None
    hora_salida2=None

    # Imprimir los valores obtenidos para verificar que no son None
    print(f"[DEBUG] hora_salida_final (getattr): {hora_salida_final}")
    print(f"[DEBUG] hora_salida_comida (getattr): {hora_salida_comida}")

    if excepcion:
        print(f"[DEBUG] Excepción encontrada: {excepcion}")

        if now < excepcion.salida:
            print(f"[DEBUG] Hora actual es antes de la hora de salida final: {now} < {excepcion.salida}")
        if excepcion.salida_comida:
            print(f"[DEBUG] Pausa comida registrada: {excepcion.salida_comida}")

        if now < excepcion.salida and excepcion.salida_comida:
            hora_salida = excepcion.salida_comida
            print(f"[DEBUG] Se asigna hora_salida con salida_comida: {hora_salida}")
        elif now >= excepcion.salida and excepcion.salida:
            hora_salida = excepcion.salida
            print(f"[DEBUG] Se asigna hora_salida con salida: {hora_salida}")
    else:
        if hora_salida_final:
            if hora_salida_comida and now < hora_salida_comida:
                hora_salida = hora_salida_comida
                print(f"[DEBUG] Hora actual antes de la salida a comer: {now} < {hora_salida_comida} => Usando salida a comer.")
            else:
                hora_salida = hora_salida_final
                print(f"[DEBUG] Hora actual igual o después de salida a comer => Usando salida final: {hora_salida_final}")


    if hora_salida:
        # Imprimir la hora de salida final antes de convertirla a string
        print(f"[DEBUG] Hora de salida (final): {hora_salida}")

        # Convertir la hora de salida a string y mostrar el resultado
        hora_salida_str = hora_salida.strftime("%H:%M:%S") if hora_salida else None
        print(f"[DEBUG] Hora de salida (formato string): {hora_salida_str}")

        hora_salida2 = datetime.strptime(hora_salida_str[:5], "%H:%M").time()

    print("DEBUG: today =", today)
    bajas = Absence.query.filter(
        Absence.user_id == current_user.id,
        Absence.absence_type == 'Baja laboral'
    ).all()
    print("DEBUG: Bajas laborales totales del usuario:", len(bajas))
    for b in bajas:
        print(f"DEBUG: Baja {b.id} | {b.start_date} - {b.end_date}")

    baja_activa = Absence.query.filter(
        Absence.user_id == current_user.id,
        Absence.absence_type == 'Baja laboral',
        Absence.start_date <= today,
        Absence.end_date >= today
    ).first()

    has_active_baja = bool(baja_activa)
    print("DEBUG: Baja activa hoy?", has_active_baja)
    print("DEBUG: baja_activa =", baja_activa)



        # Eventos de bajas laborales del usuario (siempre visibles)
    sick_leave_events = []
    for baja in bajas:
        sick_leave_events.append({
            'start': baja.start_date.strftime('%Y-%m-%d'),
            'end': (baja.end_date + timedelta(days=1)).strftime('%Y-%m-%d'),
            'title': "Baja laboral",
            'rendering': 'background',
            'color': '#ff8c00',  # NARANJA
            'user_id': current_user.id
        })

    print("[DEBUG] Eventos de bajas:", sick_leave_events)


    return render_template('user_page.html', active_menu='user_page',
                           sick_leave_events=sick_leave_events,
                           absences_count_mensual=absences_count_mensual,
                           has_active_baja=has_active_baja,
                           absences_count_anual=absences_count_anual,
                           notificaciones=notificaciones,
                           current_day_vacation_names=current_day_vacation_names,
                           has_pending_requests=1 if pending_requests else 0,
                           current_user_email=current_user_email,
                           events=events,
                           non_workdays=non_workdays,
                           users=users, user=user, useri=useri, es_entrada=es_entrada,
                           hora_entrada=hora_entrada, pausa_activa=pausa_activa, empresa=empresa,
                           hora_entrar=hora_entrar, hora_salida=hora_salida_str,now=now,hora_salida_modal=hora_salida2)





@app.route('/calcular_diferencia', methods=['POST'])
def calcular_diferencia():
    datos = request.get_json()
    print(f"📩 Datos recibidos: {datos}")  # Debug: ver datos recibidos

    hora_actual = datos['hora_actual']
    hora_salida = datos['hora_salida']
    print(f"🕒 Hora actual recibida: {hora_actual}")
    print(f"🏁 Hora salida recibida: {hora_salida}")

    # Buscar el último registro del usuario
    ultimo_registro = Registro.query.filter_by(user_id=current_user.id).order_by(Registro.entrada.desc()).first()
    print(f"🔍 Último registro encontrado: {ultimo_registro}")  # Debug: ver si se encuentra un registro

    if ultimo_registro:
        print(f"📅 Última entrada registrada: {ultimo_registro.entrada}")
        print(f"🚪 Última salida registrada: {ultimo_registro.salida}")

    if ultimo_registro and not ultimo_registro.salida:
        fecha_entrada = ultimo_registro.entrada.date()  # Obtener solo la fecha
        fecha_hoy = datetime.now().date()  # Fecha actual sin hora
        print(f"📆 Fecha de entrada: {fecha_entrada}, 📅 Fecha actual: {fecha_hoy}")

        if fecha_entrada < fecha_hoy:  # Si la entrada es de un día anterior
            print("⚠️ El usuario olvidó registrar salida el día anterior, devolviendo 5 horas (300 minutos)")
            return jsonify({'diferencia_minutos': 5 * 60})  # Devuelve 5 horas en minutos

    # Calcular la diferencia normal si no aplica la condición anterior
    diferencia_minutos = calcular_diferencia_en_minutos(hora_actual, hora_salida)
    print(f"✅ Diferencia calculada en minutos: {diferencia_minutos} minutos")

    return jsonify({'diferencia_minutos': diferencia_minutos})

def calcular_diferencia_en_minutos(hora_actual, hora_salida):
    formato = "%H:%M:%S"
    try:
        hora_actual_obj = datetime.strptime(hora_actual, formato)
        hora_salida_obj = datetime.strptime(hora_salida, formato)
        print(f"🔢 Conversión de horas: 🕒 {hora_actual_obj} - 🏁 {hora_salida_obj}")

        diferencia_segundos = (hora_actual_obj - hora_salida_obj).total_seconds()
        diferencia_minutos = diferencia_segundos / 60  # Convertir a minutos
        print(f"🕐 Diferencia en segundos: {diferencia_segundos}, en minutos: {diferencia_minutos} minutos")

        return diferencia_minutos
    except Exception as e:
        print(f"❌ Error al calcular diferencia de tiempo: {e}")  # Captura errores
        return 0  # Retorna 0 en caso de error

def obtener_fechas_ultimas_semanas():
    hoy = datetime.today()
    semanas = []
    for i in range(4):
        # Calcular el inicio y el fin de la semana (lunes a domingo)
        fin_semana = hoy - timedelta(days=hoy.weekday()) - timedelta(weeks=i)
        inicio_semana = fin_semana - timedelta(days=6)

        # Mantener el formato completo (como objetos datetime) para las comparaciones
        semanas.append((inicio_semana, fin_semana))
    return semanas



def obtener_horas_totales_por_semana():
    hoy = datetime.today()
    semanas = obtener_fechas_ultimas_semanas()  # Obtener las fechas de las últimas 4 semanas

    horas_totales_por_semana = [timedelta() for _ in range(4)]  # Inicializar horas totales por semana

    # Obtener todos los registros de todos los usuarios
    registros = Registro.query.all()

    for registro in registros:
        if registro.salida:  # Asegurarse de que haya salida
            duracion = registro.duracion_neta  # Duración sin pausas
            for i, (inicio, fin) in enumerate(semanas):
                # Verificar si el registro está dentro de la semana
                if inicio <= registro.entrada <= fin and registro.salida:
                    horas_totales_por_semana[i] += duracion

    # Convertir las horas totales trabajadas en cada semana a horas enteras
    horas_totales = [int(horas.total_seconds() // 3600) for horas in horas_totales_por_semana]

    # Formatear las fechas solo con mes y día para la visualización (en el frontend)
    semanas_formateadas = [(inicio.strftime('%d-%m'), fin.strftime('%d-%m')) for inicio, fin in semanas]

    return horas_totales, semanas_formateadas


def obtener_duracion_media_pausas_ultimos_7_dias():
    hoy = datetime.today()

    # Crear listas para almacenar la duración de las pausas y la cantidad de pausas por día
    duracion_pausas_por_dia = [timedelta() for _ in range(7)]
    cantidad_pausas_por_dia = [0 for _ in range(7)]  # Para contar las pausas cada día

    # Obtener todos los registros de todos los usuarios
    registros = Registro.query.all()

    for registro in registros:
        # Iterar sobre las pausas de este registro
        for pausa in registro.pausas_registro:
            if pausa.inicio and pausa.fin:  # Asegurarse de que la pausa tenga inicio y fin
                # Calcular la duración de la pausa
                duracion_pausa = pausa.duracion  # Esto ya devuelve un timedelta

                # Comprobar en qué día cae esta pausa
                for i in range(7):
                    dia = hoy - timedelta(days=i)
                    # Verificar si la pausa cae dentro del día (en el mismo día de la entrada del registro)
                    if dia.date() == pausa.inicio.date():
                        duracion_pausas_por_dia[i] += duracion_pausa
                        cantidad_pausas_por_dia[i] += 1
                        break

    # Calcular la duración media de las pausas por día (en minutos)
    duracion_media_pausas_por_dia = []
    for i in range(7):
        if cantidad_pausas_por_dia[i] > 0:
            # Calcular el promedio en minutos
            duracion_media = duracion_pausas_por_dia[i].total_seconds() / 60 / cantidad_pausas_por_dia[i]
        else:
            duracion_media = 0  # Si no hay pausas, la duración media es 0
        duracion_media_pausas_por_dia.append(duracion_media)

    # Crear las fechas de los últimos 7 días
    fechas = [(hoy - timedelta(days=i)).strftime('%d-%m') for i in range(7)]

    return duracion_media_pausas_por_dia, fechas




# Rutas de usuario
@app.route('/user_admin')

def user_page_admin():
    return render_template('user_page_admin.html')

from sqlalchemy import or_, and_



from sqlalchemy import extract

@app.route('/admin', methods=['GET'])
@admin_required
def admin_page():
    print("ENTRAMOS EN ADMIN PAGE")
    actualizar_fechas_verano()
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_email = current_user.email if current_user.is_authenticated else None

    from sqlalchemy.orm import joinedload

    if current_user.has_role('admin'):
        pending_requests_query = VacationRequest.query.filter(
            VacationRequest.status == 'PENDIENTE'
        )
        accepted_requests_query = VacationRequest.query.filter(
            VacationRequest.is_accepted == True
        )
    else:
        pending_requests_query = VacationRequest.query.join(User).filter(
            VacationRequest.status == 'PENDIENTE',
            User.empresa_id == current_user.empresa_id
        )
        accepted_requests_query = VacationRequest.query.join(User).filter(
            VacationRequest.is_accepted == True,
            User.empresa_id == current_user.empresa_id
        )


    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        pending_requests_query = pending_requests_query.join(User).filter(User.Centro_trabajo == 'CEEI Guadalajara')
        accepted_requests_query = accepted_requests_query.join(User).filter(User.Centro_trabajo == 'CEEI Guadalajara')
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        pending_requests_query = pending_requests_query.join(User).filter(User.Centro_trabajo == 'CEOE Formación')
        accepted_requests_query = accepted_requests_query.join(User).filter(User.Centro_trabajo == 'CEOE Formación')

    pending_requests = pending_requests_query.all()
    accepted_requests = accepted_requests_query.all()
    users = User.query.filter(
        (User.email != 'admin@gestionaempresa.es') &
        (User.email != 'inspector@gestionaempresa.es')
    ).all()
    empresa = Empresa.query.first()

    # Obtener el día de la semana actual (0 = lunes, 6 = domingo)
    today_weekday = datetime.today().weekday()

        # Filtrar usuarios según el rol
    if current_user.has_role('admin'):
        users = User.query.all()
    else:
        users = User.query.filter_by(empresa_id=current_user.empresa_id).all()

    # Filtrar usuarios que trabajan hoy
    users_working_today = []
    for user in users:
        if today_weekday == 0 and user.trabaja_lunes:
            users_working_today.append(user)
        elif today_weekday == 1 and user.trabaja_martes:
            users_working_today.append(user)
        elif today_weekday == 2 and user.trabaja_miercoles:
            users_working_today.append(user)
        elif today_weekday == 3 and user.trabaja_jueves:
            users_working_today.append(user)
        elif today_weekday == 4 and user.trabaja_viernes:
            users_working_today.append(user)
        elif today_weekday == 5 and user.trabaja_sabado:
            users_working_today.append(user)
        elif today_weekday == 6 and user.trabaja_domingo:
            users_working_today.append(user)

    # Crear una lista con usuarios y si han fichado hoy
    today_date = datetime.today().date()
    users_status = []
    for user in users_working_today:
        registros = Registro.query.filter(
            Registro.user_id == user.id,
            extract('day', Registro.entrada) == today_date.day,
            extract('month', Registro.entrada) == today_date.month,
            extract('year', Registro.entrada) == today_date.year
        ).all()

        users_status.append({'user': user, 'has_checked_in': bool(registros)})

    # Eventos aceptados (ya filtrados antes por empresa si no es admin)
    events = []
    for request in accepted_requests:
        start_date = request.start_date.strftime('%Y-%m-%d')
        end_date = (request.end_date + timedelta(days=1)).strftime('%Y-%m-%d')
        events.append({
            'start': start_date,
            'end': end_date,
            'rendering': 'background',
            'color': '#6173D6',
            'title': f"{request.user_name} {request.user_apellido}"
        })

    # Eventos pendientes (ya filtrados antes por empresa si no es admin)
    pending_events = []
    for request in pending_requests:
        start_date = request.start_date.strftime('%Y-%m-%d')
        end_date = (request.end_date + timedelta(days=1)).strftime('%Y-%m-%d')
        pending_events.append({
            'start': start_date,
            'end': end_date,
            'rendering': '',
            'color': 'orange',
            'title': ''
        })

    # Obtener el año actual
    current_year = datetime.now().year

    # Obtener los días festivos de la base de datos
    festivos = Festivo.query.all()

    # Convertir la información de días festivos en formato de texto ("YYYY-MM-DD")
    non_workdays = [f"{festivo.año}-{festivo.mes:02d}-{festivo.dia:02d}" for festivo in festivos]

    today = datetime.today().date()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = (today.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    absences_count_mensual = Absence.query.filter(
        Absence.end_date >= first_day_of_month,
        Absence.end_date <= last_day_of_month
    ).count()

    first_day_of_year = today.replace(month=1, day=1)
    last_day_of_year = today.replace(month=12, day=31)
    absences_count_anual = Absence.query.filter(
        Absence.end_date >= first_day_of_year,
        Absence.end_date <= last_day_of_year
    ).count()

    current_day_vacations = VacationRequest.query.filter_by(is_accepted=True).filter(
        extract('year', VacationRequest.start_date) == today.year,
        extract('month', VacationRequest.start_date) == today.month,
        extract('day', VacationRequest.start_date) == today.day
    ).all()

    current_day_vacation_names = [
        f"{request.user_name} {request.user_apellido}"
        for request in current_day_vacations
    ]

    user_history = VacationRequest.query.all()

    today = datetime.today().date()
    mes = today.month
    año = today.year

    def calcular_horas_trabajadas(usuario, mes, año):
        registros = Registro.query.filter(
            Registro.user_id == usuario.id,
            extract('month', Registro.entrada) == mes,
            extract('year', Registro.entrada) == año
        ).all()

        total_horas = timedelta()

        for registro in registros:
            if registro.duracion_neta:
                total_horas += registro.duracion_neta

        horas_trabajadas = total_horas.total_seconds() // 3600
        return int(horas_trabajadas)

    # Obtener empresa del usuario o la primera si no tiene
    empresa = current_user.empresa if current_user.empresa else Empresa.query.first()

    # Obtener IDs de los usuarios de esa empresa
    user_ids = [u.id for u in User.query.filter_by(empresa_id=empresa.id).all()]

    # Filtrar los registros por usuarios de esa empresa
    # Filtrar los usuarios válidos
    users_validos = User.query.filter(
        (User.email != 'admin@gestionaempresa.es') &
        (User.email != 'inspector@gestionaempresa.es') &
        (User.email != 'gestor@gestionaempresa.es') &
        (User.email != 'encargado@gestionaempresa.es') &
        (User.approved == True)
    ).all()

    # Obtener los IDs válidos
    user_ids = [u.id for u in users_validos]

    # Buscar los últimos registros solo de esos usuarios
    ultimos_registros = Registro.query.filter(
        Registro.user_id.in_(user_ids)
    ).order_by(Registro.entrada.desc()).limit(5).all()

    ultimos_vacation_requests = VacationRequest.query.filter(VacationRequest.user_id.in_(user_ids)).order_by(VacationRequest.dia_solicitud.desc()).limit(5).all()
    ultimos_absences = Absence.query.filter(Absence.user_id.in_(user_ids)).order_by(Absence.end_date.desc()).limit(5).all()

    return render_template('admin_page.html',
                           absences_count_mensual=absences_count_mensual,
                           absences_count_anual=absences_count_anual,
                           empresa=empresa,
                           current_day_vacation_names=current_day_vacation_names,
                           has_pending_requests=1 if pending_requests else 0,
                           current_user_email=current_user_email,
                           events=events,
                           pending_events=pending_events,
                           non_workdays=non_workdays,
                           users_status=users_status,  # Pasamos la lista de usuarios con estado de fichaje
                           user_history=user_history,
                           notificaciones=notificaciones,
                           current_user=current_user,
                           ultimos_registros=ultimos_registros,
                           ultimos_vacation_requests=ultimos_vacation_requests,
                           ultimos_absences=ultimos_absences,
                           active_menu='admin_page')




# Lista global de días de la semana que no son laborables
weekly_non_workdays = [
    5,  # Viernes
    6,  # Sábado
    # Agrega más días según sea necesario
]

@app.route('/get_all_vacation_dates', methods=['GET'])
def get_all_vacation_dates():
    # Obtén el correo del usuario actual
    current_user_email = current_user.email if current_user.is_authenticated else None

    # Inicializa las listas de fechas de vacaciones para este y el año pasado
    vacation_dates_this_year = []
    vacation_dates_last_year = []

    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        # Filtra las fechas de vacaciones aceptadas por centro de trabajo 'CEEI Guadalajara'
        all_vacation_dates = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEEI Guadalajara',
            VacationRequest.is_accepted == True
        ).all()
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        # Filtra las fechas de vacaciones aceptadas por centro de trabajo 'CEOE Formación'
        all_vacation_dates = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEOE Formación',
            VacationRequest.is_accepted == True
        ).all()
    else:
        # Si el usuario no tiene un correo específico o no está autenticado, muestra todas las fechas de vacaciones aceptadas
        all_vacation_dates = VacationRequest.query.filter_by(is_accepted=True).all()

    # Formatea las fechas como listas de diccionarios
    for request in all_vacation_dates:
        start_date = request.start_date
        end_date = request.end_date

        # Calcula la cantidad de días de vacaciones utilizados el año pasado y este año
        days_used_last_year = min(request.dias_utilizados_pasado, (end_date - start_date).days + 1)
        days_used_this_year = max((end_date - start_date).days + 1 - days_used_last_year, 0)

        # Agregar todas las fechas entre start_date y end_date (inclusive)
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() not in weekly_non_workdays:
                if days_used_last_year > 0:
                    vacation_dates_last_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_last_year -= 1
                elif days_used_this_year > 0:
                    vacation_dates_this_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_this_year -= 1
            current_date += timedelta(days=1)

    return jsonify({
        'vacation_dates_this_year': vacation_dates_this_year,
        'vacation_dates_last_year': vacation_dates_last_year
    })


from flask import jsonify

@app.route('/get_me_all_vacation_dates', methods=['GET'])
def get_me_all_vacation_dates():
    # Obtén el correo del usuario actual
    current_user_email = current_user.email if current_user.is_authenticated else None
    user_id = current_user.id
    # Inicializa las listas de fechas de vacaciones para este y el año pasado
    vacation_dates_this_year = []
    vacation_dates_last_year = []

    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        # Filtra las fechas de vacaciones aceptadas por centro de trabajo 'CEEI Guadalajara'
        all_vacation_dates = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEEI Guadalajara',
            VacationRequest.is_accepted == True
        ).all()
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        # Filtra las fechas de vacaciones aceptadas por centro de trabajo 'CEOE Formación'
        all_vacation_dates = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEOE Formación',
            VacationRequest.is_accepted == True
        ).all()
    else:
        # Si el usuario no tiene un correo específico o no está autenticado, muestra todas las fechas de vacaciones aceptadas
        all_vacation_dates = VacationRequest.query.filter_by(user_id=user_id,is_accepted=True).all()

    # Formatea las fechas como listas de diccionarios
    for request in all_vacation_dates:
        start_date = request.start_date
        end_date = request.end_date

        # Calcula la cantidad de días de vacaciones utilizados el año pasado y este año
        days_used_last_year = min(request.dias_utilizados_pasado, (end_date - start_date).days + 1)
        days_used_this_year = max((end_date - start_date).days + 1 - days_used_last_year, 0)

        # Agregar todas las fechas entre start_date y end_date (inclusive)
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() not in weekly_non_workdays:
                if days_used_last_year > 0:
                    vacation_dates_last_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_last_year -= 1
                elif days_used_this_year > 0:
                    vacation_dates_this_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_this_year -= 1
            current_date += timedelta(days=1)

    return jsonify({
        'vacation_dates_this_year': vacation_dates_this_year,
        'vacation_dates_last_year': vacation_dates_last_year
    })


from flask import jsonify



@app.route('/get_me_vacation_dates_pending', methods=['GET'])
def get_me_vacation_dates_pending():
    # Obtén el correo del usuario actual
    current_user_email = current_user.email if current_user.is_authenticated else None
    user_id = current_user.id
    # Inicializa la lista de fechas de vacaciones pendientes para los años pasado y actual
    vacation_dates_pending_last_year = []
    vacation_dates_pending_this_year = []

    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        # Filtra las fechas de vacaciones pendientes por centro de trabajo 'CEEI Guadalajara'
        user_vacation_dates_pending = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEEI Guadalajara',

            VacationRequest.status == 'PENDIENTE'
        ).all()
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        # Filtra las fechas de vacaciones pendientes por centro de trabajo 'CEOE Formación'
        user_vacation_dates_pending = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEOE Formación',

            VacationRequest.status == 'PENDIENTE'
        ).all()
    else:
        # Si el usuario no tiene un correo específico, muestra todas las fechas de vacaciones pendientes


        # Obtén las fechas de vacaciones pendientes para el usuario actual
        user_vacation_dates_pending = VacationRequest.query.filter_by(user_id=user_id, status='PENDIENTE').all()

    # Obtén el año actual y el año pasado
    current_year = datetime.now().year
    last_year = current_year - 1

    # Formatea las fechas como lista de diccionarios
    for request in user_vacation_dates_pending:
        start_date = request.start_date
        end_date = request.end_date

        # Encuentra al usuario que hizo la solicitud
        user = User.query.get(request.user_id)

        # Chequea los días restantes del empleado para el año pasado
        remaining_days_last_year = user.dias_restantes_pasado

        # Calcula los días utilizados en cada año
        days_used_last_year = min(remaining_days_last_year, (end_date - start_date).days + 1)
        days_used_this_year = max((end_date - start_date).days + 1 - days_used_last_year, 0)

        # Agregar fechas de vacaciones pendientes del año pasado
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() not in weekly_non_workdays:
                if days_used_last_year > 0:
                    vacation_dates_pending_last_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_last_year -= 1
                elif days_used_this_year > 0:
                    vacation_dates_pending_this_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_this_year -= 1
            current_date += timedelta(days=1)

    return jsonify({
        'vacation_dates_pending_last_year': vacation_dates_pending_last_year,
        'vacation_dates_pending_this_year': vacation_dates_pending_this_year
    })

@app.route('/get_vacation_dates_pending', methods=['GET'])
def get_vacation_dates_pending():
    # Obtén el correo del usuario actual
    current_user_email = current_user.email if current_user.is_authenticated else None

    # Inicializa la lista de fechas de vacaciones pendientes para los años pasado y actual
    vacation_dates_pending_last_year = []
    vacation_dates_pending_this_year = []

    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        # Filtra las fechas de vacaciones pendientes por centro de trabajo 'CEEI Guadalajara'
        user_vacation_dates_pending = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEEI Guadalajara',
            VacationRequest.status == 'PENDIENTE'
        ).all()
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        # Filtra las fechas de vacaciones pendientes por centro de trabajo 'CEOE Formación'
        user_vacation_dates_pending = VacationRequest.query.join(User).filter(
            User.Centro_trabajo == 'CEOE Formación',
            VacationRequest.status == 'PENDIENTE'
        ).all()
    else:
        # Si el usuario no tiene un correo específico, muestra todas las fechas de vacaciones pendientes
        user_vacation_dates_pending = VacationRequest.query.filter_by(status='PENDIENTE').all()

    # Obtén el año actual y el año pasado
    current_year = datetime.now().year
    last_year = current_year - 1

    # Formatea las fechas como lista de diccionarios
    for request in user_vacation_dates_pending:
        start_date = request.start_date
        end_date = request.end_date

        # Encuentra al usuario que hizo la solicitud
        user = User.query.get(request.user_id)

        # Chequea los días restantes del empleado para el año pasado
        remaining_days_last_year = user.dias_restantes_pasado

        # Calcula los días utilizados en cada año
        days_used_last_year = min(remaining_days_last_year, (end_date - start_date).days + 1)
        days_used_this_year = max((end_date - start_date).days + 1 - days_used_last_year, 0)

        # Agregar fechas de vacaciones pendientes del año pasado
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() not in weekly_non_workdays:
                if days_used_last_year > 0:
                    vacation_dates_pending_last_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_last_year -= 1
                elif days_used_this_year > 0:
                    vacation_dates_pending_this_year.append({'date': current_date.strftime('%Y-%m-%d')})
                    days_used_this_year -= 1
            current_date += timedelta(days=1)

    return jsonify({
        'vacation_dates_pending_last_year': vacation_dates_pending_last_year,
        'vacation_dates_pending_this_year': vacation_dates_pending_this_year
    })



@app.route('/get_legend_data', methods=['GET'])
def get_legend_data():
    # Lógica para obtener los datos de la leyenda (festivos, vacaciones, etc.) para el mes actual
    # ...

    # Obtén el correo del usuario actual
    current_user_email = current_user.email if current_user.is_authenticated else None

    # Obtén el centro de trabajo del usuario actual
    current_user_workplace = current_user.Centro_trabajo if current_user.is_authenticated else None

    # Obtén el mes y año actual desde la URL o desde la solicitud POST
    year = (request.args.get('year'))
    month = (request.args.get('month'))

    if year is None:
        year = date.today().year
    else:
        year = int(year)

    if month is None:
        month = date.today().month
    else:
        month = int(month)
    month += 1

    # Obtén el primer día del mes actual o del mes seleccionado
    first_day_of_month = date(year, month, 1)

    # Filtrar las fechas de vacaciones pendientes y aceptadas en función del correo y el centro de trabajo
    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        # Filtra por centro de trabajo si el usuario es emprendedores@ceeiguadalajara.es
        vacation_dates_pending = VacationRequest.query.filter_by(status='PENDIENTE').join(User).filter(User.Centro_trabajo == 'CEEI Guadalajara').filter(
            extract('year', VacationRequest.start_date) == year,
            extract('month', VacationRequest.start_date) == month
        ).all()

        all_vacation_dates = VacationRequest.query.filter_by(is_accepted=True).join(User).filter(User.Centro_trabajo == 'CEEI Guadalajara').filter(
            extract('year', VacationRequest.start_date) == year,
            extract('month', VacationRequest.start_date) == month
        ).all()

    elif current_user_email == 'formacion@ceoeguadalajara.es':
        # Filtra por centro de trabajo si el usuario es formacion@ceoeguadalajara.es
        vacation_dates_pending = VacationRequest.query.filter_by(status='PENDIENTE').join(User).filter(User.Centro_trabajo == 'CEOE Formación').filter(
            extract('year', VacationRequest.start_date) == year,
            extract('month', VacationRequest.start_date) == month
        ).all()

        all_vacation_dates = VacationRequest.query.filter_by(is_accepted=True).join(User).filter(User.Centro_trabajo == 'CEOE Formación').filter(
            extract('year', VacationRequest.start_date) == year,
            extract('month', VacationRequest.start_date) == month
        ).all()

    else:
        # Si el usuario no está autenticado o no coincide con ninguno de los correos específicos, muestra todas las fechas de vacaciones
        vacation_dates_pending = VacationRequest.query.filter_by(status='PENDIENTE').filter(
            extract('year', VacationRequest.start_date) == year,
            extract('month', VacationRequest.start_date) == month
        ).all()

        all_vacation_dates = VacationRequest.query.filter_by(is_accepted=True).filter(
            extract('year', VacationRequest.start_date) == year,
            extract('month', VacationRequest.start_date) == month
        ).all()

    # Filtrar los festivos en función del mes
    nonWorkdays = [day for day in non_workdays if day.month == month]

    # Formatear las fechas de vacaciones pendientes
    pending_vacations = [
        {
            'user_name': request.user_name,
            'user_apellido': request.user_apellido,
            'start_date': request.start_date.strftime('%Y-%m-%d'),
            'end_date': request.end_date.strftime('%Y-%m-%d')
        }
        for request in vacation_dates_pending
    ]

    # Formatear las fechas de vacaciones aceptadas
    accepted_vacations = [
        {
            'user_name': request.user_name,
            'user_apellido': request.user_apellido,
            'start_date': request.start_date.strftime('%Y-%m-%d'),
            'end_date': request.end_date.strftime('%Y-%m-%d')
        }
        for request in all_vacation_dates
    ]

    legend_data = {
        'nonWorkdays': nonWorkdays,
        'acceptedVacations': accepted_vacations,
        'pendingVacations': pending_vacations,
    }

    return jsonify(legend_data)


from flask import redirect, url_for
from flask_security import logout_user, current_user


@app.route('/cerrar_sesion')
@login_required
def cerrar_sesion():
    # Obtener el email del usuario actual antes de cerrar la sesión
    email = current_user.email
    print("correo del usuario: ", email)

    # Registrar el cierre de sesión en el log
    log_logout_event(email)

    # Cerrar sesión del usuario
    logout_user()

    # Mostrar mensaje de éxito
    flash('Has cerrado sesión con éxito.', 'success')

    # Redirigir al login
    return redirect(url_for('login'))


def get_admin_users():
    base_query = User.query.join(roles_users).join(Role)

    admins = base_query.filter(Role.name == 'admin').all()

    gestores = base_query.filter(
        Role.name == 'gestor',
        User.empresa_id == current_user.empresa_id,
        User.vacaciones.is_(True)
    ).all()

    encargados = base_query.filter(
        Role.name == 'encargado',
        User.empresa_id == current_user.empresa_id
    ).all()

    # Unimos y eliminamos duplicados
    return list({user.id: user for user in admins + gestores + encargados}.values())

def get_admin_emails():
    users = get_admin_users()
    return [user.email for user in users]





@app.route('/user/request_vacation', methods=['GET', 'POST'])
def request_vacation():
    print("\n🔍 Iniciando solicitud de vacaciones...")

    current_user_centro = current_user.Centro_trabajo if current_user.is_authenticated else None
    vacation_history = VacationHistory.query.filter_by(user_id=current_user.id).all()
    remaining_days = current_user.dias_restantes_este
    remaining_days_pasado = current_user.dias_restantes_pasado
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    form = MyForm()

    today = datetime.today().date()

    if form.validate_on_submit():
        start_date = form.start_date.data
        end_date = form.end_date.data

        if start_date < today:
            flash("No puedes solicitar vacaciones en fechas pasadas.", "danger")
            return redirect(url_for('request_vacation'))

        if end_date < start_date:
            flash("La fecha de fin debe ser igual o posterior a la fecha de inicio.", "danger")
            return redirect(url_for('request_vacation'))

        # Contar todos los días como laborables
        total_days = (end_date - start_date).days + 1

        available_days = remaining_days + remaining_days_pasado
        if total_days > available_days:
            flash("No dispones de suficientes días de vacaciones.", "danger")
            return redirect(url_for('request_vacation'))

        new_request = VacationRequest(
            start_date=start_date,
            end_date=end_date,
            user_id=current_user.id,
            user_email=current_user.email,
            user_name=current_user.nombre,
            user_apellido=current_user.primer_apellido,
            dias_vacas=total_days,
            dias_utilizados_pasado=0.0,
            dias_utilizados_este=0.0,
            dia_solicitud=today,
            centro=current_user.Centro_trabajo
        )

        new_history_entry = VacationHistory(
            user_id=current_user.id,
            start_date=start_date,
            end_date=end_date,
            status='PENDIENTE',
            is_accepted=False,
            dia_solicitud=today
        )

        db.session.add(new_request)
        db.session.add(new_history_entry)
        db.session.commit()

        # Notificaciones a admins
        admin_users = get_admin_users()
        for admin in admin_users:
            notificacion = Notificacion(
                usuario_id=admin.id,
                concepto=f"{current_user.nombre} {current_user.primer_apellido} ha solicitado vacaciones.",
                fecha_hora=datetime.now()
            )
            db.session.add(notificacion)
        db.session.commit()

        # Intentar enviar emails sin romper la ejecución si falla
        recipients = get_admin_emails()
        for email in recipients:
            try:
                send_email_to_user(email, 'Solicitud de vacaciones', f'{current_user.nombre} {current_user.primer_apellido} ha solicitado vacaciones. Revisa la aplicación.')
            except Exception as e:
                print(f"⚠️ Error enviando email a {email}: {e}")

        flash("Solicitud de vacaciones enviada exitosamente.", "success")
        return redirect(url_for('request_vacation'))

    return render_template(
        'request_vacation.html',
        form=form,
        remaining_days=remaining_days,
        remaining_days_pasado=remaining_days_pasado,
        notificaciones=notificaciones,
        vacation_history=vacation_history,
        current_user_centro=current_user_centro,
        active_menu='request_vacation'
    )


@app.route('/admin/assign_vacation', methods=['GET', 'POST'])
@admin_required
def admin_assign_vacation():
    print("\n🔧 Iniciando asignación de vacaciones por parte del ADMIN...")

    form = AdminAssignVacationForm()
    users = User.query.order_by(User.nombre.asc()).all()
    today = datetime.today().date()
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).all()
    # Correos que NO deben aparecer en el select
    excluded_emails = [
        "admin@gestionaempresa.es",
        "gestor@gestionaempresa.es",
        "inspector@gestionaempresa.es",
        "encargado@gestionaempresa.es"
    ]

    # Filtrar usuarios por nombre y excluir emails concretos
    users = User.query \
        .filter(~User.email.in_(excluded_emails)) \
        .order_by(User.nombre.asc()) \
        .all()


    # Llenar choices del select de usuarios
    form.user_id.choices = [(u.id, f"{u.nombre} {u.primer_apellido}") for u in users]

    if form.validate_on_submit():
        user_id = form.user_id.data
        start_date = form.start_date.data
        end_date = form.end_date.data
        user = User.query.get(user_id)

        if start_date < today:
            flash("No puedes asignar vacaciones en fechas pasadas.", "danger")
            return redirect(url_for('admin_assign_vacation'))

        if end_date < start_date:
            flash("La fecha de fin debe ser igual o posterior al inicio.", "danger")
            return redirect(url_for('admin_assign_vacation'))

        # Contar todos los días como laborables
        total_days = (end_date - start_date).days + 1

        available_days = user.dias_restantes_este
        if total_days > available_days:
            flash("El usuario no tiene suficientes días disponibles.", "danger")
            return redirect(url_for('admin_assign_vacation'))
        total_days = (end_date - start_date).days + 1  # contar todos los días del periodo

        if total_days > user.dias_restantes_este:
            flash("El usuario no tiene suficientes días disponibles.", "danger")
            return redirect(url_for('admin_assign_vacation'))

        # Restar los días de vacaciones del año actual
        user.dias_restantes_este -= total_days

        # Crear la solicitud
        vacation_request = VacationRequest(
            start_date=start_date,
            end_date=end_date,
            user_id=user.id,
            user_email=user.email,
            user_name=user.nombre,
            user_apellido=user.primer_apellido,
            dias_vacas=total_days,
            dias_utilizados_este=total_days,  # registrar los días utilizados
            dia_solicitud=today,
            centro=user.Centro_trabajo,
            status="ACEPTADA",
            is_accepted=1,
        )

        vacation_history = VacationHistory(
            user_id=user.id,
            start_date=start_date,
            end_date=end_date,
            status="ACEPTADO",
            is_accepted=True,
            dia_solicitud=today
        )

        db.session.add(vacation_request)
        db.session.add(vacation_history)

        # Notificación interna
        noti = Notificacion(
            usuario_id=user.id,
            concepto=f"El administrador te ha asignado vacaciones del {start_date} al {end_date}. Total: {total_days} días.",
            fecha_hora=datetime.now()
        )
        db.session.add(noti)

        # Enviar email sin romper la función si falla
        try:
            send_email_to_user(
                user.email,
                "Vacaciones asignadas",
                f"El administrador te ha asignado vacaciones del {start_date} al {end_date}, usando un total de {total_days} días."
            )
        except Exception as e:
            print(f"⚠️ No se pudo enviar email a {user.email}: {e}")

        db.session.commit()
        flash("Vacaciones asignadas correctamente.", "success")
        return redirect(url_for('admin_assign_vacation'))

    return render_template(
        "admin_assign_vacation.html",
        form=form,
        users=users,
        notificaciones=notificaciones,
        active_menu="admin_assign_vacation"
    )









def handle_successful_request(start_date, end_date, total_non_workdays_count):
    """Maneja una solicitud de vacaciones exitosa."""
    current_user.remaining_vacation_days -= total_non_workdays_count

    new_request = VacationRequest(
        start_date=start_date,
        end_date=end_date,
        user_id=current_user.id,
        user_email=current_user.email
    )

    new_history_entry = VacationHistory(
        user_id=current_user.id,
        start_date=start_date,
        end_date=end_date,
        status='PENDIENTE',
        is_accepted=False,
    )

    db.session.add(new_request)
    db.session.add(new_history_entry)
    db.session.commit()



from flask import request

@app.route('/admin/requests')
@admin_required
def admin_requests():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_email = current_user.email if current_user.is_authenticated else None
    page = request.args.get('page', default=1, type=int)

    # Base de las queries
    pending_requests_query = VacationRequest.query.filter(VacationRequest.status == 'PENDIENTE').order_by(VacationRequest.id.desc())
    accepted_requests_query = VacationRequest.query.filter(VacationRequest.is_accepted == True).order_by(VacationRequest.id.desc())
    rejected_requests_query = VacationRequest.query.filter(VacationRequest.is_accepted == False, VacationRequest.status == 'RECHAZADA').order_by(VacationRequest.id.desc())
    cancel_requests_query = VacationRequest.query.filter(VacationRequest.status == 'Cancelada').order_by(VacationRequest.id.desc())

    # Filtrar por roles y empresa, si el usuario tiene rol 'gestor' o 'responsable'
    if current_user.has_role('gestor') or current_user.has_role('encargado'):
        pending_requests_query = pending_requests_query.join(User).filter(User.empresa_id == current_user.empresa_id)
        accepted_requests_query = accepted_requests_query.join(User).filter(User.empresa_id == current_user.empresa_id)
        rejected_requests_query = rejected_requests_query.join(User).filter(User.empresa_id == current_user.empresa_id)
        cancel_requests_query = cancel_requests_query.join(User).filter(User.empresa_id == current_user.empresa_id)

    pending_requests = pending_requests_query.all()
    accepted_requests = accepted_requests_query.paginate(page=page, per_page=10, error_out=False)
    rejected_requests = rejected_requests_query.all()
    cancel_requests = cancel_requests_query.all()

    for request_list in [pending_requests, accepted_requests.items, rejected_requests, cancel_requests]:
        for req in request_list:
            req.start_date_str = req.start_date.strftime('%d/%m/%Y')
            req.end_date_str = req.end_date.strftime('%d/%m/%Y')
            req.dia_solicitud_str = req.dia_solicitud.strftime('%d/%m/%Y')
            req.dia_decision_str = req.dia_decision.strftime('%d/%m/%Y') if req.dia_decision else None


    return render_template(
        'admin_request.html',
        notificaciones=notificaciones,
        pending_requests=pending_requests,
        accepted_requests=accepted_requests,
        rejected_requests=rejected_requests,
        cancel_requests=cancel_requests,
        current_user_email=current_user_email,
        active_menu='combo_vacacione'
    )

@app.route('/admin/requests/accepted')
@admin_required
def paginate_accepted_requests():
    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1

    accepted_requests_query = VacationRequest.query.filter(
        VacationRequest.is_accepted == True
    ).order_by(VacationRequest.id.desc())

    # Filtrar por roles y empresa, si el usuario tiene rol 'gestor' o 'responsable'
    if current_user.has_role('gestor') or current_user.has_role('encargado'):
        accepted_requests_query = accepted_requests_query.join(User).filter(User.empresa_id == current_user.empresa_id)

    accepted_requests = accepted_requests_query.paginate(page=page, per_page=10, error_out=False)

    for req in accepted_requests.items:
        req.start_date = req.start_date.strftime('%d/%m/%Y') if req.start_date else None
        req.end_date = req.end_date.strftime('%d/%m/%Y') if req.end_date else None
        req.dia_solicitud = req.dia_solicitud.strftime('%d/%m/%Y') if req.dia_solicitud else None
        if req.dia_decision:
            req.dia_decision = req.dia_decision.strftime('%d/%m/%Y')

    return render_template('partials/accepted_requests_table.html', accepted_requests=accepted_requests)


import pytz

def notificar_usuario(user, titulo, mensaje, enviar_email=True):
    # Obtener hora actual en zona horaria de Madrid
    madrid_tz = pytz_timezone('Europe/Madrid')
    ahora_madrid = datetime.now(madrid_tz)

    # Crear la notificación interna
    notificacion = Notificacion(usuario_id=user.id, concepto=mensaje, fecha_hora=ahora_madrid)

    db.session.add(notificacion)
    db.session.commit()


    # Enviar correo si se solicita
    if enviar_email:
        send_email_to_user(user.email, titulo, mensaje)





@app.route('/admin/accept_request/<int:request_id>')
@admin_required
def accept_request(request_id):
    print(f"Iniciando proceso de aceptación de solicitud ID: {request_id}")
    today = datetime.today().date()

    vacation_request = VacationRequest.query.get_or_404(request_id)
    print(f"Solicitud encontrada: {vacation_request}")

    vacation_request.is_accepted = True
    vacation_request.status = 'Aceptada'
    vacation_request.dia_decision = today
    db.session.commit()
    print(f"Solicitud actualizada como aceptada en la base de datos. Fecha de decisión: {today}")

    if vacation_request:
        user = User.query.get(vacation_request.user_id)

        if user:
            print(f"Usuario encontrado: {user.id} - {user.email}")
            print(f"Días restantes del año pasado: {user.dias_restantes_pasado}, Días restantes de este año: {user.dias_restantes_este}")
            print(f"Días solicitados: {vacation_request.dias_vacas}")

            # Restar días del año pasado primero
            restar_del_ano_pasado = min(user.dias_restantes_pasado, vacation_request.dias_vacas)
            user.dias_restantes_pasado -= restar_del_ano_pasado
            dias_faltantes = vacation_request.dias_vacas - restar_del_ano_pasado
            print(f"Días restados del año pasado: {restar_del_ano_pasado}, Días faltantes por restar: {dias_faltantes}")

            # Si aún quedan días por restar, restar del año actual
            if dias_faltantes > 0:
                user.dias_restantes_este = max(0, user.dias_restantes_este - dias_faltantes)
                print(f"Días restantes de este año tras ajuste: {user.dias_restantes_este}")

            # Guardar información de los días utilizados
            vacation_request.dias_utilizados_pasado = restar_del_ano_pasado
            vacation_request.dias_utilizados_este = vacation_request.dias_vacas - restar_del_ano_pasado
            db.session.commit()
            print("Días utilizados actualizados en la solicitud.")

            correo = user.email

            # Actualizar estado en VacationHistory si existe
            history_entry = VacationHistory.query.filter_by(

                user_id=vacation_request.user_id,
                start_date=vacation_request.start_date,
                end_date=vacation_request.end_date
            ).first()

            if history_entry:
                print("Entrada de VacationHistory encontrada, actualizando estado...")
                history_entry.status = 'ACEPTADA'
                history_entry.is_accepted = True
                history_entry.dia_decision = today
                db.session.commit()
                print(f"Estado de VacationHistory actualizado. Fecha de decisión: {history_entry.dia_decision}")

                # Enviar correo al usuario
                send_email_to_user(correo, 'Solicitud de vacaciones aprobada',
                                   f'El responsable ha aceptado su solicitud de vacaciones.')

                notificar_usuario(user, 'Solicitud de vacaciones aprobada',
                  'El responsable ha aceptado su solicitud de vacaciones. Puede comprobar el estado en la aplicación.')

                print(f"Correo de notificación enviado a {correo}")
            else:
                print("Error: No se encontró la entrada en VacationHistory.")

        else:
            print("Error: No se encontró el usuario asociado a la solicitud de vacaciones.")
    else:
        print("Error: No se encontró la solicitud de vacaciones.")

    print("Redirigiendo a la vista de solicitudes de administrador...")
    return redirect(url_for('admin_requests'))


@app.route('/admin/reject_request/<int:request_id>', methods=['GET', 'POST'])
@admin_required
def reject_request(request_id):
    request_to_reject = VacationRequest.query.get(request_id)

    if request.method == 'POST':
        today = datetime.today().date()
        rejection_reason = request.form.get('rejection_reason')
        request_to_reject.status = 'RECHAZADA'
        request_to_reject.rejection_reason = rejection_reason
        request_to_reject.dia_decision = today
        db.session.commit()


        if request_to_reject:
            user = User.query.get(request_to_reject.user_id)
            if user:
                    # Actualiza el estado en VacationHistory (si existe)
                    history_entry = VacationHistory.query.filter_by(
                        user_id=request_to_reject.user_id,
                        start_date=request_to_reject.start_date,
                        end_date=request_to_reject.end_date,

                    ).first()

                    if history_entry:
                        history_entry.status = 'Rechazada'
                        history_entry.is_accepted = False
                        history_entry.rejection_reason = rejection_reason
                        history_entry.dia_decision=today
                        db.session.commit()
                    # Enviar correo al usuario notificando el rechazo
                    send_email_to_user(user.email, 'Solicitud de vacaciones rechazada', f'El responsable ha denegado su solicitud de vacaciones, entre en la aplicación para comprobar su situación')
                    notificar_usuario(user, 'Solicitud de vacaciones rechazada',
                    'El responsable ha denegado su solicitud de vacaciones. Puede consultar la situación en la pestaña Mis vacaciones.')



        return redirect(url_for('admin_requests'))

    return render_template('reject_request.html', request=request_to_reject)

@app.route('/admin/cancel_request/<int:request_id>', methods=['GET', 'POST'])
@admin_required
def cancel_request(request_id):
    print("000000000000")
    # Obtener la solicitud de la base de datos
    request_to_cancel = VacationRequest.query.get(request_id)

    possible_requests = VacationHistory.query.filter_by(
        user_id=request_to_cancel.user_id,
        start_date=request_to_cancel.start_date,
        end_date=request_to_cancel.end_date
    ).all()
    print("BUSCANDO PETICIONES...")
    print(possible_requests)
    # Obtener la primera solicitud cuyo status no sea 'Cancelada'
    request_to_cancelh = next(
        (req for req in possible_requests if req.status != 'Cancelada'),
        None
    )
    print("FILTRANDOPETICIONES...")
    print(request_to_cancelh)
    print("1111111111111111111")

    if request_to_cancel:
        print("222222222222222")
        # Convertir la fecha de inicio de la solicitud a un objeto date si es necesario
        start_date = request_to_cancel.start_date
        if isinstance(start_date, datetime):
            start_date = start_date.date()

        # Comprobar si la fecha de inicio es posterior a hoy
        if start_date > datetime.today().date():
            print("4444444444444444444444444")
            if request.method == 'POST':
                print("55555555555555555555555555555")
                # Lógica para procesar la cancelación
                cancel_reason = request.form.get('cancel_reason')
                request_to_cancel.status = 'Cancelada'
                request_to_cancelh.status = 'Cancelada'
                request_to_cancel.is_accepted = False
                request_to_cancelh.is_accepted = False
                request_to_cancel.cancel_reason = cancel_reason
                request_to_cancelh.cancel_reason = cancel_reason
                request_to_cancel.dia_decision = datetime.today().date()
                request_to_cancelh.dia_decision = datetime.today().date()
                user = User.query.get(request_to_cancel.user_id)
                print(user.nombre)
                if user:


                    print(request_to_cancel.dias_utilizados_este)
                    user.dias_restantes_este += request_to_cancel.dias_utilizados_este
                    print(request_to_cancel.dias_utilizados_pasado)
                    print("antes:", user.dias_restantes_pasado)
                    user.dias_restantes_pasado += request_to_cancel.dias_utilizados_pasado
                    print("después:",user.dias_restantes_pasado)
                    send_email_to_user(user.email, 'Solicitud de vacaciones Cancelada', f'El responsable ha cancelado su solicitud de vacaciones, entre en la aplicación para comprobar su situación')
                db.session.commit()

                # Redireccionar o mostrar mensaje de éxito aquí
                flash('La solicitud ha sido cancelada exitosamente.', 'success')
                return redirect(url_for('admin_requests'))
            # Mostrar el formulario de cancelación si el método es GET y la condición se cumple
            return render_template('cancel_request.html', request=request_to_cancel)
        else:
            # Si la fecha de inicio no es posterior a hoy, mostrar mensaje de error
            flash('Las solicitudes de vacaciones solo pueden cancelarse si la fecha de inicio es futura.', 'error')
    else:
        # Si no se encuentra la solicitud, mostrar mensaje de error
        flash('Solicitud de vacaciones no encontrada.', 'error')

    # Redirigir al usuario si la condición no se cumple o la solicitud no existe
    return redirect(url_for('admin_requests'))



from flask import jsonify, request

@app.route('/request_cancellation/<int:request_id>', methods=['GET', 'POST'])

def request_cancellation(request_id):

    vacation_history = db.session.get(VacationHistory, request_id)

    def json_response(messages):

        return jsonify({'messages': messages})

    if vacation_history:

        start_date = vacation_history.start_date

        today = datetime.today().date()

        vacation_request = VacationRequest.query.filter_by(

            user_id=vacation_history.user_id,

            start_date=vacation_history.start_date,

            end_date=vacation_history.end_date

        ).first()

        if vacation_request and vacation_history.status in ['ACEPTADA', 'PENDIENTE'] and start_date > today:

            if request.method == 'POST':

                cancel_reason = request.form.get('cancel_reason')

                if cancel_reason:

                    vacation_request.cancel_reason = cancel_reason

                    vacation_history.cancel_reason = cancel_reason

                    db.session.commit()

                    # 🔹 Notificar a administradores, gestores y encargados

                    try:

                        admin_users = get_admin_users()

                        admin_emails = get_admin_emails()

                        if admin_users:

                            print(f"[DEBUG] Enviando notificación de cancelación a: {admin_emails}")

                            for email in admin_emails:

                                send_cancellation_email_to_admin(

                                    email=email,

                                    request_id=vacation_request.id,

                                    employee_name=f"{current_user.nombre} {current_user.primer_apellido}",

                                    start_date=vacation_history.start_date,

                                    end_date=vacation_history.end_date,

                                    cancel_reason=cancel_reason

                                )

                            madrid_tz = pytz_timezone('Europe/Madrid')

                            ahora_madrid = datetime.now(madrid_tz)

                            for admin in admin_users:

                                nueva_notificacion = Notificacion(

                                    usuario_id=admin.id,

                                    concepto=f"{current_user.nombre} {current_user.primer_apellido} ha solicitado cancelar sus vacaciones del {vacation_history.start_date} al {vacation_history.end_date}.",

                                    fecha_hora=ahora_madrid

                                )

                                db.session.add(nueva_notificacion)

                            db.session.commit()

                            print("[DEBUG] Notificaciones de cancelación creadas para los administradores.")

                        else:

                            print("[ERROR] No hay administradores, gestores o encargados registrados")

                    except Exception as e:

                        print(f"[ERROR] Error al enviar notificaciones de cancelación: {e}")

                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                        return json_response([{

                            'message': 'Tu solicitud de cancelación ha sido enviada.',

                            'category': 'success'

                        }])

                    flash('Tu solicitud de cancelación ha sido enviada.', 'success')

                    return redirect(url_for('user_history'))

                else:

                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

                        return json_response([{

                            'message': 'Por favor, proporciona un motivo para la cancelación.',

                            'category': 'error'

                        }])

                    flash('Por favor, proporciona un motivo para la cancelación.', 'error')

            return render_template('cancel_request2.html', request=vacation_history)

        msg = 'La cancelación no está permitida para solicitudes ya iniciadas o con estados no permitidos.'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

            return json_response([{

                'message': msg,

                'category': 'error'

            }])

        flash(msg, 'error')

        return redirect(url_for('user_history'))

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':

        return json_response([{

            'message': 'Solicitud no encontrada.',

            'category': 'error'

        }])

    flash('Solicitud no encontrada.', 'error')

    return redirect(url_for('user_history'))

from flask import render_template_string


from itsdangerous import URLSafeTimedSerializer

from flask import abort

serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])


@app.route('/admin/email_cancel_action/<token>/<action>')

def email_cancel_action(token, action):

    try:

        request_id = serializer.loads(token, max_age=86400)  # válido 24h

    except Exception:

        abort(403)

    vacation_request = VacationRequest.query.get(request_id)

    if not vacation_request:

        abort(404)

    vacation_history = VacationHistory.query.filter_by(

        user_id=vacation_request.user_id,

        start_date=vacation_request.start_date,

        end_date=vacation_request.end_date

    ).first()

    if action == "approve":

        vacation_request.status = 'Cancelada'

        vacation_request.is_accepted = False

        vacation_request.dia_decision = datetime.today().date()

        if vacation_history:

            vacation_history.status = 'Cancelada'

            vacation_history.is_accepted = False

            vacation_history.dia_decision = datetime.today().date()

        user = User.query.get(vacation_request.user_id)

        if user:

            user.dias_restantes_este += vacation_request.dias_utilizados_este

            user.dias_restantes_pasado += vacation_request.dias_utilizados_pasado

            send_email_to_user(

                user.email,

                'Cancelación aprobada',

                'Tu solicitud de cancelación ha sido aprobada.'

            )

    elif action == "reject":

        vacation_request.cancel_reason = None

        if vacation_history:

            vacation_history.cancel_reason = None

        user = User.query.get(vacation_request.user_id)

        if user:

            send_email_to_user(

                user.email,

                'Cancelación rechazada',

                'Tu solicitud de cancelación ha sido rechazada.'

            )

    db.session.commit()

    return "<h2>Acción realizada correctamente. Puedes cerrar esta ventana.</h2>"

def send_cancellation_email_to_admin(email, request_id, employee_name, start_date, end_date, cancel_reason):

    token = serializer.dumps(request_id)

    approve_url = url_for(

        'email_cancel_action',

        token=token,

        action='approve',

        _external=True

    )

    reject_url = url_for(

        'email_cancel_action',

        token=token,

        action='reject',

        _external=True

    )

    html_body = f"""
<html>
<body style="font-family: Arial, sans-serif;">

            <h3>Solicitud de cancelación de vacaciones</h3>

            <p>

                El empleado <strong>{employee_name}</strong>

                ha solicitado cancelar sus vacaciones

                del <strong>{start_date}</strong> al <strong>{end_date}</strong>.
</p>

            <div style="

                background-color:#fff3cd;

                border-left:5px solid #ffc107;

                padding:15px;

                margin:15px 0;

                border-radius:6px;

            ">
<strong>Motivo indicado por el empleado:</strong>
<p style="margin-top:8px;">

                    {cancel_reason}
</p>
</div>

            <p style="color:#b02a37; font-weight:600;">

                ⚠️ Si apruebas la cancelación, este será el motivo definitivo que quedará registrado.
</p>

            <br>

            <a href="{approve_url}"

               style="background-color:#198754;

                      color:white;

                      padding:12px 20px;

                      text-decoration:none;

                      border-radius:6px;

                      margin-right:10px;

                      font-weight:bold;">

                ✅ Aprobar cancelación
</a>

            <a href="{reject_url}"

               style="background-color:#dc3545;

                      color:white;

                      padding:12px 20px;

                      text-decoration:none;

                      border-radius:6px;

                      font-weight:bold;">

                ❌ Rechazar cancelación
</a>

            <br><br>

            <p style="font-size:12px;color:gray;">

                Este enlace caduca en 24 horas.
</p>

        </body>
</html>

    """

    msg = Message(

        "Solicitud de cancelación de vacaciones",

        recipients=[email],

        html=html_body

    )

    mail.send(msg)


from flask import render_template_string

# Función para enviar correo al usuario
def send_email_to_user(email, subject, body):
    # Crear el cuerpo del correo como HTML
    html_body = render_template_string(
        """\
        <html>
            <body>
                <p>{{ body }}</p>
                <p><a href="{{ link_href }}">{{ link_text }}</a></p>
            </body>
        </html>
        """,
        body=body,
        link_text="Link a la APP",
        link_href="https://herdelsalud.gestionaempresa.es/"
    )

    # Crear el mensaje
    msg = Message(subject, recipients=[email], html=html_body)

    # Enviar el mensaje
    mail.send(msg)
    print("Enviado")


@app.route('/register', methods=['GET', 'POST'])

def register():
    if request.method == 'POST':
        email = request.form.get('Correo')
        password = request.form.get('Contraseña')
        role = 'user'  # Como está registrado por el administrador, el nuevo usuario tiene el rol 'user' por defecto
        primer_apellido=request.form.get('apellido1')
        segundo_apellido=request.form.get('apellido2')
        nombre=request.form.get('Nombre')
        user = user_datastore.find_user(email=email)

        if user:
            return jsonify({'success': False})
        else:
            user_datastore.create_user(nombre=nombre,primer_apellido=primer_apellido,segundo_apellido=segundo_apellido,email=email, password=password, approved=False, roles=[role])
            db.session.commit()
            send_email_to_user('admin@gestionaempresa.es', 'Nuevo empleado', f'Se ha solicitado crear un nuevo perfil de empleado, accede a la sección de usuarios para validarlo')
            return jsonify({'success': True, 'message': 'Se ha registrado un nuevo usuario, entra en la aplicación para validarlo.'})





        # Redirigir a una página específica después del registro exitoso
        return jsonify({'redirect': url_for('login')})

    return render_template('register.html')
@app.route('/validate_admin', methods=['POST'])
def validate_admin():
    admin_email = request.form.get('admin_email')
    admin_password = request.form.get('admin_password')

    # Verificar si las credenciales son iguales a los valores predefinidos
    valid_admin_credentials = check_admin_credentials(admin_email, admin_password)

    if valid_admin_credentials:
        return jsonify('valid')
    else:
        return jsonify('invalid')

# Función para verificar las credenciales del administrador
def check_admin_credentials(email, password):
    # Comparar con los valores predefinidos
    return email == 'admin@gestionaempresa.es' and password == 'admin'

from flask_login import current_user
# Obtener los registros del usuario y sumar las horas trabajadas

from sqlalchemy import extract

@app.route('/admin/users')
@admin_required
def admin_users():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_email = current_user.email if current_user.is_authenticated else None

    # Filtrar tanto usuarios como administradores
    users_query = User.query.filter(~User.roles.any(Role.name == 'inspector_role'))

    # Filtros por correo específicos
    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        users_query = users_query.filter(User.Centro_trabajo == 'CEEI Guadalajara')
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        users_query = users_query.filter(User.Centro_trabajo == 'CEOE Formación')

    # Si el usuario tiene rol de gestor, filtrar por su empresa
    if current_user.has_role('gestor'):
        users_query = users_query.filter(User.empresa_id == current_user.empresa_id)

    # Excluir los correos especiales
    users = users_query.filter(
        (User.email != 'admin@gestionaempresa.es') &
        (User.email != 'inspector@gestionaempresa.es') &
        (User.email != 'gestor@gestionaempresa.es') &
        (User.email != 'encargado@gestionaempresa.es') &
        (User.approved == True)

    ).all()


    jornada_irregular_por_usuario = {}

    # Obtener las fechas de inicio y fin de verano de la empresa
    empresa = Empresa.query.first()  # Suponiendo que solo hay una empresa en la base de datos
    fecha_inicio_verano = empresa.fecha_inicio_verano
    fecha_fin_verano = empresa.fecha_fin_verano

    # Obtener los días festivos
    festivos = Festivo.query.all()
    dias_festivos = [(festivo.dia, festivo.mes) for festivo in festivos]

    for user in users:
        jornada_irregular = timedelta()
        total_horas_esperadas = timedelta()
        total_horas_trabajadas = timedelta()

        # Calcular el total de horas que debería haber trabajado el usuario desde el 1 de enero hasta hoy
        year = datetime.now().year
        fecha_inicio = datetime(year, 1, 1).date()
        fecha_fin = date.today() - timedelta(days=1) # Fecha actual como fecha de fin

        for dia in range((fecha_fin - fecha_inicio).days + 1):
            fecha_actual = fecha_inicio + timedelta(days=dia)
            dia_semana = fecha_actual.weekday()

            # Verificar si el día es festivo
            if (fecha_actual.day, fecha_actual.month) in dias_festivos:
                continue  # Saltar los días festivos
            if empresa.fecha_inicio_verano:

                es_verano = fecha_inicio_verano <= fecha_actual <= fecha_fin_verano
            else:
                es_verano = False

            entrada = salida = salida_comida = entrada_comida = None
            if dia_semana == 0 and user.trabaja_lunes:
                if es_verano and user.trabaja_verano_lunes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_lunes, user.salida_verano_lunes, user.salida_comida_verano_lunes, user.entrada_comida_verano_lunes
                elif not es_verano and user.trabaja_lunes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_lunes, user.salida_lunes, user.salida_comida_lunes, user.entrada_comida_lunes
            elif dia_semana == 1 and user.trabaja_martes:
                if es_verano and user.trabaja_verano_martes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_martes, user.salida_verano_martes, user.salida_comida_verano_martes, user.entrada_comida_verano_martes
                elif not es_verano and user.trabaja_martes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_martes, user.salida_martes, user.salida_comida_martes, user.entrada_comida_martes
            elif dia_semana == 2 and user.trabaja_miercoles:
                if es_verano and user.trabaja_verano_miercoles:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_miercoles, user.salida_verano_miercoles, user.salida_comida_verano_miercoles, user.entrada_comida_verano_miercoles
                elif not es_verano and user.trabaja_miercoles:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_miercoles, user.salida_miercoles, user.salida_comida_miercoles, user.entrada_comida_miercoles
            elif dia_semana == 3 and user.trabaja_jueves:
                if es_verano and user.trabaja_verano_jueves:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_jueves, user.salida_verano_jueves, user.salida_comida_verano_jueves, user.entrada_comida_verano_jueves
                elif not es_verano and user.trabaja_jueves:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_jueves, user.salida_jueves, user.salida_comida_jueves, user.entrada_comida_jueves
            elif dia_semana == 4 and user.trabaja_viernes:
                if es_verano and user.trabaja_verano_viernes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_viernes, user.salida_verano_viernes, user.salida_comida_verano_viernes, user.entrada_comida_verano_viernes
                elif not es_verano and user.trabaja_viernes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_viernes, user.salida_viernes, user.salida_comida_viernes, user.entrada_comida_viernes
            elif dia_semana == 5 and user.trabaja_sabado:
                if es_verano and user.trabaja_verano_sabado:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_sabado, user.salida_verano_sabado, user.salida_comida_verano_sabado, user.entrada_comida_verano_sabado
                elif not es_verano and user.trabaja_sabado:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_sabado, user.salida_sabado, user.salida_comida_sabado, user.entrada_comida_sabado
            elif dia_semana == 6 and user.trabaja_domingo:
                if es_verano and user.trabaja_verano_domingo:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_domingo, user.salida_verano_domingo, user.salida_comida_verano_domingo, user.entrada_comida_verano_domingo
                elif not es_verano and user.trabaja_domingo:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_domingo, user.salida_domingo, user.salida_comida_domingo, user.entrada_comida_domingo
            else:
                continue

            def time_to_timedelta(time_obj):
                return timedelta(hours=time_obj.hour, minutes=time_obj.minute, seconds=time_obj.second)

            if entrada and salida:
                entrada_timedelta = time_to_timedelta(entrada)
                salida_timedelta = time_to_timedelta(salida)
                salida_comida_timedelta = time_to_timedelta(salida_comida) if salida_comida else timedelta()
                entrada_comida_timedelta = time_to_timedelta(entrada_comida) if entrada_comida else timedelta()

                horas_diarias_esperadas = (salida_timedelta - entrada_timedelta) - (entrada_comida_timedelta - salida_comida_timedelta)
                total_horas_esperadas += horas_diarias_esperadas



        current_year = datetime.now().year

        registros = Registro.query.filter(
            Registro.user_id == user.id,
            extract('year', Registro.entrada) == current_year
        ).all()

        for registro in registros:
            if registro.duracion_neta:
                total_horas_trabajadas += registro.duracion_neta

        jornada_irregular = total_horas_trabajadas - total_horas_esperadas
        jornada_irregular_en_horas = jornada_irregular.total_seconds() / 3600

        if total_horas_esperadas.total_seconds() > 0:
            porcentaje_irregular = (jornada_irregular.total_seconds() / total_horas_esperadas.total_seconds()) * 100
        else:
            porcentaje_irregular = 0

        jornada_irregular_por_usuario[user.id] = {
            'total_horas_esperadas': total_horas_esperadas.total_seconds() / 3600 if total_horas_esperadas else 0,
            'total_horas_trabajadas': total_horas_trabajadas.total_seconds() / 3600 if total_horas_trabajadas else 0,
            'jornada_irregular_horas': jornada_irregular_en_horas,
            'porcentaje_irregular': porcentaje_irregular
        }

    print(jornada_irregular_por_usuario)

    return render_template('admin_users.html', users=users, current_user_email=current_user_email, notificaciones=notificaciones,  jornada_irregular_por_usuario=jornada_irregular_por_usuario, active_menu='combo_config')


@app.route('/admin/usuarios_bloqueados')
@admin_required
def usuarios_bloqueados():
    users_bloqueados = User.query.filter(
        (User.email != 'admin@gestionaempresa.es') &
        (User.email != 'inspector@gestionaempresa.es') &
        (User.approved == False)
    ).order_by(User.email.asc()).all()

    return render_template('admin_users_bloqueados.html', users=users_bloqueados)


from flask import jsonify

@app.route('/admin/unblock_user/<int:user_id>', methods=['POST'])
@admin_required
def unblock_user(user_id):
    user = User.query.get_or_404(user_id)
    if not user.approved:
        user.approved = True
        user.bloqueado = 0
        user.bloqueado_hasta = None
        user.intentos_fallidos = 0
        user.bloqueos_temporales = 0
        db.session.commit()
    return jsonify({'redirect': url_for('usuarios_bloqueados')})


@app.route('/admin/delete_user/<int:user_id>')

def delete_user(user_id):
    print(f"[DEBUG] Intentando eliminar usuario con ID: {user_id}")

    user_to_delete = User.query.get_or_404(user_id)

    if user_to_delete:
        print(f"[DEBUG] Usuario encontrado: {user_to_delete.email}")

        # Importar modelos relacionados si no están en el mismo archivo
        # 1. Eliminar notificaciones
        notificaciones = Notificacion.query.filter_by(usuario_id=user_to_delete.id).all()
        for n in notificaciones:
            db.session.delete(n)

        # 2. Eliminar registros de horario
        registros = RegistroHorario.query.filter_by(user_id=user_to_delete.id).all()
        for r in registros:
            db.session.delete(r)

        # 3. Eliminar usuario
        db.session.delete(user_to_delete)
        db.session.commit()

        print(f"[DEBUG] Usuario con ID {user_id} eliminado correctamente")

    return jsonify({'redirect': url_for('admin_users')})





@app.route('/admin/configure_user/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def configure_user(user_id):
    user = User.query.get_or_404(user_id)
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False) \
    .order_by(Notificacion.fecha_hora.desc()).all()

    empresas = Empresa.query.all()



    # Prepara una lista de roles del usuario
    user_roles = [role.name for role in user.roles]

    return render_template('configure_user.html', user=user, notificaciones=notificaciones, empresas=empresas, user_roles=user_roles)




@app.route('/admin/guardar_configure_user/<int:user_id>', methods=['POST'])
@roles_required('admin')
def guardar_configure_user(user_id):
    print(f"Intentando configurar usuario con ID: {user_id}")
    user = User.query.get_or_404(user_id)

    if not request.is_json:
        print("Error: La solicitud no es JSON")
        return jsonify({'error': 'Contenido no es JSON válido'}), 400

    data = request.get_json()
    print(f"Datos recibidos: {data}")

    # Función para convertir valores vacíos en None
    def convertir_a_null(valor):
        return valor if valor not in ["", None] else None

    # Actualizar días restantes y tipo de jornada
    user.dias_restantes_este = convertir_a_null(data.get('dias_restantes_este'))
    user.dias_restantes_pasado = convertir_a_null(data.get('dias_restantes_pasado'))
    user.tipo_jornada = convertir_a_null(data.get('tipo_jornada'))

    print(f"Días restantes este año: {user.dias_restantes_este}, pasado: {user.dias_restantes_pasado}, tipo jornada: {user.tipo_jornada}")

    if 'empresa_id' in data:
        user.empresa_id = convertir_a_null(data.get('empresa_id'))
        print(f"Empresa asignada: {user.empresa_id}")

    # Validar y actualizar el rol del usuario
    nuevo_rol = data.get('rol')

    user.vacaciones = data.get('vacaciones', False)
    user.registros_horarios = data.get('registros', False)
    user.ausencias = data.get('ausencias', False)

    print("vacaciones", user.vacaciones)
    print("registros_horarios", user.registros_horarios)

    print("ausencias", user.ausencias)



    print(f"Rol recibido: {nuevo_rol}")

    if not nuevo_rol:
        print("Error: No se proporcionó un rol")
        return jsonify({'error': 'Campo \"role\" es requerido'}), 400

    roles_permitidos = ['admin', 'user', 'gestor', 'encargado']
    if nuevo_rol not in roles_permitidos:
        print(f"Error: Rol no válido - {nuevo_rol}")
        return jsonify({'error': 'Rol no válido'}), 400

    rol = Role.query.filter_by(name=nuevo_rol).first()
    if not rol:
        print(f"Error: Rol '{nuevo_rol}' no encontrado en la base de datos")
        return jsonify({'error': 'Rol no encontrado en la base de datos'}), 400

    print(f"Eliminando roles actuales del usuario: {[r.name for r in user.roles]}")
    for role in user.roles[:]:
        user_datastore.remove_role_from_user(user, role)

    print(f"Asignando nuevo rol: {rol.name}")
    user_datastore.add_role_to_user(user, rol)

    # Función para convertir string a time o None si está vacío
    def convertir_a_time(hora_str):
        if hora_str and hora_str.strip():
            if len(hora_str) == 5:
                hora_str = f"{hora_str}:00"
            try:
                return datetime.strptime(hora_str, '%H:%M:%S').time()
            except ValueError:
                print(f"Error al convertir hora: {hora_str}")
                return None
        return None

    # Función para actualizar horarios
    def actualizar_horarios(user, horarios, prefijo=None):
        dias_semana = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
        horarios_dict = {h["dia"]: h for h in horarios}
        print(f"Actualizando horarios {'verano' if prefijo else 'invierno'}")

        for dia in dias_semana:
            horario = horarios_dict.get(dia, {})
            trabaja = horario.get("trabaja", False)
            entrada = convertir_a_time(horario.get("entrada", ""))
            salida_comida = convertir_a_time(horario.get("salidaComida", ""))
            entrada_comida = convertir_a_time(horario.get("entradaComida", ""))
            salida = convertir_a_time(horario.get("salida", ""))

            print(f"{dia.title()}: trabaja={trabaja}, entrada={entrada}, salida comida={salida_comida}, entrada comida={entrada_comida}, salida={salida}")

            if prefijo is None:
                setattr(user, f'trabaja_{dia}', trabaja)
                setattr(user, f'entrada_{dia}', entrada)
                setattr(user, f'salida_comida_{dia}', salida_comida)
                setattr(user, f'entrada_comida_{dia}', entrada_comida)
                setattr(user, f'salida_{dia}', salida)
            else:
                setattr(user, f'trabaja_{prefijo}_{dia}', trabaja)
                setattr(user, f'entrada_{prefijo}_{dia}', entrada)
                setattr(user, f'salida_comida_{prefijo}_{dia}', salida_comida)
                setattr(user, f'entrada_comida_{prefijo}_{dia}', entrada_comida)
                setattr(user, f'salida_{prefijo}_{dia}', salida)

    if 'horarios_invierno' in data:
        actualizar_horarios(user, data['horarios_invierno'])

    if 'horarios_verano' in data:
        actualizar_horarios(user, data['horarios_verano'], 'verano')

    try:
        db.session.commit()
        print("Cambios guardados exitosamente en la base de datos")
    except Exception as e:
        print(f"Error al guardar en la base de datos: {e}")
        db.session.rollback()
        return jsonify({'error': 'Error al guardar los cambios'}), 500

    return jsonify({'success': True})


@app.route('/admin/validate_user/<int:user_id>', methods=['POST'])
@admin_required
def validate_user(user_id):
    user = User.query.get_or_404(user_id)

    try:
        # Validar usuario y desbloquearlo si es necesario
        if not user.approved:
            user.approved = True

        # Desbloquear al usuario si estaba bloqueado
        if user.bloqueado:
            user.bloqueado = 0
            user.intentos_fallidos = 0  # Reiniciar contador de intentos fallidos

        db.session.commit()

        # Enviar un correo notificando que el usuario ha sido validado
        send_email_to_user(user.email, 'Usuario validado', f'Su usuario ha sido validado por el responsable.')

        # Respuesta JSON para AJAX
        return jsonify({
            "status": "success",
            "message": "Usuario validado exitosamente",
            "approved": user.approved,
            "bloqueado": user.bloqueado
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Error al validar al usuario: {str(e)}"
        })



@app.route('/admin/block_user/<int:user_id>', methods=['POST'])
@admin_required
def block_user(user_id):
    print(f"[INFO] Intentando bloquear al usuario con ID: {user_id}")

    user = User.query.get_or_404(user_id)
    print(f"[INFO] Usuario encontrado: {user.email} - Aprobado: {user.approved}")

    if user.approved:
        user.approved = False
        db.session.commit()
        print(f"[SUCCESS] Usuario {user.email} ha sido bloqueado (approved=False).")
    else:
        print(f"[WARNING] Usuario {user.email} ya estaba bloqueado.")

    return redirect(url_for('admin_users', user_id=user_id))







# Ruta para eliminar una solicitud
@app.route('/admin/delete_request/<int:request_id>', methods=['DELETE'])
@admin_required
def delete_request(request_id):
    request_to_delete = VacationRequest.query.get_or_404(request_id)

    db.session.delete(request_to_delete)
    db.session.commit()
    return jsonify({'message': 'Solicitud eliminada exitosamente'})


from flask import request

from flask import render_template, request
from flask_login import current_user



@app.route('/user_history')
def user_history():
    print("👉 Entrando en /user_history")

    today = datetime.today().date()
    print(f"📅 Hoy es: {today}")

    current_user_centro = current_user.Centro_trabajo if current_user.is_authenticated else None
    print(f"🏢 Centro de trabajo: {current_user_centro}")

    # Obtener el número de página desde la URL (por defecto, la página es 1)
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Número de registros por página
    print(f"📄 Página actual: {page}, registros por página: {per_page}")

    # Obtener el historial del usuario con paginación
    pagination = VacationHistory.query.filter_by(user_id=current_user.id) \
        .order_by(VacationHistory.dia_solicitud.desc()) \
        .paginate(page=page, per_page=per_page, error_out=False)

    user_history = pagination.items  # Registros en la página actual
    print(f"📚 Total de registros en la página actual: {len(user_history)}")

    # Obtener las notificaciones
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False) \
        .order_by(Notificacion.fecha_hora.desc()).all()
    print(f"🔔 Notificaciones sin leer: {len(notificaciones)}")

    # Formatear las fechas
    print("📅 Formateando fechas de los registros:")
    for idx, entry in enumerate(user_history, start=1):
        print(f"➡️ Registro {idx}:")
        print(f"    Original start_date: {entry.start_date}")
        print(f"    Original end_date: {entry.end_date}")
        print(f"    Original dia_solicitud: {entry.dia_solicitud}")
        print(f"    Original dia_decision: {entry.dia_decision}")

        entry.formatted_start_date = entry.start_date.strftime('%d/%m/%Y')
        entry.formatted_end_date = entry.end_date.strftime('%d/%m/%Y')
        entry.formatted_dia_solicitud = entry.dia_solicitud.strftime('%d/%m/%Y')
        entry.formatted_dia_decision = entry.dia_decision.strftime('%d/%m/%Y') if entry.dia_decision else None

        print(f"    📌 Formateado start_date: {entry.formatted_start_date}")

    print("🧾 Preparando datos para renderizar la plantilla")

    return render_template('user_view_history.html',
                           user_history=user_history,
                           pagination=pagination,
                           notificaciones=notificaciones,
                           remaining_days=current_user.dias_restantes_este,
                           remaining_days_pasado=current_user.dias_restantes_pasado,
                           current_user_centro=current_user_centro,
                           today=today,
                           active_menu='user_history')


@app.route('/admin/history', methods=['GET', 'POST'])
@admin_required

def admin_history():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_email = current_user.email
    if request.method == 'POST':
        # Obtén el ID del usuario seleccionado desde el formulario
        user_id = request.form.get('user_id')

        # Redirige a la página de historial del usuario seleccionado
        return redirect(url_for('admin_user_history', user_id=user_id, current_user_email = current_user_email))

    # Obtén el correo del usuario actual
    current_user_email = current_user.email

    # Obtén el centro de trabajo del usuario actual
    current_user_workplace = current_user.Centro_trabajo

    # Filtrar la lista de usuarios en función del correo y el centro de trabajo
    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        users = User.query.filter_by(Centro_trabajo='CEEI Guadalajara').all()
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        users = User.query.filter_by(Centro_trabajo='CEOE Formación').all()
    else:
        # Si el correo no coincide con ninguno de los correos específicos, muestra todos los usuarios
        users = User.query.filter(User.email != 'admin@gestionaempresa.es' and User.email != 'inspector@gestionaempresa.es').all()

    return render_template('admin_history.html', notificaciones=notificaciones, users=users,current_user_email = current_user_email, active_menu='combo_vacacione')



# Rutas de administrador

@app.route('/admin/user_history', methods=['GET'])
@admin_required
def admin_user_history():
    current_user_email = current_user.email
    notificaciones = Notificacion.query.filter_by(
        usuario_id=current_user.id, leida=False
    ).order_by(Notificacion.fecha_hora.desc()).all()

    # Obtener el filtro de usuario desde la URL
    usuario_filtro = request.args.get('usuario')

    print(f"Usuario Filtro: {usuario_filtro}")

    # Obtener la lista de usuarios, aplicando el filtro de empresa si el usuario tiene rol 'gestor' o 'responsable'
    users_query = User.query.filter(User.email != 'admin@gestionaempresa.es', User.email != 'inspector@gestionaempresa.es')

    # Filtrar por empresa si el usuario tiene rol 'gestor' o 'responsable'
    if current_user.has_role('gestor') or current_user.has_role('encargado'):
        users_query = users_query.filter(User.empresa_id == current_user.empresa_id)

    users = users_query.all()

    # Obtener todo el historial de vacaciones, aplicando el filtro de empresa si es necesario
    user_history_query = VacationHistory.query

    if current_user.has_role('gestor') or current_user.has_role('encargado'):
        user_history_query = user_history_query.join(User).filter(User.empresa_id == current_user.empresa_id)

    # Obtener todo el historial de vacaciones
    user_history = user_history_query.all()

    # Aplicar filtro si hay un usuario seleccionado
    if usuario_filtro:
        try:
            usuario_filtro_id = int(usuario_filtro)
            user_history = [entry for entry in user_history if entry.user_id == usuario_filtro_id]
        except ValueError:
            user_history = []  # Si el filtro no es válido, devolver lista vacía

    # Formatear las fechas de los eventos
    events = []
    for entry in user_history:
        # Solo realizar la conversión de fechas para visualización (si es necesario)
        start_date = entry.start_date.strftime('%Y-%m-%d')
        end_date = (entry.end_date + timedelta(days=1)).strftime('%Y-%m-%d')  # para tener el día completo
        dia_solicitud = entry.dia_solicitud.strftime('%Y-%m-%d')
        dia_decision = entry.dia_decision.strftime('%Y-%m-%d') if entry.dia_decision else None

        # Aplicar lógica de eventos
        if entry.is_accepted and entry.status.lower() == "aceptada":
            events.append({
                'start': start_date,
                'end': end_date,
                'rendering': 'background',  # Para colorear el fondo de la celda
                'color': '#6173D6'
            })

    current_year = datetime.now().year

    # Consultar los festivos desde la base de datos
    festivos = Festivo.query.all()

    # Convertir la información de días festivos en formato de texto ("YYYY-MM-DD")
    non_workdays = [f"{festivo.año}-{festivo.mes:02d}-{festivo.dia:02d}" for festivo in festivos]

    return render_template(
        'admin_user_history.html',
        user_history=user_history,
        events=events,
        current_user_email=current_user_email,
        non_workdays=non_workdays,
        users=users,
        usuario_filtro=usuario_filtro,
        notificaciones=notificaciones
    )



@app.route('/manage_absences')
def manage_absences_user():
    return render_template('manage_absences_user.html')




@app.route('/user/absence_history')
def user_absence_history():
    print("[DEBUG] Entrando a user_absence_history")

    # Obtener notificaciones
    notificaciones = Notificacion.query.filter_by(
        usuario_id=current_user.id, leida=False
    ).order_by(Notificacion.fecha_hora.desc()).all()

    # Obtener centro de trabajo del usuario
    current_user_centro = current_user.Centro_trabajo if current_user.is_authenticated else None

    # Número de página actual (por defecto 1)
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Número de ausencias por página

    try:
        # Obtener ausencias del usuario con paginación
        user_absences = Absence.query.filter_by(user_id=current_user.id) \
            .order_by(Absence.start_date.desc()) \
            .paginate(page=page, per_page=per_page, error_out=False)
    except Exception as e:
        print(f"[ERROR] Error al obtener ausencias: {e}")
        return jsonify({"status": "error", "message": "Error al cargar el historial"}), 500

    # Renderizar la plantilla con datos paginados
    return render_template(
        'user_absence_history.html',
        notificaciones=notificaciones,
        user_absences=user_absences,
        current_user_centro=current_user_centro,
        active_menu='user_absence_history'
    )

@app.route('/edit_absence', methods=['POST'])
def edit_absence():
    try:
        print("📌 Recibiendo datos del formulario...")

        # Obtener los datos del formulario
        absence_id = request.form.get('absenceId')
        absence_date = request.form.get('absenceDate')  # Fecha inicio
        absence_end_date = request.form.get('absenceEndDate')  # Fecha fin
        absence_type = request.form.get('absenceType')
        absence_description = request.form.get('absenceDescription')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        absence_file = request.files.get('absenceFile')

        print(f"Datos recibidos: ID {absence_id}, Inicio {absence_date}, Fin {absence_end_date}, Tipo {absence_type}, Descripción {absence_description}, Inicio {start_time}, Fin {end_time}")

        # Validar que los datos obligatorios están presentes
        if not absence_id or not absence_date:
            return jsonify({'status': 'error', 'message': 'Faltan datos obligatorios.'}), 400

        # Convertir fechas
        try:
            start_date = datetime.strptime(absence_date, '%Y-%m-%d').date()
            if absence_end_date:
                end_date = datetime.strptime(absence_end_date, '%Y-%m-%d').date()
            else:
                end_date = start_date  # Si no hay fecha de fin, se asume el mismo día

            print(f"✔ Fechas convertidas correctamente: {start_date} → {end_date}")
        except ValueError as e:
            return jsonify({'status': 'error', 'message': f'Error en el formato de fecha: {e}'}), 400

        # Convertir horas si existen
        if start_time and end_time:
            try:
                start_time_obj = datetime.strptime(start_time, '%H:%M').time()
                end_time_obj = datetime.strptime(end_time, '%H:%M').time()
                print(f"✔ Horas convertidas correctamente: {start_time_obj} - {end_time_obj}")
            except ValueError as e:
                return jsonify({'status': 'error', 'message': f'Error en el formato de hora: {e}'}), 400
        else:
            start_time_obj = None
            end_time_obj = None

        # Buscar la ausencia en la base de datos
        absence = Absence.query.get(absence_id)
        if not absence:
            return jsonify({'status': 'error', 'message': 'Ausencia no encontrada'}), 404

        # Calcular duración si hay horas
        duration = None
        if start_time_obj and end_time_obj:
            duration = calculate_duration(start_time_obj, end_time_obj)
            print(f"⏱ Duración calculada: {duration}")

        # Actualizar los datos de la ausencia
        absence.start_date = start_date
        absence.end_date = end_date
        absence.start_time = start_time_obj
        absence.end_time = end_time_obj
        absence.absence_type = absence_type
        absence.description = absence_description
        absence.duration_hours = duration

        # Manejo del archivo adjunto
        if absence_file and absence_file.filename:
            upload_folder = os.path.join(app.root_path, 'static', 'uploads')
            os.makedirs(upload_folder, exist_ok=True)

            filename = secure_filename(absence_file.filename)
            file_path = os.path.join(upload_folder, filename)
            absence_file.save(file_path)

            absence.file_name = filename
            absence.file_path = os.path.relpath(file_path, app.root_path)

        # Guardar cambios en la base de datos
        db.session.commit()

        print("✅ Ausencia actualizada correctamente.")
        return jsonify({'status': 'success', 'message': 'Ausencia actualizada exitosamente.', 'duration': duration}), 200

    except Exception as e:
        print(f"❌ Error inesperado: {e}")
        return jsonify({'status': 'error', 'message': 'Error interno del servidor.'}), 500

def calculate_duration(start_time, end_time):
    """Calcula la duración en formato HH:MM"""
    start_dt = datetime.combine(datetime.today(), start_time)
    end_dt = datetime.combine(datetime.today(), end_time)

    duration = end_dt - start_dt
    hours, remainder = divmod(duration.seconds, 3600)
    minutes = remainder // 60

    return f"{hours:02}:{minutes:02}"


from flask import jsonify, abort

def calculateDuration(start, end):
    # Parsea las horas en objetos datetime
    startTime = datetime.strptime(start, '%H:%M')
    endTime = datetime.strptime(end, '%H:%M')

    # Calcula la diferencia como un objeto timedelta
    duration = endTime - startTime

    return duration
from flask import jsonify, abort
from flask import jsonify, render_template, request
from datetime import datetime
from pytz import timezone as pytz_timezone
from werkzeug.utils import secure_filename
import os

@app.route('/report_absence', methods=['GET', 'POST'])
def report_absence():
    print("[DEBUG] Entrando a report_absence")

    # Verificar autenticación
    if not current_user.is_authenticated:
        print("[ERROR] Usuario no autenticado")
        return jsonify({"status": "error", "message": "No estás autenticado"}), 401

    print(f"[DEBUG] Método HTTP recibido: {request.method}")

    # Obtener notificaciones y centro de trabajo
    try:
        notificaciones = Notificacion.query.filter_by(
            usuario_id=current_user.id, leida=False
        ).order_by(Notificacion.fecha_hora.desc()).all()
        current_user_centro = current_user.Centro_trabajo
    except Exception as e:
        print(f"[ERROR] No se pudieron obtener notificaciones o centro de trabajo: {e}")
        return jsonify({"status": "error", "message": "Error al cargar datos del usuario"}), 500

    # === POST ===
    if request.method == 'POST':
        print("[DEBUG] Se detectó una solicitud POST")

        try:
            # Obtener datos del formulario
            start_date_str = request.form.get('start_date', '').strip()
            end_date_str = request.form.get('end_date', '').strip()
            absence_type = request.form.get('absence_type', '').strip()
            description = request.form.get('description', '').strip()
            start_time = request.form.get('start_time', '08:00').strip()
            end_time = request.form.get('end_time', '16:00').strip()
            absence_category = request.form.get('type', '').strip()

            # PRINTS DE DEPURACIÓN
            print(f"[DEBUG] start_date_str: {start_date_str}")
            print(f"[DEBUG] end_date_str: {end_date_str}")
            print(f"[DEBUG] absence_type recibido: '{absence_type}' (tipo: {type(absence_type)})")
            print(f"[DEBUG] description: {description}")
            print(f"[DEBUG] start_time: {start_time}, end_time: {end_time}")
            print(f"[DEBUG] absence_category: '{absence_category}'")

            # Validaciones
            if not start_date_str or not end_date_str or not absence_type or not description:
                print("[ERROR] Faltan datos obligatorios")
                return jsonify({"status": "error", "message": "Faltan datos obligatorios"}), 400

            # Conversión de fechas/horas
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            start_time_obj = datetime.strptime(start_time, '%H:%M').time()
            end_time_obj = datetime.strptime(end_time, '%H:%M').time()

            # Calcular duración
            duration = calculateDuration(start_time, end_time)
            duration_hours, remainder = divmod(duration.seconds, 3600)
            duration_minutes = remainder // 60
            formatted_duration = f"{duration_hours}:{duration_minutes:02d}"

            # === ARCHIVO ADJUNTO ===
            file_path = None
            file_name = None
            mime_type = None
            if 'file' in request.files:
                file = request.files['file']
                if file and file.filename:
                    file_name = secure_filename(file.filename)
                    file_extension = os.path.splitext(file_name)[1].lower()
                    ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.pdf', '.docx'}

                    if file_extension not in ALLOWED_EXTENSIONS:
                        print(f"[ERROR] Extensión de archivo no permitida: {file_extension}")
                        return jsonify({"status": "error", "message": "Formato de archivo no permitido"}), 400

                    file_path = os.path.join(UPLOAD_FOLDER, file_name)
                    file.save(file_path)
                    mime_type = file.mimetype
                    print(f"[DEBUG] Archivo guardado correctamente en: {file_path}")

            # === GUARDAR EN BASE DE DATOS ===
            try:
                new_absence = Absence(
                    start_date=start_date,
                    end_date=end_date,
                    absence_type=absence_type,
                    description=description,
                    user_id=current_user.id,
                    start_time=start_time_obj,
                    type=absence_category,
                    duration_hours=formatted_duration,
                    file_name=file_name,
                    file_path=file_path,
                    mime_type=mime_type
                )
                db.session.add(new_absence)
                db.session.commit()
                print("[DEBUG] Ausencia guardada correctamente en la base de datos.")
            except Exception as e:
                db.session.rollback()
                print(f"[ERROR] Error al guardar la ausencia en BD: {e}")
                return jsonify({"status": "error", "message": "Error al guardar en la base de datos"}), 500

            # === ENVIAR CORREOS Y NOTIFICACIONES ===
            try:
                print("[DEBUG] Obteniendo administradores y correos...")
                admin_users = get_admin_users()
                admin_emails = get_admin_emails()
                print(f"[DEBUG] admin_users: {admin_users}")
                print(f"[DEBUG] admin_emails: {admin_emails}")

                if not admin_users:
                    print("[ERROR] No hay administradores, gestores o encargados registrados")
                else:
                    print(f"[DEBUG] Enviando notificaciones a: {admin_emails}")

                    for email in admin_emails:
                        try:
                            print(f"[DEBUG] Enviando correo a {email}...")
                            send_email_to_user(
                                email,
                                'Nuevo informe de ausencia',
                                f'El empleado {current_user.nombre} {current_user.primer_apellido} ha reportado una ausencia. Revisa la aplicación para más detalles.'
                            )
                            print(f"[DEBUG] ✅ Correo enviado correctamente a {email}")
                        except Exception as e:
                            print(f"[ERROR] ❌ Falló el envío de correo a {email}: {e}")

                    madrid_tz = pytz_timezone('Europe/Madrid')
                    ahora_madrid = datetime.now(madrid_tz)

                    for admin in admin_users:
                        try:
                            nueva_notificacion = Notificacion(
                                usuario_id=admin.id,
                                concepto=f"{current_user.nombre} {current_user.primer_apellido} ha reportado una nueva ausencia.",
                                fecha_hora=ahora_madrid
                            )
                            db.session.add(nueva_notificacion)
                        except Exception as e:
                            print(f"[ERROR] Falló la creación de notificación para admin_id={admin.id}: {e}")

                    db.session.commit()
                    print("[DEBUG] ✅ Notificaciones creadas correctamente para todos los administradores.")

                return jsonify({"status": "success", "message": "Formulario enviado correctamente"}), 200

            except Exception as e:
                print(f"[ERROR] Error al enviar correos o crear notificaciones: {e}")
                return jsonify({"status": "error", "message": "Error al enviar notificación"}), 500

        except Exception as e:
            print(f"[ERROR] Excepción general en el bloque POST: {e}")
            return jsonify({"status": "error", "message": "Ocurrió un error inesperado"}), 500

    # === GET ===
    print("[DEBUG] Renderizando la página con GET.")
    return render_template(
        'report_absence.html',
        notificaciones=notificaciones,
        current_user_centro=current_user_centro,
        active_menu='report_absence'
    )

from sqlalchemy import and_



@app.route('/admin/absence_view')
@admin_required
def admin_absence_view():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_email = current_user.email

    # Obtener la lista de usuarios, aplicando el filtro de empresa si el usuario tiene rol 'gestor' o 'responsable'
    usuarios_query = User.query.filter(User.email != 'admin@gestionaempresa.es')

    # Filtrar por empresa si el usuario tiene rol 'gestor' o 'responsable'
    if current_user.has_role('gestor') or current_user.has_role('encargado'):
        usuarios_query = usuarios_query.filter(User.empresa_id == current_user.empresa_id)

    usuarios = usuarios_query.all()

    usuario_id_filtro = request.args.get('usuario_id')
    mes_filtro = request.args.get('mes')
    anio_filtro = request.args.get('anio')

    print(f"Filtros recibidos - Nombre/Apellido: {usuario_id_filtro}, Mes: {mes_filtro}, Año: {anio_filtro}")

    year_now = datetime.now().year
    absences_query = Absence.query

    # Filtrar ausencias por usuario si hay un filtro de usuario
    if usuario_id_filtro:
        absences_query = absences_query.filter_by(user_id=usuario_id_filtro)

    # Filtrar ausencias por mes si se ha proporcionado
    if mes_filtro:
        start_date = datetime.strptime(f'{year_now}-{mes_filtro}-01', '%Y-%m-%d')
        next_month = start_date.replace(month=start_date.month % 12 + 1)
        print(f"Filtrando por mes. Desde: {start_date}, Hasta: {next_month}")
        absences_query = absences_query.filter(
            Absence.start_date >= start_date,
            Absence.start_date < next_month
        )

    # Filtrar ausencias por año si se ha proporcionado
    if anio_filtro:
        start_date = datetime.strptime(f'{anio_filtro}-01-01', '%Y-%m-%d')
        end_date = datetime.strptime(f'{anio_filtro}-12-31', '%Y-%m-%d')
        print(f"Filtrando por año. Desde: {start_date}, Hasta: {end_date}")
        absences_query = absences_query.filter(
            Absence.start_date >= start_date,
            Absence.start_date <= end_date
        )

    # Filtrar ausencias por empresa si el usuario tiene rol 'gestor' o 'responsable'
    if current_user.has_role('gestor') or current_user.has_role('encargado'):
        absences_query = absences_query.join(User).filter(User.empresa_id == current_user.empresa_id)

    absences_query = absences_query.order_by(Absence.start_date.desc())

    # Paginación
    page = request.args.get('page', 1, type=int)
    per_page = 10

    absences_paginated = absences_query.paginate(page=page, per_page=per_page, error_out=False)

    print(f"Mostrando página {page}. Total ausencias en esta página: {len(absences_paginated.items)}")

    return render_template(
        'admin_absence_view.html',
        absences=absences_paginated.items,
        usuarios=usuarios,
        current_user_email=current_user_email,
        notificaciones=notificaciones,
        active_menu='combo_ausencias',
        pagination=absences_paginated
    )



@app.route('/admin_user_absence_history/<int:user_id>')
def admin_user_absence_history(user_id):
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_email = current_user.email
    # Asegúrate de tener la lógica para verificar si el usuario actual es un administrador
    # Puedes usar Flask-Login o cualquier otra forma de autenticación

    # Obtén las ausencias del usuario con el ID proporcionado
    user = User.query.get_or_404(user_id)

    user_absences = Absence.query.filter_by(user_id=user.id).all()


    print(user_absences)
    return render_template('admin_user_absence_history.html', notificaciones=notificaciones, user=user, user_absences=user_absences,current_user_email = current_user_email)


@app.route('/admin_absence_history', methods=['GET', 'POST'])
@admin_required
def admin_absence_history():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    # Obtén el correo del usuario actual
    current_user_email = current_user.email

    # Obtén el centro de trabajo del usuario actual
    current_user_workplace = current_user.Centro_trabajo

    if request.method == 'POST':
        # Obtén el ID del usuario seleccionado desde el formulario
        user_id = request.form.get('user_id')
        print(user_id)

        # Redirige a la página de historial del usuario seleccionado
        return redirect(url_for('admin_user_absence_history', user_id=user_id))

    # Filtrar la lista de usuarios en función del correo y el centro de trabajo

    users = User.query.filter(User.email != 'admin@gestionaempresa.es' and User.email != 'inspector@gestionaempresa.es').all()

    return render_template('admin_absence_history.html', notificaciones= notificaciones, users=users,current_user_email=current_user_email, active_menu='combo_ausencias')


"""
  iban.py -  ProInf.net 2014-01-24

  IBAN

  Es una estándar de homogeneización bancaria,
  creada por el Comité Europeo de Estándares Bancarios (ECSB).
  Está regulado en las normas ISO 13616 y EBS204.
  Su formato puede variar teniendo un máximo de 34 caracteres,
  pudiendo ser tanto números como letras.
  En España, está formado por 24 caracteres.

  Su composición es la siguiente:
   - Primeros dos dígitos: código del país según la norma ISO 3166-1
   - 2 dígitos de control, calculados según la norma ISO 13616
   - BBAN, es el número de cuenta bancaria básica.
     En España, se corresponde con el CCC (Código Cuenta Cliente).

  Ejemplos de uso:
      IBAN.convertir("1234-5678-06-1234567890")      --> "ES68 1234 5678 0612 3456 7890"
      IBAN.calcular("1234-5678-??-1234567890")       --> "ES6812345678061234567890" (68 y 06)
      IBAN.validar("ES68 1234 5678 0612 3456 7890")  --> true (68)
      IBAN.validar("1234-5678-06-1234567890")        --> true (06)
      IBAN.formatear("12345678061234567890")         --> "1234-5678-06-1234567890" (guiones)
      IBAN.formatear("ES6812345678061234567890")     --> "ES68 1234 5678 0612 3456 7890" (espacios)


  Referencias:
   http://queaprendemoshoy.com/como-se-interpretan-los-digitos-de-ccc-y-el-iban/
   http://www.integrasistemas.es/blog/general/calculo-del-iban/
   http://www.lawebdelprogramador.com/foros/Visual_Basic/1409866-Calculo_IBAN.html#i1409890
   http://es.ibancalculator.com/bic_und_iban.html
   http://www.cnb.cz/miranda2/export/sites/www.cnb.cz/cs/platebni_styk/iban/download/EBS204.pdf
"""


paisOmision = "es"


##################################################
# PUBLISHED


"""
  El parámetro número puede ser un CCC o un IBAN
  Si es un CCC retorna el IBAN correspondiente
  Si es un IBAN lo formatea
  Avisa si es un CCC incorrecto o un IBAN incorrecto

  Ejemplo1: IBAN.convertir("12345") --> "Error: No es IBAN ni CCC"
  Ejemplo2: IBAN.convertir("ES0012345678061234567890") --> "Error: IBAN incorrecto"
  Ejemplo3: IBAN.convertir("ES5212345678001234567890") --> "Error: CCC incorrecto"
  Ejemplo4: IBAN.convertir("ES6812345678061234567890") --> "ES68 1234 5678 0612 3456 7890"
  Ejemplo5: IBAN.convertir("1234-5678-06-1234567890") --> "ES68 1234 5678 0612 3456 7890"
"""
def convertir(numero, pais="es"):
    numero = limpiar(numero)
    iban = numero[-24:]
    ccc = numero[-20:]
    if not esIBAN(numero) and not esCCC(numero):
        return "Error: No es IBAN ni CCC"
    elif esIBAN(numero) and not validarIBAN(iban):
        return "Error: IBAN incorrecto"
    elif not validarCCC(ccc):
        return "Error: CCC incorrecto"
    elif esIBAN(numero):
        return formatearIBAN(iban)
    else:
        return formatearIBAN(calcularIBAN(ccc, pais))


# Ejemplo: IBAN.calcular("1234-5678-??-1234567890") --> "ES6812345678061234567890" (68 y 06)
def calcular(numero, pais="es"):
    numero = limpiar(numero)
    if esCCC(numero):
        dc = numero[8:10]
        if not dc.isdigit():
            numero = calcularCCC(numero)
        return calcularIBAN(numero, pais)
    else:
        return numero


# Ejemplo1: IBAN.validar("ES68 1234 5678 0612 3456 7890") --> True (68)
# Ejemplo2: IBAN.validar("1234-5678-06-1234567890") --> True (06)
def validar(numero):
    numero = limpiar(numero)
    if esIBAN(numero):
        return validarIBAN(numero)
    elif esCCC(numero):
        return validarCCC(numero)
    else:
        return False


# Ejemplo: IBAN.formatear("12345678061234567890") --> "1234-5678-06-1234567890"
# Ejemplo: IBAN.formatear("ES6812345678061234567890") --> "ES68 1234 5678 0612 3456 7890"
def formatear(numero, separador=None):
    numero = limpiar(numero)
    if esIBAN(numero):
        return formatearIBAN(numero, separador)
    elif esCCC(numero):
        return formatearCCC(numero, separador)
    else:
        return ""


##################################################
# HIGH LEVEL


"""
  Como se calcula los dígitos de control del IBAN
  a) Se añade al final de la BBAN, el código del país
     según la norma ISO 3166-1 y dos ceros.
  b) Si en el BBAN hay letras, convierte estas letras en números del 10 al 35,
     siguiendo el orden del abecedario A=10 y Z=35.
  c) Divide el número por 97, y quédate con el resto.
  d) Restale a 98 el resto que te quede
  e) Ya tenemos los dígitos de control, si la diferencia es menor a 10,
     añade un 0 a la izquierda.
"""

# Ejemplo: calcularIBAN("1234-5678-06-1234567890", "es") --> "ES6812345678061234567890"
def calcularIBAN(ccc, pais="es"):
    ccc = limpiar(ccc)
    pais = pais.upper()
    cifras = ccc + valorCifras(pais) + "00"
    resto = modulo(cifras, 97)
    return pais + cerosIzquierda(str(98 - resto), 2) + ccc


# Ejemplo1: validarIBAN("ES00 1234 5678 0612 3456 7890") --> False
# Ejemplo2: validarIBAN("ES68 1234 5678 0612 3456 7890") --> True
def validarIBAN(iban):
    iban = limpiar(iban)
    pais = iban[0:2]
    dc = iban[2:4]
    cifras = iban[4:] + valorCifras(pais) + dc
    resto = modulo(cifras, 97)
    return resto == 1


# Ejemplo1: validarCCC("1234-5678-00-1234567890") --> False
# Ejemplo2: validarCCC("1234-5678-06-1234567890") --> True
def validarCCC(ccc):
    ccc = limpiar(ccc)
    items = formatearCCC(ccc, " ").split()
    dc = str(modulo11(items[0] + items[1])) + str(modulo11(items[3]))
    return dc == items[2]


# Ejemplo: calcularCCC("1234-5678-??-1234567890") --> "12345678061234567890"
def calcularCCC(ccc):
    ccc = limpiar(ccc)
    return ccc[0:8] + calcularDC(ccc) + ccc[10:20]


# Ejemplo: calcularDC("1234-5678-??-1234567890") --> "06"
def calcularDC(ccc):
    ccc = limpiar(ccc)
    items = formatearCCC(ccc, " ").split()
    return str(modulo11(items[0] + items[1])) + str(modulo11(items[3]))


# Ejemplo: formatearCCC("12345678061234567890") --> "1234-5678-06-1234567890"
def formatearCCC(ccc, separador=None):
    ccc = limpiar(ccc)
    if separador == None: separador = "-"
    return ccc[0:4] + separador + ccc[4:8] + separador + ccc[8:10] + separador + ccc[10:20]


# Ejemplo: formatearIBAN("ES6812345678061234567890") --> "ES68 1234 5678 0612 3456 7890"
def formatearIBAN(iban, separador=None):
    iban = limpiar(iban)
    if separador == None: separador = " "
    items = []
    for i in range(6): items.append(iban[i*4: (i+1)*4])
    return separador.join(items)


##################################################
# LOW LEVEL


def esCCC(cifras):
    return len(cifras) == 20


def esIBAN(cifras):
    return len(cifras) == 24


# Ejemplo: limpiar("IBAN1234 5678-90") --> "1234567890"
def limpiar(numero):
    return numero \
      .replace("IBAN", "") \
      .replace(" ", "") \
      .replace("-", "")


# Ejemplo: modulo("12345678061234567890142800", 97) --> 30
def modulo(cifras, divisor):
    """
    El entero más grande en Python es 9.223.372.036.854.775.807 (2**63-1)
    que tiene 19 cifras, de las cuales las 18 últimas pueden tomar cualquier valor.
    El divisor y el resto tendrán 2 cifras. Por lo tanto CUENTA como tope
    puede ser de 16 cifras (18-2) y como mínimo de 1 cifra.
    """
    CUENTA, resto, i = 13, 0, 0
    while i < len(cifras):
        dividendo = str(resto) + cifras[i: i+CUENTA]
        resto = int(dividendo) % divisor
        i += CUENTA
    return resto


# Ejemplo1: modulo11("12345678") --> "0"
# Ejemplo2: modulo11("1234567890") --> "6"
def modulo11(cifras):
    modulos = [(2**x)%11 for x in range(10)]
    suma = 0
    cifras = cerosIzquierda(cifras, 10)
    for cifra, modulo in zip(cifras, modulos):
        suma += int(cifra) * modulo
    control = suma % 11
    return control if control < 2 else 11 - control


# Ejemplo: cerosIzquierda("7", 3) --> "007"
def cerosIzquierda(cifras, largo):
    cantidad = largo - len(cifras)
    ceros = "0"*cantidad
    return ceros + cifras


# Ejemplo: valorCifras("es") --> "1428"
def valorCifras(cifras):
    LETRAS = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ" # A=10, B=11, ... Z=35
    items = []
    for cifra in cifras:
        posicion = LETRAS.find(cifra)
        items.append(str(posicion) if posicion >= 0 else "-")
    return "".join(items)


@app.route('/edit_profile', methods=['GET', 'POST'])
@login_required
def edit_profile():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_centro = current_user.Centro_trabajo if current_user.is_authenticated else None
    form = UserProfileForm(obj=current_user)
    password_form = ChangePasswordForm()

    # Poblar las opciones del campo Centro_trabajo
    centros_trabajo = CentroTrabajo.query.all()
    form.Centro_trabajo.choices = [(c.id, c.nombre) for c in centros_trabajo]

    # Poblar las opciones del campo Departamento
    departamentos = Departamento.query.all()
    form.Departamento.choices = [(d.id, d.nombre) for d in departamentos]

    if form.validate_on_submit():
        form.populate_obj(current_user)
        db.session.commit()
        return jsonify({"status": "success", "message": "Cambios en el perfil guardados exitosamente."})

    return render_template(
        'edit_profile.html',
        form=form,
        notificaciones=notificaciones,
        password_form=password_form,
        current_user_centro=current_user_centro,
        user=current_user,
        active_menu='edit_profile'
    )


@app.route('/admin/editar_usuario/<int:user_id>', methods=['GET', 'POST'])
@admin_required  # Asume que tienes un decorador para verificar si el usuario es admin
def editar_usuario(user_id):
    user = User.query.get_or_404(user_id)
    form = UserProfileForm2(obj=user)
    password_form = ChangePasswordForm()
    empresa_id = SelectField('Empresa', coerce=int)

    # Poblar las opciones del campo Centro_trabajo
    centros_trabajo = CentroTrabajo.query.all()
    form.Centro_trabajo.choices = [(c.id, c.nombre) for c in centros_trabajo]

    # Poblar las opciones del campo Departamento
    departamentos = Departamento.query.all()
    form.Departamento.choices = [(d.id, d.nombre) for d in departamentos]
    empresas = Empresa.query.all()
    form.empresa_id.choices = [(e.id, e.nombre) for e in empresas]

    if form.validate_on_submit():
        form.populate_obj(user)
        user.empresa_id = form.empresa_id.data
        db.session.commit()
        return jsonify({"status": "success", "message": f"Perfil de {user.nombre} actualizado correctamente."})

    return render_template(
        'admin_editar_usuario.html',  # Usa una plantilla distinta si quieres
        form=form,
        password_form=password_form,
        user=user,
        active_menu='usuarios'  # Puedes cambiar esto según el menú actual
    )


from flask import render_template, redirect, url_for, flash
from flask_login import current_user

@app.route('/edit_password', methods=['GET', 'POST'])
@login_required
def edit_password():
    password_form = ChangePasswordForm()

    print("Form loaded:", password_form)

    if password_form.validate_on_submit():
        current_password = password_form.current_password.data
        new_password = password_form.new_password.data

        print("Form submitted with current_password:", current_password, "new_password:", new_password)
        print("Stored hashed password:", current_user.password)

        # Verify the hashed current password
        if check_password_hash(current_user.password, current_password):
            print("Password verification succeeded.")
            # Hash the new password before saving it to the database
            hashed_new_password = generate_password_hash(new_password)
            current_user.password = hashed_new_password
            db.session.commit()
            print("Password updated and changes committed to the database.")
            return jsonify({"status": "success", "message": "Contraseña modificada."})
        else:
            print("Password verification failed.")
            flash('La contraseña actual es incorrecta.', 'danger')
    else:
        print("Form validation failed:", password_form.errors)

    return render_template('edit_password.html', password_form=password_form)







def send_reset_password_email(user):

    reset_url = url_for('reset_password', _external=True)

    subject = 'Recuperar Contraseña'
    body = f'Haz clic en el siguiente enlace para cambiar tu contraseña: {reset_url}'

    message = Message(subject, recipients=[user.email], body=body)
    mail.send(message)


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    form = ForgotPasswordForm()

    if form.validate_on_submit():
        email = form.email.data
        user = User.query.filter_by(email=email).first()

        if user:
            send_reset_password_email(user)
            # Aquí puedes enviar un correo electrónico con un enlace único para restablecer la contraseña.
            # Este enlace debe contener algún identificador único del usuario (por ejemplo, el ID del usuario).
            # Por ahora, simplemente redirigiremos al usuario a una página de éxito.

            return redirect(url_for('login'))



    return render_template('forgot_password.html', form=form)

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    form = ResetPasswordForm()

    if form.validate_on_submit():
        # Obtén el email y busca al usuario
        email = form.email.data
        user = User.query.filter_by(email=email).first()

        if user:
            # Hashea la nueva contraseña antes de guardarla
            hashed_password = generate_password_hash(form.new_password.data)
            user.password = hashed_password

            # Guarda los cambios en la base de datos
            db.session.commit()

            return redirect(url_for('login'))

        else:
            # Si el usuario no existe, puedes agregar un mensaje de error
            flash('No se encontró un usuario con ese correo electrónico.', 'danger')

    return render_template('reset_password.html', form=form)

@app.route('/view_absence_file/<int:absence_id>', methods=['GET'])
def view_absence_file(absence_id):
    absence = Absence.query.get(absence_id)
    if absence and absence.file_path:
        print(f"Descargando archivo desde: {absence.file_path}")
        return send_file(absence.file_path,
                         as_attachment=True,
                         download_name=absence.file_name,
                         mimetype=absence.mime_type)

    print(f"Error: No se encontró el archivo con ID {absence_id}")
    return jsonify({"status": "error", "message": "Archivo no encontrado."})


@app.route('/solicitar_usuarios', methods=['GET', 'POST'])
def solicitar_usuarios():
    if request.method == 'POST':
        # Obtener datos del formulario
        admin_users = int(request.form['admin_users'])
        normal_users = int(request.form['normal_users'])
        description = request.form['description']
        message = f'Solicitud de usuarios desde CEOE\n\nNúmero de administradores: {admin_users}\nNúmero de usuarios normales: {normal_users}\n\nDescripción: {description}'
        send_email_to_user('notificaciones@gestionaempresa.es', 'Solicitud de usuarios desde CEOE', message)

        # Aquí puedes realizar las acciones necesarias con los datos (por ejemplo, almacenarlos en la base de datos)
        # ...

        # Redirigir a una página de confirmación o realizar otras acciones según tus necesidades
        return render_template('report_problem.html')

    # Si el método es GET, simplemente renderiza el formulario
    return render_template('solicitar_usuarios.html')

@app.route('/actualizar_ausencia', methods=['POST'])
def actualizar_ausencia():
    # Recibir los datos del formulario
    absence_id = request.form['absenceId']
    absence_date_str = request.form['absenceDate']

    # Convertir la cadena de fecha a objeto de fecha de Python
    absence_date = datetime.strptime(absence_date_str, '%Y-%m-%d')

    absence_type = request.form['absenceType']
    absence_description = request.form['absenceDescription']
    absence_duration = request.form['absenceDuration']
    file_name = None
    file_data = None
    mime_type = None
    print(request.form)

    if 'absenceFile' in request.files:
        print("lalala")
        file = request.files['absenceFile']
        print("lololo")
        if file.filename != '':
            print("lululu")
            # Leer los datos del archivo y obtener el tipo MIME
            file_data = file.read()
            mime_type = file.mimetype

            # Guardar el archivo en la carpeta de subida
            filename = secure_filename(file.filename)

            file_name=filename
    print(file_name)
    print(file_data)
    print(mime_type)
    # Actualizar la entrada en la base de datos
    ausencia = Absence.query.get(absence_id)
    ausencia.start_date = absence_date
    ausencia.absence_type = absence_type
    ausencia.description = absence_description
    ausencia.duration_hours = absence_duration
    if file.filename != '':
        ausencia.file_name=file_name
        ausencia.file_data=file_data
        ausencia.mime_type=mime_type

    # Guardar los cambios en la base de datos
    db.session.commit()

    # Redireccionar u ofrecer alguna otra respuesta al usuario
    return 'La ausencia ha sido actualizada exitosamente.'

import fitz  # PyMuPDF
import re

def procesar_documento_pdf(archivo):
    texto_pdf = extraer_texto_de_pdf(archivo)
    dni_encontrado = buscar_dni(texto_pdf)
    fecha_encontrada = buscar_fecha(texto_pdf)
    fecha_encontrada=fecha_encontrada[0]
    fecha_convertida = convertir_fecha(fecha_encontrada)
    return dni_encontrado, fecha_convertida

def extraer_texto_de_pdf(archivo):
    texto = ""
    # Abre el archivo PDF con PyMuPDF
    documento = fitz.open(archivo)
    for pagina in documento:
        texto += pagina.get_text()
    documento.close()
    return texto

def buscar_dni(texto):
    patron_dni = r"\b\d{8}[A-Za-z]\b"
    dnis_encontrados = re.findall(patron_dni, texto)
    return dnis_encontrados

def buscar_fecha(texto):
    patron_fecha = r"MENS \d{2} [A-Z]{3} \d{2} a \d{2} [A-Z]{3} \d{2}"
    fechas_encontradas = re.findall(patron_fecha, texto)
    print("FECHAS FUCNION",fechas_encontradas)
    return fechas_encontradas

#############################################

def procesar_documento_pdf_contrato(archivo):
    texto_pdf = extraer_texto_de_pdf_contrato(archivo)
    dni_encontrado = buscar_dni_contrato(texto_pdf)

    return dni_encontrado

def extraer_texto_de_pdf_contrato(archivo):
    texto = ""
    # Abre el archivo PDF con PyMuPDF
    documento = fitz.open(archivo)
    for pagina in documento:
        texto += pagina.get_text()
    documento.close()
    return texto

def buscar_dni_contrato(texto):
    patron_dni = r"\b\d{8}[A-Za-z]\b"
    dnis_encontrados = re.findall(patron_dni, texto)
    # Devuelve el segundo DNI encontrado si hay al menos dos DNIs, de lo contrario, devuelve None
    return dnis_encontrados[1] if len(dnis_encontrados) >= 2 else None


################################################

def procesar_documento_pdf_certificado(archivo):
    texto_pdf = extraer_texto_de_pdf_certificado(archivo)
    dni_encontrado = buscar_dni_certificado(texto_pdf)

    return dni_encontrado

def extraer_texto_de_pdf_certificado(archivo):
    texto = ""
    # Abre el archivo PDF con PyMuPDF
    documento = fitz.open(archivo)
    for pagina in documento:
        texto += pagina.get_text()
    documento.close()
    return texto

def buscar_dni_certificado(texto):
    patron_dni = r"\b\d{8}[A-Za-z]\b"
    dnis_encontrados = re.findall(patron_dni, texto)
    return dnis_encontrados





def convertir_fecha(fecha_str):
    # Dividir la cadena de fecha en partes usando el espacio como separador
    partes_fecha = fecha_str.split(' ')

    # Obtener el mes de la fecha
    mes = partes_fecha[2]

    # Mapear el mes a su equivalente numérico
    meses = {
        'ENE': '01', 'ENERO': '01',
        'FEB': '02', 'FEBRERO': '02',
        'MAR': '03', 'MARZO': '03',
        'ABR': '04', 'ABRIL': '04',
        'MAY': '05', 'MAYO': '05',
        'JUN': '06', 'JUNIO': '06',
        'JUL': '07', 'JULIO': '07',
        'AGO': '08', 'AGOSTO': '08',
        'SEP': '09', 'SEPTIEMBRE': '09',
        'OCT': '10', 'OCTUBRE': '10',
        'NOV': '11', 'NOVIEMBRE': '11',
        'DIC': '12', 'DICIEMBRE': '12'
    }
    mes_numero = meses[mes]

    # Construir la fecha en el formato 'YYYY-MM-DD'
    fecha_converted_str = f"20{partes_fecha[3]}-{mes_numero}-{partes_fecha[1]}"

    # Convertir la cadena de fecha al objeto de fecha de Python
    fecha_converted = datetime.strptime(fecha_converted_str, '%Y-%m-%d').date()

    return fecha_converted





import requests



def obtener_informacion_geografica(ip_address):
    api_key = "16e64e7141094053b2308f103b9e7621"
    url = f"https://ipgeolocation.abstractapi.com/v1/?api_key={api_key}&ip_address={ip_address}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return data
    else:
        return None


import requests
import folium
import os
import tempfile
# Función para obtener las coordenadas de una dirección IP
def obtener_coordenadas_ip(ip):
    url = f"http://ip-api.com/json/{ip}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        latitud = data['lat']
        longitud = data['lon']
        return latitud, longitud
    else:
        return None, None

import webbrowser

@app.route('/mostrar_mapa')
def mostrar_mapa():
    latitud = request.args.get('latitud')
    longitud = request.args.get('longitud')
    if latitud is not None and longitud is not None:
        # Crear un mapa en base a las coordenadas proporcionadas
        mapa = folium.Map(location=[float(latitud), float(longitud)], zoom_start=10)
        folium.Marker(location=[float(latitud), float(longitud)], popup='Ubicación').add_to(mapa)

        # Generar HTML del mapa en lugar de guardarlo en un archivo
        mapa_html = mapa._repr_html_()  # Genera el HTML en una cadena

        return mapa_html
    else:
        return jsonify({'error': 'No se proporcionaron coordenadas de ubicación.'}), 400

import os

@app.route('/mostrar_mapa_user')
def mostrar_mapa_user():
    latitud = request.args.get('latitud')
    longitud = request.args.get('longitud')
    if latitud is not None and longitud is not None:
        # Crear un mapa en base a las coordenadas proporcionadas
        mapa = folium.Map(location=[float(latitud), float(longitud)], zoom_start=10)
        folium.Marker(location=[float(latitud), float(longitud)], popup='Ubicación').add_to(mapa)

        # Generar HTML del mapa en lugar de guardarlo en un archivo
        mapa_html = mapa._repr_html_()  # Genera el HTML en una cadena

        return mapa_html
    else:
        return jsonify({'error': 'No se proporcionaron coordenadas de ubicación.'}), 400







# Zona horaria de España
timezone_spain = pytz_timezone('Europe/Madrid')
@app.route('/registro', methods=['POST'])
def registro():
    print("ENTRAAAAA en la función de registro")
    hora_actual = datetime.now().astimezone(timezone_spain)
    print(f"[DEBUG] Hora actual en España: {hora_actual}")

    if current_user.is_authenticated:
        print("Usuario autenticado")
        user_id = current_user.id


        # --- NUEVO BLOQUE: evitar duplicados recientes ---
        ultimo_registro = (
            Registro.query
            .filter_by(user_id=user_id)
            .order_by(Registro.entrada.desc())
            .first()
        )

        if ultimo_registro:
            ultima_hora = ultimo_registro.salida or ultimo_registro.entrada
            if ultima_hora:
                hora_actual_naive = hora_actual.replace(tzinfo=None)
                # Evitar si el último registro es de hace menos de 3 segundos
                if (hora_actual_naive - ultima_hora) < timedelta(seconds=3):
                    print(f"[DEBUG] Registro duplicado detectado para user_id={user_id}")
                    return jsonify({'error': 'Registro duplicado detectado'}), 429
        # --- FIN BLOQUE NUEVO ---

        # Verificar si hay un registro de entrada sin salida
        entrada_sin_salida = Registro.query.filter_by(user_id=user_id, salida=None).first()
        print("Consulta de entrada sin salida completada")

        try:
            if not request.is_json:
                print("Solicitud no es JSON")
                print(f"Contenido recibido: {request.data}")
                return jsonify({'error': 'Solicitud debe ser en formato JSON'}), 400

            data = request.get_json()
            latitud = data.get('latitud') if data else None
            longitud = data.get('longitud') if data else None
            print(f"Latitud: {latitud}, Longitud: {longitud}")

            ip_address = request.headers.get('X-Real-IP')

            if entrada_sin_salida:
                print("Registrando salida")
                entrada_sin_salida.salida = hora_actual
                entrada_sin_salida.ip_address_salida = ip_address
                entrada_sin_salida.latitud_salida = latitud
                entrada_sin_salida.longitud_salida = longitud
                db.session.commit()
                print("Salida registrada correctamente")
                return jsonify({'message': 'Salida registrada correctamente.', 'entrada_activa': False}), 200
            else:
                print("Registrando entrada")
                registro = Registro(
                    user_id=user_id,
                    entrada=hora_actual,
                    ip_address_entrada=ip_address,
                    latitud=latitud,
                    longitud=longitud
                )
                db.session.add(registro)
                db.session.commit()
                print("Entrada registrada correctamente")
                return jsonify({'message': 'Entrada registrada correctamente.', 'entrada_activa': True}), 200
        except Exception as e:
            print(f"Error procesando la solicitud: {e}")
            return jsonify({'error': f'Error procesando la solicitud: {e}'}), 500
    else:
        print("Usuario no autenticado")
        return jsonify({'error': 'Usuario no autenticado'}), 401


import pytz

from flask import render_template

@app.route('/registro_qr', methods=['GET'])
def registro_qr():
    try:
        hora_actual = datetime.now().astimezone(timezone_spain)
        user_id = request.args.get('user_id')
        centro_trabajo_id = request.args.get('centro_id')

        print(f"[DEBUG] user_id: {user_id}, centro_trabajo_id: {centro_trabajo_id}")

        if not user_id or not centro_trabajo_id:
            return render_template('registro_error.html', mensaje="Faltan parámetros obligatorios."), 400

        usuario = User.query.get(user_id)
        if not usuario:
            return render_template('registro_error.html', mensaje="Usuario no encontrado."), 404

        ip_address = request.headers.get('X-Real-IP', request.remote_addr)
        print(f"[DEBUG] IP del usuario: {ip_address}")

        # --- Verificación de IP ---
        if usuario.bloquear_por_ip:
            if not usuario.empresa or not usuario.empresa.ip_empresa:
                return render_template('registro_error.html',
                                       mensaje="No se puede verificar la IP: la empresa no tiene IP configurada."), 400
            if ip_address != usuario.empresa.ip_empresa:
                print(f"[DEBUG] Acceso denegado: {ip_address} != {usuario.empresa.ip_empresa}")
                return render_template('registro_error.html',
                                       mensaje="Tu IP no coincide con la IP registrada de la empresa.",
                                       ip=ip_address), 403
        # ---------------------------

        # --- NUEVO BLOQUE: evitar duplicados recientes ---
        ultimo_registro = (
            Registro.query
            .filter_by(user_id=user_id)
            .order_by(Registro.entrada.desc())
            .first()
        )

        if ultimo_registro:
            ultima_hora = ultimo_registro.salida or ultimo_registro.entrada
            if ultima_hora:
                hora_actual_naive = hora_actual.replace(tzinfo=None)
                # Evitar si el último registro es de hace menos de 3 segundos
                if (hora_actual_naive - ultima_hora) < timedelta(seconds=3):
                    print(f"[DEBUG] Registro duplicado detectado para user_id={user_id}")
                    return jsonify({'error': 'Registro duplicado detectado'}), 429
        # --- FIN BLOQUE NUEVO ---

        entrada_sin_salida = Registro.query.filter_by(user_id=user_id, salida=None).first()

        if entrada_sin_salida:
            entrada_sin_salida.salida = hora_actual
            entrada_sin_salida.ip_address_salida = ip_address
            db.session.commit()
            duracion_formateada = entrada_sin_salida.duracion_formateada

            return redirect(url_for('registro_resultado',
                                    user_id=user_id,
                                    centro_id=centro_trabajo_id,
                                    tipo='salida',
                                    duracion=duracion_formateada))
        else:
            nuevo_registro = Registro(
                user_id=user_id,
                entrada=hora_actual,
                ip_address_entrada=ip_address,
                centro_trabajo_id=centro_trabajo_id
            )
            db.session.add(nuevo_registro)
            db.session.commit()

            return redirect(url_for('registro_resultado',
                                    user_id=user_id,
                                    centro_id=centro_trabajo_id,
                                    tipo='entrada'))

    except Exception as e:
        print(f"[QR] Error en registro QR: {e}")
        return render_template('registro_error.html',
                               mensaje=f"Error procesando la solicitud: {e}"), 500


@app.route('/registro_corregido', methods=['POST'])
@login_required
def registro_corregido():
    print("Entra en registro corregidooooooooooooo")

    try:
        data = request.get_json()
        hora_salida_text = data.get('hora_salida')
        latitud = data.get('latitud')
        longitud = data.get('longitud')

        timezone_spain = pytz_timezone("Europe/Madrid")
        now = datetime.now().astimezone(timezone_spain)

        # Buscar la última entrada sin salida
        entrada_sin_salida = Registro.query.filter_by(user_id=current_user.id, salida=None).order_by(Registro.entrada.desc()).first()

        if not entrada_sin_salida:
            return jsonify({'error': 'No se encontró una entrada sin salida'}), 404

        # Convertir entrada a aware datetime si es naive
        if entrada_sin_salida.entrada.tzinfo is None:
            entrada_sin_salida.entrada = timezone_spain.localize(entrada_sin_salida.entrada)

        tiempo_transcurrido = now - entrada_sin_salida.entrada

        # Si han pasado más de 12 horas desde la entrada y no hay salida, se asigna entrada + 8 horas como salida
        if tiempo_transcurrido > timedelta(hours=12):
            hora_salida = entrada_sin_salida.entrada + timedelta(hours=8)
            print("Salida automática asignada:", hora_salida)
        else:
            if not hora_salida_text:
                return jsonify({'error': 'No se proporcionó la hora de salida'}), 400

            hora_salida = datetime.strptime(hora_salida_text, "%H:%M:%S").time()
            fecha_registro = entrada_sin_salida.entrada.date()
            hora_salida = timezone_spain.localize(datetime.combine(fecha_registro, hora_salida))

        # Guardar la salida en el registro
        entrada_sin_salida.salida = hora_salida
        entrada_sin_salida.latitud_salida = latitud
        entrada_sin_salida.longitud_salida = longitud
        entrada_sin_salida.ip_address_salida = request.headers.get('X-Real-IP')

        db.session.commit()
        print("Salida corregida correctamente")

        return jsonify({'success': True, 'message': 'Salida corregida correctamente'}), 200

    except Exception as e:
        print(f"Error en registro_corregido: {e}")
        return jsonify({'error': f'Error al corregir la salida: {e}'}), 500


@app.route('/estado_registro', methods=['GET'])
def estado_registro():
    if current_user.is_authenticated:
        user_id = current_user.id
        entrada_sin_salida = Registro.query.filter_by(user_id=user_id, salida=None).first()
        entrada_activa = entrada_sin_salida is not None
        return jsonify({'entrada_activa': entrada_activa}), 200
    else:
        return jsonify({'error': 'Usuario no autenticado'}), 401


# Ruta para descargar el mapa del registro
@app.route('/descargar_mapa/<registro_id>')
def descargar_mapa(registro_id):
    registro = Registro.query.get(registro_id)
    if registro:
        archivo_mapa = mostrar_mapa(registro.latitud, registro.longitud)
        if archivo_mapa:
            return send_file(archivo_mapa, as_attachment=True)
        else:
            return jsonify({'error': 'No se pudo generar el mapa.'}), 500
    else:
        return jsonify({'error': 'Registro no encontrado.'}), 404



@app.route('/registros', methods=['GET', 'POST'])
def obtener_registros():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()

    # Obtener los filtros desde los parámetros de la URL
    mes = request.args.get('mess')
    anio = request.args.get('anio')
    usuario = request.args.get('usuario')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)  # Número de registros por página
    centro_trabajo = request.args.get('centro_trabajo')


    if not anio or anio == "":  # Si no se especifica un año o está vacío
        anio = datetime.now().year
    else:
        anio = int(anio)

    # Si 'mes' es vacío o "Todos", no asignar el mes actual
    if not mes or mes == "":
        mes = None  # No se filtra por mes
    else:
        mes = int(mes)

    # Consulta base para los registros
    query = Registro.query.order_by(Registro.entrada.desc())

    # Filtrar por año si se selecciona un año específico
    if anio:
        query = query.filter(extract('year', Registro.entrada) == anio)

    # Filtrar por mes si se selecciona un mes específico
    if mes and mes != "":
        query = query.filter(extract('month', Registro.entrada) == mes)

    # Filtrar por usuario si se selecciona un usuario específico
    if usuario and usuario != ' ':
        query = query.filter(Registro.user_id == int(usuario))

    # Filtrar por centro de trabajo si se selecciona uno
    if centro_trabajo and centro_trabajo != "":
        query = query.filter(Registro.centro_trabajo_id == int(centro_trabajo))


    # Verificar el rol del usuario actual
    if current_user.has_role('gestor') or current_user.has_role('encargado'):
        # Filtrar por la empresa del usuario actual
        empresa_usuario = current_user.empresa_id  # Suponiendo que el usuario tiene un campo 'empresa_id'
        query = query.join(User).filter(User.empresa_id == empresa_usuario)

    # Obtener los registros filtrados con paginación
    registros_paginados = query.paginate(page=page, per_page=per_page)
    registros = registros_paginados.items

    # --- Calcular las horas trabajadas ---
    def calcular_horas(registros):
        total_horas = sum(
            ((registro.salida - registro.entrada).total_seconds() / 3600)
            for registro in registros
            if registro.entrada and registro.salida
        )
        return round(total_horas)

    # --- Horas trabajadas en el año actual ---
    query_anio = query.filter(extract('year', Registro.entrada) == datetime.now().year)
    registros_anuales = query_anio.all()
    total_horas_anuales = calcular_horas(registros_anuales)

    # --- Horas trabajadas en el mes actual ---
    query_mes_actual = query.filter(
        extract('year', Registro.entrada) == datetime.now().year,
        extract('month', Registro.entrada) == datetime.now().month
    )
    registros_mes_actual = query_mes_actual.all()
    total_horas_mes_actual = calcular_horas(registros_mes_actual)

    # --- Horas trabajadas en la semana actual ---
    hoy = datetime.now()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    fin_semana = inicio_semana + timedelta(days=6)

    query_semana = query.filter(Registro.entrada >= inicio_semana, Registro.entrada <= fin_semana)
    registros_semana_actual = query_semana.all()
    total_horas_semana_actual = calcular_horas(registros_semana_actual)

    # --- Horas trabajadas en el mes seleccionado ---
    if mes and mes != "":
        mes = int(mes)
        query_mes_anio = query.filter(
            extract('year', Registro.entrada) == anio,
            extract('month', Registro.entrada) == mes
        )
    else:
        mes_actual = datetime.now().month
        query_mes_anio = query.filter(
            extract('year', Registro.entrada) == anio,
            extract('month', Registro.entrada) == mes_actual
        )

    # Obtener los registros del mes seleccionado
    registros_mes_anio = query_mes_anio.all()
    total_horas_mes = calcular_horas(registros_mes_anio)

    # --- Desglose de horas trabajadas por semana del mes seleccionado ---
    semanas_mes = {}
    for registro in registros_mes_anio:
        if registro.entrada and registro.salida:
            semana_num = registro.entrada.isocalendar()[1]
            if semana_num not in semanas_mes:
                semanas_mes[semana_num] = 0
            semanas_mes[semana_num] += (registro.salida - registro.entrada).total_seconds() / 3600

    semanas_mes = dict(sorted(semanas_mes.items()))
    semanas_mes = {semana: round(horas) for semana, horas in semanas_mes.items()}

    # --- Cálculo para "Todos los usuarios" ---
    if not usuario:
        total_horas_empleados = 0
        users = User.query.all()
        for user in users:
            registros_usuario = query.filter(Registro.user_id == user.id).all()
            horas_usuario = calcular_horas(registros_usuario)
            total_horas_empleados += horas_usuario

        total_horas_deberian_trabajar = 160 * len(users)
    else:
        total_horas_empleados = total_horas_mes
        total_horas_deberian_trabajar = 160

    total_horas_mes_actual_completa = convertir_a_hms(total_horas_mes_actual)

    # Obtener todos los usuarios
    users = User.query.filter(User.email != 'admin@gestionaempresa.es' and User.email != 'inspector@gestionaempresa.es').all()

    # Contar las pausas asociadas a cada registro
    for registro in registros:
        registro.num_pausas = Pausa.query.filter_by(registro_id=registro.id).count()

    current_year = datetime.now().year
    rol = 'inspector_role' if current_user.email == 'inspector@gestionaempresa.es' else None

    empresa = Empresa.query.first()
    if current_user.has_role('encargado'):
        user_role="encargado"

    centros_trabajo = CentroTrabajo.query.all()


    return render_template('registros.html',
                           registros=registros,
                           current_year=current_year,
                           centros_trabajo=centros_trabajo,
                           users=users,
                           obtener_nombre_usuario_por_id=obtener_nombre_usuario_por_id,
                           active_menu='obtener_registros',
                           mes=mes,
                           anio=anio,
                           usuario=usuario,
                           total_horas_anuales=total_horas_anuales,
                           total_horas_mes_actual=total_horas_mes_actual,
                           total_horas_semana_actual=total_horas_semana_actual,
                           total_horas_mes=total_horas_mes,
                           semanas_mes=semanas_mes,
                           notificaciones=notificaciones,
                           total_horas_empleados=total_horas_empleados,
                           total_horas_deberian_trabajar=total_horas_deberian_trabajar,
                           total_horas_mes_actual_completa=total_horas_mes_actual_completa,
                           registros_paginados=registros_paginados,
                           empresa=empresa,

                           rol=rol)






def convertir_a_hms(horas_decimal):
    horas = int(horas_decimal)
    minutos = int((horas_decimal - horas) * 60)
    segundos = int(((horas_decimal - horas) * 60 - minutos) * 60)
    return f"{horas:02}:{minutos:02}:{segundos:02}"


@app.route('/registros_user', methods=['GET', 'POST'])
def obtener_registros_user():
    print("🔍 Entrando en la vista /registros_user")
    print(f"👤 Usuario actual: {current_user.id} - {current_user.email}")

    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    user_id = current_user.id

    mes = request.args.get('mes')
    anio = request.args.get('anio')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)

    print(f"📥 Parámetros recibidos -> mes: {mes}, año: {anio}, página: {page}, por página: {per_page}")

    if not anio:
        anio = datetime.now().year
    else:
        anio = int(anio)
    print(f"📅 Año usado: {anio}")

    query = Registro.query.filter(Registro.user_id == user_id).order_by(Registro.entrada.desc())

    # Horas del año actual
    query_anio = query.filter(extract('year', Registro.entrada) == datetime.now().year)
    registros_anuales = query_anio.all()
    print(f"📊 Registros del año actual encontrados: {len(registros_anuales)}")

    total_horas_anuales = sum(
        ((r.salida - r.entrada).total_seconds() / 3600)
        for r in registros_anuales if r.entrada and r.salida
    )
    print(f"⏱ Total horas anuales: {total_horas_anuales}")

    # Horas del mes actual
    query_mes_actual = query.filter(
        extract('year', Registro.entrada) == datetime.now().year,
        extract('month', Registro.entrada) == datetime.now().month
    )
    registros_mes_actual = query_mes_actual.all()
    print(f"📆 Registros del mes actual: {len(registros_mes_actual)}")

    total_horas_mes_actual = sum(
        ((r.salida - r.entrada).total_seconds() / 3600)
        for r in registros_mes_actual if r.entrada and r.salida
    )
    print(f"⏱ Total horas mes actual: {total_horas_mes_actual}")

    # Semana actual
    hoy = datetime.now()
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    fin_semana = inicio_semana + timedelta(days=6)

    query_semana = query.filter(Registro.entrada >= inicio_semana, Registro.entrada <= fin_semana)
    registros_semana_actual = query_semana.all()
    print(f"📅 Registros semana actual: {len(registros_semana_actual)}")

    total_horas_semana_actual = sum(
        ((r.salida - r.entrada).total_seconds() / 3600)
        for r in registros_semana_actual if r.entrada and r.salida
    )
    print(f"⏱ Total horas semana actual: {total_horas_semana_actual}")

    # Mes seleccionado o mes actual
    # Mes seleccionado o todos los meses del año
    if mes:
        mes = int(mes)
        print(f"🗓 Mes seleccionado: {mes}")
        query_mes_anio = query.filter(
            extract('year', Registro.entrada) == anio,
            extract('month', Registro.entrada) == mes
        )
    else:
        print(f"🗓 Mostrando TODOS los meses del año: {anio}")
        query_mes_anio = query.filter(
            extract('year', Registro.entrada) == anio
        )


    registros_mes_anio = query_mes_anio.all()
    print(f"📂 Registros en mes/año seleccionado: {len(registros_mes_anio)}")

    total_horas_mes = sum(
        ((r.salida - r.entrada).total_seconds() / 3600)
        for r in registros_mes_anio if r.entrada and r.salida
    )
    print(f"⏱ Total horas mes seleccionado: {total_horas_mes}")

    semanas_mes = {}
    for r in registros_mes_anio:
        if r.entrada and r.salida:
            semana_num = r.entrada.isocalendar()[1]
            semanas_mes[semana_num] = semanas_mes.get(semana_num, 0) + (r.salida - r.entrada).total_seconds() / 3600

    semanas_mes = dict(sorted(semanas_mes.items()))
    semanas_mes = {sem: round(h) for sem, h in semanas_mes.items()}
    print(f"📊 Horas por semana en mes seleccionado: {semanas_mes}")

    total_horas_mes_actual_completa = convertir_a_hms(total_horas_mes_actual)

    total_horas_anuales = round(total_horas_anuales)
    total_horas_mes_actual = round(total_horas_mes_actual)
    total_horas_semana_actual = round(total_horas_semana_actual)
    total_horas_mes = round(total_horas_mes)

    for r in registros_mes_anio:
        r.num_pausas = Pausa.query.filter_by(registro_id=r.id).count()

    registros_paginados = query_mes_anio.paginate(page=page, per_page=per_page)
    print(f"📄 Mostrando página {page} de registros paginados (por página: {per_page})")

    empresa = Empresa.query.first()

    print("✅ Preparando render_template con todos los datos")

    return render_template(
        'registros_user.html',
        registros=registros_paginados.items,
        registros_paginados=registros_paginados,
        active_menu='obtener_registros_user',
        mes=mes,
        anio=anio,
        total_horas_anuales=total_horas_anuales,
        total_horas_mes_actual=total_horas_mes_actual,
        total_horas_semana_actual=total_horas_semana_actual,
        total_horas_mes=total_horas_mes,
        semanas_mes=semanas_mes,
        notificaciones=notificaciones,
        total_horas_mes_actual_completa=total_horas_mes_actual_completa,
        empresa=empresa,
        current_user=current_user
    )



from flask import request, make_response
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


from reportlab.platypus import PageTemplate, BaseDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus.frames import Frame
from reportlab.lib.styles import ParagraphStyle


from reportlab.platypus import  PageTemplate, Frame, Paragraph
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super(NumberedCanvas, self).__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        """Add page info to each page (page x of y)"""
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super(NumberedCanvas, self).showPage()

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.drawRightString(200*inch, 0.75*inch,
                             "Page %d of %d" % (self._pageNumber, page_count))
import locale

from reportlab.lib import colors

import calendar

@app.route('/generar_pdf', methods=['POST'])
def generar_pdf():
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
             'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    usuario_id = request.form['usuario']
    mes = request.form['mes'].strip()

    if len(mes.split('-')) > 2:
        partes = mes.split('-')
        if len(partes) >= 3:
            mes = f"{partes[0]}-{partes[2]}"

    if not re.match(r"^\d{4}-\d{2}$", mes):
        return "El formato del mes es inválido. Debe ser 'YYYY-MM'.", 400

    mes_datetime = datetime.strptime(mes, "%Y-%m")
    year = mes_datetime.year
    month = mes_datetime.month

    from calendar import monthrange
    _, num_days = monthrange(year, month)
    dias_mes = [datetime(year, month, day) for day in range(1, num_days + 1)]

    registros = Registro.query.filter_by(user_id=usuario_id).filter(
        db.func.strftime("%Y-%m", Registro.entrada) == mes
    ).all()

    usuario = User.query.get(usuario_id)
    nombre_usuario = usuario.nombre
    apellidos_usuario = usuario.primer_apellido

    empresa = current_user.empresa if current_user.empresa else Empresa.query.first()
    nombre_empresa = empresa.nombre

    # Traer festivos de la empresa
    festivos = Festivo.query.filter_by(empresa_id=empresa.id).all()
    festivos_set = {(f.dia, f.mes) for f in festivos}  # más rápido para consultar

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    content = []

    sample_styles = getSampleStyleSheet()
    title = Paragraph("Listado mensual de registro laboral", sample_styles['Title'])
    content.append(title)
    content.append(Paragraph("<br/><br/>", sample_styles['Normal']))

    subtitle_style = ParagraphStyle('Subtitle', parent=sample_styles['Normal'], fontSize=8, alignment=0)
    subtitle_text = (
        f"Mes: {meses[mes_datetime.month - 1]} {mes_datetime.year}<br/>"
        f"Empleado: {nombre_usuario} {apellidos_usuario}<br/>"
        f"Empresa: {nombre_empresa}<br/>"
    )
    content.append(Paragraph(subtitle_text, subtitle_style))

    data = [["Fecha", "Hora de entrada", "Hora de salida", "Duración", "Número de pausas", "Observaciones"]]
    ausencias = Absence.query.filter_by(user_id=usuario_id).all()
    vacaciones = VacationRequest.query.filter_by(user_id=usuario_id, is_accepted=True).all()

    # Guardar los estilos dinámicos
    table_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]

    for idx, dia in enumerate(dias_mes, start=1):
        fecha_dia = dia.date()
        fecha_entrada = fecha_dia.strftime("%d-%m-%Y")

        registros_dia = [r for r in registros if r.entrada.date() == fecha_dia]

        if registros_dia:
            for registro in registros_dia:
                data.append([
                    fecha_entrada,
                    registro.entrada.strftime("%H:%M:%S"),
                    registro.salida.strftime("%H:%M:%S") if registro.salida else "",
                    registro.duracion_formateada,
                    Pausa.query.filter_by(registro_id=registro.id).count(),
                    registro.observaciones or ""
                ])
        else:
            ausencia = next((a for a in ausencias if a.start_date <= fecha_dia <= a.end_date), None)
            if ausencia:
                tipo_ausencia = "Ausencia justificada" if ausencia.file_name else "Ausencia no justificada"
                data.append([fecha_entrada, "", "", "", "", tipo_ausencia])
            else:
                vacaciones_dia = next((v for v in vacaciones if v.start_date <= fecha_dia <= v.end_date), None)
                if vacaciones_dia:
                    data.append([fecha_entrada, "", "", "", "", "Vacaciones"])
                else:
                    data.append([fecha_entrada, "", "", "", "", ""])

        # Comprobar si es fin de semana o festivo
        fila_index = len(data) - 1  # índice en data
        if dia.weekday() >= 5:  # sábado o domingo
            table_styles.append(('BACKGROUND', (0, fila_index), (-1, fila_index), colors.lightgrey))
        elif (dia.day, dia.month) in festivos_set:
            table_styles.append(('BACKGROUND', (0, fila_index), (-1, fila_index), colors.lightpink))

    col_widths = [inch * 1.5, inch * 1.5, inch * 1.5, inch * 1.5, inch * 1.5, inch * 2.5]
    table = Table(data, colWidths=col_widths)
    table.setStyle(table_styles)
    content.append(table)

    doc.build(content)
    pdf_data = buffer.getvalue()
    buffer.close()

    filename = f"{apellidos_usuario}_{nombre_usuario}_{meses[mes_datetime.month - 1]}_{mes_datetime.year}.pdf"
    response = make_response(pdf_data)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response




@app.route('/previsualizar_pdf')
def previsualizar_pdf():
    usuario_id = request.args.get('usuario')
    mes = request.args.get('mes').strip()  # Elimina espacios en blanco adicionales
    año = request.args.get('año').strip()

    # Validar y corregir el formato del mes si es necesario
    if len(mes.split('-')) > 2:
        partes = mes.split('-')
        if len(partes) >= 3:
            mes = f"{partes[0]}-{partes[2]}"

    # Validar el formato resultante
    if not re.match(r"^\d{4}-\d{2}$", mes):
        return "El formato del mes es inválido. Debe ser 'YYYY-MM'.", 400

    # Convertir a objeto datetime
    mes_datetime = datetime.strptime(mes, "%Y-%m")
    year = mes_datetime.year
    month = mes_datetime.month

    # Generar lista de todos los días del mes
    from calendar import monthrange
    _, num_days = monthrange(year, month)
    dias_mes = [datetime(year, month, day) for day in range(1, num_days + 1)]

    # Consultar los registros del usuario para el mes especificado
    registros = Registro.query.filter_by(user_id=usuario_id).filter(
        db.func.strftime("%Y-%m", Registro.entrada) == mes
    ).all()

    # Obtener el nombre y apellidos del usuario
    usuario = User.query.get(usuario_id)
    nombre_usuario = usuario.nombre
    apellidos_usuario = usuario.primer_apellido

    # Obtener los datos de la empresa
    empresa = current_user.empresa if current_user.empresa else Empresa.query.first()  # Ajusta esto si necesitas filtrar por un criterio específico
    nombre_empresa = empresa.nombre
    CIF_empresa = empresa.CIF
    domicilio_empresa = empresa.Domicilio
    provincia_empresa = empresa.Provincia
    localidad_empresa = empresa.Localidad
    CP_empresa = empresa.CP

    # Crear un objeto BytesIO para almacenar el PDF en memoria
    buffer = BytesIO()

    # Crear el PDF con tamaño A4 y orientación horizontal
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))

    # Crear el contenido del PDF
    content = []

    # Agregar el título
    sample_styles = getSampleStyleSheet()
    title_style = sample_styles['Title']
    title_text = "Listado mensual de registro laboral"
    title = Paragraph(title_text, title_style)
    content.append(title)
    content.append(Paragraph("<br/><br/>", title_style))

    # Agregar el subtítulo con los datos del usuario y la empresa
    subtitle_style = ParagraphStyle('Subtitle', parent=title_style, fontSize=8, alignment=0)
    subtitle_text = (
        f"Mes: {mes_datetime.strftime('%B %Y')}<br/>"
        f"Empleado: {nombre_usuario} {apellidos_usuario}<br/>"
        f"Nombre de la empresa: {nombre_empresa}<br/>"
        f"CIF: {CIF_empresa}<br/>"
    )
    subtitle = Paragraph(subtitle_text, subtitle_style)
    content.append(subtitle)

    # Construir el contenido de la tabla
    data = [["Fecha", "Hora de entrada", "Hora de salida", "Duración", "Número de pausas", "Observaciones"]]

    # Buscar ausencias y vacaciones
    ausencias = Absence.query.filter_by(user_id=usuario_id).all()
    vacaciones = VacationRequest.query.filter_by(user_id=usuario_id, is_accepted=True).all()

    # Construir las filas del mes
    for dia in dias_mes:
        fecha_dia = dia.date()
        fecha_entrada = fecha_dia.strftime("%d-%m-%Y")

        # Verificar si el día tiene un registro
        registros_dia = [r for r in registros if r.entrada.date() == fecha_dia]

        if registros_dia:
            for registro in registros_dia:
                hora_entrada = registro.entrada.strftime("%H:%M:%S")
                hora_salida = registro.salida.strftime("%H:%M:%S") if registro.salida else ""
                duracion = registro.duracion_formateada
                num_pausas = Pausa.query.filter_by(registro_id=registro.id).count()
                observaciones = registro.observaciones
                data.append([fecha_entrada, hora_entrada, hora_salida, duracion, num_pausas, observaciones])
        else:
            # Verificar si hay ausencia para ese día
            ausencia = next((a for a in ausencias if a.start_date <= fecha_dia <= a.end_date), None)
            if ausencia:
                tipo_ausencia = "Ausencia justificada" if ausencia.file_name else "Ausencia no justificada"
                data.append([fecha_entrada, "", "", "", "", tipo_ausencia])
            else:
                # Verificar si hay vacaciones para ese día
                vacaciones_dia = next((v for v in vacaciones if v.start_date <= fecha_dia <= v.end_date), None)
                if vacaciones_dia:
                    data.append([fecha_entrada, "", "", "", "", "Vacaciones"])
                else:
                    # Línea vacía cuando no haya datos
                    data.append([fecha_entrada, "", "", "", "", ""])

    # Establecer las anchuras de columna
    col_widths = [inch * 1.5, inch * 1.5, inch * 1.5, inch * 1.5, inch * 1.5, inch * 2.5]
    table = Table(data, colWidths=col_widths)

    # Aplicar estilo a la tabla
    style = [('BACKGROUND', (0, 0), (-1, 0), colors.grey),
             ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
             ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
             ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
             ('GRID', (0, 0), (-1, -1), 1, colors.black)]
    table.setStyle(style)

    # Agregar la tabla al contenido
    content.append(table)

    # Construir el PDF
    doc.build(content)

    # Obtener los datos del PDF del búfer
    pdf_data = buffer.getvalue()
    buffer.close()

    # Enviar el PDF como contenido en línea
    return Response(pdf_data, mimetype='application/pdf')








import io
import csv
from flask import Response, request

import locale
import xlsxwriter


@app.route('/generar_csv', methods=['POST'])
def generar_csv():
    usuario_id = request.form['usuario']
    mes = request.form['mes'].strip()

    # Validar y corregir el formato del mes
    if len(mes.split('-')) > 2:
        partes = mes.split('-')
        if len(partes) >= 3:
            mes = f"{partes[0]}-{partes[2]}"

    if not re.match(r"^\d{4}-\d{2}$", mes):
        return "El formato del mes es inválido. Debe ser 'YYYY-MM'.", 400

    mes_datetime = datetime.strptime(mes, "%Y-%m")
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

    registros = Registro.query.filter_by(user_id=usuario_id).filter(
        db.func.strftime("%Y-%m", Registro.entrada) == mes
    ).all()

    usuario = User.query.get(usuario_id)
    empresa = current_user.empresa if current_user.empresa else Empresa.query.first()

    # Crear archivo Excel en memoria
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Registros Mensuales")

    # Formatos de celdas
    header_format = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9', 'align': 'center'})
    normal_format = workbook.add_format({'align': 'center'})
    red_format = workbook.add_format({'bg_color': '#FFCCCC', 'align': 'center'})  # Sin registros
    orange_format = workbook.add_format({'bg_color': '#FFD966', 'align': 'center'})  # Ausencias
    blue_format = workbook.add_format({'bg_color': '#CCE5FF', 'align': 'center'})  # Vacaciones
    total_format = workbook.add_format({'bold': True, 'bg_color': '#C6E0B4', 'align': 'center'})  # Total horas

    # Encabezados de la empresa
    worksheet.write(0, 0, "Nombre de la empresa", header_format)
    worksheet.write(0, 1, empresa.nombre, normal_format)
    worksheet.write(1, 0, "CIF", header_format)
    worksheet.write(1, 1, empresa.CIF, normal_format)
    worksheet.write(3, 0, "Mes", header_format)
    worksheet.write(3, 1, mes_datetime.strftime('%B %Y'), normal_format)
    worksheet.write(4, 0, "Empleado", header_format)
    worksheet.write(4, 1, f"{usuario.nombre} {usuario.primer_apellido}", normal_format)

    # Encabezados de la tabla
    headers = ["Fecha", "Hora de entrada", "Hora de salida", "Duración", "Número de pausas", "Observaciones"]
    for col, header in enumerate(headers):
        worksheet.write(6, col, header, header_format)

    # Variables para procesar días del mes
    first_day_of_month = mes_datetime.replace(day=1)
    last_day_of_month = (first_day_of_month.replace(month=first_day_of_month.month + 1) - timedelta(days=1))
    current_day = first_day_of_month
    row = 7  # Inicia la tabla en la fila 7
    total_duracion = timedelta()

    # Iterar sobre cada día del mes
    while current_day <= last_day_of_month:
        registros_del_dia = [r for r in registros if r.entrada.date() == current_day.date()]
        ausencias = Absence.query.filter(
            Absence.start_date <= current_day.date(),
            Absence.end_date >= current_day.date(),
            Absence.user_id == usuario_id
        ).all()
        vacaciones = VacationRequest.query.filter(
            VacationRequest.start_date <= current_day.date(),
            VacationRequest.end_date >= current_day.date(),
            VacationRequest.user_id == usuario_id,
            VacationRequest.is_accepted == True
        ).all()

        observaciones = []
        formato = normal_format  # Formato por defecto

        # **Orden de prioridad para el formato de color**
        if vacaciones:
            observaciones.append("Vacaciones")
            formato = blue_format  # Color azul para vacaciones
        elif ausencias:
            justificada = "justificada" if ausencias[0].file_path else "no justificada"
            observaciones.append(f"Ausencia ({justificada})")
            formato = orange_format  # Color amarillo para ausencias
        elif not registros_del_dia:
            formato = red_format  # Color rojo si no hay registros

        if registros_del_dia:
            for registro in registros_del_dia:
                fecha_entrada = registro.entrada.strftime("%d-%m-%Y")
                hora_entrada = registro.entrada.strftime("%H:%M:%S")
                hora_salida = registro.salida.strftime("%H:%M:%S") if registro.salida else ""
                duracion = registro.duracion_formateada  # "HH:MM:SS"
                num_pausas = Pausa.query.filter_by(registro_id=registro.id).count()

                # Convertir duración a timedelta para sumarla
                h, m, s = map(int, duracion.split(":"))
                total_duracion += timedelta(hours=h, minutes=m, seconds=s)

                # Insertar datos en la fila actual
                data = [fecha_entrada, hora_entrada, hora_salida, duracion, num_pausas, ", ".join(observaciones)]
                for col, value in enumerate(data):
                    worksheet.write(row, col, value, formato if col == 5 else normal_format)
                row += 1
        else:
            # Si no hay registros, mostrar la fila vacía con el color correspondiente
            data = [current_day.strftime("%d-%m-%Y"), "", "", "", "", ", ".join(observaciones)]
            for col, value in enumerate(data):
                worksheet.write(row, col, value, formato)
            row += 1

        current_day += timedelta(days=1)

    # Insertar la suma total de duraciones
    total_horas = total_duracion.total_seconds() // 3600
    total_minutos = (total_duracion.total_seconds() % 3600) // 60
    total_segundos = total_duracion.total_seconds() % 60
    total_format_texto = f"{int(total_horas):02}:{int(total_minutos):02}:{int(total_segundos):02}"

    worksheet.write(row, 2, "Total de horas trabajadas:", total_format)
    worksheet.write(row, 3, total_format_texto, total_format)

    # Ajustar el tamaño de columnas
    worksheet.set_column(0, 0, 15)  # Fecha
    worksheet.set_column(1, 2, 20)  # Entrada y salida
    worksheet.set_column(3, 5, 25)  # Duración, pausas y observaciones

    workbook.close()
    output.seek(0)

    filename = usuario.primer_apellido + '_' + usuario.nombre + '_' + mes_datetime.strftime('%B_%Y') + '.xlsx'

    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response




@app.route('/subir_archivos', methods=['GET'])
def subir_archivos():

    return render_template('subir_archivos.html')



@app.route('/marcar_leida/<int:notificacion_id>', methods=['POST'])
@login_required
def marcar_leida(notificacion_id):

    notificacion = Notificacion.query.get_or_404(notificacion_id)
    if notificacion.usuario_id != current_user.id:
        return jsonify({'status': 'error'}), 403
    if notificacion:

        notificacion.leida = True
    db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/marcar_notificaciones_leidas', methods=['POST'])
@login_required
def marcar_notificaciones_leidas():
    try:
        # Actualiza todas las notificaciones del usuario actual para marcarlas como leídas
        Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).update({"leida": True})
        db.session.commit()
        flash('Todas las notificaciones han sido marcadas como leídas.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al marcar las notificaciones como leídas: {str(e)}', 'error')

    # Redirigir a la página anterior
    return redirect(request.referrer)

@app.route('/actualizar_registro', methods=['POST'])
def actualizar_registro():
    data = request.json
    registro_id = data['id']
    observaciones = data['observaciones']

    # Verifica que se reciben los datos correctos
    print(f"ID de registro: {registro_id}, Observaciones: {observaciones}")

    # Actualiza el registro en la base de datos
    registro = Registro.query.get(registro_id)
    if registro:
        registro.observaciones = observaciones
        db.session.commit()
        return 'Registro actualizado correctamente', 200
    else:
        return 'Registro no encontrado', 404



    # LOGICA PARA ENVIO DE CHAT


@app.route('/chats')
@login_required
def chats():

    # Obtener los usuarios que sigue el usuario actual
    usuarios = User.query.filter(User.email != 'admin@gestionaempresa.es').all()

    usuario_id = request.args.get('usuario_id')
    destinatario = None
    mensajes = []
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()

    for usuario in usuarios:


        # Contar mensajes no leídos de este usuario
        mensajes_no_leidos = Mensaje.query.filter(
            ((Mensaje.remitente_id == usuario.id) & (Mensaje.destinatario_id == current_user.id) & (Mensaje.leido == False)) |
            ((Mensaje.remitente_id == current_user.id) & (Mensaje.destinatario_id == usuario.id) & (Mensaje.leido == False))
        ).count()



        # Añadir la cuenta a cada usuario
        usuario.mensajes_no_leidos = mensajes_no_leidos

    if usuario_id:
        destinatario = User.query.get(usuario_id)
        mensajes = Mensaje.query.filter(
            ((Mensaje.remitente_id == current_user.id) & (Mensaje.destinatario_id == usuario_id)) |
            ((Mensaje.remitente_id == usuario_id) & (Mensaje.destinatario_id == current_user.id))
        ).order_by(Mensaje.fecha_envio).all()

        # Formatear la fecha de envío de cada mensaje y convertir URLs en enlaces
        for mensaje in mensajes:
            # Sumar 2 horas a la fecha de envío
            mensaje.fecha_envio = mensaje.fecha_envio
            # Formatear la fecha al formato 'dd-mm-yyyy hh:mm'
            mensaje.fecha_envio = mensaje.fecha_envio.strftime('%d-%m-%Y %H:%M')
            # Convertir las URLs en enlaces
            mensaje.contenido = convertir_urls_a_enlaces(mensaje.contenido)

    # Generar URLs de las imágenes
    current_user_image_url = url_for('static', filename='logo1.png')
    destinatario_image_url = url_for('static', filename='logo1.png')

    return render_template('chat.html',
                           usuarios=usuarios,
                           destinatario=destinatario,
                           mensajes=mensajes,

                           notificaciones=notificaciones,
                           current_user=current_user,
                           current_user_image_url=current_user_image_url,
                           destinatario_image_url=destinatario_image_url,
                           active_menu='chats')

@app.route('/marcar_leidos', methods=['POST'])
@login_required
def marcar_mensajes_leidos():
    destinatario_id = request.json.get('destinatario_id')
    if not destinatario_id:
        return jsonify({"error": "Destinatario no especificado"}), 400

    # Marcar los mensajes como leídos
    mensajes_no_leidos = Mensaje.query.filter_by(remitente_id=destinatario_id, destinatario_id=current_user.id, leido=False).all()
    mensajes_no_leidos2 = Mensaje.query.filter_by(remitente_id=current_user.id, destinatario_id=destinatario_id, leido=False).all()
    for mensaje in mensajes_no_leidos:
        mensaje.leido = True
    for mensaje3 in mensajes_no_leidos2:
        mensaje3.leido = True
    db.session.commit()

    return jsonify({"success": True})



import re
from flask import Markup

def convertir_urls_a_enlaces(texto):
    """
    Detecta URLs en el texto, incluyendo las que empiezan con 'http://', 'https://', y 'www.'.
    Convierte esas URLs en enlaces clicables.
    Si se encuentra la URL 'https://8x8.vc/', se reemplaza con un enlace específico.
    """
    # Patrón para URLs con http, https o www
    url_pattern = re.compile(r'((http|https):\/\/[^\s]+)|(www\.[^\s]+)', re.IGNORECASE)

    # Función para reemplazar las URLs detectadas
    def reemplazar_url(match):
        url = match.group(0)
        # Verifica si la URL es 'https://8x8.vc/' para hacer el reemplazo
        if url.startswith("https://8x8.vc/"):
            return '<a href="{}" target="_blank">enlace a la videollamada</a>'.format(url)
        # Para otras URLs, añade 'http://' si es necesario
        return '<a href="{}" target="_blank">{}</a>'.format(url, url if match.group(1) else "http://" + url)

    # Convertimos las URLs detectadas en enlaces clicables
    return Markup(url_pattern.sub(reemplazar_url, texto))

@app.route('/get_messages/<int:usuario_id>', methods=['GET'])
@login_required
def get_messages(usuario_id):
    mensajes = Mensaje.query.filter(
        ((Mensaje.remitente_id == current_user.id) & (Mensaje.destinatario_id == usuario_id)) |
        ((Mensaje.remitente_id == usuario_id) & (Mensaje.destinatario_id == current_user.id))
    ).order_by(Mensaje.fecha_envio).all()

    mensajes_dict = [{
        'contenido': convertir_urls_a_enlaces(mensaje.contenido),  # Convertir URLs aquí
        'fecha_envio': (mensaje.fecha_envio + timedelta(hours=2)).strftime('%d-%m-%Y %H:%M'),  # Sumar 2 horas y formatear
        'remitente_id': mensaje.remitente_id
    } for mensaje in mensajes]

    return jsonify(mensajes_dict)


@app.route('/send_message', methods=['POST'])
@login_required
def send_message():
    usuario_id = request.form.get('usuario_id')
    contenido = request.form.get('contenido')
    hora_espana1 = datetime.now().astimezone(timezone_spain)
    if usuario_id and contenido:
        # Crear y guardar el nuevo mensaje
        nuevo_mensaje = Mensaje(
            remitente_id=current_user.id,
            destinatario_id=usuario_id,
            contenido=contenido,  # Guardar el contenido tal cual se envía
            hora_envio=hora_espana1
        )
        db.session.add(nuevo_mensaje)
        db.session.commit()

        # Verificar si ya existe una notificación reciente para el destinatario
        tiempo_limite = hora_espana1 - timedelta(minutes=10)  # Ajusta el tiempo según tus necesidades
        notificacion_reciente = Notificacion.query.filter_by(usuario_id=usuario_id).filter(Notificacion.fecha_hora > tiempo_limite).first()

        if notificacion_reciente:
            # Si hay una notificación reciente, no crear una nueva
            return redirect(url_for('chats', usuario_id=usuario_id))

        # Crear una nueva notificación si no existe una reciente
        remitente = Usuario.query.get(current_user.id)  # Obtener el usuario remitente
        destinatario = Usuario.query.get(usuario_id)  # Obtener el usuario destinatario

        if remitente and destinatario:
            nueva_notificacion = Notificacion(
                usuario_id=usuario_id,
                concepto=f"{remitente.nombre} te ha enviado un mensaje.",
                fecha_hora=hora_espana1
            )
            subject = "Gestionaempresa informa: Has recibido un mensaje en el chat"
            body = f"{remitente.nombre} te ha enviado un mensaje al chat. Entra en el chat para responder."

            # Enviar correo a usuario destinatario
            send_email_to_user(destinatario.email, subject, body)
            db.session.add(nueva_notificacion)
            db.session.commit()

    return redirect(url_for('chats', usuario_id=usuario_id))

from flask import flash, redirect, url_for

@app.route('/iniciar_pausa', methods=['POST'])
@login_required
def iniciar_pausa():
    data = request.get_json()
    tipo_pausa = data.get('tipo_pausa')

    if not tipo_pausa:
        return jsonify({'success': False, 'message': 'Tipo de pausa es requerido.'}), 400

    # Obtener el registro actual del usuario
    registro = Registro.query.filter_by(user_id=current_user.id, salida=None).first()
    if not registro:
        return jsonify({'success': False, 'message': 'No se encontró un registro de entrada activo.'}), 400

    # Crear una nueva pausa
    nueva_pausa = Pausa(tipo=tipo_pausa, inicio=datetime.now(), registro_id=registro.id)
    db.session.add(nueva_pausa)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Pausa iniciada correctamente.'})

@app.route('/finalizar_pausa', methods=['POST'])
@login_required
def finalizar_pausa():
    # Obtener el registro actual del usuario
    registro = Registro.query.filter_by(user_id=current_user.id, salida=None).first()
    if not registro:
        return jsonify({'success': False, 'message': 'No se encontró un registro de entrada activo.'}), 400

    # Obtener la pausa activa
    pausa = Pausa.query.filter_by(registro_id=registro.id, fin=None).first()
    if not pausa:
        return jsonify({'success': False, 'message': 'No se encontró una pausa activa.'}), 400

    # Finalizar la pausa
    pausa.fin = datetime.now()
    db.session.commit()

    return jsonify({'success': True, 'message': 'Pausa finalizada correctamente.'})





@app.route('/get_pause_details/<int:registro_id>', methods=['GET'])
def get_pause_details(registro_id):
    pausas = Pausa.query.filter_by(registro_id=registro_id).all()
    pausa_list = [{
        'tipo': pausa.tipo,
        'inicio': pausa.inicio.strftime('%d-%m-%Y %H:%M:%S'),
        'fin': pausa.fin.strftime('%d-%m-%Y %H:%M:%S') if pausa.fin else "N/A"
    } for pausa in pausas]
    return jsonify(pausas=pausa_list)



def generar_password(longitud=6):
    caracteres = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choices(caracteres, k=longitud))

import os

from openpyxl import Workbook, load_workbook

@app.route('/crear_usuario', methods=['POST'])
def crear_usuario():
    data = request.json  # Recibe los datos desde el frontend
    print(f"[DEBUG] Datos recibidos: {data}")

    email = data.get('email')
    role = data.get('role')

    if not email or not role:
        print("[DEBUG] Falta email o rol en la solicitud")
        return jsonify({'error': 'Email y rol son obligatorios'}), 400

    # Generar contraseña aleatoria
    password = generar_password()
    print(f"[DEBUG] Contraseña generada: {password}")
    #password='123456'

    # Verificar si el usuario ya existe
    if User.query.filter_by(email=email).first():
        print(f"[DEBUG] Usuario con email {email} ya existe")
        return jsonify({'error': 'El usuario ya existe'}), 400

    # Crear el usuario
    nuevo_usuario = User(
        email=email,
        password=generate_password_hash(password),  # Encriptar contraseña
        empresa_id=current_user.empresa_id if current_user.empresa_id else None,

    )
    db.session.add(nuevo_usuario)
    db.session.commit()
    print(f"[DEBUG] Usuario {email} creado y guardado en la base de datos")

    # Asignar el rol al usuario después de la creación
    if role == 'admin':
        admin_role = user_datastore.find_or_create_role(name='admin', description='Administrator')
        user_datastore.add_role_to_user(nuevo_usuario, admin_role)
    elif role == 'user':
        user_role = user_datastore.find_or_create_role(name='user', description='Regular User')
        user_datastore.add_role_to_user(nuevo_usuario, user_role)
    elif role == 'gestor':
        gestor_role = user_datastore.find_or_create_role(name='gestor')
        user_datastore.add_role_to_user(nuevo_usuario, gestor_role)
    elif role == 'encargado':
        user_role = user_datastore.find_or_create_role(name='encargado')
        user_datastore.add_role_to_user(nuevo_usuario, user_role)

    db.session.commit()
    print(f"[DEBUG] Rol {role} asignado al usuario {email}")

    # Enviar correo al nuevo usuario con su contraseña
    cuerpo_mensaje = f"Hola! Tu cuenta ha sido creada exitosamente. Aquí está tu contraseña temporal."
    cuerpo_mensaje2 = f"Nuevo usuario creado en Herdel. Email: {nuevo_usuario.email}"
    send_email_to_user_pssw(email, "Bienvenido a la aplicación", cuerpo_mensaje, password)
    send_email_to_user('alfonsomunoz@adaptasystem.com', "Nuevo usuario en Herdel", cuerpo_mensaje2)

    # Guardar en el archivo Excel
    guardar_usuario_en_excel(email, role)
    print("[DEBUG] Usuario guardado en Excel")

    return jsonify({'mensaje': 'Usuario creado con éxito'}), 201

def guardar_usuario_en_excel(email, role):
    # Obtener la fecha y hora actual
    fecha_creacion = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f"Fecha de creación del usuario: {fecha_creacion}")  # Imprimir la fecha

    # Ruta donde se almacenará el archivo Excel

    ruta_archivo = os.path.join(app.static_folder, 'usuarios.xlsx')
    print(f"Ruta del archivo Excel: {ruta_archivo}")  # Imprimir la ruta del archivo

    # Si el archivo no existe, crearlo
    if not os.path.exists(ruta_archivo):
        print(f"El archivo {ruta_archivo} no existe. Creando uno nuevo.")  # Imprimir que se crea un archivo nuevo
        # Crear un nuevo archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.append(["Email", "Rol", "Fecha de Creación"])  # Encabezados
        print("Archivo Excel creado y encabezados añadidos.")  # Confirmar que se añadieron los encabezados
    else:
        print(f"El archivo {ruta_archivo} ya existe. Cargando el archivo existente.")  # Imprimir que el archivo ya existe
        # Si ya existe, cargarlo
        wb = load_workbook(ruta_archivo)
        ws = wb.active

    # Añadir la nueva fila con los datos del usuario
    print(f"Añadiendo los datos del usuario: {email}, {role}, {fecha_creacion}")  # Imprimir los datos que se añaden
    ws.append([email, role, fecha_creacion])

    # Guardar el archivo Excel
    wb.save(ruta_archivo)
    print(f"Información del usuario {email} guardada correctamente en {ruta_archivo}")  # Confirmar que los datos se guardaron correctamente



def send_email_to_user_pssw(email, subject, body, password):
    """
    Esta función envía un correo al usuario con su contraseña.

    :param email: Dirección de correo electrónico del usuario.
    :param subject: Asunto del correo.
    :param body: Cuerpo principal del correo (puede incluir un mensaje personalizado).
    :param password: Contraseña generada para el usuario.
    """
    # Crear el cuerpo del correo como HTML
    html_body = render_template_string(
        """\
        <html>
            <body>
                <p>{{ body }}</p>
                <p><strong>Tu contraseña es:</strong> {{ password }}</p>
                <p><a href="{{ link_href }}">{{ link_text }}</a></p>
            </body>
        </html>
        """,
        body=body,
        password=password,  # Incluir la contraseña en el correo
        link_text="Link a la APP",
        link_href="https://herdelsalud.gestionaempresa.es/"
    )

    # Crear el mensaje
    msg = Message(subject, recipients=[email], html=html_body)

    # Enviar el mensaje
    mail.send(msg)
    print(f"Correo enviado a {email} con la contraseña.")

@app.route('/guardar_empresa', methods=['GET', 'POST'])
@admin_required
def guardar_empresa():
    print("\n[DEBUG] ---- ENTRANDO EN guardar_empresa ----")

    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    es_gestor = any(role.name == 'gestor' for role in current_user.roles)
    empresas = Empresa.query.all() if not es_gestor else []

    empresa_id = request.args.get('empresa_id', type=int)
    print(f"[DEBUG] es_gestor: {es_gestor}")
    print(f"[DEBUG] empresa_id recibido: {empresa_id}")

    # Determinar empresa según el rol y parámetros
    if es_gestor:
        if current_user.empresa_id:
            empresa = Empresa.query.get(current_user.empresa_id)
            print(f"[DEBUG] Empresa del gestor: {empresa.nombre if empresa else 'No encontrada'}")
        else:
            empresa = Empresa.query.first()
            print("[DEBUG] No hay empresa asociada al gestor, se usa la primera")
    else:
        if empresa_id:
            empresa = Empresa.query.get(empresa_id)
            print(f"[DEBUG] Empresa cargada por ID: {empresa.nombre if empresa else 'No encontrada'}")
        elif current_user.empresa_id:
            empresa = Empresa.query.get(current_user.empresa_id)
            print(f"[DEBUG] Empresa del usuario actual: {empresa.nombre if empresa else 'No encontrada'}")
        else:
            empresa = Empresa.query.first()
            print("[DEBUG] Empresa por defecto (primera en BD)")

    # Crear empresa si no existe
    if not empresa:
        print("[DEBUG] No existe ninguna empresa, creando nueva...")
        empresa = Empresa(tipo_jornada='intensiva')
        db.session.add(empresa)
        db.session.commit()
        print(f"[DEBUG] Empresa creada con ID: {empresa.id}")

    print(f"[DEBUG] Empresa actual: {empresa.nombre if empresa else 'Desconocida'} (ID: {empresa.id})")
    print(f"[DEBUG] IP actual registrada en empresa: {empresa.ip_empresa}")

    # Crear formulario
    form = EmpresaForm(obj=empresa)
    print(f"[DEBUG] Método HTTP: {request.method}")
    print(f"[DEBUG] Formulario recibido con datos: {request.form.to_dict()}")

    # Validación y guardado
    if form.validate_on_submit():
        print("\n[DEBUG] ---- Formulario validado correctamente ----")
        print(f"[DEBUG] Fecha inicio verano: {form.fecha_inicio_verano.data}")
        print(f"[DEBUG] Fecha fin verano: {form.fecha_fin_verano.data}")
        print(f"[DEBUG] IP recibida desde formulario: {form.ip_empresa.data}")

        # Copiar datos del form al modelo
        form.populate_obj(empresa)

        print(f"[DEBUG] IP guardada en objeto empresa antes de commit: {empresa.ip_empresa}")

        db.session.commit()
        print("[DEBUG] Cambios guardados en la base de datos correctamente ✅")
    else:
        if request.method == 'POST':
            print("\n[DEBUG] ---- Error de validación en formulario ----")
            print(f"[DEBUG] Errores: {form.errors}")

    festivos = Festivo.query.filter_by(empresa_id=empresa.id).all()
    MESES = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    festivos_con_nombres = [
        {
            'id': festivo.id,
            'dia': festivo.dia,
            'mes': MESES[festivo.mes]
        } for festivo in festivos
    ]

    current_year = datetime.now().year

    print("[DEBUG] ---- Renderizando plantilla admin_empresa.html ----\n")

    return render_template(
        'admin_empresa.html',
        form=form,
        empresa=empresa,
        notificaciones=notificaciones,
        festivos=festivos_con_nombres,
        current_year=current_year,
        empresas=empresas,
        active_menu='combo_config'
    )

@app.route('/crear_empresa_ajax', methods=['POST'])
@admin_required
def crear_empresa_ajax():
    data = request.get_json()
    nombre = data.get('nombre')

    if not nombre:
        return jsonify(success=False, message="Nombre requerido")

    nueva = Empresa(nombre=nombre, tipo_jornada='intensiva')  # Añade más campos por defecto si quieres
    db.session.add(nueva)
    db.session.commit()

    return jsonify(success=True, empresa_id=nueva.id)


@app.route('/eliminar_empresa_ajax', methods=['POST'])
@admin_required
def eliminar_empresa_ajax():
    data = request.get_json()
    empresa_id = data.get('empresa_id')
    empresa = Empresa.query.get(empresa_id)

    if not empresa:
        return jsonify(success=False, message="Empresa no encontrada")

    # Verifica si tiene relaciones críticas antes de eliminar (opcional)
    try:
        db.session.delete(empresa)
        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, message=str(e))



@app.route("/update_ver_vacaciones", methods=["POST"])
def update_ver_vacaciones():
    data = request.get_json()
    empresa_id = data.get("empresa_id")
    if not empresa_id:
        return jsonify({"success": False, "message": "No se proporcionó empresa_id"}), 400

    empresa = Empresa.query.get(empresa_id)
    if empresa:
        empresa.ver_vacaciones = bool(data["ver_vacaciones"])
        db.session.commit()
        return jsonify({"success": True})
    return jsonify({"success": False, "message": "Empresa no encontrada"}), 404


@app.route("/update_ver_pausas", methods=["POST"])
def update_ver_pausas():
    data = request.get_json()
    empresa_id = data.get("empresa_id")
    if not empresa_id:
        return jsonify({"success": False, "message": "No se proporcionó empresa_id"}), 400

    empresa = Empresa.query.get(empresa_id)
    if empresa:
        empresa.ver_pausas = bool(data.get("ver_pausas"))
        db.session.commit()
        return jsonify({"success": True})
    return jsonify({"success": False, "message": "Empresa no encontrada"}), 404


from flask import request, jsonify


@app.route('/guardar_empresa_save', methods=['POST'])
def guardar_empresa_save():
    try:
        print("ENTRAAAAAAA")

        data = request.get_json()
        print("Datos recibidos:", data)

        def convertir_a_null(valor):
            return valor if valor not in ["", None] else None

        # Obtener el ID de la empresa desde los datos
        empresa_id = data.get('empresa_id')

        # Buscar la empresa por ID
        empresa = Empresa.query.get(empresa_id)
        if not empresa:
            return jsonify({'success': False, 'error': 'Empresa no encontrada'}), 404

        # Asignar los campos de texto y numéricos
        empresa.nombre = convertir_a_null(data.get('nombre'))
        empresa.CIF = convertir_a_null(data.get('CIF'))
        empresa.Domicilio = convertir_a_null(data.get('Domicilio'))
        empresa.email = convertir_a_null(data.get('email'))
        empresa.Telefono = convertir_a_null(data.get('Telefono'))
        empresa.Localidad = convertir_a_null(data.get('Localidad'))
        empresa.Provincia = convertir_a_null(data.get('Provincia'))
        empresa.CP = convertir_a_null(data.get('CP'))
        empresa.dias_vacaciones = convertir_a_null(data.get('vacaciones'))
        empresa.tiempo_cortesia = convertir_a_null(data.get('tiempo_cortesia'))
        empresa.jornada = convertir_a_null(data.get('jornada'))

        # 🆕 Nuevo campo: IP de la empresa
        empresa.ip_empresa = convertir_a_null(data.get('ip_empresa'))
        print("📡 IP Empresa recibida:", data.get('ip_empresa'))
        print("📦 IP Empresa asignada al modelo:", empresa.ip_empresa)

        # Conversión de fechas
        def convertir_a_fecha(valor):
            try:
                return datetime.strptime(valor, '%Y-%m-%d').date() if valor else None
            except ValueError:
                print(f"⚠️ Error al convertir fecha: {valor}")
                return None

        if data.get('fecha_inicio_verano'):
            empresa.fecha_inicio_verano = convertir_a_fecha(data.get('fecha_inicio_verano'))
        if data.get('fecha_fin_verano'):
            empresa.fecha_fin_verano = convertir_a_fecha(data.get('fecha_fin_verano'))

        print("Vacaciones:", empresa.dias_vacaciones)
        print("Tiempo de cortesía:", empresa.tiempo_cortesia)

        # Convertir y asignar los horarios
        dias_semana = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]

        for temporada in ['invierno', 'verano']:
            for horario in data.get('horarios', []):  # Asegurar que horarios sea una lista
                if temporada in horario:
                    horario_temporada = horario[temporada]

                    dia = horario_temporada.get("dia")
                    if not dia or dia not in dias_semana:
                        print(f"⚠️ Error: 'dia' no encontrado o inválido en {temporada}, datos: {horario_temporada}")
                        continue

                    trabaja = horario_temporada.get("trabaja", False)
                    entrada = convert_to_time(horario_temporada.get("entrada"))
                    salida_comida = convert_to_time(horario_temporada.get("salidaComida"))
                    entrada_comida = convert_to_time(horario_temporada.get("entradaComida"))
                    salida = convert_to_time(horario_temporada.get("salida"))

                    setattr(empresa, f'trabaja_{temporada}_{dia}', trabaja)
                    setattr(empresa, f'entrada_{temporada}_{dia}', entrada)
                    setattr(empresa, f'salida_comida_{temporada}_{dia}', salida_comida)
                    setattr(empresa, f'entrada_comida_{temporada}_{dia}', entrada_comida)
                    setattr(empresa, f'salida_{temporada}_{dia}', salida)

        db.session.commit()
        print("✅ GUARDADO CORRECTAMENTE 🚀")
        print("🧾 Empresa guardada con IP:", empresa.ip_empresa)

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        print(f"❌ Error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/toggle_bloquear_ip/<int:user_id>', methods=['POST'])
@admin_required
def toggle_bloquear_ip(user_id):
    try:
        data = request.get_json()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404

        user.bloquear_por_ip = data.get('bloquear_por_ip', False)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/festivos', methods=['GET'])
def obtener_festivos():
    empresa_id = request.args.get('empresa_id')
    print(f"[DEBUG] Parámetro empresa_id recibido: {empresa_id}")

    if not empresa_id:
        print("[ERROR] Falta el parámetro empresa_id en la petición.")
        return jsonify({'error': 'Falta el parámetro empresa_id'}), 400

    # Filtrar los festivos por empresa
    festivos = Festivo.query.filter_by(empresa_id=empresa_id).all()
    print(f"[DEBUG] Festivos encontrados para empresa_id {empresa_id}: {[f'{f.dia}/{f.mes}' for f in festivos]}")

    return jsonify([{'id': f.id, 'dia': f.dia, 'mes': f.mes} for f in festivos])


@app.route('/agregar_festivo', methods=['POST'])
def agregar_festivo():
    data = request.get_json()

    if not all(k in data for k in ('dia', 'mes', 'empresa_id')):
        return jsonify({'success': False, 'error': 'Faltan campos requeridos'}), 400

    nuevo_festivo = Festivo(
        dia=data['dia'],
        mes=data['mes'],
        empresa_id=data['empresa_id']
    )
    db.session.add(nuevo_festivo)
    db.session.commit()

    return jsonify({'success': True})


@app.route('/eliminar_festivo', methods=['POST'])
def eliminar_festivo():
    data = request.get_json()
    festivo = Festivo.query.get(data['id'])
    if festivo:
        db.session.delete(festivo)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Festivo no encontrado'})



from flask import render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import io
from openpyxl.styles import Alignment, Font, PatternFill
import calendar
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from calendar import monthrange
from openpyxl.utils import get_column_letter


@app.route('/export_calendar', methods=['GET', 'POST'])
def export_calendar():
    if request.method == 'POST':
        year = int(request.form['year'])
        empresa_id = 1  # Reemplaza con la lógica para obtener la empresa correcta
        empresa = Empresa.query.get(empresa_id)
        if not empresa:
            return "Empresa no encontrada.", 404

        # Obtener días festivos
        current_year = year
        festivos = Festivo.query.all()
        non_workdays = [datetime(current_year, festivo.mes, festivo.dia).date() for festivo in festivos]

        # Crear un nuevo libro de Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f'Calendario Laboral {year}'

        # Estilos
        festivo_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        no_laboral_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Configurar la hoja en horizontal y ajustar tamaños
        sheet.page_setup.orientation = "landscape"
        sheet.column_dimensions['A'].width = 12  # Columna de los meses
        for col_idx in range(2, 34):  # Columnas de días
            col_letter = get_column_letter(col_idx)  # Convertir índice a letra usando openpyxl
            sheet.column_dimensions[col_letter].width = 3
            sheet.column_dimensions['AG'].width = 15

        # Encabezado general
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=32)
        header_cell = sheet.cell(row=1, column=1)
        header_cell.value = f'Calendario Laboral {year}'
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.font = Font(size=14, bold=True)

        # Encabezados de días
        for dia in range(1, 32):
            dia_cell = sheet.cell(row=2, column=dia + 1)
            dia_cell.value = dia
            dia_cell.alignment = Alignment(horizontal='center')
            dia_cell.fill = header_fill
            dia_cell.border = border

        # Agregar encabezado de "Totales" en la columna 34
        sheet.cell(row=2, column=33).value = 'Totales'
        sheet.cell(row=2, column=33).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=33).fill = header_fill
        sheet.cell(row=2, column=33).border = border

        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

        # Generar las filas del calendario
        fila_inicial = 3
        total_horas_anuales = 0  # Variable para acumular las horas anuales
        dias_laborables_totales = 0

        for mes_idx, mes in enumerate(meses, start=1):
            mes_cell = sheet.cell(row=fila_inicial, column=1)
            mes_cell.value = mes
            mes_cell.alignment = Alignment(horizontal='center')
            mes_cell.border = border

            fecha_inicial = datetime(year, mes_idx, 1)
            _, dias_mes = calendar.monthrange(year, mes_idx)

            total_horas_mes = 0  # Variable para acumular las horas del mes

            for dia in range(1, 32):
                columna_dia = dia + 1
                cell = sheet.cell(row=fila_inicial, column=columna_dia)
                cell.border = border

                if dia > dias_mes:
                    cell.fill = black_fill
                else:
                    fecha_actual = fecha_inicial.replace(day=dia)
                    dia_semana = fecha_actual.weekday()

                    # Calcular las horas trabajadas
                    horas_trabajadas = None
                    es_festivo = fecha_actual.date() in non_workdays
                    if not es_festivo:
                        if empresa.fecha_inicio_verano:
                            fecha_inicio_verano = datetime.combine(empresa.fecha_inicio_verano, datetime.min.time())
                            fecha_fin_verano = datetime.combine(empresa.fecha_fin_verano, datetime.min.time())

                            if fecha_inicio_verano <= fecha_actual <= fecha_fin_verano:
                                temporada = 'verano'
                            else:
                                temporada = 'invierno'
                        else:
                            temporada = 'invierno'

                        dias_temporada = {
                            0: 'lunes', 1: 'martes', 2: 'miercoles', 3: 'jueves', 4: 'viernes', 5: 'sabado', 6: 'domingo'
                        }
                        dia_nombre = dias_temporada[dia_semana]

                        if getattr(empresa, f'trabaja_{temporada}_{dia_nombre}'):
                            horas_trabajadas = calcular_horas_trabajadas(empresa, temporada, dia_nombre)

                    if horas_trabajadas is not None:
                        total_horas_mes += horas_trabajadas
                        dias_laborables_totales += 1
                        cell.value = round(horas_trabajadas)  # Redondear y mostrar solo el número de horas
                    else:
                        cell.value = ''

                    # Aplicar colores
                    if es_festivo:
                        cell.fill = festivo_fill
                    elif horas_trabajadas is None:
                        cell.fill = no_laboral_fill

            # Colocar el total de horas trabajadas en la columna de "Totales" en la fila correspondiente al mes
            sheet.cell(row=fila_inicial, column=33).value = round(total_horas_mes)  # Total de horas en el mes
            sheet.cell(row=2, column=33).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=2, column=33).fill = header_fill
            sheet.cell(row=2, column=33).border = border


            total_horas_anuales += total_horas_mes  # Sumar las horas mensuales al total anual
            fila_inicial += 1

        # Calcular horas según vacaciones
        horas_por_dia = total_horas_anuales / dias_laborables_totales if dias_laborables_totales else 0
        horas_vacaciones = horas_por_dia * empresa.dias_vacaciones
        horas_anuales_netas = total_horas_anuales - horas_vacaciones

        # Agregar fila con el total anual de horas debajo de diciembre
        # Títulos en las filas 15, 16, y 17 combinando celdas AA a AF
        titulos = ['Total Horas Año', 'Horas Vacaciones', 'Total Horas Netas']
        valores = [round(total_horas_anuales), round(horas_vacaciones), round(horas_anuales_netas)]

        for idx, row in enumerate(range(15, 18)):  # Iterar sobre las filas 15, 16 y 17
            cell_range = f'AA{row}:AF{row}'  # Rango de celdas a combinar
            sheet.merge_cells(cell_range)  # Combina las celdas
            title_cell = sheet[f'AA{row}']  # Selecciona la celda inicial del rango
            title_cell.value = titulos[idx]  # Asigna el título correspondiente
            title_cell.alignment = Alignment(horizontal='center', vertical='center')  # Centra el texto
            title_cell.font = Font(size=12, bold=True)  # Ajusta el formato de la fuente
            title_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Fondo azul claro

            # Colocar el valor en la columna AG de la misma fila
            valor_cell = sheet[f'AG{row}']  # Celda de valores en la columna AG
            valor_cell.value = valores[idx]
            valor_cell.alignment = Alignment(horizontal='center', vertical='center')
            valor_cell.font = Font(size=12)

        sheet.cell(row=2, column=33).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=2, column=33).fill = header_fill
        sheet.cell(row=2, column=33).border = border

        # Guardar el archivo en un objeto BytesIO
        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        return send_file(file_stream, as_attachment=True, download_name=f'Calendario_Laboral_{year}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return redirect(url_for('guardar_empresa'))

@app.route('/preview_calendar', methods=['GET'])
def preview_calendar():
    year = int(request.args.get('year', datetime.now().year))
    empresa_id = 1  # Ajusta según tu lógica
    empresa = Empresa.query.get(empresa_id)
    if not empresa:
        return jsonify({"error": "Empresa no encontrada"}), 404

    festivos = Festivo.query.all()
    non_workdays = [datetime(year, f.mes, f.dia).date() for f in festivos]

    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    preview_data = []

    for mes_idx, mes in enumerate(meses, start=1):
        fecha_inicial = datetime(year, mes_idx, 1)
        _, dias_mes = calendar.monthrange(year, mes_idx)

        mes_info = {"nombre": mes, "dias": {}}

        for dia in range(1, dias_mes + 1):
            fecha_actual = fecha_inicial.replace(day=dia)
            es_festivo = fecha_actual.date() in non_workdays
            horas_trabajadas = None

            if not es_festivo:
                if empresa.fecha_inicio_verano:
                    fecha_inicio_verano = datetime.combine(empresa.fecha_inicio_verano, datetime.min.time())
                    fecha_fin_verano = datetime.combine(empresa.fecha_fin_verano, datetime.min.time())

                    temporada = 'verano' if fecha_inicio_verano <= fecha_actual <= fecha_fin_verano else 'invierno'
                else:
                    temporada = 'invierno'
                dia_nombre = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"][fecha_actual.weekday()]

                if getattr(empresa, f'trabaja_{temporada}_{dia_nombre}', False):
                    horas_trabajadas = calcular_horas_trabajadas(empresa, temporada, dia_nombre)

            mes_info["dias"][dia] = round(horas_trabajadas) if horas_trabajadas is not None else 'F' if es_festivo else '-'

        preview_data.append(mes_info)

    return jsonify(preview_data)



def calcular_horas_trabajadas(empresa, temporada, dia):
    # Obtener los valores de entrada y salida
    entrada = getattr(empresa, f'entrada_{temporada}_{dia}', None)
    salida_comida = getattr(empresa, f'salida_comida_{temporada}_{dia}', None)
    entrada_comida = getattr(empresa, f'entrada_comida_{temporada}_{dia}', None)
    salida = getattr(empresa, f'salida_{temporada}_{dia}', None)

    # Comprobar si todos los valores son válidos
    if entrada is None or salida is None:
        return None  # Si faltan entrada o salida, no calculamos las horas

    if salida_comida and entrada_comida:
        # Jornada partida
        if entrada_comida and salida_comida:
            try:
                horas_trabajadas = (datetime.combine(datetime.min, salida) - datetime.combine(datetime.min, entrada)) + \
                                   (datetime.combine(datetime.min, salida_comida) - datetime.combine(datetime.min, entrada_comida))
            except TypeError:
                return None  # Manejar errores si las conversiones no son válidas
        else:
            return None  # Si faltan datos de comida, no calculamos las horas
    else:
        # Jornada intensiva
        try:
            horas_trabajadas = datetime.combine(datetime.min, salida) - datetime.combine(datetime.min, entrada)
        except TypeError:
            return None  # Manejar errores si las conversiones no son válidas

    return horas_trabajadas.total_seconds() / 3600  # Convertir a horas






# Función para convertir cadenas de hora en objetos datetime.time
def convert_to_time(time_str):
    if time_str:
        try:
            # Si la cadena no tiene segundos, añadir "00" para que sea válido
            if len(time_str) == 5:  # Formato HH:MM
                time_str = time_str + ":00"

            return datetime.strptime(time_str, '%H:%M:%S').time()
        except ValueError:
            print(f"Error de formato de hora: {time_str}")
            return None  # Si la cadena no es válida, devolvemos None
    return None

@app.route('/api/usuarios', methods=['GET'])
def obtener_usuarios():
    try:
        empresa_id = request.args.get('empresa_id', type=int)
        if not empresa_id:
            return jsonify({'success': False, 'error': 'No se proporcionó empresa_id'}), 400

        usuarios = User.query.filter(
            (User.email != 'admin@gestionaempresa.es') &
            (User.email != 'inspector@gestionaempresa.es') &
            (User.email != 'gestor@gestionaempresa.es') &
            (User.email != 'encargado@gestionaempresa.es'),
            (User.empresa_id == empresa_id)

        ).all()


        usuarios_data = [{
            'id': user.id,
            'nombre': user.nombre,
            'primer_apellido': user.primer_apellido
        } for user in usuarios]

        return jsonify({'usuarios': usuarios_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/asignar_horario_todos', methods=['POST'])
def asignar_horario_todos():
    try:
        # DEBUG: mostrar el usuario que hace la asignación
        print("DEBUG: usuario actual ->", current_user)

        # Obtener la empresa del usuario que hace la asignación
        empresa = getattr(current_user, 'empresa', None)
        print("DEBUG: empresa del usuario ->", empresa)

        if not empresa:
            print("DEBUG: El usuario no tiene empresa asignada")
            return jsonify({'success': False, 'error': 'El usuario no tiene empresa asignada'}), 400

        # Obtener los usuarios de esa empresa
        users = User.query.filter_by(empresa_id=empresa.id).all()
        print("DEBUG: usuarios encontrados ->", users)

        if not users:
            print("DEBUG: No hay usuarios para esta empresa")
            return jsonify({'success': False, 'error': 'No hay usuarios para esta empresa'}), 404

        # Verificar si estamos dentro del periodo de verano
        today = datetime.today().date()
        print("DEBUG: fecha de hoy ->", today)

        # Usar horarios de invierno (puedes agregar lógica de verano si quieres)
        horarios = {
            'lunes': {'trabaja': empresa.trabaja_invierno_lunes, 'entrada': empresa.entrada_invierno_lunes,
                      'salida_comida': empresa.salida_comida_invierno_lunes, 'entrada_comida': empresa.entrada_comida_invierno_lunes, 'salida': empresa.salida_invierno_lunes},
            'martes': {'trabaja': empresa.trabaja_invierno_martes, 'entrada': empresa.entrada_invierno_martes,
                       'salida_comida': empresa.salida_comida_invierno_martes, 'entrada_comida': empresa.entrada_comida_invierno_martes, 'salida': empresa.salida_invierno_martes},
            'miercoles': {'trabaja': empresa.trabaja_invierno_miercoles, 'entrada': empresa.entrada_invierno_miercoles,
                          'salida_comida': empresa.salida_comida_invierno_miercoles, 'entrada_comida': empresa.entrada_comida_invierno_miercoles, 'salida': empresa.salida_invierno_miercoles},
            'jueves': {'trabaja': empresa.trabaja_invierno_jueves, 'entrada': empresa.entrada_invierno_jueves,
                       'salida_comida': empresa.salida_comida_invierno_jueves, 'entrada_comida': empresa.entrada_comida_invierno_jueves, 'salida': empresa.salida_invierno_jueves},
            'viernes': {'trabaja': empresa.trabaja_invierno_viernes, 'entrada': empresa.entrada_invierno_viernes,
                        'salida_comida': empresa.salida_comida_invierno_viernes, 'entrada_comida': empresa.entrada_comida_invierno_viernes, 'salida': empresa.salida_invierno_viernes},
            'sabado': {'trabaja': empresa.trabaja_invierno_sabado, 'entrada': empresa.entrada_invierno_sabado,
                       'salida_comida': empresa.salida_comida_invierno_sabado, 'entrada_comida': empresa.entrada_comida_invierno_sabado, 'salida': empresa.salida_invierno_sabado},
            'domingo': {'trabaja': empresa.trabaja_invierno_domingo, 'entrada': empresa.entrada_invierno_domingo,
                        'salida_comida': empresa.salida_comida_invierno_domingo, 'entrada_comida': empresa.entrada_comida_invierno_domingo, 'salida': empresa.salida_invierno_domingo}
        }

        # Asignar los horarios a todos los usuarios
        for user in users:
            print(f"DEBUG: asignando horario a usuario -> {user.id} - {user.nombre}")
            for dia, horario in horarios.items():
                if dia == 'lunes':
                    user.trabaja_lunes = horario['trabaja']
                    user.entrada_lunes = horario['entrada']
                    user.salida_comida_lunes = horario['salida_comida']
                    user.entrada_comida_lunes = horario['entrada_comida']
                    user.salida_lunes = horario['salida']
                elif dia == 'martes':
                    user.trabaja_martes = horario['trabaja']
                    user.entrada_martes = horario['entrada']
                    user.salida_comida_martes = horario['salida_comida']
                    user.entrada_comida_martes = horario['entrada_comida']
                    user.salida_martes = horario['salida']
                elif dia == 'miercoles':
                    user.trabaja_miercoles = horario['trabaja']
                    user.entrada_miercoles = horario['entrada']
                    user.salida_comida_miercoles = horario['salida_comida']
                    user.entrada_comida_miercoles = horario['entrada_comida']
                    user.salida_miercoles = horario['salida']
                elif dia == 'jueves':
                    user.trabaja_jueves = horario['trabaja']
                    user.entrada_jueves = horario['entrada']
                    user.salida_comida_jueves = horario['salida_comida']
                    user.entrada_comida_jueves = horario['entrada_comida']
                    user.salida_jueves = horario['salida']
                elif dia == 'viernes':
                    user.trabaja_viernes = horario['trabaja']
                    user.entrada_viernes = horario['entrada']
                    user.salida_comida_viernes = horario['salida_comida']
                    user.entrada_comida_viernes = horario['entrada_comida']
                    user.salida_viernes = horario['salida']
                elif dia == 'sabado':
                    user.trabaja_sabado = horario['trabaja']
                    user.entrada_sabado = horario['entrada']
                    user.salida_comida_sabado = horario['salida_comida']
                    user.entrada_comida_sabado = horario['entrada_comida']
                    user.salida_sabado = horario['salida']
                elif dia == 'domingo':
                    user.trabaja_domingo = horario['trabaja']
                    user.entrada_domingo = horario['entrada']
                    user.salida_comida_domingo = horario['salida_comida']
                    user.entrada_comida_domingo = horario['entrada_comida']
                    user.salida_domingo = horario['salida']

            # Guardar cambios después de asignar horarios a cada usuario
            db.session.commit()

        return jsonify({'success': True, 'message': 'Horarios asignados correctamente a todos los usuarios'})

    except Exception as e:
        db.session.rollback()
        print("DEBUG: excepción ->", str(e))
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/asignar_horario_usuario/<int:user_id>', methods=['POST'])
def asignar_horario_usuario(user_id):
    try:
        # Obtener el usuario
        user = User.query.get_or_404(user_id)

        # Obtener la empresa asociada al usuario
        empresa = user.empresa if user.empresa else Empresa.query.first() # Puedes cambiar esto si el usuario tiene una relación con una empresa

        # Verificar si estamos dentro del periodo de verano
        today = datetime.today().date()


                # Usar horarios de invierno
        horarios = {
            'lunes': {'trabaja': empresa.trabaja_invierno_lunes, 'entrada': empresa.entrada_invierno_lunes,
                      'salida_comida': empresa.salida_comida_invierno_lunes, 'entrada_comida': empresa.entrada_comida_invierno_lunes, 'salida': empresa.salida_invierno_lunes},
            'martes': {'trabaja': empresa.trabaja_invierno_martes, 'entrada': empresa.entrada_invierno_martes,
                       'salida_comida': empresa.salida_comida_invierno_martes, 'entrada_comida': empresa.entrada_comida_invierno_martes, 'salida': empresa.salida_invierno_martes},
            'miercoles': {'trabaja': empresa.trabaja_invierno_miercoles, 'entrada': empresa.entrada_invierno_miercoles,
                          'salida_comida': empresa.salida_comida_invierno_miercoles, 'entrada_comida': empresa.entrada_comida_invierno_miercoles, 'salida': empresa.salida_invierno_miercoles},
            'jueves': {'trabaja': empresa.trabaja_invierno_jueves, 'entrada': empresa.entrada_invierno_jueves,
                       'salida_comida': empresa.salida_comida_invierno_jueves, 'entrada_comida': empresa.entrada_comida_invierno_jueves, 'salida': empresa.salida_invierno_jueves},
            'viernes': {'trabaja': empresa.trabaja_invierno_viernes, 'entrada': empresa.entrada_invierno_viernes,
                        'salida_comida': empresa.salida_comida_invierno_viernes, 'entrada_comida': empresa.entrada_comida_invierno_viernes, 'salida': empresa.salida_invierno_viernes},
            'sabado': {'trabaja': empresa.trabaja_invierno_sabado, 'entrada': empresa.entrada_invierno_sabado,
                       'salida_comida': empresa.salida_comida_invierno_sabado, 'entrada_comida': empresa.entrada_comida_invierno_sabado, 'salida': empresa.salida_invierno_sabado},
            'domingo': {'trabaja': empresa.trabaja_invierno_domingo, 'entrada': empresa.entrada_invierno_domingo,
                        'salida_comida': empresa.salida_comida_invierno_domingo, 'entrada_comida': empresa.entrada_comida_invierno_domingo, 'salida': empresa.salida_invierno_domingo}
        }

        # Asignar los horarios al usuario
        for dia, horario in horarios.items():
            if dia == 'lunes':
                user.trabaja_lunes = horario['trabaja']
                user.entrada_lunes = horario['entrada']
                user.salida_comida_lunes = horario['salida_comida']
                user.entrada_comida_lunes = horario['entrada_comida']
                user.salida_lunes = horario['salida']
            elif dia == 'martes':
                user.trabaja_martes = horario['trabaja']
                user.entrada_martes = horario['entrada']
                user.salida_comida_martes = horario['salida_comida']
                user.entrada_comida_martes = horario['entrada_comida']
                user.salida_martes = horario['salida']
            elif dia == 'miercoles':
                user.trabaja_miercoles = horario['trabaja']
                user.entrada_miercoles = horario['entrada']
                user.salida_comida_miercoles = horario['salida_comida']
                user.entrada_comida_miercoles = horario['entrada_comida']
                user.salida_miercoles = horario['salida']
            elif dia == 'jueves':
                user.trabaja_jueves = horario['trabaja']
                user.entrada_jueves = horario['entrada']
                user.salida_comida_jueves = horario['salida_comida']
                user.entrada_comida_jueves = horario['entrada_comida']
                user.salida_jueves = horario['salida']
            elif dia == 'viernes':
                user.trabaja_viernes = horario['trabaja']
                user.entrada_viernes = horario['entrada']
                user.salida_comida_viernes = horario['salida_comida']
                user.entrada_comida_viernes = horario['entrada_comida']
                user.salida_viernes = horario['salida']
            elif dia == 'sabado':
                user.trabaja_sabado = horario['trabaja']
                user.entrada_sabado = horario['entrada']
                user.salida_comida_sabado = horario['salida_comida']
                user.entrada_comida_sabado = horario['entrada_comida']
                user.salida_sabado = horario['salida']
            elif dia == 'domingo':
                user.trabaja_domingo = horario['trabaja']
                user.entrada_domingo = horario['entrada']
                user.salida_comida_domingo = horario['salida_comida']
                user.entrada_comida_domingo = horario['entrada_comida']
                user.salida_domingo = horario['salida']

        # Guardar cambios
        db.session.commit()

        return jsonify({'success': True, 'message': 'Horario asignado correctamente'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500



@app.route('/asignar_horario_verano_todos', methods=['POST'])
def asignar_horario_verano_todos():
    try:
        data = request.get_json()
        empresa_id = data.get("empresa_id")
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Falta el ID de empresa'}), 400

        # Obtener la empresa por ID
        empresa = Empresa.query.get(empresa_id)
        if not empresa:
            return jsonify({'success': False, 'error': 'Empresa no encontrada'}), 404

        # Obtener los usuarios de esa empresa
        users = User.query.filter_by(empresa_id=empresa.id).all()

        # Verificar si hay usuarios
        if not users:
            return jsonify({'success': False, 'error': 'No hay usuarios para esta empresa'}), 404


        # Verificar si estamos dentro del periodo de verano
        today = datetime.today().date()


        # Usar horarios de verano
        horarios = {
            'lunes': {'trabaja': empresa.trabaja_verano_lunes, 'entrada': empresa.entrada_verano_lunes,
                      'salida_comida': empresa.salida_comida_verano_lunes, 'entrada_comida': empresa.entrada_comida_verano_lunes, 'salida': empresa.salida_verano_lunes},
            'martes': {'trabaja': empresa.trabaja_verano_martes, 'entrada': empresa.entrada_verano_martes,
                       'salida_comida': empresa.salida_comida_verano_martes, 'entrada_comida': empresa.entrada_comida_verano_martes, 'salida': empresa.salida_verano_martes},
            'miercoles': {'trabaja': empresa.trabaja_verano_miercoles, 'entrada': empresa.entrada_verano_miercoles,
                          'salida_comida': empresa.salida_comida_verano_miercoles, 'entrada_comida': empresa.entrada_comida_verano_miercoles, 'salida': empresa.salida_verano_miercoles},
            'jueves': {'trabaja': empresa.trabaja_verano_jueves, 'entrada': empresa.entrada_verano_jueves,
                       'salida_comida': empresa.salida_comida_verano_jueves, 'entrada_comida': empresa.entrada_comida_verano_jueves, 'salida': empresa.salida_verano_jueves},
            'viernes': {'trabaja': empresa.trabaja_verano_viernes, 'entrada': empresa.entrada_verano_viernes,
                        'salida_comida': empresa.salida_comida_verano_viernes, 'entrada_comida': empresa.entrada_comida_verano_viernes, 'salida': empresa.salida_verano_viernes},
            'sabado': {'trabaja': empresa.trabaja_verano_sabado, 'entrada': empresa.entrada_verano_sabado,
                       'salida_comida': empresa.salida_comida_verano_sabado, 'entrada_comida': empresa.entrada_comida_verano_sabado, 'salida': empresa.salida_verano_sabado},
            'domingo': {'trabaja': empresa.trabaja_verano_domingo, 'entrada': empresa.entrada_verano_domingo,
                        'salida_comida': empresa.salida_comida_verano_domingo, 'entrada_comida': empresa.entrada_comida_verano_domingo, 'salida': empresa.salida_verano_domingo}
        }

        # Asignar los horarios de verano a todos los usuarios
        for user in users:
            for dia, horario in horarios.items():
                if dia == 'lunes':
                    user.trabaja_verano_lunes = horario['trabaja']
                    user.entrada_verano_lunes = horario['entrada']
                    user.salida_comida_verano_lunes = horario['salida_comida']
                    user.entrada_comida_verano_lunes = horario['entrada_comida']
                    user.salida_verano_lunes = horario['salida']
                elif dia == 'martes':
                    user.trabaja_verano_martes = horario['trabaja']
                    user.entrada_verano_martes = horario['entrada']
                    user.salida_comida_verano_martes = horario['salida_comida']
                    user.entrada_comida_verano_martes = horario['entrada_comida']
                    user.salida_verano_martes = horario['salida']
                elif dia == 'miercoles':
                    user.trabaja_verano_miercoles = horario['trabaja']
                    user.entrada_verano_miercoles = horario['entrada']
                    user.salida_comida_verano_miercoles = horario['salida_comida']
                    user.entrada_comida_verano_miercoles = horario['entrada_comida']
                    user.salida_verano_miercoles = horario['salida']
                elif dia == 'jueves':
                    user.trabaja_verano_jueves = horario['trabaja']
                    user.entrada_verano_jueves = horario['entrada']
                    user.salida_comida_verano_jueves = horario['salida_comida']
                    user.entrada_comida_verano_jueves = horario['entrada_comida']
                    user.salida_verano_jueves = horario['salida']
                elif dia == 'viernes':
                    user.trabaja_verano_viernes = horario['trabaja']
                    user.entrada_verano_viernes = horario['entrada']
                    user.salida_comida_verano_viernes = horario['salida_comida']
                    user.entrada_comida_verano_viernes = horario['entrada_comida']
                    user.salida_verano_viernes = horario['salida']
                elif dia == 'sabado':
                    user.trabaja_verano_sabado = horario['trabaja']
                    user.entrada_verano_sabado = horario['entrada']
                    user.salida_comida_verano_sabado = horario['salida_comida']
                    user.entrada_comida_verano_sabado = horario['entrada_comida']
                    user.salida_verano_sabado = horario['salida']
                elif dia == 'domingo':
                    user.trabaja_verano_domingo = horario['trabaja']
                    user.entrada_verano_domingo = horario['entrada']
                    user.salida_comida_verano_domingo = horario['salida_comida']
                    user.entrada_comida_verano_domingo = horario['entrada_comida']
                    user.salida_verano_domingo = horario['salida']

            # Guardar cambios
            db.session.commit()

        return jsonify({'success': True, 'message': 'Horarios de verano asignados correctamente a todos los usuarios'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/asignar_horario_verano_usuario/<int:user_id>', methods=['POST'])
def asignar_horario_verano_usuario(user_id):
    try:
        # Obtener el usuario
        user = User.query.get_or_404(user_id)

        # Obtener la empresa asociada al usuario
        empresa = user.empresa if user.empresa else Empresa.query.first()  # Puedes cambiar esto si el usuario tiene una relación con una empresa

        # Verificar si estamos dentro del periodo de verano
        today = datetime.today().date()

        if True:
            if True:
                # Usar horarios de verano
                horarios = {
                    'lunes': {'trabaja': empresa.trabaja_verano_lunes, 'entrada': empresa.entrada_verano_lunes,
                              'salida_comida': empresa.salida_comida_verano_lunes, 'entrada_comida': empresa.entrada_comida_verano_lunes, 'salida': empresa.salida_verano_lunes},
                    'martes': {'trabaja': empresa.trabaja_verano_martes, 'entrada': empresa.entrada_verano_martes,
                               'salida_comida': empresa.salida_comida_verano_martes, 'entrada_comida': empresa.entrada_comida_verano_martes, 'salida': empresa.salida_verano_martes},
                    'miercoles': {'trabaja': empresa.trabaja_verano_miercoles, 'entrada': empresa.entrada_verano_miercoles,
                                  'salida_comida': empresa.salida_comida_verano_miercoles, 'entrada_comida': empresa.entrada_comida_verano_miercoles, 'salida': empresa.salida_verano_miercoles},
                    'jueves': {'trabaja': empresa.trabaja_verano_jueves, 'entrada': empresa.entrada_verano_jueves,
                               'salida_comida': empresa.salida_comida_verano_jueves, 'entrada_comida': empresa.entrada_comida_verano_jueves, 'salida': empresa.salida_verano_jueves},
                    'viernes': {'trabaja': empresa.trabaja_verano_viernes, 'entrada': empresa.entrada_verano_viernes,
                                'salida_comida': empresa.salida_comida_verano_viernes, 'entrada_comida': empresa.entrada_comida_verano_viernes, 'salida': empresa.salida_verano_viernes},
                    'sabado': {'trabaja': empresa.trabaja_verano_sabado, 'entrada': empresa.entrada_verano_sabado,
                               'salida_comida': empresa.salida_comida_verano_sabado, 'entrada_comida': empresa.entrada_comida_verano_sabado, 'salida': empresa.salida_verano_sabado},
                    'domingo': {'trabaja': empresa.trabaja_verano_domingo, 'entrada': empresa.entrada_verano_domingo,
                                'salida_comida': empresa.salida_comida_verano_domingo, 'entrada_comida': empresa.entrada_comida_verano_domingo, 'salida': empresa.salida_verano_domingo}
                }

        # Asignar los horarios de verano al usuario
        for dia, horario in horarios.items():
            if dia == 'lunes':
                user.trabaja_verano_lunes = horario['trabaja']
                user.entrada_verano_lunes = horario['entrada']
                user.salida_comida_verano_lunes = horario['salida_comida']
                user.entrada_comida_verano_lunes = horario['entrada_comida']
                user.salida_verano_lunes = horario['salida']
            elif dia == 'martes':
                user.trabaja_verano_martes = horario['trabaja']
                user.entrada_verano_martes = horario['entrada']
                user.salida_comida_verano_martes = horario['salida_comida']
                user.entrada_comida_verano_martes = horario['entrada_comida']
                user.salida_verano_martes = horario['salida']
            elif dia == 'miercoles':
                user.trabaja_verano_miercoles = horario['trabaja']
                user.entrada_verano_miercoles = horario['entrada']
                user.salida_comida_verano_miercoles = horario['salida_comida']
                user.entrada_comida_verano_miercoles = horario['entrada_comida']
                user.salida_verano_miercoles = horario['salida']
            elif dia == 'jueves':
                user.trabaja_verano_jueves = horario['trabaja']
                user.entrada_verano_jueves = horario['entrada']
                user.salida_comida_verano_jueves = horario['salida_comida']
                user.entrada_comida_verano_jueves = horario['entrada_comida']
                user.salida_verano_jueves = horario['salida']
            elif dia == 'viernes':
                user.trabaja_verano_viernes = horario['trabaja']
                user.entrada_verano_viernes = horario['entrada']
                user.salida_comida_verano_viernes = horario['salida_comida']
                user.entrada_comida_verano_viernes = horario['entrada_comida']
                user.salida_verano_viernes = horario['salida']
            elif dia == 'sabado':
                user.trabaja_verano_sabado = horario['trabaja']
                user.entrada_verano_sabado = horario['entrada']
                user.salida_comida_verano_sabado = horario['salida_comida']
                user.entrada_comida_verano_sabado = horario['entrada_comida']
                user.salida_verano_sabado = horario['salida']
            elif dia == 'domingo':
                user.trabaja_verano_domingo = horario['trabaja']
                user.entrada_verano_domingo = horario['entrada']
                user.salida_comida_verano_domingo = horario['salida_comida']
                user.entrada_comida_verano_domingo = horario['entrada_comida']
                user.salida_verano_domingo = horario['salida']

        # Guardar cambios
        db.session.commit()

        return jsonify({'success': True, 'message': 'Horario de verano asignado correctamente al usuario'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/eliminar_centro_trabajo', methods=['POST'])
def eliminar_centro_trabajo():
    centro_id = request.form.get('id')  # Obtener el ID desde los datos del formulario
    if not centro_id:
        return jsonify({'success': False, 'message': 'ID del centro no proporcionado'}), 400

    centro = CentroTrabajo.query.get(centro_id)

    if centro is None:
        return jsonify({'success': False, 'message': 'Centro de trabajo no encontrado'}), 404

    db.session.delete(centro)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Centro de trabajo eliminado con éxito'}), 200


@app.route('/eliminar_departamento', methods=['POST'])
def eliminar_departamento():
    departamento_id = request.form.get('id')  # Obtener el ID del departamento
    if not departamento_id:
        return jsonify({'success': False, 'message': 'ID del departamento no proporcionado'}), 400

    departamento = Departamento.query.get(departamento_id)

    if departamento is None:
        return jsonify({'success': False, 'message': 'Departamento no encontrado'}), 404

    db.session.delete(departamento)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Departamento eliminado con éxito'}), 200

@app.route('/guardar_centros_departamentos', methods=['POST'])
def guardar_centros_departamentos():
    try:
        # Obtenemos los datos enviados desde el frontend
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se enviaron datos en formato JSON'}), 400

        centros_trabajo = data.get('centros_trabajo', [])
        departamentos = data.get('departamentos', [])


        empresa_id = data.get('empresa_id')
        if not empresa_id:
            return jsonify({'success': False, 'message': 'No se proporcionó empresa_id'}), 400

        empresa = Empresa.query.get(empresa_id)
        if not empresa:
            return jsonify({'success': False, 'message': 'Empresa no encontrada'}), 404


        # Obtenemos los centros de trabajo y departamentos actuales de la base de datos
        centros_actuales = CentroTrabajo.query.filter(CentroTrabajo.empresa_id == empresa.id).all()
        departamentos_actuales = Departamento.query.filter(Departamento.empresa_id == empresa.id).all()
        for c in centros_actuales:
            print(c.nombre)

        # Limpiar los centros de trabajo y departamentos antiguos que no están en la lista de los nuevos
        #for centro in centros_actuales:
            #if centro.nombre not in [centro_data['nombre'] for centro_data in centros_trabajo]:
                #db.session.delete(centro)

        #for departamento in departamentos_actuales:
            #if departamento.nombre not in [departamento_data['nombre'] for departamento_data in departamentos]:
                #db.session.delete(departamento)

        CentroTrabajo.query.filter_by(empresa_id=empresa.id).delete()
        Departamento.query.filter_by(empresa_id=empresa.id).delete()
        for centro in centros_trabajo:
            print(centro['nombre'])
            nuevo_centro = CentroTrabajo(nombre=centro['nombre'], empresa_id=empresa.id)
            db.session.add(nuevo_centro)

        # Agregar los nuevos departamentos a la base de datos
        for departamento in departamentos:
            nuevo_departamento = Departamento(nombre=departamento['nombre'], empresa_id=empresa.id)
            db.session.add(nuevo_departamento)

        # Guardar los cambios en la base de datos
        db.session.commit()

        return jsonify({'success': True, 'message': 'Centros de trabajo y departamentos actualizados correctamente'}), 200

    except Exception as e:
        db.session.rollback()  # Rollback de la sesión en caso de error
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/guardar_logos/<int:empresa_id>', methods=['POST'])
def guardar_logos(empresa_id):
    empresa = Empresa.query.get_or_404(empresa_id)

    if 'logo_favicon' in request.files:
        file = request.files['logo_favicon']
        if file and file.filename:
            filename = f"logo_favicon_empresa_{empresa.id}.png"
            filepath = os.path.join(UPLOAD_FOLDER2, filename)
            file.save(filepath)
            empresa.logo_favicon = filename  # <--- solo nombre del archivo

    if 'logo_login' in request.files:
        file2 = request.files['logo_login']
        if file2 and file2.filename:
            filename2 = f"logo_login_empresa_{empresa.id}.png"
            filepath2 = os.path.join(UPLOAD_FOLDER2, filename2)
            file2.save(filepath2)
            empresa.logo_login = filename2  # <--- solo nombre del archivo

    if 'logo_menu' in request.files:
        file3 = request.files['logo_menu']
        if file3 and file3.filename:
            filename3 = f"logo_menu_empresa_{empresa.id}.png"
            filepath3 = os.path.join(UPLOAD_FOLDER2, filename3)
            file3.save(filepath3)
            empresa.logo_menu = filename3  # <--- solo nombre del archivo

    db.session.commit()
    flash('Los logos se han guardado correctamente.')
    return redirect(url_for('guardar_empresa'))

import io
import pyotp
import qrcode
from flask import send_file
@app.route('/enable_2fa', methods=['GET', 'POST'])
@login_required
def enable_2fa():
    if request.method == 'GET':
        # Si no existe un secreto previamente asignado, se genera uno nuevo
        if not current_user.two_factor_secret:
            secret = pyotp.random_base32()  # Generar una nueva clave secreta
            current_user.two_factor_secret = secret
            db.session.commit()
        else:
            secret = current_user.two_factor_secret  # Si ya existe, se usa el existente

        # Crear el código QR
        issuer = "Gestionaempresa"  # Cambia este nombre por el de tu aplicación
        qr_code_url = pyotp.TOTP(secret).provisioning_uri(name=current_user.email, issuer_name=issuer)

        # Generar el código QR y devolverlo como una imagen
        qr = qrcode.make(qr_code_url)
        buf = io.BytesIO()
        qr.save(buf, format='PNG')
        buf.seek(0)

        return send_file(buf, mimetype='image/png')

    elif request.method == 'POST':
        otp_code = request.form.get('otp_code')  # Código OTP ingresado por el usuario
        try:
            # Verificar el código OTP
            totp = pyotp.TOTP(current_user.two_factor_secret)
            if totp.verify(otp_code):
                current_user.two_factor_enabled = True  # Habilitar 2FA
                db.session.commit()
                flash("2FA habilitado correctamente.", "success")
                return redirect(url_for('edit_profile'))  # Redirigir a la página del perfil
            else:
                flash("Código 2FA incorrecto. Inténtalo de nuevo.", "danger")
                return redirect(url_for('edit_profile'))  # Redirigir al perfil si el código es incorrecto
        except Exception as e:
            # Capturar cualquier error y mostrarlo
            flash(f"Error al verificar 2FA: {str(e)}", "danger")
            return redirect(url_for('edit_profile'))  # Redirigir al perfil en caso de error



@app.route('/disable_2fa', methods=['POST'])
@login_required
def disable_2fa():
    print("TRALARALARALAAAAA")
    # Deshabilitar el 2FA para el usuario actual
    current_user.two_factor_enabled = False
    current_user.two_factor_secret = None  # Opcional: limpiar el secreto
    db.session.commit()
    flash("2FA deshabilitado correctamente.", "success")
    return redirect(url_for('edit_profile'))


@app.route('/verify_2fa', methods=['POST'])
@login_required
def verify_2fa():
    user = current_user
    otp_code = request.form.get('otp_code')

    if user.two_factor_secret:
        totp = pyotp.TOTP(user.two_factor_secret)
        if totp.verify(otp_code):  # Verifica el código ingresado
            user.two_factor_enabled = True
            user.two_factor_verified = True
            db.session.commit()
            flash("2FA activado correctamente.", "success")
            return redirect(url_for('edit_profile'))
        else:
            flash("Código 2FA incorrecto.", "error")
            return redirect(url_for('edit_profile'))
    flash("No se pudo activar el 2FA.", "error")
    return redirect(url_for('edit_profile'))


from flask_mail import Message


from flask import jsonify

from flask import jsonify, request

@app.route('/resend_2fa_qr', methods=['POST'])
def resend_2fa_qr():
    print("Recibida solicitud para reenviar el QR.")

    # Obtener el correo del formulario (se espera que sea enviado en el cuerpo de la solicitud)
    data = request.get_json()
    email = data.get('email')
    print(f"Correo recibido: {email}")

    # Verificar si se proporcionó un correo electrónico
    if not email:
        print("No se proporcionó correo electrónico.")
        return jsonify({"success": False, "message": "El correo electrónico es necesario."}), 400

    # Buscar al usuario en la base de datos usando el correo electrónico
    user = User.query.filter_by(email=email).first()
    if not user:
        print(f"No se encontró un usuario con el correo {email}.")
        return jsonify({"success": False, "message": "No se encontró un usuario con ese correo electrónico."}), 404

    # Verificar si el usuario tiene un secreto 2FA
    secret = user.two_factor_secret
    if not secret:
        print(f"El usuario {email} no tiene un código 2FA.")
        return jsonify({"success": False, "message": "No se pudo encontrar un código 2FA para tu cuenta."}), 400

    # Crear el código QR
    print("Generando el código QR...")
    issuer = "Gestionaempresa"
    qr_code_url = pyotp.TOTP(secret).provisioning_uri(name=email, issuer_name=issuer)
    print(f"QR Code URL: {qr_code_url}")

    # Generar el código QR como una imagen
    qr = qrcode.make(qr_code_url)
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    buf.seek(0)

    # Obtener la URL de la imagen de la firma
    image_url = url_for('static', filename='firma.png', _external=True)

    email_body = f"""
    Hola, {user.nombre}!<br>


    Abra el código QR adjunto en el correo electrónico y abra la aplicación de autenticación e insértelo para ingresar el código 2FA en la aplicación.<br>

    <b><i>Elimine el correo una vez recuperado el código 2FA.</i></b><br><br>
    <!-- Aquí incluimos la imagen de la firma -->
    <img src="{image_url}" alt="Firma de Ecoinntech" style="max-width: 200px; height: auto;">
    """

    # Enviar el código QR por correo
    print(f"Enviando el QR al correo: {email}")
    msg = Message("Recuperar código 2FA", recipients=[email])
    msg.body = "Este correo solo es legible en formato HTML, por favor habilita HTML en tu cliente de correo."
    msg.html = email_body  # Usamos el campo .html para enviar el correo con formato HTML
    msg.attach("qr_code.png", "image/png", buf.read())

    try:
        mail.send(msg)
        print(f"Correo enviado exitosamente a {email}.")
        return jsonify({"success": True, "message": "Te hemos enviado un nuevo código QR a tu correo."}), 200
    except Exception as e:
        print(f"Error al enviar el correo: {e}")
        return jsonify({"success": False, "message": "Ocurrió un error al enviar el correo. Intenta nuevamente."}), 500



# Función para convertir las horas en formato legible
def format_time(value):
    """Formatea la hora a 'HH:MM:SS' o retorna None si el valor es None."""
    if value:
        return value.strftime('%H:%M:%S')
    return None

def serialize_work_schedule(work_schedule):
    for day, schedule in work_schedule.items():
        # Utiliza la función `format_time` para convertir las horas de manera eficiente
        schedule["entrada"] = format_time(schedule["entrada"])
        schedule["salida_comida"] = format_time(schedule["salida_comida"])
        schedule["entrada_comida"] = format_time(schedule["entrada_comida"])
        schedule["salida"] = format_time(schedule["salida"])
    return work_schedule

def serialize_excepciones(excepciones_list):
    for excepcion in excepciones_list:
        excepcion["entrada"] = format_time(excepcion["entrada"])
        excepcion["salida"] = format_time(excepcion["salida"])
        excepcion["entrada_comida"] = format_time(excepcion["entrada_comida"])
        excepcion["salida_comida"] = format_time(excepcion["salida_comida"])
    return excepciones_list






@app.route('/calendario_user/<int:user_id>')
def calendario_user(user_id):
    if not any(role.name in ["admin", "gestor", "encargado"] for role in current_user.roles):
        user_id = current_user.id
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()

    view_mode = request.args.get('view', 'user')
    # Supón que el usuario ya está autenticado
    user = User.query.get(user_id)
    if not user:
        return "Usuario no encontrado", 404

    current_year = datetime.now().year

    # Obtener los días festivos de la base de datos
    festivos = Festivo.query.all()

    # Convertir la información de días festivos en formato de texto ("YYYY-MM-DD")
    non_workdays = [f"{festivo.año}-{festivo.mes:02d}-{festivo.dia:02d}" for festivo in festivos]

    # Obtener los horarios de trabajo de la semana
    work_schedule = {
        "lunes": {
            "trabaja": user.trabaja_lunes,
            "entrada": user.entrada_lunes,
            "salida_comida": user.salida_comida_lunes,
            "entrada_comida": user.entrada_comida_lunes,
            "salida": user.salida_lunes
        },
        "martes": {
            "trabaja": user.trabaja_martes,
            "entrada": user.entrada_martes,
            "salida_comida": user.salida_comida_martes,
            "entrada_comida": user.entrada_comida_martes,
            "salida": user.salida_martes
        },
        "miercoles": {
            "trabaja": user.trabaja_miercoles,
            "entrada": user.entrada_miercoles,
            "salida_comida": user.salida_comida_miercoles,
            "entrada_comida": user.entrada_comida_miercoles,
            "salida": user.salida_miercoles
        },
        "jueves": {
            "trabaja": user.trabaja_jueves,
            "entrada": user.entrada_jueves,
            "salida_comida": user.salida_comida_jueves,
            "entrada_comida": user.entrada_comida_jueves,
            "salida": user.salida_jueves
        },
        "viernes": {
            "trabaja": user.trabaja_viernes,
            "entrada": user.entrada_viernes,
            "salida_comida": user.salida_comida_viernes,
            "entrada_comida": user.entrada_comida_viernes,
            "salida": user.salida_viernes
        },
        "sabado": {
            "trabaja": user.trabaja_sabado,
            "entrada": user.entrada_sabado,
            "salida_comida": user.salida_comida_sabado,
            "entrada_comida": user.entrada_comida_sabado,
            "salida": user.salida_sabado
        },
        "domingo": {
            "trabaja": user.trabaja_domingo,
            "entrada": user.entrada_domingo,
            "salida_comida": user.salida_comida_domingo,
            "entrada_comida": user.entrada_comida_domingo,
            "salida": user.salida_domingo
        }
    }

    # Convertir las horas de trabajo a formato de texto
    work_schedule = serialize_work_schedule(work_schedule)

    work_schedule_verano = {
        "lunes": {
            "trabaja": user.trabaja_verano_lunes,
            "entrada": user.entrada_verano_lunes,
            "salida_comida": user.salida_comida_verano_lunes,
            "entrada_comida": user.entrada_comida_verano_lunes,
            "salida": user.salida_verano_lunes
        },
        "martes": {
            "trabaja": user.trabaja_verano_martes,
            "entrada": user.entrada_verano_martes,
            "salida_comida": user.salida_comida_verano_martes,
            "entrada_comida": user.entrada_comida_verano_martes,
            "salida": user.salida_verano_martes
        },
        "miercoles": {
            "trabaja": user.trabaja_verano_miercoles,
            "entrada": user.entrada_verano_miercoles,
            "salida_comida": user.salida_comida_verano_miercoles,
            "entrada_comida": user.entrada_comida_verano_miercoles,
            "salida": user.salida_verano_miercoles
        },
        "jueves": {
            "trabaja": user.trabaja_verano_jueves,
            "entrada": user.entrada_verano_jueves,
            "salida_comida": user.salida_comida_verano_jueves,
            "entrada_comida": user.entrada_comida_verano_jueves,
            "salida": user.salida_verano_jueves
        },
        "viernes": {
            "trabaja": user.trabaja_verano_viernes,
            "entrada": user.entrada_verano_viernes,
            "salida_comida": user.salida_comida_verano_viernes,
            "entrada_comida": user.entrada_comida_verano_viernes,
            "salida": user.salida_verano_viernes
        },
        "sabado": {
            "trabaja": user.trabaja_verano_sabado,
            "entrada": user.entrada_verano_sabado,
            "salida_comida": user.salida_comida_verano_sabado,
            "entrada_comida": user.entrada_comida_verano_sabado,
            "salida": user.salida_verano_sabado
        },
        "domingo": {
            "trabaja": user.trabaja_verano_domingo,
            "entrada": user.entrada_verano_domingo,
            "salida_comida": user.salida_comida_verano_domingo,
            "entrada_comida": user.entrada_comida_verano_domingo,
            "salida": user.salida_verano_domingo
        }
    }
    work_schedule_verano = serialize_work_schedule(work_schedule_verano)
    # Obtener las excepciones del usuario (por ejemplo, festivos, cambios de jornada)
    excepciones = RegistroHorario.query.filter_by(user_id=user.id, validada=True).all()
    print("AQUI VAN LAS EXCEPCIONES")
    print(excepciones)
    for e in excepciones:
        print(e.fecha)

    excepciones_list = []
    for excepcion in excepciones:
        excepciones_list.append({
            "fecha": excepcion.fecha,
            "entrada": excepcion.entrada,
            "salida": excepcion.salida,
            "entrada_comida": excepcion.entrada_comida,
            "salida_comida": excepcion.salida_comida
        })

    # Convertir las horas en las excepciones a formato de texto
    excepciones_list = serialize_excepciones(excepciones_list)

    events = []
    pending_events = []
    work_schedule = work_schedule or {}
    excepciones_list = excepciones_list or []

    empresa = Empresa.query.first()
    if not empresa:
        return "Empresa no encontrada", 404

    fecha_inicio_verano = empresa.fecha_inicio_verano
    fecha_fin_verano = empresa.fecha_fin_verano

    if 'admin' in [role.name for role in user.roles]:
        rol='admin'
    else:
        rol='user'

    # Renderizar la plantilla HTML y pasar los datos del horario y las excepciones
    return render_template('calendario.html',view_mode=view_mode,rol=rol,notificaciones=notificaciones, user=user, work_schedule=work_schedule, work_schedule_verano=work_schedule_verano,fecha_inicio_verano=fecha_inicio_verano, fecha_fin_verano=fecha_fin_verano, excepciones=excepciones_list, non_workdays=non_workdays, events=events, pending_events=pending_events, user_id=user.id, current_year=current_year)




@app.route('/set_work_schedule/<int:user_id>', methods=['POST'])
def set_work_schedule(user_id):
    # Obtener el usuario de la base de datos
    user = User.query.get(user_id)
    if not user:
        return "Usuario no encontrado", 404

    # Establecer la jornada semanal para el usuario (esto puede variar según los valores que des)
    user.trabaja_lunes = True
    user.entrada_lunes = time(9, 0)  # 9:00 AM
    user.salida_comida_lunes = time(13, 0)  # 1:00 PM
    user.entrada_comida_lunes = time(14, 0)  # 2:00 PM
    user.salida_lunes = time(18, 0)  # 6:00 PM

    user.trabaja_martes = True
    user.entrada_martes = time(9, 0)
    user.salida_comida_martes = time(13, 0)
    user.entrada_comida_martes = time(14, 0)
    user.salida_martes = time(18, 0)

    # Repite este proceso para el resto de los días de la semana...
    # Establecer los valores para miércoles, jueves, viernes, etc.

    # Guardar los cambios en la base de datos
    db.session.commit()

    return "Jornada semanal actualizada correctamente"

@app.route('/set_exception/<int:user_id>', methods=['POST'])
def set_exception(user_id):
    # Obtener el usuario de la base de datos
    user = User.query.get(user_id)
    if not user:
        return "Usuario no encontrado", 404

    # Definir la fecha de la excepción (por ejemplo, el 25 de diciembre)
    fecha_excepcion = "2024-10-18"

    # Convertir la cadena a un objeto datetime
    fecha_objeto = datetime.strptime(fecha_excepcion, "%Y-%m-%d").date()
    entrada_excepcion = time(10, 0)  # Hora de entrada diferente
    salida_excepcion = time(16, 0)  # Hora de salida diferente

    # Crear un registro de excepción
    excepcion = RegistroHorario(
        user_id=user.id,
        fecha=fecha_objeto,
        entrada=entrada_excepcion,
        salida=salida_excepcion
    )

    # Añadir la excepción a la base de datos
    db.session.add(excepcion)
    db.session.commit()

    return f"Excepción agregada para el {fecha_excepcion}"




from babel.dates import format_date
from flask_login import current_user

@app.route('/save_exception', methods=['POST'])
def save_exception():
    try:
        hora_espana1 = datetime.now().astimezone(timezone_spain)
        print(f"Hora actual en España: {hora_espana1}")

        data = request.get_json()
        print(f"Datos recibidos: {data}")

        user_id = data.get('user_id')
        fecha_str = data.get('fecha')
        entrada_str = data.get('entrada')
        salida_str = data.get('salida')
        entrada_comida_str = data.get('entrada_comida')  # Nuevo campo para entrada comida
        salida_comida_str = data.get('salida_comida')    # Nuevo campo para salida comida

        print(f"ID Usuario: {user_id}, Fecha: {fecha_str}, Entrada: {entrada_str}, Salida: {salida_str}")
        print(f"Entrada comida: {entrada_comida_str}, Salida comida: {salida_comida_str}")

        # Convertir la cadena de fecha del formato 'dd/mm/yyyy' a 'yyyy-mm-dd'
        try:
            # Cambiar el formato de la fecha
            fecha = datetime.strptime(fecha_str, '%d/%m/%Y').date()
            print(f"Fecha convertida: {fecha}")
        except ValueError:
            print(f"Error en el formato de fecha: {fecha_str}")
            return jsonify({'status': 'error', 'message': 'Invalid date format'}), 400

        # Convertir las cadenas de tiempo a objetos time de Python
        try:
            entrada = datetime.strptime(entrada_str, '%H:%M').time()
            print(f"Entrada convertida: {entrada}")
        except ValueError:
            try:
                entrada = datetime.strptime(entrada_str, '%H:%M:%S').time()
                print(f"Entrada convertida (segundos): {entrada}")
            except ValueError:
                print(f"Error en el formato de entrada: {entrada_str}")
                return jsonify({'status': 'error', 'message': 'Invalid time format for entrada'}), 400

        try:
            salida = datetime.strptime(salida_str, '%H:%M').time()
            print(f"Salida convertida: {salida}")
        except ValueError:
            try:
                salida = datetime.strptime(salida_str, '%H:%M:%S').time()
                print(f"Salida convertida (segundos): {salida}")
            except ValueError:
                print(f"Error en el formato de salida: {salida_str}")
                return jsonify({'status': 'error', 'message': 'Invalid time format for salida'}), 400

        # Convertir los horarios de comida
        entrada_comida = None
        salida_comida = None
        if entrada_comida_str and salida_comida_str:
            try:
                entrada_comida = datetime.strptime(entrada_comida_str, '%H:%M').time()
                print(f"Entrada comida convertida: {entrada_comida}")
            except ValueError:
                try:
                    entrada_comida = datetime.strptime(entrada_comida_str, '%H:%M:%S').time()
                    print(f"Entrada comida convertida (segundos): {entrada_comida}")
                except ValueError:
                    print(f"Error en el formato de entrada comida: {entrada_comida_str}")
                    return jsonify({'status': 'error', 'message': 'Invalid time format for entrada_comida'}), 400

            try:
                salida_comida = datetime.strptime(salida_comida_str, '%H:%M').time()
                print(f"Salida comida convertida: {salida_comida}")
            except ValueError:
                try:
                    salida_comida = datetime.strptime(salida_comida_str, '%H:%M:%S').time()
                    print(f"Salida comida convertida (segundos): {salida_comida}")
                except ValueError:
                    print(f"Error en el formato de salida comida: {salida_comida_str}")
                    return jsonify({'status': 'error', 'message': 'Invalid time format for salida_comida'}), 400

        # Verificar si ya existe una excepción para esta fecha y usuario
        excepcion = RegistroHorario.query.filter_by(user_id=user_id, fecha=fecha).first()
        print(f"Excepción existente: {excepcion}")

        if excepcion:
            # Eliminar la excepción existente
            db.session.delete(excepcion)
            print(f"Excepción eliminada para la fecha {fecha} y usuario {user_id}")

        # Crear una nueva excepción
        nueva_excepcion = RegistroHorario(
            user_id=user_id,
            fecha=fecha,
            entrada=entrada,
            salida=salida,
            entrada_comida=entrada_comida,
            salida_comida=salida_comida,
            validada=False  # Al principio está como no validada
        )
        db.session.add(nueva_excepcion)
        print(f"Nueva excepción creada para el usuario {user_id}, fecha {fecha}")

        # Verificar si el current_user es el dueño del calendario (user_id)
        if current_user.id == user_id:
            fecha_str_es = format_date(fecha, "EEEE, d 'de' MMMM 'de' y", locale='es_ES').capitalize()
            concepto = f"Se ha solicitado un nuevo cambio de horario para el día {fecha_str_es} por parte de {current_user.nombre} {current_user.primer_apellido}"

            print(f"Notificación general: {concepto}")

            # 1. Notificar a administradores
            admins = User.query.filter(User.roles.any(name='admin')).all()
            for admin in admins:
                print(f"Notificando admin: {admin.nombre}")
                db.session.add(Notificacion(
                    usuario_id=admin.id,
                    concepto=concepto,
                    fecha_hora=hora_espana1
                ))

            # 2. Notificar a gestores de la misma empresa con registros_horarios=True
            gestores = User.query.filter(
                User.roles.any(name='gestor'),
                User.empresa_id == current_user.empresa_id,
                User.registros_horarios.is_(True)
            ).all()
            for gestor in gestores:
                print(f"Notificando gestor: {gestor.nombre}")
                db.session.add(Notificacion(
                    usuario_id=gestor.id,
                    concepto=concepto,
                    fecha_hora=hora_espana1
                ))

            # 3. Notificar a encargados de la misma empresa
            encargados = User.query.filter(
                User.roles.any(name='encargado'),
                User.empresa_id == current_user.empresa_id
            ).all()
            for encargado in encargados:
                print(f"Notificando encargado: {encargado.nombre}")
                db.session.add(Notificacion(
                    usuario_id=encargado.id,
                    concepto=concepto,
                    fecha_hora=hora_espana1
                ))

            db.session.commit()

        else:
            fecha_str_es = format_date(fecha, "EEEE, d 'de' MMMM 'de' y", locale='es_ES').capitalize()
            nueva_notificacion = Notificacion(
                usuario_id=user_id,
                concepto=f"Se ha solicitado un nuevo cambio de horario para el día {fecha_str_es} por parte del administrador.",

                fecha_hora=hora_espana1
            )
            db.session.add(nueva_notificacion)

        # Enviar un correo electrónico a los administradores o al usuario según corresponda
        if current_user.id == user_id:
            # 1. Admins
            admins = User.query.filter(User.roles.any(name='admin')).all()
            for admin in admins:
                print(f"Enviando correo a admin: {admin.nombre} ({admin.email})")
                send_exception_to_admin(admin.email, nueva_excepcion.id, fecha, user_id, entrada, salida, entrada_comida, salida_comida)

            # 2. Gestores de la misma empresa con registros_horarios=True
            gestores = User.query.filter(
                User.roles.any(name='gestor'),
                User.empresa_id == current_user.empresa_id,
                User.registros_horarios.is_(True)
            ).all()
            for gestor in gestores:
                print(f"Enviando correo a gestor: {gestor.nombre} ({gestor.email})")
                send_exception_to_admin(gestor.email, nueva_excepcion.id, fecha, user_id, entrada, salida, entrada_comida, salida_comida)

            # 3. Encargados de la misma empresa
            encargados = User.query.filter(
                User.roles.any(name='encargado'),
                User.empresa_id == current_user.empresa_id
            ).all()
            for encargado in encargados:
                print(f"Enviando correo a encargado: {encargado.nombre} ({encargado.email})")
                send_exception_to_admin(encargado.email, nueva_excepcion.id, fecha, user_id, entrada, salida, entrada_comida, salida_comida)

            # 4. Correo adicional fijo
            send_exception_to_admin('alfonsomunoz@adaptasystem.com', nueva_excepcion.id, fecha, user_id, entrada, salida, entrada_comida, salida_comida)

        else:
            user = User.query.get(user_id)
            if user:
                print(f"Enviando correo al usuario: {user.nombre} ({user.email})")
                send_exception_to_user(user.email, nueva_excepcion.id, fecha, user_id, entrada, salida, entrada_comida, salida_comida)

        try:
            db.session.commit()
            print("Excepción guardada correctamente en la base de datos.")
            return jsonify({'status': 'success'})
        except Exception as e:
            print(f"Error al guardar en la base de datos: {e}")
            db.session.rollback()
            return jsonify({'status': 'error', 'message': str(e)}), 500

    except Exception as e:
        print(f"Error inesperado: {e}")
        return jsonify({'status': 'error', 'message': 'An unexpected error occurred'}), 500


import locale

# Función para enviar correo a los administradores
# Función para enviar correo a los administradores
def send_exception_to_admin(admin_email, exception_id, fecha, user_id, entrada, salida, entrada_comida, salida_comida):
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    accept_url = url_for('accept_exception', exception_id=exception_id, _external=True)
    reject_url = url_for('reject_exception', exception_id=exception_id, _external=True)
    empresa = Empresa.query.first()
    if not empresa:
        return "Empresa no encontrada", 404

    user = User.query.filter_by(id=user_id).first()
    if not user:
        return "Usuario no encontrado", 404
    nombre_usuario=user.nombre
    apellidos=user.primer_apellido

    # Determinar si es verano o invierno
    if empresa.fecha_inicio_verano and (empresa.fecha_inicio_verano <= fecha <= empresa.fecha_fin_verano):
        temporada = "verano"
        sufijo = "verano_"
    else:
        temporada = "invierno"
        sufijo = ""

    # Obtener el día de la semana (0: Lunes, 6: Domingo)
    dia_semana = fecha.weekday()
    dias = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
    dia_nombre = dias[dia_semana]

    # Seleccionar los horarios específicos basados en el día de la semana y la temporada
    trabaja = getattr(user, f"trabaja_{sufijo}{dia_nombre}")
    entrada_predeterminada = getattr(user, f"entrada_{sufijo}{dia_nombre}")
    salida_predeterminada = getattr(user, f"salida_{sufijo}{dia_nombre}")
    entrada_comida_predeterminada = getattr(user, f"entrada_comida_{sufijo}{dia_nombre}")
    salida_comida_predeterminada = getattr(user, f"salida_comida_{sufijo}{dia_nombre}")

    # Si no trabaja en el día especificado, devolver mensaje
    if not trabaja:
        return f"El usuario no trabaja los {fecha.strftime('%A')}s en temporada de {temporada}", 400


    def format_time(time):
        return time.strftime('%H:%M') if time else "N/A"

    entrada_predeterminada = format_time(entrada_predeterminada)
    salida_predeterminada = format_time(salida_predeterminada)
    entrada_comida_predeterminada = format_time(entrada_comida_predeterminada)
    salida_comida_predeterminada = format_time(salida_comida_predeterminada)

    entrada = format_time(entrada)
    salida = format_time(salida)
    entrada_comida = format_time(entrada_comida)
    salida_comida = format_time(salida_comida)

    # Preparar el correo
    subject = "Nueva excepción de horario para validación"
    html_body = f"""
    <html>
    <body>
        <p>Hola administrador,</p>
        <p>El usuario {nombre_usuario} {apellidos} ha solicitado una modificación en su horario para el día {fecha.strftime('%d de %B de %Y')} ({temporada}). Por favor, revisa los detalles y valida la excepción.</p>
        <p><strong>Horario predeterminado:</strong></p>
        <ul>
            <li>Entrada: {entrada_predeterminada}</li>
            <li>Salida: {salida_predeterminada}</li>
            <li>Entrada comida: {entrada_comida_predeterminada}</li>
            <li>Salida comida: {salida_comida_predeterminada}</li>
        </ul>
        <p><strong>Horario solicitado:</strong></p>
        <ul>
            <li>Entrada: {entrada}</li>
            <li>Salida: {salida}</li>
            <li>Entrada comida: {entrada_comida}</li>
            <li>Salida comida: {salida_comida}</li>
        </ul>
        <p>
            <a href="{accept_url}" style="display: inline-block; padding: 20px 30px; margin: 10px; color: white; background-color: green; text-decoration: none; border-radius: 20px;">Aceptar</a>
            <a href="{reject_url}" style="display: inline-block; padding: 20px 30px; margin: 10px; color: white; background-color: red; text-decoration: none; border-radius: 20px;">Rechazar</a>
        </p>
        <p>Saludos,</p>
        <p>El equipo de Gestionaempresa.</p>
    </body>
    </html>
    """

    msg = Message(subject, recipients=[admin_email])
    msg.html = html_body

    # Enviar el correo
    try:
        mail.send(msg)
        print("Correo enviado con éxito")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")


# Función para enviar correo al usuario dueño del calendario (como ya estaba definida)
def send_exception_to_user(user_email, exception_id,fecha, user_id, entrada, salida, entrada_comida, salida_comida):
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    accept_url = url_for('accept_exception', exception_id=exception_id, _external=True)
    reject_url = url_for('reject_exception', exception_id=exception_id, _external=True)
    empresa = Empresa.query.first()
    if not empresa:
        return "Empresa no encontrada", 404

    user = User.query.filter_by(id=user_id).first()
    if not user:
        return "Usuario no encontrado", 404
    nombre_usuario=user.nombre
    apellidos=user.primer_apellido

    # Determinar si es verano o invierno
    if empresa.fecha_inicio_verano and (empresa.fecha_inicio_verano <= fecha <= empresa.fecha_fin_verano):
        temporada = "verano"
        sufijo = "verano_"
    else:
        temporada = "invierno"
        sufijo = ""

    # Obtener el día de la semana (0: Lunes, 6: Domingo)
    dia_semana = fecha.weekday()
    dias = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
    dia_nombre = dias[dia_semana]

    # Seleccionar los horarios específicos basados en el día de la semana y la temporada
    trabaja = getattr(user, f"trabaja_{sufijo}{dia_nombre}")
    entrada_predeterminada = getattr(user, f"entrada_{sufijo}{dia_nombre}")
    salida_predeterminada = getattr(user, f"salida_{sufijo}{dia_nombre}")
    entrada_comida_predeterminada = getattr(user, f"entrada_comida_{sufijo}{dia_nombre}")
    salida_comida_predeterminada = getattr(user, f"salida_comida_{sufijo}{dia_nombre}")

    # Si no trabaja en el día especificado, devolver mensaje
    if not trabaja:
        return f"El usuario no trabaja los {fecha.strftime('%A')}s en temporada de {temporada}", 400


    def format_time(time):
        return time.strftime('%H:%M') if time else "N/A"

    entrada_predeterminada = format_time(entrada_predeterminada)
    salida_predeterminada = format_time(salida_predeterminada)
    entrada_comida_predeterminada = format_time(entrada_comida_predeterminada)
    salida_comida_predeterminada = format_time(salida_comida_predeterminada)

    entrada = format_time(entrada)
    salida = format_time(salida)
    entrada_comida = format_time(entrada_comida)
    salida_comida = format_time(salida_comida)


    subject = "Nueva excepción de horario para validación"
    html_body = f"""
    <html>
    <body>
        <p>Hola,</p>
        <p>Se ha generado una nueva excepción en tu horario. Por favor, revisa los detalles y valida la excepción.</p>
        <p>El administrador ha solicitado una modificación en su horario para el día {fecha.strftime('%d de %B de %Y')} ({temporada}). Por favor, revisa los detalles y valida/rechaza la excepción.</p>
        <p><strong>Horario predeterminado:</strong></p>
        <ul>
            <li>Entrada: {entrada_predeterminada}</li>
            <li>Salida: {salida_predeterminada}</li>
            <li>Entrada comida: {entrada_comida_predeterminada}</li>
            <li>Salida comida: {salida_comida_predeterminada}</li>
        </ul>
        <p><strong>Horario solicitado:</strong></p>
        <ul>
            <li>Entrada: {entrada}</li>
            <li>Salida: {salida}</li>
            <li>Entrada comida: {entrada_comida}</li>
            <li>Salida comida: {salida_comida}</li>
        </ul>
        <p>
            <a href="{accept_url}" style="display: inline-block; padding: 20px 30px; margin: 10px; color: white; background-color: green; text-decoration: none; border-radius: 20px;">Aceptar</a>
            <a href="{reject_url}" style="display: inline-block; padding: 20px 30px; margin: 10px; color: white; background-color: red; text-decoration: none; border-radius: 20px;">Rechazar</a>
        </p>
        <p>Saludos,</p>
        <p>El equipo de Gestionaempresa.</p>
    </body>
    </html>
    """

    msg = Message(subject, recipients=[user_email])
    msg.html = html_body

    # Enviar el correo
    try:
        mail.send(msg)
        print("Correo enviado con éxito")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

@app.route('/accept_exception/<int:exception_id>')
def accept_exception(exception_id):
    """Ruta para aceptar una excepción y cambiar el campo 'validada' a True"""
    exception = RegistroHorario.query.get(exception_id)
    if exception:
        exception.validada = True
        db.session.commit()

        # Obtener el correo del usuario propietario de la excepción
        owner = User.query.get(exception.user_id)
        if not owner:
            return "Error: Propietario de la excepción no encontrado", 404

        owner_email = owner.email

        # Verificar si el current_user está autenticado
        if owner.is_authenticated:
            current_user_email = owner.email
            current_user_name = owner.nombre
        else:
            # Manejar el caso donde no hay current_user
            current_user_email = None
            current_user_name = "Un usuario no autenticado"
            print("Error: Usuario no autenticado al aceptar la excepción")
            return "Error: Usuario no autenticado al aceptar la excepción", 401

        # Formatear la fecha en español
        fecha_formateada = formatear_fecha_espanol(exception.fecha)

        subject = "Confirmación de excepción aceptada"
        message = f"Se ha aceptado la excepción para el día {fecha_formateada}."

        # Enviar correo al usuario que aceptó la excepción si está autenticado
        if current_user_email:
            send_confirmation_email(current_user_email, subject, message)

        # Enviar correo al dueño del calendario si es diferente
        if current_user_email != owner_email:
            send_confirmation_email(owner_email, subject, message)

        # Notificar a los administradores
        admins = get_admin_users()
        for admin in admins:
            send_confirmation_email(admin.email, subject, message)

        return redirect(url_for('login'))
    return "Excepción no encontrada."


@app.route('/reject_exception/<int:exception_id>')
def reject_exception(exception_id):
    """Ruta para rechazar una excepción y notificar al usuario"""
    exception = RegistroHorario.query.get(exception_id)
    if exception:
        exception.validada = False  # Actualiza el estado si es necesario
        db.session.commit()

        # Obtener el correo del usuario propietario de la excepción
        owner = User.query.get(exception.user_id)
        if not owner:
            return "Error: Propietario de la excepción no encontrado", 404

        owner_email = owner.email

        # Verificar si el current_user está autenticado
        if owner.is_authenticated:
            current_user_email = owner.email
            current_user_name = owner.nombre
        else:
            # Manejar el caso donde no hay current_user
            current_user_email = None
            current_user_name = "Un usuario no autenticado"
            print("Error: Usuario no autenticado al rechazar la excepción")
            return "Error: Usuario no autenticado al rechazar la excepción", 401

        # Formatear la fecha en español
        fecha_formateada = formatear_fecha_espanol(exception.fecha)

        subject = "Notificación de excepción rechazada"
        message = f"Se ha rechazado la excepción para el día {fecha_formateada}."

        # Enviar correo al usuario que rechazó la excepción si está autenticado
        if current_user_email:
            send_confirmation_email(current_user_email, subject, message)

        # Enviar correo al dueño del calendario si es diferente
        if current_user_email != owner_email:
            send_confirmation_email(owner_email, subject, message)

        # Notificar a los administradores
        admins = get_admin_users()
        for admin in admins:
            send_confirmation_email(admin.email, subject, message)

        return redirect(url_for('login'))
    return "Excepción no encontrada."


def send_confirmation_email(email, subject, message):
    """Función para enviar un correo de confirmación"""
    html_body = f"""
    <html>
    <body>
        <p>{message}</p>
        <p>Saludos,</p>
        <p>El equipo de Gestión de Empresa.</p>
    </body>
    </html>
    """

    msg = Message(subject, recipients=[email])
    msg.html = html_body

    # Enviar el correo
    try:
        mail.send(msg)
        print(f"📧 Correo de confirmación enviado con éxito a {email}")
    except Exception as e:
        print(f"⚠️ Error al enviar el correo de confirmación a {email}: {e}")


def formatear_fecha_espanol(fecha):
    """Convierte una fecha en formato 'día de mes de año' en español"""
    meses_espanol = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]
    return f"{fecha.day} de {meses_espanol[fecha.month - 1]} de {fecha.year}"
def eliminar_todos_los_festivos():
    # Eliminar todos los registros en la tabla Festivo
    Festivo.query.delete()
    db.session.commit()


def agregar_festivos():
    for festivo in non_workdays:
        # Comprobar si el festivo ya existe en la base de datos (solo por día y mes)
        empresas = Empresa.query.all()
        existe_festivo = Festivo.query.filter_by(dia=festivo.day, mes=festivo.month).first()
        if not existe_festivo:
            for empresa in empresas:
                # Si no existe, agregarlo a la base de datos
                nuevo_festivo = Festivo(dia=festivo.day, mes=festivo.month, empresa_id=empresa.id)
                db.session.add(nuevo_festivo)
                db.session.commit()  # Asegúrate de agregarlo a la base de datos

        # Asociar el festivo a todas las empresas


@app.route('/editar_hora', methods=['POST'])
def editar_hora():
    try:
        data = request.get_json()
        registro_id = data.get('registroId')
        campo = data.get('campo')
        nueva_hora = data.get('nuevaHora')
        observacion = data.get('observacion', '')

        if not registro_id or not campo or not nueva_hora:
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400

        # Asegurarse de que nueva_hora tiene formato HH:MM:SS o YYYY-MM-DDTHH:MM
        try:
            if 'T' in nueva_hora:
                nueva_fecha = datetime.strptime(nueva_hora, '%Y-%m-%dT%H:%M')
            else:
                nueva_fecha = datetime.strptime(nueva_hora, '%H:%M:%S')
        except ValueError:
            return jsonify({'success': False, 'error': 'Formato de hora no válido'}), 400

        registro = Registro.query.get(registro_id)
        if not registro:
            return jsonify({'success': False, 'error': 'Registro no encontrado'}), 404

        # Guardar valor anterior
        fecha_anterior = registro.entrada if campo == 'entrada' else registro.salida

        # Actualizar el campo correspondiente
        if campo == 'entrada':
            registro.entrada = nueva_fecha
        elif campo == 'salida':
            registro.salida = nueva_fecha

        # Actualizar información de modificación
        registro.modificado = 'usuario'
        registro.fecha_modificacion = datetime.now()
        registro.campo_modificado = campo
        registro.valor_anterior = fecha_anterior.strftime('%Y-%m-%d %H:%M:%S') if fecha_anterior else None
        registro.observaciones = observacion
        registro.aprobado_modificacion = False  # 🔴 pendiente de validación

        db.session.commit()

        return jsonify({
            'success': True,
            'nuevaHoraCompleta': nueva_fecha.strftime('%d-%m-%Y %H:%M:%S'),
            'pendiente': True
        })
    except Exception as e:
        print(f"Error al editar hora: {e}")
        return jsonify({'success': False, 'error': 'Error interno del servidor'}), 500





@app.route('/registrar_salida', methods=['POST'])
def registrar_salida():
    data = request.get_json()

    hora_real_salida = data.get('hora_real_salida')
    comentarios = data.get('comentarios')
    latitud = data.get('latitud')
    longitud = data.get('longitud')

    if not hora_real_salida:
        return jsonify({"status": "error", "message": "Hora real de salida es requerida"}), 400

    return registrar_salida_backend(hora_real_salida, comentarios, latitud, longitud)


from flask import url_for
from flask_mail import Message
import locale
from pytz import timezone as pytz_timezone

def registrar_salida_backend(hora_real_salida, comentarios, latitud, longitud):
    try:
        usuario_id = current_user.id
        usuario = User.query.get(usuario_id)
        if not usuario:
            print(f"[ERROR] Usuario no encontrado: {usuario_id}")
            return "Usuario no encontrado", 404

        fecha_hoy = datetime.now().date()
        registro = Registro.query.filter_by(user_id=usuario_id, salida=None).first()
        if not registro:
            print(f"[ERROR] Registro no encontrado para el usuario {usuario_id} en la fecha {fecha_hoy}")
            return "Registro no encontrado", 404

                # Establece el locale en español
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

        # Obtén el día de la semana en español
        dias_espanol = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
        dia_actual = datetime.today().weekday()  # 0 = lunes, 6 = domingo
        dia_nombre = dias_espanol[dia_actual]

        empresa = current_user.empresa if current_user.empresa else Empresa.query.first()
        if not empresa:
            print(f"[ERROR] No se pudo obtener la información de la empresa.")
            return "Empresa no encontrada", 404

        temporada = "verano" if empresa.fecha_inicio_verano and (empresa.fecha_inicio_verano <= fecha_hoy <= empresa.fecha_fin_verano) else "invierno"
        print(f"[INFO] Temporada actual: {temporada}")

        registro_horario = RegistroHorario.query.filter_by(user_id=usuario_id, fecha=fecha_hoy).first()

        if registro_horario and registro_horario.salida:
            print(f"[INFO] Se encontró una excepción de horario. Se usará la hora de salida: {registro_horario.salida}")
            salida_predeterminada = registro_horario.salida
        else:
            salida_predeterminada = getattr(usuario, f"salida_{'verano_' if temporada == 'verano' else ''}{dia_nombre}")

        if not isinstance(salida_predeterminada, time):
            print(f"[ERROR] El valor de salida_predeterminada no es un objeto time: {salida_predeterminada}")
            return "Formato de hora de salida incorrecto", 500

        salida_predeterminada_dt = datetime.combine(fecha_hoy, salida_predeterminada)
        registro.salida = salida_predeterminada_dt
        registro.latitud_salida=latitud
        registro.longitud_salida=longitud
        ip_address = request.headers.get('X-Real-IP')
        registro.ip_address_salida = ip_address

        # ✅ Se usa la hora actual como hora real de salida
        hora_real_salida = datetime.now(pytz_timezone('Europe/Madrid')).time()

        print(f"[INFO] Registrando excepción: salida tardía de {hora_real_salida}")

        excepcion = ExcepcionSalidaTardia(
            registro_id=registro.id,
            fecha=fecha_hoy,
            hora_salida_real=hora_real_salida,
            comentarios=comentarios
        )
        db.session.add(excepcion)
        db.session.commit()
        print(f"[INFO] Excepción registrada con éxito: {excepcion.id}")

        try:
            admins = User.query.filter(User.roles.any(name='admin')).all()
            for admin in admins:
                send_salida_tardia_exception(
                    admin_email=admin.email,
                    exception_id=excepcion.id,
                    fecha=fecha_hoy,
                    user_id=usuario_id,
                    salida_real=hora_real_salida,
                    comentarios=comentarios  # ✅ Se incluyen los comentarios
                )
                print(f"[INFO] Correo de excepción enviado a administrador.")
        except Exception as e:
            print(f"[ERROR] Error al enviar correo de excepción: {str(e)}")
            return "Error al enviar correo de excepción", 500

        db.session.commit()
        print(f"[INFO] Cambios confirmados en la base de datos para el usuario {usuario_id}.")
        return {"success": True, "message": "Salida registrada con excepción"}, 200

    except Exception as e:
        print(f"[ERROR] Ocurrió un error en el proceso de registrar la salida: {str(e)}")
        return f"Error inesperado: {str(e)}", 500


@app.route('/accept_salida_tardia_exception/<int:exception_id>')
def accept_salida_tardia_exception(exception_id):
    excepcion = ExcepcionSalidaTardia.query.get(exception_id)
    if not excepcion:
        return "Excepción no encontrada", 404

    # Combinar la fecha de la excepción con la hora de salida real
    fecha_excepcion = excepcion.fecha
    hora_salida = excepcion.hora_salida_real
    excepcion.registro.salida = datetime.combine(fecha_excepcion, hora_salida)  # Combina fecha y hora

    excepcion.aceptada = True
    db.session.commit()

    return redirect(url_for('login'))


@app.route('/reject_salida_tardia_exception/<int:exception_id>')
def reject_salida_tardia_exception(exception_id):
    excepcion = ExcepcionSalidaTardia.query.get(exception_id)
    if not excepcion:
        return "Excepción no encontrada", 404

    excepcion.aceptada = False
    db.session.commit()

    return redirect(url_for('login'))

def send_salida_tardia_exception(admin_email, exception_id, fecha, user_id, salida_real, comentarios):
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

    accept_url = url_for('accept_salida_tardia_exception', exception_id=exception_id, _external=True)
    reject_url = url_for('reject_salida_tardia_exception', exception_id=exception_id, _external=True)

    empresa = current_user.empresa if current_user.empresa else Empresa.query.first()
    user = User.query.filter_by(id=user_id).first()

    if not empresa or not user:
        return "Error en los datos", 404

    nombre_usuario = user.nombre
    apellidos = user.primer_apellido
    temporada = "verano" if empresa.fecha_inicio_verano and  (empresa.fecha_inicio_verano <= fecha <= empresa.fecha_fin_verano) else "invierno"
    sufijo = "verano_" if temporada == "verano" else ""

    # Verificar si el usuario tiene una excepción en su registro de horario para ese día
    registro_horario = RegistroHorario.query.filter_by(user_id=user.id, fecha=fecha).first()

    if registro_horario and registro_horario.salida:
        print(f"[INFO] Se encontró una excepción de horario. Se usará la hora de salida registrada en RegistroHorario: {registro_horario.salida}")
        salida_predeterminada = registro_horario.salida
    else:
        # Si no hay excepción, usar la salida predeterminada
        dia_semana = fecha.weekday()
        dias = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo"]
        dia_nombre = dias[dia_semana]
        salida_predeterminada = getattr(user, f"salida_{sufijo}{dia_nombre}")


    def format_time(time_obj):
        return time_obj.strftime('%H:%M') if isinstance(time_obj, time) else "N/A"

    salida_predeterminada = format_time(salida_predeterminada)
    salida_real = format_time(salida_real)

    subject = "Nueva excepción de salida tardía para validación"
    html_body = f"""
    <html>
    <body>
        <p>Hola administrador,</p>
        <p>El usuario {nombre_usuario} {apellidos} ha solicitado una modificación en su salida para el día {fecha.strftime('%d de %B de %Y')} ({temporada}). Por favor, revisa los detalles y valida la excepción.</p>
        <p><strong>Horario predeterminado para hoy:</strong> {salida_predeterminada}</p>
        <p><strong>Horario real de salida:</strong> {salida_real}</p>
        <p><strong>Comentarios del usuario:</strong></p>
        <p>{comentarios}</p>
        <p>
            <a href="{accept_url}" style="display: inline-block; padding: 15px 25px; margin: 10px; color: white; background-color: green; text-decoration: none; border-radius: 5px;">Aceptar</a>
            <a href="{reject_url}" style="display: inline-block; padding: 15px 25px; margin: 10px; color: white; background-color: red; text-decoration: none; border-radius: 5px;">Rechazar</a>
        </p>
        <p>Saludos,</p>
        <p>El equipo de Gestionaempresa.</p>
    </body>
    </html>
    """

    msg = Message(subject, recipients=[admin_email])
    msg.html = html_body

    try:
        mail.send(msg)
        print("Correo enviado con éxito")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")





@app.route('/calcular_diferencia_entrada', methods=['POST'])
def calcular_diferencia_entrada():
    datos = request.json
    hora_actual = datetime.strptime(datos['hora_actual'], '%H:%M:%S')
    hora_esperada = datetime.strptime(datos['hora_esperada'], '%H:%M:%S')

    diferencia_minutos = (hora_actual - hora_esperada).total_seconds() / 60

    return jsonify({'diferencia_minutos': diferencia_minutos})


@app.route('/registro_entrada_corregida', methods=['POST'])
def registro_entrada_corregida():
    datos = request.json
    hora_entrada_corregida = datos['hora_entrada']

    # Aquí envías la solicitud al administrador para aprobación
    enviar_solicitud_aprobacion("entrada", hora_entrada_corregida)

    return jsonify({'success': True})






@app.route('/vaciar_verano', methods=['POST'])
@login_required
def vaciar_verano():
    try:
        data = request.get_json()
        empresa_id = data.get('empresa_id')

        if not empresa_id:
            return jsonify({'error': 'ID de empresa no proporcionado'}), 400

        empresa = Empresa.query.get(empresa_id)

        if not empresa:
            return jsonify({'error': 'Empresa no encontrada'}), 404

        empresa.fecha_inicio_verano = None
        empresa.fecha_fin_verano = None
        db.session.commit()

        return jsonify({'success': True, 'message': 'Horario de verano vaciado con éxito'}), 200
    except Exception as e:
        return jsonify({'error': f'Error al vaciar el horario de verano: {e}'}), 500


@app.route('/admin/metricas/por_persona')
@admin_required
@admin_required
def generar_lista_empleados():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    current_user_email = current_user.email if current_user.is_authenticated else None

    # Filtrar tanto usuarios como administradores
    users_query = User.query.filter(~User.roles.any(Role.name == 'inspector_role'))

    if current_user_email == 'emprendedores@ceeiguadalajara.es':
        users_query = users_query.filter(User.Centro_trabajo == 'CEEI Guadalajara')
    elif current_user_email == 'formacion@ceoeguadalajara.es':
        users_query = users_query.filter(User.Centro_trabajo == 'CEOE Formación')

    users = users_query.filter(
        (User.email != 'admin@gestionaempresa.es') &
        (User.email != 'inspector@gestionaempresa.es')
    ).all()

    jornada_irregular_por_usuario = {}

    # Obtener las fechas de inicio y fin de verano de la empresa
    empresa = Empresa.query.first()  # Suponiendo que solo hay una empresa en la base de datos
    fecha_inicio_verano = empresa.fecha_inicio_verano
    fecha_fin_verano = empresa.fecha_fin_verano

    # Obtener los días festivos
    festivos = Festivo.query.all()
    dias_festivos = [(festivo.dia, festivo.mes) for festivo in festivos]

    for user in users:
        jornada_irregular = timedelta()
        total_horas_esperadas = timedelta()
        total_horas_trabajadas = timedelta()

        primer_registro = Registro.query.filter(
            Registro.user_id == user.id,  # Filtrar por el usuario
            ).order_by(Registro.entrada.asc()).first()

        if primer_registro:
            fecha_inicio = primer_registro.entrada.date()
        else:
            year = datetime.now().year
            fecha_inicio = datetime(year, 1, 1).date()

        # Calcular el total de horas que debería haber trabajado el usuario

        fecha_fin = date.today() - timedelta(days=1) # Fecha actual como fecha de fin

        for dia in range((fecha_fin - fecha_inicio).days + 1):
            fecha_actual = fecha_inicio + timedelta(days=dia)
            dia_semana = fecha_actual.weekday()

            # Verificar si el día es festivo
            if (fecha_actual.day, fecha_actual.month) in dias_festivos:
                continue  # Saltar los días festivos
            if empresa.fecha_inicio_verano:

                es_verano = fecha_inicio_verano <= fecha_actual <= fecha_fin_verano
            else:
                es_verano = False

            entrada = salida = salida_comida = entrada_comida = None
            if dia_semana == 0 and user.trabaja_lunes:
                if es_verano and user.trabaja_verano_lunes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_lunes, user.salida_verano_lunes, user.salida_comida_verano_lunes, user.entrada_comida_verano_lunes
                elif not es_verano and user.trabaja_lunes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_lunes, user.salida_lunes, user.salida_comida_lunes, user.entrada_comida_lunes
            elif dia_semana == 1 and user.trabaja_martes:
                if es_verano and user.trabaja_verano_martes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_martes, user.salida_verano_martes, user.salida_comida_verano_martes, user.entrada_comida_verano_martes
                elif not es_verano and user.trabaja_martes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_martes, user.salida_martes, user.salida_comida_martes, user.entrada_comida_martes
            elif dia_semana == 2 and user.trabaja_miercoles:
                if es_verano and user.trabaja_verano_miercoles:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_miercoles, user.salida_verano_miercoles, user.salida_comida_verano_miercoles, user.entrada_comida_verano_miercoles
                elif not es_verano and user.trabaja_miercoles:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_miercoles, user.salida_miercoles, user.salida_comida_miercoles, user.entrada_comida_miercoles
            elif dia_semana == 3 and user.trabaja_jueves:
                if es_verano and user.trabaja_verano_jueves:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_jueves, user.salida_verano_jueves, user.salida_comida_verano_jueves, user.entrada_comida_verano_jueves
                elif not es_verano and user.trabaja_jueves:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_jueves, user.salida_jueves, user.salida_comida_jueves, user.entrada_comida_jueves
            elif dia_semana == 4 and user.trabaja_viernes:
                if es_verano and user.trabaja_verano_viernes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_viernes, user.salida_verano_viernes, user.salida_comida_verano_viernes, user.entrada_comida_verano_viernes
                elif not es_verano and user.trabaja_viernes:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_viernes, user.salida_viernes, user.salida_comida_viernes, user.entrada_comida_viernes
            elif dia_semana == 5 and user.trabaja_sabado:
                if es_verano and user.trabaja_verano_sabado:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_sabado, user.salida_verano_sabado, user.salida_comida_verano_sabado, user.entrada_comida_verano_sabado
                elif not es_verano and user.trabaja_sabado:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_sabado, user.salida_sabado, user.salida_comida_sabado, user.entrada_comida_sabado
            elif dia_semana == 6 and user.trabaja_domingo:
                if es_verano and user.trabaja_verano_domingo:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_verano_domingo, user.salida_verano_domingo, user.salida_comida_verano_domingo, user.entrada_comida_verano_domingo
                elif not es_verano and user.trabaja_domingo:
                    entrada, salida, salida_comida, entrada_comida = user.entrada_domingo, user.salida_domingo, user.salida_comida_domingo, user.entrada_comida_domingo
            else:
                continue

            def time_to_timedelta(time_obj):
                return timedelta(hours=time_obj.hour, minutes=time_obj.minute, seconds=time_obj.second)

            if entrada and salida:
                entrada_timedelta = time_to_timedelta(entrada)
                salida_timedelta = time_to_timedelta(salida)
                salida_comida_timedelta = time_to_timedelta(salida_comida) if salida_comida else timedelta()
                entrada_comida_timedelta = time_to_timedelta(entrada_comida) if entrada_comida else timedelta()

                horas_diarias_esperadas = (salida_timedelta - entrada_timedelta) - (entrada_comida_timedelta - salida_comida_timedelta)
                total_horas_esperadas += horas_diarias_esperadas

        # Obtener los registros del usuario y sumar las horas trabajadas
        registros = Registro.query.filter_by(user_id=user.id).all()
        for registro in registros:
            if registro.duracion_neta:
                total_horas_trabajadas += registro.duracion_neta

        jornada_irregular = total_horas_trabajadas - total_horas_esperadas
        jornada_irregular_en_horas = jornada_irregular.total_seconds() / 3600

        if total_horas_esperadas.total_seconds() > 0:
            porcentaje_irregular = (jornada_irregular.total_seconds() / total_horas_esperadas.total_seconds()) * 100
        else:
            porcentaje_irregular = 0

        jornada_irregular_por_usuario[user.id] = {
            'total_horas_esperadas': total_horas_esperadas.total_seconds() / 3600 if total_horas_esperadas else 0,
            'total_horas_trabajadas': total_horas_trabajadas.total_seconds() / 3600 if total_horas_trabajadas else 0,
            'jornada_irregular_horas': jornada_irregular_en_horas,
            'porcentaje_irregular': porcentaje_irregular
        }

    print(jornada_irregular_por_usuario)

    return render_template('metricas_listar_por_persona.html', users=users, current_user_email=current_user_email, notificaciones=notificaciones, active_menu='por_persona', jornada_irregular_por_usuario=jornada_irregular_por_usuario)

def convertir_hora_a_decimal(hora_str):
    if ':' in hora_str:
        horas, minutos = map(int, hora_str.split(':'))
        return horas + minutos / 60
    elif '.' in hora_str:
        return float(hora_str)
    else:
        return float(hora_str)


def get_horas_absentismo(dia_inicio, dia_fin, user_id):
    total_horas_absentismo = 0

    absentismos = db.session.query(Absence).filter(
        Absence.user_id == user_id,
        func.date(Absence.start_date) >= dia_inicio,
        func.date(Absence.end_date) <= dia_fin
    ).all()

    for absentismo in absentismos:
        if absentismo.duration_hours:
            total_horas_absentismo += convertir_hora_a_decimal(absentismo.duration_hours)

    return total_horas_absentismo

def get_vacaciones(user_id):
    dias_totales = 0
    user = User.query.get(user_id)
    centro_trabajo = CentroTrabajo.query.filter(CentroTrabajo.nombre.like(f'%{user.Centro_trabajo}%')).first()

    if centro_trabajo:
        empresa = Empresa.query.get(centro_trabajo.empresa.id)
        dias_totales = empresa.dias_vacaciones

    return user.dias_restantes_este, dias_totales



def get_horas_trabajadas(dia_inicio, dia_fin, user_id):
    total_horas_trabajadas = timedelta()

    registros = db.session.query(Registro).filter(
        Registro.user_id == user_id,
        func.date(Registro.entrada) >= dia_inicio,
        func.date(Registro.entrada) <= dia_fin
    ).all()

    for registro in registros:
        if registro.duracion_neta:
            total_horas_trabajadas += registro.duracion_neta

    return total_horas_trabajadas.total_seconds() / 3600

def get_minutos_pausas(registros):
    total_tiempo = 0
    registro_ids = [r.id for r in registros] #Sacamos los id de registro para hacer la busqueda por ellos
    pausas = db.session.query(Pausa).filter(
        Pausa.registro_id.in_(registro_ids)
    )

    for pausa in pausas:
        total_tiempo += pausa.duracion_minutos

    return total_tiempo


def calcular_horas_esperadas(dia_inicio, dia_fin, user_id):
    total_dias = 0
    total_horas_esperadas = timedelta()
    user = User.query.get(user_id)
    if not user: return 0.0

    # Obtener las fechas de inicio y fin de verano de la empresa
    empresa = Empresa.query.first()  # Suponiendo que solo hay una empresa en la base de datos
    fecha_inicio_verano = empresa.fecha_inicio_verano
    fecha_fin_verano = empresa.fecha_fin_verano

    # Obtener los días festivos
    festivos = Festivo.query.all()
    dias_festivos = [(festivo.dia, festivo.mes) for festivo in festivos]

    vacaciones = VacationHistory.query.filter(
        VacationHistory.user_id == user.id,
        VacationHistory.is_accepted == True,
        VacationHistory.end_date >= dia_inicio,
        VacationHistory.start_date <= dia_fin
    ).all()

    dias_vacaciones = set()
    for vac in vacaciones:
        current = vac.start_date
        while current <= vac.end_date:
            dias_vacaciones.add(current)
            current += timedelta(days=1)

    for dia in range((dia_fin - dia_inicio).days + 1):
        fecha_actual = dia_inicio + timedelta(days=dia)
        dia_semana = fecha_actual.weekday()

        # Verificar si el día es festivo o ha estado de vacaciones
        if ((fecha_actual.day, fecha_actual.month) in dias_festivos) or (fecha_actual in dias_vacaciones):
            continue  # Saltar los días festivos

        if empresa.fecha_inicio_verano:
            es_verano = fecha_inicio_verano <= fecha_actual <= fecha_fin_verano
        else:
            es_verano = False
        entrada = salida = salida_comida = entrada_comida = None
        if dia_semana == 0 and user.trabaja_lunes:
            if es_verano and user.trabaja_verano_lunes:
                entrada, salida, salida_comida, entrada_comida = user.entrada_verano_lunes, user.salida_verano_lunes, user.salida_comida_verano_lunes, user.entrada_comida_verano_lunes
            elif not es_verano and user.trabaja_lunes:
                entrada, salida, salida_comida, entrada_comida = user.entrada_lunes, user.salida_lunes, user.salida_comida_lunes, user.entrada_comida_lunes
        elif dia_semana == 1 and user.trabaja_martes:
            if es_verano and user.trabaja_verano_martes:
                entrada, salida, salida_comida, entrada_comida = user.entrada_verano_martes, user.salida_verano_martes, user.salida_comida_verano_martes, user.entrada_comida_verano_martes
            elif not es_verano and user.trabaja_martes:
                entrada, salida, salida_comida, entrada_comida = user.entrada_martes, user.salida_martes, user.salida_comida_martes, user.entrada_comida_martes
        elif dia_semana == 2 and user.trabaja_miercoles:
            if es_verano and user.trabaja_verano_miercoles:
                entrada, salida, salida_comida, entrada_comida = user.entrada_verano_miercoles, user.salida_verano_miercoles, user.salida_comida_verano_miercoles, user.entrada_comida_verano_miercoles
            elif not es_verano and user.trabaja_miercoles:
                entrada, salida, salida_comida, entrada_comida = user.entrada_miercoles, user.salida_miercoles, user.salida_comida_miercoles, user.entrada_comida_miercoles
        elif dia_semana == 3 and user.trabaja_jueves:
            if es_verano and user.trabaja_verano_jueves:
                entrada, salida, salida_comida, entrada_comida = user.entrada_verano_jueves, user.salida_verano_jueves, user.salida_comida_verano_jueves, user.entrada_comida_verano_jueves
            elif not es_verano and user.trabaja_jueves:
                entrada, salida, salida_comida, entrada_comida = user.entrada_jueves, user.salida_jueves, user.salida_comida_jueves, user.entrada_comida_jueves
        elif dia_semana == 4 and user.trabaja_viernes:
            if es_verano and user.trabaja_verano_viernes:
                entrada, salida, salida_comida, entrada_comida = user.entrada_verano_viernes, user.salida_verano_viernes, user.salida_comida_verano_viernes, user.entrada_comida_verano_viernes
            elif not es_verano and user.trabaja_viernes:
                entrada, salida, salida_comida, entrada_comida = user.entrada_viernes, user.salida_viernes, user.salida_comida_viernes, user.entrada_comida_viernes
        elif dia_semana == 5 and user.trabaja_sabado:
            if es_verano and user.trabaja_verano_sabado:
                entrada, salida, salida_comida, entrada_comida = user.entrada_verano_sabado, user.salida_verano_sabado, user.salida_comida_verano_sabado, user.entrada_comida_verano_sabado
            elif not es_verano and user.trabaja_sabado:
                entrada, salida, salida_comida, entrada_comida = user.entrada_sabado, user.salida_sabado, user.salida_comida_sabado, user.entrada_comida_sabado
        elif dia_semana == 6 and user.trabaja_domingo:
            if es_verano and user.trabaja_verano_domingo:
                entrada, salida, salida_comida, entrada_comida = user.entrada_verano_domingo, user.salida_verano_domingo, user.salida_comida_verano_domingo, user.entrada_comida_verano_domingo
            elif not es_verano and user.trabaja_domingo:
                entrada, salida, salida_comida, entrada_comida = user.entrada_domingo, user.salida_domingo, user.salida_comida_domingo, user.entrada_comida_domingo
        else:
            continue
        def time_to_timedelta(time_obj):
            return timedelta(hours=time_obj.hour, minutes=time_obj.minute, seconds=time_obj.second)
        if entrada and salida:
            entrada_timedelta = time_to_timedelta(entrada)
            salida_timedelta = time_to_timedelta(salida)
            salida_comida_timedelta = time_to_timedelta(salida_comida) if salida_comida else timedelta()
            entrada_comida_timedelta = time_to_timedelta(entrada_comida) if entrada_comida else timedelta()
            horas_diarias_esperadas = (salida_timedelta - entrada_timedelta) - (entrada_comida_timedelta - salida_comida_timedelta)
            total_horas_esperadas += horas_diarias_esperadas
            total_dias += 1

    return total_horas_esperadas.total_seconds() / 3600 , total_dias


@app.route('/admin/metricas/por_persona/<int:user_id>')
@admin_required
@admin_required
def ver_metricas_user(user_id):
    return render_template("metrica_persona.html", user_id = user_id)

from collections import defaultdict
from sqlalchemy import func

@app.route('/api/datos')
@admin_required
@admin_required
def obtener_datos_user():

    rango = request.args.get("rango")
    fecha_param = request.args.get("fecha")
    user_id = request.args.get("id")

    # Función para convertir "HH:MM" a número de horas (float)
    def hora_a_float(hora_str):
        hora = datetime.strptime(hora_str, "%H:%M")
        return hora.hour + hora.minute / 60

    # Pasa los registros al formato de array que se va a usar para poder trabajar con ellos
    def formatear_registros(registros: list) -> list:
        registros_formatedos = []
        for registro in registros:
            data = {
                "fecha": registro.entrada.date().strftime("%Y-%m-%d"),  # Convertimos a string con el formato "YYYY-MM-DD"
                "valor": [registro.entrada.time().strftime("%H:%M")]   # Convertimos a string con el formato "HH:MM:SS"
            }
            if registro.salida:
                data["valor"].append(registro.salida.time().strftime("%H:%M"))
            registros_formatedos.append(data)

        return registros_formatedos

    def formatear_registros_mensuales(registros: list, entradas_name: str, salidas_name: str) -> dict:
        entradas_formateadas = []
        salidas_formateadas = []
        for registro in registros:
            data_entrada = {
                "fecha": registro.entrada.date().strftime("%Y-%m-%d"),  # Convertimos a string con el formato "YYYY-MM-DD"
                "valor": registro.entrada.time().strftime("%H:%M")   # Convertimos a string con el formato "HH:MM:SS"
            }
            data_salida = {
                "fecha": registro.salida.date().strftime("%Y-%m-%d"),  # Convertimos a string con el formato "YYYY-MM-DD"
                "valor": registro.salida.time().strftime("%H:%M")   # Convertimos a string con el formato "HH:MM:SS"
            }
            entradas_formateadas.append(data_entrada)
            salidas_formateadas.append(data_salida)

        return {entradas_name : entradas_formateadas, salidas_name : salidas_formateadas}

    def formatear_pausas(pausas: list) -> dict:
        pausas_formateadas = []
        for pausa in pausas:
            data = {
                "fecha": pausa.inicio.date().strftime("%Y-%m-%d"),  # Convertimos a string con el formato "YYYY-MM-DD"
                "valor": [pausa.inicio.time().strftime("%H:%M")]   # Convertimos a string con el formato "HH:MM:SS"
            }
            if registro.fin:
                data["valor"].append(pausa.fin.time().strftime("%H:%M"))
            pausas_formateadas.append(data)

        return pausas_formateadas


    resultado = []
    metricas = {}
    resultado_dia2 = []
    datos_finales =[]
    datos_pausas = []

    match rango:
        case "dia":
            fecha_busqueda = datetime.strptime(fecha_param, "%Y-%m-%d").date()
            registros = db.session.query(Registro).filter(
                Registro.user_id == user_id,
                func.date(Registro.entrada) == fecha_busqueda
                ).all()

            registro_ids = [r.id for r in registros] #Sacamos los id de registro para hacer la busqueda por ellos
            pausas = db.session.query(Pausa).filter(
                Pausa.registro_id.in_(registro_ids)
            )

            if registros: datos_finales = formatear_registros(registros)
            if pausas: datos_pausas = formatear_pausas(pausas)

            horas_esperadas, total_dias = calcular_horas_esperadas(fecha_busqueda, fecha_busqueda, user_id)
            horas_trabajadas = get_horas_trabajadas(fecha_busqueda, fecha_busqueda, user_id)
            horas_absentismo = get_horas_absentismo(fecha_busqueda, fecha_busqueda, user_id)
            minutos_pausas = get_minutos_pausas(registros)
            dias_vacaciones, total_dias_vacaciones = get_vacaciones(user_id)

            metricas["horas_esperadas"] = horas_esperadas
            metricas["horas_trabajadas"] = horas_trabajadas
            metricas["horas_absentismo"] = horas_absentismo
            metricas["total_dias"] = total_dias
            metricas["minutos_pausas"] = minutos_pausas
            metricas["dias_vacaciones"] = dias_vacaciones
            metricas["total_dias_vacaciones"] = total_dias_vacaciones

            resultado = jsonify({"grafico" : {"Registro": datos_finales, "Pausas": datos_pausas},
                                 "metricas" : metricas})

        case "semana":
            fecha_busqueda = datetime.strptime(fecha_param + "-1", "%Y-W%W-%w")
            año, semana = map(int, fecha_param.split("-W"))

            # Obtener lunes de esa semana
            inicio_semana_dt = datetime.fromisocalendar(año, semana, 1)
            fin_semana_dt = inicio_semana_dt + timedelta(days=6)

            # Convertimos a date (sin hora)
            inicio_semana = inicio_semana_dt.date()
            fin_semana = fin_semana_dt.date()

            print("Fecha busqueda: ", fecha_busqueda)
            print("Fecha inicio: ", inicio_semana)
            print("Fecha fin: ", fin_semana)

            registros = db.session.query(Registro).filter(
                Registro.user_id == user_id,
                func.date(Registro.entrada) >= inicio_semana,
                func.date(Registro.entrada) <= fin_semana
                ).all()

            if registros: datos_finales = formatear_registros(registros)

            horas_esperadas, total_dias = calcular_horas_esperadas(inicio_semana, fin_semana, user_id)
            horas_trabajadas = get_horas_trabajadas(inicio_semana, fin_semana, user_id)
            horas_absentismo = get_horas_absentismo(inicio_semana, fin_semana, user_id)
            minutos_pausas = get_minutos_pausas(registros)
            dias_vacaciones, total_dias_vacaciones = get_vacaciones(user_id)

            metricas["horas_esperadas"] = horas_esperadas
            metricas["horas_trabajadas"] = horas_trabajadas
            metricas["horas_absentismo"] = horas_absentismo
            metricas["total_dias"] = total_dias
            metricas["minutos_pausas"] = minutos_pausas
            metricas["dias_vacaciones"] = dias_vacaciones
            metricas["total_dias_vacaciones"] = total_dias_vacaciones

            resultado = jsonify({"grafico" : {"Registro": datos_finales},
                                 "metricas" : metricas})

        case "mes":
            fecha_busqueda = datetime.strptime(fecha_param, "%Y-%m")
            primer_dia_mes = fecha_busqueda.replace(day=1).date()
            ultimo_dia_mes = fecha_busqueda.replace(
                day=calendar.monthrange(fecha_busqueda.year, fecha_busqueda.month)[1]
            ).date()

            all_registros = db.session.query(Registro).filter(
                Registro.user_id == user_id,
                func.date(Registro.entrada) >= primer_dia_mes,
                func.date(Registro.entrada) <= ultimo_dia_mes
            ).all()

            subquery_min = db.session.query(
                func.min(Registro.entrada).label("primera_entrada"),
                func.date(Registro.entrada).label("solo_fecha")
            ).filter(
                Registro.user_id == user_id,
                func.extract('year', Registro.entrada) == fecha_busqueda.year,
                func.extract('month', Registro.entrada) == fecha_busqueda.month
            ).group_by(
                func.date(Registro.entrada)
            ).subquery()

            # Sacamos todos los registros que coincidan con la subquery
            registros_tramo_1 = db.session.query(Registro).join(
                subquery_min,
                db.and_(
                    Registro.entrada == subquery_min.c.primera_entrada,
                    Registro.user_id == user_id
                )
            ).distinct(Registro.entrada).all()

            subquery_max = db.session.query(
                func.max(Registro.entrada).label("primera_entrada"),
                func.date(Registro.entrada).label("solo_fecha")
            ).filter(
                Registro.user_id == user_id,
                func.extract('year', Registro.entrada) == fecha_busqueda.year,
                func.extract('month', Registro.entrada) == fecha_busqueda.month
            ).group_by(
                func.date(Registro.entrada)
            ).subquery()

            # Sacamos todos los registros que coincidan con la subquery
            registros_tramo_2 = db.session.query(Registro).join(
                subquery_max,
                db.and_(
                    Registro.entrada == subquery_max.c.primera_entrada,
                    Registro.user_id == user_id
                )
            ).distinct(Registro.entrada).all()

            datos_finales_tramo_1 = datos_finales_tramo_2 = {}

            if registros_tramo_1: datos_finales_tramo_1 = formatear_registros_mensuales(registros_tramo_1, "Entradas tramo 1", "Salidas tramo 1")
            if registros_tramo_2: datos_finales_tramo_2 = formatear_registros_mensuales(registros_tramo_2, "Entradas tramo 2", "Salidas tramo 2")

            # LAS SIGUIENTES LINEAS CREAN UNA TUPLA PARA DESPUES VER SI LOS VALORES SE ENCUENTRAN EN ELLA Y ELIMINARLOS
            # ES PARA QUITAR LOS PUNTOS REPETIDOS DEL TRAMO 1 Y TRAMO 2

            # Crear un conjunto de tuplas (fecha, valor) para las entradas y salidas de tramo 1
            if "Entradas tramo 1" in datos_finales_tramo_1:
                entradas_tramo1_set = {(entrada["fecha"], entrada["valor"]) for entrada in datos_finales_tramo_1["Entradas tramo 1"]}
            else:
                entradas_tramo1_set = set()

            if "Salidas tramo 1" in datos_finales_tramo_1:
                salidas_tramo1_set = {(salida["fecha"], salida["valor"]) for salida in datos_finales_tramo_1["Salidas tramo 1"]}
            else:
                salidas_tramo1_set = set()

            if "Entradas tramo 2" in datos_finales_tramo_2 and "Salidas tramo 2" in datos_finales_tramo_2:
                # Filtrar las entradas y salidas del tramo 2 que NO estén en el conjunto de tramo 1
                new_entradas_tramo2 = [entrada for entrada in datos_finales_tramo_2["Entradas tramo 2"]
                                       if (entrada["fecha"], entrada["valor"]) not in entradas_tramo1_set]

                new_salidas_tramo2 = [salida for salida in datos_finales_tramo_2["Salidas tramo 2"]
                                      if (salida["fecha"], salida["valor"]) not in salidas_tramo1_set]

                # Si quieres actualizar las listas originales:
                datos_finales_tramo_2["Entradas tramo 2"] = new_entradas_tramo2
                datos_finales_tramo_2["Salidas tramo 2"] = new_salidas_tramo2
            else:
                datos_finales_tramo_2["Entradas tramo 2"] = []
                datos_finales_tramo_2["Salidas tramo 2"] = []



            horas_esperadas, total_dias = calcular_horas_esperadas(primer_dia_mes, ultimo_dia_mes, user_id)
            horas_trabajadas = get_horas_trabajadas(primer_dia_mes, ultimo_dia_mes, user_id)
            horas_absentismo = get_horas_absentismo(primer_dia_mes, ultimo_dia_mes, user_id)
            minutos_pausas = get_minutos_pausas(all_registros)
            dias_vacaciones, total_dias_vacaciones = get_vacaciones(user_id)


            metricas["horas_esperadas"] = horas_esperadas
            metricas["horas_trabajadas"] = horas_trabajadas
            metricas["horas_absentismo"] = horas_absentismo
            metricas["total_dias"] = total_dias
            metricas["minutos_pausas"] = minutos_pausas
            metricas["dias_vacaciones"] = dias_vacaciones
            metricas["total_dias_vacaciones"] = total_dias_vacaciones

            resultado = jsonify({"grafico" : datos_finales_tramo_1 | datos_finales_tramo_2,
                                 "metricas" : metricas})

        case _:
            resultado = jsonify(resultado)
            raise ValueError("Not a point")

    return resultado





import os
import zipfile

from flask import request, redirect, flash, url_for, render_template
from werkzeug.utils import secure_filename
from pdf2image import convert_from_path
import pytesseract
import re

UPLOAD_FOLDER_nominas = 'static/uploads/nominas'
ALLOWED_EXTENSIONS_nominas = {'pdf', 'zip'}

# OCR para extraer DNI
def clean_text(text):
    """Limpia caracteres invisibles y normaliza el texto."""
    # Quitar caracteres no imprimibles
    text = ''.join(c for c in text if c.isprintable())
    # Reemplazar múltiples espacios o saltos de línea por uno solo
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

import fitz  # PyMuPDF

def extract_dni_from_pdf(pdf_path):
    try:
        print(f"Intentando extraer DNI del PDF: {pdf_path}")
        with fitz.open(pdf_path) as doc:
            for i, page in enumerate(doc):
                text = page.get_text()
                clean = clean_text(text)
                print(f"Texto limpio extraído de página {i + 1}:\n{clean}\n{'-'*40}")

                # Buscar todos los posibles DNIs/NIEs válidos
                posibles_dnis = re.findall(r'\b(\d{7,8}[A-Z]|[XYZ]\d{7}[A-Z])\b', clean.upper())
                print(f"Posibles DNIs encontrados: {posibles_dnis}")

                # Filtrar DNIs personales (que no empiezan con letra de empresa)
                dnis_personales = [
                    dni for dni in posibles_dnis
                    if not re.match(r'^[ABCDEFGHJNPQRSUVW]', dni)
                ]
                print(f"DNIs personales filtrados: {dnis_personales}")

                if dnis_personales:
                    print(f"DNI válido encontrado: {dnis_personales[0]}")
                    return dnis_personales[0]

    except Exception as e:
        print(f"Error al extraer texto para {pdf_path}: {e}")

    return None

from PyPDF2 import PdfReader, PdfWriter


def allowed_file(filename):
    result = '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_nominas
    print(f"¿Archivo permitido ({filename})? {result}")
    return result


@app.route('/subir_nominas', methods=['GET'])
@admin_required
def mostrar_subir_nominas():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()
    print("Mostrando formulario de subida de nóminas")
    return render_template('subir_nomina.html', notificaciones=notificaciones)


@app.route('/subir_nominas', methods=['POST'])
@admin_required
def subir_nominas():
    print("Procesando subida de nóminas...")
    archivos = request.files.getlist('archivo')

    if not archivos:
        flash("No se ha seleccionado ningún archivo.", "danger")
        return redirect(request.url)

    fecha_nomina = date.today().replace(day=1)
    os.makedirs(UPLOAD_FOLDER_nominas, exist_ok=True)
    print(f"Carpeta de destino: {UPLOAD_FOLDER_nominas}")

    mensaje_exito = []  # Variable para acumular los mensajes de éxito

    for archivo in archivos:
        print(f"Procesando archivo: {archivo.filename}")
        if archivo and allowed_file(archivo.filename):
            filename = secure_filename(archivo.filename)
            ext = filename.rsplit('.', 1)[1].lower()
            print(f"Extensión detectada: {ext}")

            if ext == 'pdf':
                print("Archivo es PDF, procesando individualmente.")
                mensaje_exito += procesar_pdf(archivo, fecha_nomina)
            elif ext == 'zip':
                print("Archivo es ZIP, extrayendo contenido...")
                with zipfile.ZipFile(archivo) as zip_ref:
                    for nombre_pdf in zip_ref.namelist():
                        print(f"Archivo dentro del ZIP: {nombre_pdf}")
                        if nombre_pdf.endswith('.pdf'):
                            ruta_temp = os.path.join(UPLOAD_FOLDER_nominas, "temp_extract.pdf")
                            with open(ruta_temp, 'wb') as f:
                                f.write(zip_ref.read(nombre_pdf))

                            print(f"Archivo extraído temporalmente en: {ruta_temp}")
                            mensaje_exito += procesar_pdf_desde_ruta(ruta_temp, fecha_nomina)
                            os.remove(ruta_temp)
                        else:
                            print(f"Ignorado (no es PDF): {nombre_pdf}")
        else:
            print(f"Archivo no permitido o vacío: {archivo.filename}")

    db.session.commit()
    print("Todos los cambios guardados en la base de datos.")

    if mensaje_exito:
        flash("Nóminas procesadas correctamente. " + " ".join(mensaje_exito), "success")
    else:
        flash("No se procesaron nóminas correctamente.", "danger")

    return redirect(url_for('subir_nominas'))


def procesar_pdf(archivo_pdf, fecha_nomina):
    print(f"Procesando PDF individual: {archivo_pdf.filename}")
    ruta_temp = os.path.join(UPLOAD_FOLDER_nominas, "temp_upload.pdf")
    archivo_pdf.save(ruta_temp)
    return procesar_pdf_desde_ruta(ruta_temp, fecha_nomina)


def procesar_pdf_desde_ruta(ruta_pdf, fecha_nomina_default):
    try:
        sub_pdfs = dividir_pdf_por_nombres(ruta_pdf)
        print(f"Detectadas {len(sub_pdfs)} posibles nóminas dentro del archivo.")
    except Exception as e:
        print(f"❌ Error al dividir el PDF por nombres: {e}")
        return []

    mensajes_exito = []  # Lista para acumular los mensajes de éxito por archivo procesado

    for idx, ruta_subpdf in enumerate(sub_pdfs):
        print(f"📂 Procesando subarchivo {idx+1}/{len(sub_pdfs)}: {ruta_subpdf}")

        try:
            texto = extract_text_from_pdf(ruta_subpdf)
            print(f"📝 Texto extraído: {texto[:100]}...")  # Muestra los primeros 100 caracteres
        except Exception as e:
            print(f"❌ Error al extraer texto del sub-PDF: {e}")
            continue

        try:
            fecha_nomina = detectar_fecha_periodo(texto) or fecha_nomina_default
            print(f"📆 Fecha de nómina detectada: {fecha_nomina.strftime('%Y-%m-%d')}")
        except Exception as e:
            print(f"❌ Error al detectar fecha de nómina: {e}")
            fecha_nomina = fecha_nomina_default

        try:
            dni = extract_dni_from_pdf(ruta_subpdf)
            print(f"🆔 DNI extraído: {dni}")
        except Exception as e:
            print(f"❌ Error al extraer DNI: {e}")
            dni = None

        if not dni:
            flash(f"No se pudo extraer el DNI del fragmento {idx+1}", "warning")
            try:
                os.remove(ruta_subpdf)
                print(f"🗑️ Subarchivo eliminado: {ruta_subpdf}")
            except Exception as e:
                print(f"❌ Error al eliminar subarchivo: {e}")
            continue

        try:
            user = User.query.filter_by(DNI=dni).first()
            if user:
                print(f"👤 Usuario encontrado: {user.nombre} (ID: {user.id})")
            else:
                print(f"⚠️ Usuario no encontrado para DNI: {dni}")
        except Exception as e:
            print(f"❌ Error al buscar usuario con DNI {dni}: {e}")
            user = None

        if not user:
            flash(f"Usuario no encontrado para DNI: {dni}", "warning")
            try:
                os.remove(ruta_subpdf)
                print(f"🗑️ Subarchivo eliminado: {ruta_subpdf}")
            except Exception as e:
                print(f"❌ Error al eliminar subarchivo: {e}")
            continue

        try:
            nombre_base = f"nomina_{dni}_{fecha_nomina.strftime('%m_%Y')}.pdf"
            ruta_archivo = os.path.join(UPLOAD_FOLDER_nominas, nombre_base)

            contador = 2
            while os.path.exists(ruta_archivo):
                nuevo_nombre = f"nomina_{dni}_{fecha_nomina.strftime('%m_%Y')}-{contador}.pdf"
                ruta_archivo = os.path.join(UPLOAD_FOLDER_nominas, nuevo_nombre)
                contador += 1

            if contador > 2:
                mensajes_exito.append(f"Nómina para {dni} renombrada a {ruta_archivo.split('/')[-1]}.")

            os.rename(ruta_subpdf, ruta_archivo)
            print(f"📁 Archivo movido a: {ruta_archivo}")

            nomina = Nomina(user_id=user.id, fecha=fecha_nomina, ruta=ruta_archivo)
            db.session.add(nomina)
            print(f"✅ Nómina registrada en la base de datos para {dni}")
        except Exception as e:
            print(f"❌ Error al guardar nómina para {dni}: {e}")

    try:
        db.session.commit()
        print("✅ Cambios guardados en la base de datos.")
    except Exception as e:
        print(f"❌ Error al hacer commit en la base de datos: {e}")
        db.session.rollback()

    return mensajes_exito

def dividir_pdf_por_nombres(ruta_pdf):
    lector = PdfReader(ruta_pdf)
    escritor = PdfWriter()
    rutas_generadas = []
    nombre_anterior = None
    paginas_actuales = []

    # Nuevo patrón más flexible para detectar nombres completos en mayúsculas (2 o 3 palabras)
    nombre_pattern = re.compile(
        r"\b([A-ZÁÉÍÓÚÑ]{2,}(?:\s+[A-ZÁÉÍÓÚÑ]{2,}){1,2})\b"
    )

    for i, pagina in enumerate(lector.pages):
        texto = pagina.extract_text() or ""
        print(f"📄 Texto extraído de la página {i + 1}: {texto[:100]}...")  # Mostrar solo primeros 100 caracteres

        match = nombre_pattern.search(texto)
        if match:
            nombre_actual = match.group(1).strip().title()
            print(f"✅ Nombre detectado en página {i + 1}: {nombre_actual}")
        else:
            nombre_actual = None
            print(f"❌ No se detectó nombre en página {i + 1}")

        if nombre_actual != nombre_anterior and paginas_actuales:
            print(f"📦 Nuevo nombre detectado: {nombre_actual}, guardando grupo anterior de páginas.")
            ruta_nueva = os.path.join(UPLOAD_FOLDER_nominas, f"nomina_temp_{len(rutas_generadas)}.pdf")
            escritor = PdfWriter()
            for p in paginas_actuales:
                escritor.add_page(p)
            with open(ruta_nueva, "wb") as f:
                escritor.write(f)
            rutas_generadas.append(ruta_nueva)
            paginas_actuales = []

        paginas_actuales.append(pagina)
        nombre_anterior = nombre_actual

    # Guardar el último grupo de páginas
    if paginas_actuales:
        print(f"💾 Guardando el último grupo de páginas para: {nombre_anterior}")
        ruta_nueva = os.path.join(UPLOAD_FOLDER_nominas, f"nomina_temp_{len(rutas_generadas)}.pdf")
        escritor = PdfWriter()
        for p in paginas_actuales:
            escritor.add_page(p)
        with open(ruta_nueva, "wb") as f:
            escritor.write(f)
        rutas_generadas.append(ruta_nueva)

    return rutas_generadas




import re

from dateutil import parser


from dateutil import parser
import re
from unidecode import unidecode

import re
from unidecode import unidecode
from dateutil import parser


def detectar_fecha_periodo(texto):
    texto = unidecode(texto.lower())

    # Diccionario de abreviaturas de meses
    meses_abrev = {
        'ene': 'enero', 'feb': 'febrero', 'mar': 'marzo', 'abr': 'abril',
        'may': 'mayo', 'jun': 'junio', 'jul': 'julio', 'ago': 'agosto',
        'sep': 'septiembre', 'oct': 'octubre', 'nov': 'noviembre', 'dic': 'diciembre'
    }

    # Reemplazar abreviaturas por nombres completos
    for abrev, completo in meses_abrev.items():
        texto = re.sub(rf'\b{abrev}\b', completo, texto)

    print(f"📝 Texto después de reemplazar abreviaturas: '{texto}'")

    patrones_fecha = [
        r"(\d{1,2}\s+de\s+[a-z]+\s+de\s+\d{4})",                      # 1 de abril de 2024
        r"([a-z]+\s+\d{4})",                                          # abril 2024
        r"(\d{1,2}/\d{1,2}/\d{4})",                                   # 01/04/2024
        r"(\d{4}-\d{2}-\d{2})",                                       # 2024-04-01
        r"(\d{1,2}\s+[a-z]+\s+\d{4})",                                # 1 abril 2024
        r"(\d{1,2}\s+[a-z]+\s+\d{2})",                                # 1 abril 24
        r"(\d{1,2}\s+[a-z]+\s+\d{2})\s+a\s+(\d{1,2}\s+[a-z]+\s+\d{2})",  # rango
    ]

    for patron in patrones_fecha:
        match = re.search(patron, texto)
        if match:
            fecha_str = match.group(1)
            print(f"🕵️ Fecha candidata: '{fecha_str}'")

            fecha_str = fecha_str.strip()

            # Normalizar años de 2 cifras
            if re.match(r"\d{1,2}\s+[a-z]+\s+\d{2}$", fecha_str):
                partes = fecha_str.split()
                partes[2] = "20" + partes[2]
                fecha_str = " ".join(partes)

            # Eliminar palabra "de" para evitar errores de parser
            fecha_str = re.sub(r"\bde\b", "", fecha_str).strip()

            try:
                fecha = parser.parse(fecha_str, dayfirst=True)
                print(f"✅ Fecha parseada: {fecha}")
                return fecha.replace(day=1)
            except Exception as e:
                print(f"❌ Fallo al parsear '{fecha_str}': {e}")

    print("❌ No se encontró fecha válida, usando la fecha actual como fallback.")
    return date.today().replace(day=1)


# En tu archivo utils.py o en flask_app.py (donde tengas helpers)
def mes_castellano(fecha):
    meses = [
        'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
        'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre'
    ]
    return f"{meses[fecha.month - 1]} {fecha.year}"
app.jinja_env.filters['mes_castellano'] = mes_castellano


from PyPDF2 import PdfReader

def extract_text_from_pdf(ruta_pdf):
    try:
        lector = PdfReader(ruta_pdf)
        texto_total = ""
        for pagina in lector.pages:
            texto_pagina = pagina.extract_text()
            if texto_pagina:
                texto_total += texto_pagina + "\n"
        return texto_total
    except Exception as e:
        print(f"❌ Error al extraer texto del PDF {ruta_pdf}: {e}")
        return ""



from flask import render_template, request
from sqlalchemy import extract
from math import ceil
from sqlalchemy import func

@app.route('/nominas', methods=['GET'])
def listar_nominas():
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id, leida=False).order_by(Notificacion.fecha_hora.desc()).all()

    page = request.args.get('page', 1, type=int)

    try:
        user_id = int(request.args.get('user_id')) if request.args.get('user_id') else None
    except ValueError:
        user_id = None

    try:
        mes = int(request.args.get('mes')) if request.args.get('mes') else None
    except ValueError:
        mes = None

    try:
        anio = int(request.args.get('anio')) if request.args.get('anio') else None
    except ValueError:
        anio = None

    print("🔍 Parámetros recibidos:")
    print(f"📄 Página: {page}, 👤 User ID: {user_id}, 🗓️ Mes: {mes}, 📅 Año: {anio}")

    # Consulta nóminas
    nominas = Nomina.query
    if user_id:
        nominas = nominas.filter_by(user_id=user_id)
        print(f"➡️ Filtrado por user_id: {user_id}")
    if mes:
        nominas = nominas.filter(extract('month', Nomina.fecha) == mes)
        print(f"➡️ Filtrado por mes: {mes}")
    if anio:
        nominas = nominas.filter(extract('year', Nomina.fecha) == anio)
        print(f"➡️ Filtrado por año: {anio}")

    nominas = nominas.all()
    print(f"📁 Nóminas encontradas: {len(nominas)}")
    for n in nominas:
        n.tipo = 'nomina'  # en minúscula para coincidir con la plantilla
        n.ruta_archivo = n.ruta
        n.user = User.query.get(n.user_id)  # asignar usuario
        print(f"📝 Nómina: ID={n.id}, Fecha={n.fecha}, Ruta={n.ruta}, Usuario={n.user.nombre if n.user else 'Desconocido'}")

    # Consulta contratos
    contratos = Contrato.query
    if user_id:
        contratos = contratos.filter_by(user_id=user_id)
        print(f"➡️ Contratos filtrados por user_id: {user_id}")
    if mes:
        contratos = contratos.filter(extract('month', Contrato.fecha) == mes)
        print(f"➡️ Contratos filtrados por mes: {mes}")
    if anio:
        contratos = contratos.filter(extract('year', Contrato.fecha) == anio)
        print(f"➡️ Contratos filtrados por año: {anio}")

    contratos = contratos.all()
    print(f"📁 Contratos encontrados: {len(contratos)}")
    for c in contratos:
        c.tipo = 'contrato'  # en minúscula también
        c.ruta_archivo = c.archivo
        c.user = User.query.get(c.user_id)
        print(f"📄 Contrato: ID={c.id}, Fecha={c.fecha}, Archivo={c.archivo}, Usuario={c.user.nombre if c.user else 'N/A'}")

    # Combinar y ordenar
    documentos = nominas + contratos
    documentos.sort(key=lambda d: d.fecha, reverse=True)
    print(f"📑 Total documentos combinados: {len(documentos)}")

    per_page = 10
    total = len(documentos)
    total_pages = ceil(total / per_page)
    documentos_pagina = documentos[(page - 1) * per_page : page * per_page]
    print(f"📃 Mostrando página {page} con {len(documentos_pagina)} documentos (Total páginas: {total_pages})")

    usuarios = User.query.order_by(User.nombre.asc()).all()
    print(f"👥 Total usuarios cargados: {len(usuarios)}")

    return render_template('admin_nominas.html',
                           notificaciones=notificaciones,
                           nominas=documentos_pagina,
                           usuarios=usuarios,
                           page=page,
                           total_pages=total_pages,
                           user_id=user_id,
                           mes=mes,
                           anio=anio,
                           int=int)


@app.route('/descargar_nomina/<int:nomina_id>')
def descargar_nomina(nomina_id):
    # Buscar la nómina por ID
    nomina = Nomina.query.get_or_404(nomina_id)

    # Ruta del archivo
    ruta_archivo = nomina.ruta

    # Verificar si el archivo existe
    if not os.path.exists(ruta_archivo):
        abort(404, description="Archivo no encontrado")

    # Enviar el archivo al navegador para descargar
    return send_file(ruta_archivo, as_attachment=True)


@app.route('/descargar_contrato/<int:contrato_id>')
def descargar_contrato(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)
    ruta_archivo = contrato.archivo

    if not os.path.exists(ruta_archivo):
        abort(404, description="Archivo no encontrado")

    return send_file(ruta_archivo, as_attachment=True)



UPLOAD_FOLDER_contratos = 'static/uploads/contratos'


@app.route('/subir_contratos', methods=['GET'])
def subir_contratos():
    print("Mostrando formulario de subida de contratos")
    return render_template('subir_contrato.html')


@app.route('/admin/subir_contrato', methods=['POST'])
def subir_contrato():
    print("Procesando subida de contratos...")
    archivos = request.files.getlist('archivo')

    if not archivos:
        flash("No se ha seleccionado ningún archivo.", "danger")
        return redirect(request.url)

    fecha_contrato = date.today()
    os.makedirs(UPLOAD_FOLDER_contratos, exist_ok=True)
    print(f"Carpeta de destino: {UPLOAD_FOLDER_contratos}")

    for archivo in archivos:
        print(f"Procesando archivo: {archivo.filename}")
        if archivo and allowed_file(archivo.filename):
            filename = secure_filename(archivo.filename)
            ext = filename.rsplit('.', 1)[1].lower()
            print(f"Extensión detectada: {ext}")

            if ext == 'pdf':
                print("Archivo es PDF, procesando individualmente.")
                ruta_temp = os.path.join(UPLOAD_FOLDER_contratos, "temp_extract.pdf")
                archivo.save(ruta_temp)

                dni = procesar_contrato_pdf(ruta_temp)
                if not dni:
                    flash(f"No se pudo extraer el DNI del archivo: {filename}", "warning")
                    continue

                print(f"Buscando usuario con DNI: {dni}")
                user = User.query.filter_by(DNI=dni).first()
                if not user:
                    flash(f"Usuario no encontrado para DNI: {dni}", "warning")
                    continue

                existente = Contrato.query.filter_by(user_id=user.id).first()
                if existente:
                    flash(f"Contrato ya existente para {dni}", "warning")
                    continue

                nuevo_nombre = f"contrato_{dni}_{fecha_contrato.strftime('%Y%m%d')}.pdf"
                ruta_archivo = os.path.join(UPLOAD_FOLDER_contratos, nuevo_nombre)
                os.rename(ruta_temp, ruta_archivo)

                print(f"Contrato guardado en: {ruta_archivo}")
                contrato = Contrato(user_id=user.id, archivo=ruta_archivo, fecha=fecha_contrato)
                db.session.add(contrato)

            elif ext == 'zip':
                print("Archivo es ZIP, extrayendo contenido...")
                with zipfile.ZipFile(archivo) as zip_ref:
                    for nombre_pdf in zip_ref.namelist():
                        print(f"Archivo dentro del ZIP: {nombre_pdf}")
                        if nombre_pdf.endswith('.pdf'):
                            ruta_temp = os.path.join(UPLOAD_FOLDER_contratos, "temp_extract.pdf")
                            with open(ruta_temp, 'wb') as f:
                                f.write(zip_ref.read(nombre_pdf))

                            dni = procesar_contrato_pdf(ruta_temp)
                            if not dni:
                                flash(f"No se pudo extraer el DNI de: {nombre_pdf}", "warning")
                                continue

                            print(f"Buscando usuario con DNI: {dni}")
                            user = User.query.filter_by(DNI=dni).first()
                            if not user:
                                flash(f"Usuario no encontrado para DNI: {dni}", "warning")
                                continue

                            existente = Contrato.query.filter_by(user_id=user.id).first()
                            if existente:
                                flash(f"Contrato ya existente para {dni}", "warning")
                                continue

                            nuevo_nombre = f"contrato_{dni}_{fecha_contrato.strftime('%Y%m%d')}.pdf"
                            ruta_archivo = os.path.join(UPLOAD_FOLDER_contratos, nuevo_nombre)
                            os.rename(ruta_temp, ruta_archivo)

                            print(f"Contrato guardado en: {ruta_archivo}")
                            contrato = Contrato(user_id=user.id, archivo=ruta_archivo, fecha=fecha_contrato)
                            db.session.add(contrato)
                        else:
                            print(f"Ignorado (no es PDF): {nombre_pdf}")
        else:
            print(f"Archivo no permitido o vacío: {archivo.filename}")

    db.session.commit()
    print("Todos los cambios guardados en la base de datos.")
    flash("Contratos procesados correctamente.", "success")
    return redirect(url_for('subir_contratos'))


def procesar_contrato_pdf(ruta_pdf):
    try:
        print(f"Procesando contrato desde el archivo: {ruta_pdf}")
        with fitz.open(ruta_pdf) as doc:
            for i, page in enumerate(doc):
                text = page.get_text()
                clean = clean_text(text)
                print(f"Texto limpio extraído de página {i + 1}:\n{clean}\n{'-'*40}")

                # Buscar DNIs con y sin espacios
                coincidencias = re.findall(r'\b(\d{7,8})\s*([A-Z])\b|\b([XYZ])(\d{7})\s*([A-Z])\b', clean.upper())
                posibles_dnis = []

                for match in coincidencias:
                    if match[0]:
                        posibles_dnis.append(f"{match[0]}{match[1]}")
                    elif match[2]:
                        posibles_dnis.append(f"{match[2]}{match[3]}{match[4]}")

                print(f"Posibles DNIs encontrados: {posibles_dnis}")

                dnis_personales = [
                    dni for dni in posibles_dnis
                    if not re.match(r'^[ABCDEFGHJNPQRSUVW]', dni)
                ]
                print(f"DNIs personales filtrados: {dnis_personales}")

                # en lugar de guardar directamente, solo devuelve el DNI
                if dnis_personales:
                    if len(dnis_personales) >= 2:
                        return dnis_personales[1]
                    else:
                        return dnis_personales[0]

                    print(f"DNI válido seleccionado: {dni}")



        print("No se encontró un DNI válido en el documento.")
    except Exception as e:
        print(f"Error al procesar el contrato {ruta_pdf}: {e}")

@app.route('/admin/contratos')
def ver_contratos():
    usuario_id = request.args.get('usuario')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    query = Contrato.query

    if usuario_id:
        query = query.filter_by(usuario_id=usuario_id)
    if fecha_inicio and fecha_fin:
        query = query.filter(Contrato.fecha_creacion.between(fecha_inicio, fecha_fin))

    contratos = query.order_by(Contrato.fecha_creacion.desc()).all()
    usuarios = User.query.all()

    return render_template('ver_contratos.html', contratos=contratos, usuarios=usuarios)



from flask import jsonify

@app.route('/centros_trabajo/<int:user_id>', methods=['GET'])
def obtener_centros_trabajo(user_id):
    # Obtener el usuario y la empresa relacionada
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'Usuario no encontrado'}), 404

    empresa_id = user.empresa_id  # Suponiendo que el usuario tiene un campo empresa_id
    centros_trabajo = CentroTrabajo.query.filter_by(empresa_id=empresa_id).all()

    centros = [{'id': centro.id, 'nombre': centro.nombre} for centro in centros_trabajo]

    return jsonify({'centros': centros})

import qrcode
import os
from io import BytesIO
from flask import send_file

@app.route('/generar_qr', methods=['POST'])
def generar_qr():
    try:
        print("\n🔹 Iniciando generación de QR 🔹")
        centro_id = request.form.get('centro_id')
        centro_nombre = request.form.get('centro_nombre')
        user_id = request.form.get('user_id')

        print(f"[DEBUG] Datos recibidos -> Centro ID: {centro_id}, Centro Nombre: {centro_nombre}, User ID: {user_id}")

        # Usuario
        user = User.query.get(user_id)
        if not user:
            print(f"[DEBUG] Usuario con ID {user_id} no encontrado.")
            return jsonify({'message': 'Usuario no encontrado'}), 404
        print(f"[DEBUG] Usuario encontrado -> {user.nombre} {user.primer_apellido}")

        # Centro de trabajo
        centro_trabajo = CentroTrabajo.query.filter_by(nombre=centro_nombre).first()
        if not centro_trabajo:
            print(f"[DEBUG] Centro de trabajo '{centro_nombre}' no encontrado.")
            return jsonify({'message': 'Centro de trabajo no encontrado'}), 404
        print(f"[DEBUG] Centro de trabajo encontrado -> {centro_trabajo.nombre}")

        # Directorio QR
        qr_dir_absoluto = '/home/adaptasystem/gestiona_empresa/herdelsalud/static/qr'
        os.makedirs(qr_dir_absoluto, exist_ok=True)
        print(f"[DEBUG] Directorio de QR asegurado -> {qr_dir_absoluto}")

        # Verificar QR existente
        qr_existente = QR.query.filter_by(user_id=user.id, centro_trabajo_id=centro_trabajo.id).first()
        if qr_existente:
            print("[DEBUG] QR ya existe, devolviendo info existente.")
            qr_url_relativa = f"/static/qr/{os.path.basename(qr_existente.imagen)}"
            print(f"[DEBUG] QR URL relativa -> {qr_url_relativa}")
            return jsonify({
                'message': 'QR ya existía',
                'qr_url': qr_url_relativa,
                'nombre': user.nombre,
                'apellido1': user.primer_apellido,
                'empresa': centro_trabajo.empresa.nombre,
                'centro_trabajo': centro_trabajo.nombre,
                'ya_existe': True
            })

        # Generar QR nuevo
        base_url = request.host_url.rstrip('/')
        qr_data = f"{base_url}/registro_qr?centro_id={centro_id}&user_id={user_id}"
        print(f"[DEBUG] Datos para QR -> {qr_data}")

        qr_filename = f"qr_{centro_id}_{user_id}.png"
        qr_path_absoluto = os.path.join(qr_dir_absoluto, qr_filename)
        img = qrcode.make(qr_data)
        img.save(qr_path_absoluto)
        print(f"[DEBUG] QR generado y guardado -> {qr_path_absoluto}")

        # Guardar en DB
        nuevo_qr = QR(user_id=user.id, centro_trabajo=centro_trabajo, imagen=qr_path_absoluto)
        db.session.add(nuevo_qr)
        db.session.commit()
        print("[DEBUG] Nuevo QR guardado en la base de datos.")

        qr_url_relativa = f"/static/qr/{qr_filename}"

        print("[DEBUG] QR generado correctamente, devolviendo JSON.\n")
        return jsonify({
            'message': 'QR generado correctamente',
            'qr_url': qr_url_relativa,
            'nombre': user.nombre,
            'apellido1': user.primer_apellido,
            'empresa': centro_trabajo.empresa.nombre,
            'centro_trabajo': centro_trabajo.nombre
        })

    except Exception as e:
        db.session.rollback()
        print("[ERROR] en generar_qr:", str(e))
        return jsonify({'message': f'Error interno: {str(e)}'}), 500


@app.route('/api/get_all_qr', methods=['GET'])
def get_all_qr():
    try:
        qr_records = QR.query.all()

        qr_list = []
        for qr in qr_records:
            user = User.query.get(qr.user_id)
            if not user:
                continue

            centro = CentroTrabajo.query.get(qr.centro_trabajo_id)
            empresa_nombre = centro.empresa.nombre if centro and centro.empresa else ''

            # Corregir la URL del QR
            qr_rel_path = qr.imagen
            if '/static/' in qr_rel_path:
                qr_url_relativa = qr_rel_path.split('/static/')[-1]
            else:
                qr_url_relativa = os.path.basename(qr_rel_path)

            qr_list.append({
                'qr_url': '/static/' + qr_url_relativa,
                'nombre': user.nombre,
                'apellido1': user.primer_apellido,
                'empresa': empresa_nombre,
                'centro_trabajo': centro.nombre if centro else ''
            })

        return jsonify({'qr_list': qr_list})

    except Exception as e:
        print("ERROR en get_all_qr:", str(e))
        return jsonify({'message': str(e), 'qr_list': []}), 500


@app.route('/registro_resultado')
def registro_resultado():
    user_id = request.args.get('user_id')
    centro_id = request.args.get('centro_id')
    tipo = request.args.get('tipo')  # 'entrada' o 'salida'
    duracion = request.args.get('duracion')

    usuario = User.query.get(user_id)
    centro = CentroTrabajo.query.get(centro_id)

    return render_template('registro_resultado.html', usuario=usuario, centro=centro, tipo=tipo, duracion=duracion)


import secrets

@app.route('/solicitar-cambio-horario', methods=['POST'])
def solicitar_cambio_horario():
    data = request.get_json()
    registro_id = data['registro_id']
    tipo = data['tipo']
    nueva_fecha_hora = datetime.strptime(data['nueva_fecha_hora'], "%Y-%m-%dT%H:%M")
    motivo = data['motivo']

    token = secrets.token_urlsafe(32)

    solicitud = SolicitudCambioHorario(
        registro_id=registro_id,
        tipo=tipo,
        nueva_fecha_hora=nueva_fecha_hora,
        motivo=motivo,
        token=token
    )
    db.session.add(solicitud)
    db.session.commit()

    # Enviar correo a admins
    admins = get_admin_users()
    for admin in admins:
        send_email_cambio_horario(admin.email, solicitud)

    return jsonify({"status": "ok"})


def send_email_cambio_horario(email, solicitud):
    subject = "Solicitud de correción de registro"
    registro = solicitud.registro
    usuario = registro.user
    nombre_completo = f"{usuario.nombre} {usuario.primer_apellido}"

    if solicitud.tipo == 'entrada':
        fecha_original = registro.entrada
        otra_fecha = registro.salida
        otro_tipo = "Salida"
    else:
        fecha_original = registro.fecha_salida
        otra_fecha = registro.fecha_entrada
        otro_tipo = "Entrada"

    body = f"""
    Se ha solicitado modificar la <b>{solicitud.tipo}</b> del registro <b>#{solicitud.registro_id}</b>.<br><br>
    <b>Empleado:</b> {nombre_completo}<br>
    <b>Fecha y hora actual de {solicitud.tipo}:</b> {fecha_original.strftime('%d-%m-%Y %H:%M')}<br>
    <b>{otro_tipo} registrada:</b> {otra_fecha.strftime('%d-%m-%Y %H:%M') if otra_fecha else 'No registrada'}<br><br>
    <b>Motivo:</b> {solicitud.motivo}<br>
    <b>Nueva fecha y hora propuesta:</b> {solicitud.nueva_fecha_hora.strftime('%d-%m-%Y %H:%M')}<br><br>
    ¿Deseas aprobar o rechazar esta solicitud?
    """


    approve_url = url_for("aprobar_cambio_horario", token=solicitud.token, _external=True)
    reject_url = url_for("rechazar_cambio_horario", token=solicitud.token, _external=True)

    html_body = render_template_string(
        """<html>
            <body>
                <p>{{ body|safe }}</p>
                <a href="{{ approve_url }}" style="color: white; background: green; padding: 8px 12px; text-decoration: none;">✅ Aprobar</a>
                <a href="{{ reject_url }}" style="color: white; background: red; padding: 8px 12px; text-decoration: none; margin-left:10px;">❌ Rechazar</a>
            </body>
        </html>""",
        body=body,
        approve_url=approve_url,
        reject_url=reject_url
    )

    msg = Message(subject, recipients=[email], html=html_body)
    mail.send(msg)


@app.route('/aprobar-cambio-horario/<token>')
def aprobar_cambio_horario(token):
    print(f"🔍 Buscando solicitud con token: {token}")
    solicitud = SolicitudCambioHorario.query.filter_by(token=token).first_or_404()

    if solicitud.estado == 'aprobada':
        print("ℹ️ La solicitud ya fue aprobada previamente.")
        return "⚠️ Esta solicitud ya ha sido aprobada anteriormente."

    if solicitud.estado != 'pendiente':
        print(f"⚠️ Estado no válido para aprobación: {solicitud.estado}")
        return f"❌ No se puede aprobar esta solicitud porque su estado es '{solicitud.estado}'."

    print(f"✅ Solicitud encontrada: ID {solicitud.id}, tipo {solicitud.tipo}, nueva fecha: {solicitud.nueva_fecha_hora}")

    registro = solicitud.registro
    print(f"📋 Registro obtenido: ID {registro.id}, entrada: {registro.entrada}, salida: {registro.salida}")

    usuario = registro.user
    print(f"👤 Usuario del registro: {usuario.nombre} {usuario.primer_apellido} - {usuario.email}")

    if solicitud.tipo == 'entrada':
        print(f"✏️ Actualizando entrada: {registro.entrada} -> {solicitud.nueva_fecha_hora}")
        registro.entrada = solicitud.nueva_fecha_hora
    else:
        print(f"✏️ Actualizando salida: {registro.salida} -> {solicitud.nueva_fecha_hora}")
        registro.salida = solicitud.nueva_fecha_hora

    solicitud.estado = 'aprobada'
    db.session.commit()
    print("💾 Cambios guardados en la base de datos.")

    # Notificar por correo
    if usuario and usuario.email:
        try:
            print(f"📧 Enviando correo a: {usuario.email}")
            msg = Message(subject="✅ Cambio de horario aprobado",
                          recipients=[usuario.email])
            msg.html = f"""
                Hola {usuario.nombre} {usuario.primer_apellido},<br><br>
                Tu solicitud de modificación de <b>{solicitud.tipo}</b> ha sido <b>aprobada</b>.<br>
                Nueva fecha y hora: <b>{solicitud.nueva_fecha_hora.strftime('%d-%m-%Y %H:%M')}</b><br><br>
                Gracias por tu paciencia.<br>
                — El equipo de Administración
            """
            mail.send(msg)
            print("✅ Correo enviado correctamente.")
        except Exception as e:
            print(f"❌ Error al enviar el correo: {e}")
            app.logger.error(f"Error al enviar notificación al usuario: {e}")
    else:
        print("⚠️ No se pudo enviar correo: usuario o email no disponible.")

    return "✅ Solicitud aprobada y horario actualizado."


@app.route('/rechazar-cambio-horario/<token>')
def rechazar_cambio_horario(token):
    solicitud = SolicitudCambioHorario.query.filter_by(token=token, estado='pendiente').first_or_404()
    solicitud.estado = 'rechazada'
    db.session.commit()
    return "❌ Solicitud rechazada correctamente."
  # ajusta el import según tu estructura

def actualizar_fechas_verano():
    today = datetime.today()
    current_year = today.year

    empresas = Empresa.query.all()

    for empresa in empresas:
        cambiado = False

        if empresa.fecha_inicio_verano:
            if empresa.fecha_inicio_verano.year != current_year:
                # mantener mes y día
                nueva_fecha_inicio = datetime(
                    year=current_year,
                    month=empresa.fecha_inicio_verano.month,
                    day=empresa.fecha_inicio_verano.day
                )
                print(f"[DEBUG] Actualizando fecha_inicio_verano de {empresa.nombre} a {nueva_fecha_inicio}")
                empresa.fecha_inicio_verano = nueva_fecha_inicio
                cambiado = True

        if empresa.fecha_fin_verano:
            if empresa.fecha_fin_verano.year != current_year:
                nueva_fecha_fin = datetime(
                    year=current_year,
                    month=empresa.fecha_fin_verano.month,
                    day=empresa.fecha_fin_verano.day
                )
                print(f"[DEBUG] Actualizando fecha_fin_verano de {empresa.nombre} a {nueva_fecha_fin}")
                empresa.fecha_fin_verano = nueva_fecha_fin
                cambiado = True

        if cambiado:
            db.session.add(empresa)

    db.session.commit()
    print("[INFO] Fechas de verano actualizadas correctamente.")

def actualizar_fechas_verano():
    today = datetime.today()
    current_year = today.year

    empresas = Empresa.query.all()

    for empresa in empresas:
        cambiado = False

        if empresa.fecha_inicio_verano:
            if empresa.fecha_inicio_verano.year != current_year:
                # mantener mes y día
                nueva_fecha_inicio = datetime(
                    year=current_year,
                    month=empresa.fecha_inicio_verano.month,
                    day=empresa.fecha_inicio_verano.day
                )
                print(f"[DEBUG] Actualizando fecha_inicio_verano de {empresa.nombre} a {nueva_fecha_inicio}")
                empresa.fecha_inicio_verano = nueva_fecha_inicio
                cambiado = True

        if empresa.fecha_fin_verano:
            if empresa.fecha_fin_verano.year != current_year:
                nueva_fecha_fin = datetime(
                    year=current_year,
                    month=empresa.fecha_fin_verano.month,
                    day=empresa.fecha_fin_verano.day
                )
                print(f"[DEBUG] Actualizando fecha_fin_verano de {empresa.nombre} a {nueva_fecha_fin}")
                empresa.fecha_fin_verano = nueva_fecha_fin
                cambiado = True

        if cambiado:
            db.session.add(empresa)

    db.session.commit()
    print("[INFO] Fechas de verano actualizadas correctamente.")


#Nueva ruta de vista de administración de usuarios
@app.route('/admin/gestion_usuarios', methods=['GET', 'POST'])
@login_required
def gestion_usuarios():
    if not current_user.has_role('admin'):
        return "Acceso denegado", 403

    if request.method == 'POST':
        user_id = request.form.get('user_id')
        action = request.form.get('action')

        user = User.query.get(user_id)
        if not user:
            flash("Usuario no encontrado", "error")
            return redirect(url_for('gestion_usuarios'))

        if action == 'bloquear':
            user.bloqueado = True
            user.bloqueado_hasta = None  # Bloqueo manual = permanente
            flash(f"Usuario {user.email} bloqueado permanentemente", "warning")

        elif action == 'desbloquear':
            user.desbloquear()
            flash(f"Usuario {user.email} desbloqueado", "success")

        db.session.commit()
        return redirect(url_for('gestion_usuarios'))

    usuarios = User.query.filter(User.email != current_user.email).all()
    return render_template('gestion_usuarios.html', usuarios=usuarios)


@app.route('/carta')
def carta():

    return render_template('carta.html')


@app.route('/toggle_baja', methods=['POST'])
@admin_required  # o según tus permisos
def toggle_baja():
    user_id = request.json.get('user_id')
    activar = request.json.get('activar')  # True o False
    user = User.query.get(user_id)

    if not user:
        return jsonify({'error': 'Usuario no encontrado'}), 404

    if activar:
        # Crear nueva baja con fecha_inicio ahora y fecha_fin None
        nueva_baja = Baja(
            user_id=user.id,
            fecha_inicio=datetime.now(timezone_spain).date(),
            fecha_fin=None
        )
        db.session.add(nueva_baja)
    else:
        # Cerrar la baja activa (la última sin fecha_fin)
        baja_activa = Baja.query.filter_by(user_id=user.id, fecha_fin=None).first()
        if baja_activa:
            baja_activa.fecha_fin = datetime.now(timezone_spain).date()
    db.session.commit()
    return jsonify({'success': True})


# Ejecutar desde el contexto de Flask (flask shell o script con app.app_context())
def looks_like_hash(pwd_str):
    """Detecta si una cadena parece ya ser un hash conocido."""
    if not pwd_str or not isinstance(pwd_str, str):
        return False
    prefixes = ('pbkdf2:', 'scrypt:', 'argon2', '$2a$', '$2b$', '$2y$')  # bcrypt variants included
    return pwd_str.startswith(prefixes)

def hash_passwords_from_id(start_id=0, default_password_if_empty=None, batch_size=100):
    """
    Hashea las contraseñas de todos los usuarios con id >= start_id.
    - Si user.password ya parece un hash (ej. empieza por 'pbkdf2:' o 'scrypt:'), se salta.
    - Si user.password está vacío y default_password_if_empty es None: se salta.
    - Si user.password está vacío y default_password_if_empty está definido: se asigna ese password y se hashea.
    """
    total_processed = 0
    total_hashed = 0
    total_skipped = 0
    try:
        with app.app_context():
            query = User.query.filter(User.id >= start_id).order_by(User.id)
            batch = []
            for user in query.yield_per(batch_size):
                total_processed += 1
                raw_pwd = user.password or ''
                # Si ya parece hash, lo saltamos
                if looks_like_hash(raw_pwd):
                    print(f"SKIP (already hashed) -> id={user.id} email={getattr(user,'email',None)} pwd_preview={raw_pwd[:30]}")
                    total_skipped += 1
                    continue

                # Si no hay contraseña
                if not raw_pwd:
                    if default_password_if_empty is None:
                        print(f"SKIP (empty password) -> id={user.id} email={getattr(user,'email',None)}")
                        total_skipped += 1
                        continue
                    else:
                        pwd_to_hash = default_password_if_empty
                        print(f"NOTICE: user id={user.id} had empty password. Using default before hashing.")
                else:
                    pwd_to_hash = raw_pwd

                # Generar hash y asignar
                hashed = generate_password_hash(pwd_to_hash)
                user.password = hashed
                batch.append(user)
                total_hashed += 1
                print(f"HASHED -> id={user.id} email={getattr(user,'email',None)}")

                # Commit por lotes
                if len(batch) >= batch_size:
                    db.session.commit()
                    print(f"COMMIT batch of {len(batch)} users (up to id={user.id})")
                    batch = []

            # Commit final si queda
            if batch:
                db.session.commit()
                print(f"FINAL COMMIT batch of {len(batch)} users")

        print("DONE.")
        print(f"Summary: processed={total_processed}, hashed={total_hashed}, skipped={total_skipped}")
        return {
            'processed': total_processed,
            'hashed': total_hashed,
            'skipped': total_skipped
        }

    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        print("ERROR durante hashing:", str(e))
        raise

# --- EJEMPLOS DE USO ---
# 1) Desde flask shell:
# >>> from path.to.hash_users import hash_passwords_from_id
# >>> hash_passwords_from_id(start_id=4)
#
# 2) Si quieres asignar una contraseña por defecto a los vacíos (ej. 'aDmin@2024%'):
# >>> hash_passwords_from_id(start_id=4, default_password_if_empty='aDmin@2024%')
#
# 3) Para ejecutarlo desde un script externo, envuelve la llamada con app.app_context():
# with app.app_context():
#     hash_passwords_from_id(start_id=4, default_password_if_empty=None)

@app.route('/registro/validar/<int:registro_id>', methods=['POST'])
def validar_registro(registro_id):
    """
    Ruta para aprobar o rechazar una modificación de registro.
    Recibe un JSON con { "aceptar": true/false }
    """
    try:
        print("📩 [DEBUG] Petición recibida en /registro/validar/")
        print(f"Registro ID recibido: {registro_id}")

        data = request.get_json()
        print(f"Datos recibidos: {data}")

        aceptar = data.get('aceptar')
        print(f"Valor de 'aceptar': {aceptar}")

        registro = Registro.query.get_or_404(registro_id)
        print(f"Registro encontrado: ID={registro.id}, "
              f"modificado={registro.modificado}, "
              f"aprobado_modificacion={registro.aprobado_modificacion}")

        # Actualizamos el estado según lo recibido
        registro.aprobado_modificacion = True if aceptar else False
        registro.modificado = "administrador"  # opcional, si quieres marcar quién validó
        print(f"Nuevo valor de aprobado_modificacion: {registro.aprobado_modificacion}")

        db.session.commit()
        print("💾 [DEBUG] Cambios guardados correctamente en la base de datos.")

        return jsonify({
            "success": True,
            "registro_id": registro.id,
            "validado": registro.aprobado_modificacion
        })

    except Exception as e:
        print("❌ [ERROR] Fallo en la validación del registro:")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

def agregar_festivos2026():
    empresas = Empresa.query.all()

    for festivo in non_workdays_2026:
        año = festivo.year

        for empresa in empresas:
            # Comprobar si el festivo ya existe por día, mes, año y empresa
            existe_festivo = Festivo.query.filter_by(
                dia=festivo.day,
                mes=festivo.month,
                año=año,
                empresa_id=empresa.id
            ).first()

            if not existe_festivo:
                nuevo_festivo = Festivo(
                    dia=festivo.day,
                    mes=festivo.month,
                    año=año,
                    empresa_id=empresa.id
                )
                db.session.add(nuevo_festivo)

    db.session.commit()

if __name__ == '__main__':
    with app.app_context():
        #db.drop_all()
        #users = User.query.all()
        #for u in users:
         #   u.tipo_jornada='verano'
        db.create_all()
        '''
        registro = Registro(
                        user_id=2,
                        entrada=datetime(2024, 11, 21, 8, 30),
                        salida=datetime(2024, 11, 21, 18, 30),
                        ip_address_entrada=1

                    )
        db.session.add(registro)
        db.session.commit()
            '''
        #sql_command = text('DROP TABLE IF EXISTS registro;')
        #db.session.execute(sql_command)
        #usuario = User.query.filter_by(email='alfonsomunoz@adaptasystem.com').first()

        #usuario.dias_restantes_este = 20

        from werkzeug.security import generate_password_hash

        # Crear roles
        admin_role = user_datastore.find_or_create_role(name='admin', description='Administrator')
        inspector_role = user_datastore.find_or_create_role(name='inspector', description='Inspector')
        user_role = user_datastore.find_or_create_role(name='user', description='Regular User')
        gestor_role = user_datastore.find_or_create_role(name='gestor', description='Gestor de empresa')
        encargado_role = user_datastore.find_or_create_role(name='encargado', description='Encargado con funciones limitadas')

        # Crear el usuario admin si no existe
        admin_user = user_datastore.find_user(email='admin@gestionaempresa.es')
        if not admin_user:
            hashed_password = generate_password_hash('aDmin@2024%')
            admin_user = user_datastore.create_user(email='admin@gestionaempresa.es', password=hashed_password, approved=True)
            user_datastore.add_role_to_user(admin_user, admin_role)

        # Crear el usuario inspector si no existe
        inspector_user = user_datastore.find_user(email='inspector@gestionaempresa.es')
        if not inspector_user:
            hashed_password2 = generate_password_hash('Inspector@2025%')
            inspector_user = user_datastore.create_user(email='inspector@gestionaempresa.es', password=hashed_password2, approved=True)
            user_datastore.add_role_to_user(inspector_user, inspector_role)
        '''
        # Crear el usuario gestor si no existe
        gestor_user = user_datastore.find_user(email='gestor@gestionaempresa.es')
        if not gestor_user:
            hashed_password3 = generate_password_hash('Gestor@2025%')
            gestor_user = user_datastore.create_user(email='gestor@gestionaempresa.es', password=hashed_password3, approved=True)
            user_datastore.add_role_to_user(gestor_user, gestor_role)

        # Crear el usuario encargado si no existe
        encargado_user = user_datastore.find_user(email='encargado@gestionaempresa.es')
        if not encargado_user:
            hashed_password4 = generate_password_hash('Encargado@2025%')
            encargado_user = user_datastore.create_user(email='encargado@gestionaempresa.es', password=hashed_password4, approved=True)
            user_datastore.add_role_to_user(encargado_user, encargado_role)
        '''








        #set_exception(2)
        #set_work_schedule(2)
        #eliminar_todos_los_festivos()
        #agregar_festivos()
        #agregar_festivos2026()
        db.session.commit()

        '''
        if admin_user:
            hashed_password = generate_password_hash('aDmin@2024%')
            admin_user.password=hashed_password
            db.session.commit()
        '''

        #hash_passwords_from_id(start_id=0, default_password_if_empty=None)


        app.run(debug=True, port=5039)